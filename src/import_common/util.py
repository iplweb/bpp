from __future__ import annotations

import re
from collections import defaultdict
from collections.abc import Generator
from typing import TYPE_CHECKING

from django.utils.functional import cached_property

from .exceptions import (
    BadNoOfSheetsException,
    DecompressionBombException,
    HeaderNotFoundException,
    ImproperFileException,
)

# Limit rozmiaru pliku XLSX PO dekompresji. Realne pliki importowe (nawet
# kilkanaście MB XLSX) rozpakowują się do najwyżej dziesiątek/setek MB XML —
# 500 MB daje spory zapas, a bomba zip (KB → GB) go przekracza i jest
# odrzucana przed załadowaniem do pamięci (obrona przed OOM workera importu).
MAX_ROZMIAR_PO_DEKOMPRESJI = 500 * 1024 * 1024


def sprawdz_bombe_dekompresji(sciezka, max_rozpakowany=MAX_ROZMIAR_PO_DEKOMPRESJI):
    """Odrzuca XLSX-owe (ZIP) bomby dekompresyjne PRZED załadowaniem do pamięci.

    XLSX to archiwum ZIP; złośliwy plik ~KB może rozpakować się do GB i zabić
    workera importu (OOM). Sumujemy deklarowane rozmiary po dekompresji z
    centralnego katalogu ZIP (bez faktycznego rozpakowywania) i odrzucamy plik
    przekraczający ``max_rozpakowany``. Pliki nie-ZIP (stary .xls OLE, .csv)
    nie są bombami tego typu → cicho przepuszczamy (inny wektor).
    """
    import zipfile

    try:
        with zipfile.ZipFile(sciezka) as zf:
            rozpakowany = sum(info.file_size for info in zf.infolist())
    except zipfile.BadZipFile:
        return  # nie-ZIP (np. .xls / .csv) — nie ten wektor

    if rozpakowany > max_rozpakowany:
        raise DecompressionBombException(
            f"Rozmiar pliku po dekompresji ({rozpakowany} B) przekracza "
            f"bezpieczny limit ({max_rozpakowany} B) — plik odrzucony jako "
            f"potencjalna bomba dekompresyjna."
        )


# openpyxl importujemy LOKALNIE (nie na poziomie modułu): ten moduł jest
# importowany eager przez modele importerów (np. import_list_if.models →
# XLSImportFile) już przy django.setup(), a openpyxl przez compat/numbers.py
# ciągnie całe numpy
# (~tens MB RSS) do KAŻDEGO procesu — także web/ASGI, który nigdy nie czyta
# xlsx. Dzięki PEP 563 (future annotations) adnotacje typu ``openpyxl.*``
# w sygnaturach poniżej są tylko łańcuchami i nie wymagają openpyxl w
# runtime; pod TYPE_CHECKING dajemy je type-checkerom/ruff. Plik xlsx
# otwierają tylko funkcje z lokalnym importem.
if TYPE_CHECKING:
    import openpyxl

DEFAULT_COL_NAMES = [
    "imię",
    "imie",
    "imiona",
    "nazwisko",
    "nazwiska",
    "orcid",
    "pesel",
    "pbn_id",
    "stanowisko",
    "wydział",
    "jednostka",
    "numer",
]

DEFAULT_MIN_POINTS = 3


def normalize_cell_header(value):
    """Normalizuje SUROWĄ wartość komórki nagłówka (str/None/liczba/datetime),
    NIE openpyxl ``Cell`` — dzięki temu ten sam kod obsługuje XLSX (openpyxl)
    i CSV (stringi)."""
    s = str(value).lower().split("\n")[0]

    s = s.replace(".", " ")
    while s.find("  ") >= 0:
        s = s.replace("  ", " ")
    s = s.strip()

    return s.replace(" ", "_").replace("/", "_").replace("\\", "_").replace("-", "_")


def find_similar_row_in_rows(rows, try_names=None, min_points=None, max_row_length=128):
    """Rdzeń fuzzy-detekcji nagłówka nad gołymi listami wartości.

    :param rows: iterowalne wierszy; każdy wiersz to lista wartości (str/None/…)
    :return: ``(znormalizowane_nazwy, n_1based)`` pierwszego wiersza z
        ``>= min_points`` trafień, albo ``None``.
    """
    if try_names is None:
        try_names = DEFAULT_COL_NAMES

    if min_points is None:
        min_points = DEFAULT_MIN_POINTS

    for n, row in enumerate(rows, start=1):
        r = [normalize_cell_header(v) for v in row[:max_row_length]]
        points = 0
        for elem in try_names:
            if elem in r:
                points += 1
        if points >= min_points:
            return r, n


def find_similar_row(sheet, try_names=None, min_points=None, max_row_length=128):
    """Wrapper zachowujący dotychczasową sygnaturę (arkusz openpyxl):
    wyciąga ``cell.value`` z każdej komórki i deleguje do
    ``find_similar_row_in_rows``. Istniejący callerzy (``znajdz_naglowek``,
    ``XLSImportFile``) nie wymagają zmian."""
    rows = ([cell.value for cell in row] for row in sheet.rows)
    return find_similar_row_in_rows(rows, try_names, min_points, max_row_length)


def znajdz_naglowek(
    sciezka,
    try_names=None,
    min_points=None,
):
    """
    :return: ([str, str...], no_row)
    """
    import openpyxl
    from openpyxl.utils.exceptions import InvalidFileException

    sprawdz_bombe_dekompresji(sciezka)
    try:
        f: openpyxl.workbook.workbook.Workbook = openpyxl.load_workbook(sciezka)
    except InvalidFileException as e:
        raise ImproperFileException(e) from e

    # Sprawdź, ile jest skoroszytów
    if len(f.worksheets) != 1:
        raise BadNoOfSheetsException()

    s = f.worksheets[0]

    res = find_similar_row(s, try_names, min_points)

    if res is None:
        raise HeaderNotFoundException()

    return res


DEFAULT_BANNED_NAMES = ["pesel", "pesel_md5", "peselmd5"]


class XLSImportFile:
    def __init__(
        self,
        xls_path,
        try_names=None,
        min_points=None,
        banned_names=None,
        only_first_sheet=False,
    ):
        """
        :param xls_path: ścieżka do pliku
        :param try_names: nazwy które będą poszukiwane jako nagłówek
        :param min_points: ile z nazw nagłówka musi się odnaleźć w wierszu, żeby był tak potraktowany
        :param banned_names: niedozwolone nazwy nagłówków - te kolumny nie będą importowane (np PESEL)
        """

        self.xls_path = xls_path
        self.try_names = try_names
        self.min_points = min_points

        if banned_names is None:
            banned_names = DEFAULT_BANNED_NAMES
        self.banned_names = banned_names

        self.only_first_sheet = only_first_sheet

    @cached_property
    def xl_workbook(self) -> openpyxl.workbook.workbook.Workbook:
        import openpyxl

        sprawdz_bombe_dekompresji(self.xls_path)
        return openpyxl.load_workbook(self.xls_path)

    @cached_property
    def sheet_limit_range_end(self):
        limit = None
        if self.only_first_sheet:
            limit = 1
        return limit

    @cached_property
    def sheet_row_cache(self):
        _cache = {}

        for _n_sheet, sheet in enumerate(
            self.xl_workbook.worksheets[: self.sheet_limit_range_end]
        ):
            res = find_similar_row(
                sheet, try_names=self.try_names, min_points=self.min_points
            )
            if res is None:
                continue
            _cache[sheet] = res
        return _cache

    @staticmethod
    def _pusty(values) -> bool:
        """Wiersz pusty = wszystkie komórki danych ``None`` albo białe znaki.
        openpyxl ``sheet.max_row`` obejmuje puste wiersze końcowe (Excel śledzi
        „użyty zakres" — stąd rozjazd `max_row` z realną liczbą danych), więc
        bez tego filtra puste wiersze trafiają do pipeline'u i wywalają
        walidację wymaganych pól (nazwisko/imię) — cały import pada na jednym
        pustym wierszu na końcu arkusza. Zwierciadło ``CSVSource._pusty``."""
        return not any((str(v).strip() if v is not None else "") for v in values)

    def count(self) -> int:
        """
        Zwraca całkowitą liczbę NIEPUSTYCH wierszy do analizy — spójne z
        ``data()`` (puste wiersze pomijamy po obu stronach, inaczej pasek
        postępu nigdy nie dobija do 100%).
        """
        total = 0
        for sheet in self.xl_workbook.worksheets:
            res = self.sheet_row_cache.get(sheet)
            if res is None:
                continue
            labels, no = res
            for n_row, row in enumerate(sheet.rows):
                if n_row < no:
                    continue
                if self._pusty(cell.value for cell in row[: len(labels)]):
                    continue
                total += 1

        return total

    def liczba_arkuszy_z_danymi(self) -> int:
        """Liczba arkuszy z rozpoznanym nagłówkiem (= arkuszy z danymi).
        Importy wymuszające „jeden arkusz = jeden import" (np. import
        pracowników) używają tego do odrzucenia plików wieloarkuszowych: dwa
        arkusze w jednym skoroszycie to zwykle dwa rozłączne zbiory (np. dwie
        uczelnie), których nie wolno po cichu skleić w jeden import."""
        return len(self.sheet_row_cache)

    def data(self) -> Generator[dict, None, None]:
        """
        Ta funkcja dla każdego arkusza:
        1) znajduje nagłówek za pomoca funkcji `find_similar_row`,
        2) wczytuje poniższe wiersze do końca arkusza,
        3) dla każdego wiersza przypisuje nazwy kolumn z wiersza nagłówkowego
        4) zwraca słowniki.
        """

        for n_sheet, sheet in enumerate(self.xl_workbook.worksheets):
            res = self.sheet_row_cache.get(sheet)
            if res is None:
                continue

            colnames, no = res

            colnames = rename_duplicate_columns(colnames)

            colnames.append("__xls_loc_sheet__")
            colnames.append("__xls_loc_row__")

            for n_row, row in enumerate(sheet.rows):
                if n_row < res[1]:
                    continue
                data = [x.value for x in row[: len(colnames) - 2]]
                if self._pusty(data):
                    continue
                data.append(n_sheet)
                data.append(n_row)

                yld = dict(zip(colnames, data, strict=False))

                for banned_name in self.banned_names:
                    if banned_name in yld:
                        del yld[banned_name]

                yield yld


def rename_duplicate_columns(s: list[str], marker: str = "_") -> list[str]:
    seen = defaultdict(lambda: 1)
    ret = []
    for elem in s:
        no_seen = seen[elem]
        if no_seen == 1:
            ret.append(elem)
        else:
            ret.append(f"{elem}{marker}{no_seen}")
        seen[elem] += 1
    return ret


doi_regexp = re.compile(r"10.\d{4,9}/[-._;()/:A-Za-z0-9]+")


def strip_doi_urls(s: str) -> str:
    if not s:
        return

    return (
        s.replace("https://dx.doi.org/", "")
        .replace("http://dx.doi.org/", "")
        .replace("https://doi.org/", "")
        .replace("http://doi.org/", "")
        .replace("dx.doi.org/", "")
        .replace("doi.org/", "")
        .strip()
    )


def check_if_doi(s: str):
    """Sprawdza, czy wpisany ciąg znaków to może być DOI, POD WARUNKIEM że
    numeryczne DOI jest na początku (bez spacji itp)"""

    if not s:
        return False

    s = strip_doi_urls(s)

    ret = doi_regexp.search(s)
    if ret is not None:
        if ret.span()[0] == 0:
            return True

    return False
