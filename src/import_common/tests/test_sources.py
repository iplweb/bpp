import pytest

from import_common.exceptions import HeaderNotFoundException
from import_common.sources import CSVSource, XLSXSource, otworz_zrodlo, wykryj_format


def test_wykryj_format_xlsx_po_magic_bytes(test1_xlsx):
    # XLSX = archiwum ZIP, zaczyna się od b"PK"
    assert wykryj_format(test1_xlsx) == "xlsx"


def test_wykryj_format_csv_gdy_nie_zip(tmp_path):
    p = tmp_path / "dane.csv"
    p.write_text("Nazwisko;Imię\nKowalski;Jan\n", encoding="utf-8")
    assert wykryj_format(str(p)) == "csv"


def test_xlsxsource_deleguje_do_xlsimportfile(default_xlsx):
    src = XLSXSource(default_xlsx)
    # count() i data() zwracają to samo, co XLSImportFile
    assert src.count() >= 0
    wiersze = list(src.data())
    assert len(wiersze) == src.count()
    if wiersze:
        # kontrakt kluczy lokalizacyjnych
        assert "__xls_loc_sheet__" in wiersze[0]
        assert "__xls_loc_row__" in wiersze[0]


_CSV_NAGLOWEK = "Numer;Nazwisko;Imię;Orcid;Stanowisko;Nazwa jednostki"


def _zapisz(tmp_path, tresc, encoding="utf-8", nazwa="dane.csv"):
    p = tmp_path / nazwa
    p.write_bytes(tresc.encode(encoding))
    return str(p)


def test_csvsource_srednik_cp1250(tmp_path):
    tresc = (
        f"{_CSV_NAGLOWEK}\n"
        "1;Kowalski;Jan;;Asystent;Katedra Chorób\n"
        "2;Wiśniewska;Zofia;;Adiunkt;Zakład Fizyki\n"
    )
    path = _zapisz(tmp_path, tresc, encoding="cp1250")
    src = CSVSource(path)
    assert src.count() == 2
    wiersze = list(src.data())
    assert wiersze[0]["nazwisko"] == "Kowalski"
    assert wiersze[0]["nazwa_jednostki"] == "Katedra Chorób"
    # klucze lokalizacyjne — CSV to jeden arkusz
    assert wiersze[0]["__xls_loc_sheet__"] == 0
    assert wiersze[1]["__xls_loc_row__"] > wiersze[0]["__xls_loc_row__"]


def test_csvsource_przecinek_delimiter(tmp_path):
    tresc = (
        "Nazwisko,Imię,Nazwa jednostki,Orcid,Stanowisko\n"
        "Nowak,Anna,Katedra Testowa,,Profesor\n"
    )
    path = _zapisz(tmp_path, tresc, encoding="utf-8")
    src = CSVSource(path)
    assert src.count() == 1
    assert list(src.data())[0]["imię"] == "Anna"


def test_csvsource_odrzuca_pesel(tmp_path):
    tresc = (
        "Nazwisko;Imię;PESEL;Nazwa jednostki;Orcid;Stanowisko\n"
        "Kowalski;Jan;12345678901;Katedra;;Asystent\n"
    )
    path = _zapisz(tmp_path, tresc, encoding="utf-8")
    wiersz = list(CSVSource(path).data())[0]
    assert "pesel" not in wiersz


def test_csvsource_brak_naglowka_rzuca(tmp_path):
    tresc = "aaa;bbb;ccc\n1;2;3\n"
    path = _zapisz(tmp_path, tresc, encoding="utf-8")
    with pytest.raises(HeaderNotFoundException):
        CSVSource(path).count()


def test_csvsource_pusty_plik_zero(tmp_path):
    # sam nagłówek, brak danych
    tresc = f"{_CSV_NAGLOWEK}\n"
    path = _zapisz(tmp_path, tresc, encoding="utf-8")
    src = CSVSource(path)
    assert src.count() == 0
    assert list(src.data()) == []


def test_csvsource_pomija_puste_linie(tmp_path):
    tresc = (
        f"{_CSV_NAGLOWEK}\n"
        "1;Kowalski;Jan;;Asystent;Katedra\n"
        "\n"
        ";;;;;\n"
        "2;Nowak;Ewa;;Adiunkt;Zakład\n"
    )
    path = _zapisz(tmp_path, tresc, encoding="utf-8")
    src = CSVSource(path)
    assert src.count() == 2


def test_csvsource_poszarpany_wiersz_zachowuje_loc_keys(tmp_path):
    # wiersz danych KRÓTSZY niż nagłówek (csv.reader nie padduje) — klucze
    # lokalizacyjne muszą i tak powstać na właściwych pozycjach
    tresc = f"{_CSV_NAGLOWEK}\n1;Kowalski;Jan\n"  # brak 3 ostatnich kolumn
    path = _zapisz(tmp_path, tresc, encoding="utf-8")
    wiersz = list(CSVSource(path).data())[0]
    assert wiersz["__xls_loc_sheet__"] == 0
    assert wiersz["__xls_loc_row__"] == 1
    assert wiersz["nazwisko"] == "Kowalski"
    # brakujące kolumny → puste, nie przesunięte
    assert wiersz["nazwa_jednostki"] == ""


def test_csvsource_srednik_wygrywa_z_przecinkiem_dziesietnym(tmp_path):
    # §13: plik `;` z przecinkiem dziesiętnym `0,5` w komórce — detekcja NIE
    # może wybrać `,` jako delimitera; komórka „0,5" zostaje w całości
    tresc = (
        "Nazwisko;Imię;Nazwa jednostki;Orcid;Stanowisko;Wymiar etatu\n"
        "Kowalski;Jan;Katedra;;Asystent;0,5\n"
        "Nowak;Ewa;Zakład;;Adiunkt;0,5\n"
    )
    path = _zapisz(tmp_path, tresc, encoding="utf-8")
    src = CSVSource(path)
    assert src.count() == 2
    wiersz = list(src.data())[0]
    assert wiersz["wymiar_etatu"] == "0,5"
    assert wiersz["nazwisko"] == "Kowalski"


def _zapisz_xlsx(tmp_path, wiersze, nazwa="dane.xlsx"):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    for w in wiersze:
        ws.append(w)
    p = tmp_path / nazwa
    wb.save(str(p))
    return str(p)


def test_xlsxsource_pomija_puste_wiersze(tmp_path):
    # openpyxl `sheet.max_row` obejmuje puste wiersze końcowe (Excel śledzi
    # „użyty zakres"), więc bez filtra pustych trafiają do pipeline'u i wywalają
    # walidację wymaganych pól (nazwisko/imię) — cały import pada na jednym
    # pustym wierszu na końcu arkusza. XLSX MUSI pomijać puste wiersze tak samo
    # jak CSVSource (symetria kontraktu). Regresja: operacja bd90660e.
    path = _zapisz_xlsx(
        tmp_path,
        [
            ["Nazwisko", "Imię", "Stanowisko"],  # nagłówek (row 1)
            ["Kowalski", "Jan", "Adiunkt"],  # dane (row 2)
            [None, None, None],  # pusty wiersz w środku (row 3)
            ["Nowak", "Ewa", "Profesor"],  # dane (row 4)
            [None, None, None],  # pusty wiersz końcowy (row 5)
            ["", "", ""],  # pusty wiersz — puste stringi (row 6)
        ],
    )
    src = XLSXSource(path)
    assert src.count() == 2
    wiersze = list(src.data())
    assert len(wiersze) == 2
    assert wiersze[0]["nazwisko"] == "Kowalski"
    assert wiersze[1]["nazwisko"] == "Nowak"
    # klucze lokalizacyjne wskazują REALNY wiersz w arkuszu (0-based) — pominięcie
    # pustego wiersza nie przesuwa numeracji kolejnych.
    assert wiersze[1]["__xls_loc_row__"] == 3


def test_xlsxsource_liczba_arkuszy_z_danymi_dwa(tmp_path):
    # Dwa arkusze, każdy z rozpoznanym nagłówkiem → 2 arkusze z danymi (baza
    # reguły „jeden arkusz = jeden import"). Odzwierciedla plik UAFM_x_MWSL.
    import openpyxl

    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "UAFM"
    ws0.append(["Nazwisko", "Imię", "Stanowisko"])
    ws0.append(["Kowalski", "Jan", "Adiunkt"])
    ws1 = wb.create_sheet("MWSL")
    ws1.append(["Nazwisko", "Imię", "Stanowisko"])
    ws1.append(["Nowak", "Ewa", "Profesor"])
    p = tmp_path / "dwa.xlsx"
    wb.save(str(p))
    assert XLSXSource(str(p)).liczba_arkuszy_z_danymi() == 2


def test_xlsxsource_liczba_arkuszy_z_danymi_jeden(tmp_path):
    path = _zapisz_xlsx(
        tmp_path,
        [["Nazwisko", "Imię", "Stanowisko"], ["Kowalski", "Jan", "Adiunkt"]],
    )
    assert XLSXSource(path).liczba_arkuszy_z_danymi() == 1


def test_xlsxsource_pusty_drugi_arkusz_nie_liczy(tmp_path):
    # Drugi arkusz bez rozpoznanego nagłówka (pusty / śmieciowy) NIE liczy się
    # jako arkusz z danymi — reguła „jeden arkusz = jeden import" go ignoruje,
    # więc plik z jednym arkuszem danych + pustym Sheet2 przechodzi.
    import openpyxl

    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.append(["Nazwisko", "Imię", "Stanowisko"])
    ws0.append(["Kowalski", "Jan", "Adiunkt"])
    wb.create_sheet("Pusty")  # bez nagłówka
    p = tmp_path / "jeden_plus_pusty.xlsx"
    wb.save(str(p))
    assert XLSXSource(str(p)).liczba_arkuszy_z_danymi() == 1


def test_csvsource_liczba_arkuszy_z_danymi_jeden(tmp_path):
    path = _zapisz(tmp_path, f"{_CSV_NAGLOWEK}\n1;Kowalski;Jan;;Asystent;Katedra\n")
    assert CSVSource(path).liczba_arkuszy_z_danymi() == 1


def test_otworz_zrodlo_xlsx(default_xlsx):
    assert isinstance(otworz_zrodlo(default_xlsx), XLSXSource)


def test_otworz_zrodlo_csv(tmp_path):
    p = tmp_path / "dane.csv"
    p.write_text("Nazwisko;Imię\nKowalski;Jan\n", encoding="utf-8")
    assert isinstance(otworz_zrodlo(str(p)), CSVSource)


def test_csvsource_crlf_realny_excel(tmp_path):
    # realny Excel zapisuje końce linii jako CRLF (\r\n) — csv.reader ścina
    # \r, więc ostatnia kolumna NIE może wyciec z doklejonym \r
    tresc = f"{_CSV_NAGLOWEK}\r\n1;Kowalski;Jan;;Asystent;Katedra Chorób\r\n"
    path = _zapisz(tmp_path, tresc, encoding="utf-8")
    src = CSVSource(path)
    assert src.count() == 1
    wiersz = list(src.data())[0]
    assert wiersz["nazwa_jednostki"] == "Katedra Chorób"
    assert wiersz["__xls_loc_row__"] == 1


def test_csvsource_cr_only_konce_linii(tmp_path):
    # Końce linii CR-only (\r, stare Maki / niektóre eksporty kadrowe) — bez
    # `newline=""` w StringIO `csv.reader` rzuca `_csv.Error: new-line character
    # seen in unquoted field` i wywala CAŁĄ analizę surowym tracebackiem. Musi
    # sparsować się poprawnie (a nie-CSV śmieć degradować do domenowego wyjątku).
    tresc = f"{_CSV_NAGLOWEK}\r1;Kowalski;Jan;;Asystent;Katedra Chorób\r"
    path = _zapisz(tmp_path, tresc, encoding="utf-8")
    src = CSVSource(path)
    assert src.count() == 1
    wiersz = list(src.data())[0]
    assert wiersz["nazwisko"] == "Kowalski"
    assert wiersz["nazwa_jednostki"] == "Katedra Chorób"


def test_csvsource_stray_cr_w_polu(tmp_path):
    # Zabłąkany \r w niecytowanym polu (mieszane końce linii po ręcznej edycji)
    # — z `newline=""` csv.reader traktuje \r jako koniec rekordu (rozbija wiersz)
    # ZAMIAST rzucać `_csv.Error`. Kluczowe: analiza NIE wywala się surowym
    # tracebackiem; degradacja (rozbity rekord) jest akceptowalna wobec crasha.
    tresc = f"{_CSV_NAGLOWEK}\n1;Kowal\rski;Jan;;Asystent;Katedra\n"
    path = _zapisz(tmp_path, tresc, encoding="utf-8")
    src = CSVSource(path)
    # nie rzuca `_csv.Error` (przed fixem: crash całej analizy)
    wiersze = list(src.data())
    assert wiersze  # coś się sparsowało bez wyjątku
