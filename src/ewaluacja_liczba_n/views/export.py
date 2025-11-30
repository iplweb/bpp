from decimal import Decimal

from braces.views import GroupRequiredMixin
from django.db.models import Q
from django.views import View
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

from bpp.const import GR_WPROWADZANIE_DANYCH

from ..excel_export import LiczbaNExcelExporter
from ..models import IloscUdzialowDlaAutoraZaCalosc, IloscUdzialowDlaAutoraZaRok


class AutorzyLiczbaNExporter(LiczbaNExcelExporter):
    """Exporter for author N-number data with yearly breakdown."""

    def get_filename(self) -> str:
        return "liczba_n_ewaluacja_2022_2025.xlsx"

    def _get_filtered_udzialy_queryset(self, request):
        """Get filtered queryset based on request parameters."""
        udzialy = IloscUdzialowDlaAutoraZaRok.objects.filter(
            rok__gte=2022, rok__lte=2025
        )

        # Apply filters from URL
        search = request.GET.get("search")
        if search:
            udzialy = udzialy.filter(
                Q(autor__nazwisko__icontains=search)
                | Q(autor__imiona__icontains=search)
            )

        dyscyplina_id = request.GET.get("dyscyplina")
        if dyscyplina_id:
            udzialy = udzialy.filter(dyscyplina_naukowa_id=dyscyplina_id)

        rok = request.GET.get("rok")
        if rok:
            udzialy = udzialy.filter(rok=rok)

        return udzialy.select_related(
            "autor",
            "dyscyplina_naukowa",
            "autor_dyscyplina",
            "autor_dyscyplina__rodzaj_autora",
        ).order_by(
            "autor__nazwisko", "autor__imiona", "rok", "dyscyplina_naukowa__nazwa"
        )

    def _calculate_wymiar_etatu_dla_dyscypliny(self, udzial, autor_dyscyplina):
        """Calculate wymiar etatu for discipline."""
        if not autor_dyscyplina or not autor_dyscyplina.wymiar_etatu:
            return None

        if udzial.dyscyplina_naukowa_id == autor_dyscyplina.dyscyplina_naukowa_id:
            return float(
                autor_dyscyplina.wymiar_etatu
                * (autor_dyscyplina.procent_dyscypliny or Decimal("0"))
                / Decimal("100")
            )
        elif udzial.dyscyplina_naukowa_id == autor_dyscyplina.subdyscyplina_naukowa_id:
            return float(
                autor_dyscyplina.wymiar_etatu
                * (autor_dyscyplina.procent_subdyscypliny or Decimal("0"))
                / Decimal("100")
            )
        return None

    def _format_autorzy_detail_worksheet(self, ws_detail, row_num):
        """Apply formatting to autorzy detail worksheet."""
        if row_num <= 2:
            return

        # Define the table range
        table_range = f"A1:I{row_num - 1}"

        # Create table style
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        # Create the table
        table = Table(
            displayName="AutorzyLiczbaN", ref=table_range, tableStyleInfo=style
        )

        # Add table to worksheet
        ws_detail.add_table(table)

        # Freeze panes to keep headers visible
        ws_detail.freeze_panes = "A2"

        # Apply number formatting for decimal columns
        for row in range(2, row_num):
            # Column 5: Wymiar Etatu (decimal with 2 places)
            if ws_detail.cell(row=row, column=5).value is not None:
                ws_detail.cell(row=row, column=5).number_format = "0.00"

            # Column 6: Wymiar Etatu dla Dyscypliny (decimal with 4 places)
            if ws_detail.cell(row=row, column=6).value is not None:
                ws_detail.cell(row=row, column=6).number_format = "0.0000"

            # Column 7: Ilość Udziałów (decimal with 2 places)
            ws_detail.cell(row=row, column=7).number_format = "0.00"

            # Column 8: Ilość Udziałów - Monografie (decimal with 2 places)
            ws_detail.cell(row=row, column=8).number_format = "0.00"

    def _create_detail_worksheet(self, wb, request):
        """Create detailed author data worksheet."""
        ws_detail = wb.create_sheet("Autorzy - Liczba N")

        # Headers
        headers = [
            "Autor",
            "Rok",
            "Dyscyplina",
            "Kod Dyscypliny",
            "Wymiar Etatu",
            "Wymiar Etatu dla Dyscypliny",
            "Ilość Udziałów",
            "Ilość Udziałów - Monografie",
            "Rodzaj Autora",
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        # Author data - with URL filters applied
        udzialy = self._get_filtered_udzialy_queryset(request)

        row_num = 2
        for udzial in udzialy:
            autor_dyscyplina = udzial.autor_dyscyplina

            ws_detail.cell(row=row_num, column=1, value=str(udzial.autor))
            ws_detail.cell(row=row_num, column=2, value=udzial.rok)
            ws_detail.cell(row=row_num, column=3, value=udzial.dyscyplina_naukowa.nazwa)
            ws_detail.cell(row=row_num, column=4, value=udzial.dyscyplina_naukowa.kod)

            if autor_dyscyplina and autor_dyscyplina.wymiar_etatu:
                ws_detail.cell(
                    row=row_num,
                    column=5,
                    value=float(autor_dyscyplina.wymiar_etatu),
                )
                wymiar_dla_dyscypliny = self._calculate_wymiar_etatu_dla_dyscypliny(
                    udzial, autor_dyscyplina
                )
                if wymiar_dla_dyscypliny is not None:
                    ws_detail.cell(
                        row=row_num,
                        column=6,
                        value=wymiar_dla_dyscypliny,
                    )

            if autor_dyscyplina:
                ws_detail.cell(
                    row=row_num,
                    column=9,
                    value=(
                        autor_dyscyplina.rodzaj_autora.nazwa
                        if autor_dyscyplina.rodzaj_autora
                        else ""
                    ),
                )

            # Number of shares
            ws_detail.cell(row=row_num, column=7, value=float(udzial.ilosc_udzialow))
            ws_detail.cell(
                row=row_num, column=8, value=float(udzial.ilosc_udzialow_monografie)
            )

            row_num += 1

        # Create Excel Table and apply formatting
        self._format_autorzy_detail_worksheet(ws_detail, row_num)


class ExportAutorzyLiczbaNView(GroupRequiredMixin, View):
    """Eksport danych autorów do pliku XLSX"""

    group_required = GR_WPROWADZANIE_DANYCH

    def get(self, request, *args, **kwargs):
        exporter = AutorzyLiczbaNExporter()
        return exporter.export(request)


class UdzialyZaCaloscExporter(LiczbaNExcelExporter):
    """Exporter for full evaluation period share data."""

    def get_filename(self) -> str:
        return "udzialy_za_calosc_ewaluacja_2022_2025.xlsx"

    def _get_filtered_udzialy_calosc_queryset(self, request):
        """Get filtered udzialy za calosc queryset."""
        udzialy = IloscUdzialowDlaAutoraZaCalosc.objects.all()

        search = request.GET.get("search")
        if search:
            udzialy = udzialy.filter(
                Q(autor__nazwisko__icontains=search)
                | Q(autor__imiona__icontains=search)
            )

        dyscyplina_id = request.GET.get("dyscyplina")
        if dyscyplina_id:
            udzialy = udzialy.filter(dyscyplina_naukowa_id=dyscyplina_id)

        rodzaj_autora_id = request.GET.get("rodzaj_autora")
        if rodzaj_autora_id:
            udzialy = udzialy.filter(rodzaj_autora_id=rodzaj_autora_id)

        return udzialy.select_related(
            "autor", "dyscyplina_naukowa", "rodzaj_autora"
        ).order_by("autor__nazwisko", "autor__imiona", "dyscyplina_naukowa__nazwa")

    def _format_calosc_detail_worksheet(self, ws_detail, row_num):
        """Apply formatting to full period detail worksheet."""
        if row_num <= 2:
            return

        table_range = f"A1:G{row_num - 1}"
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table = Table(
            displayName="UdzialyZaCalosc", ref=table_range, tableStyleInfo=style
        )
        ws_detail.add_table(table)
        ws_detail.freeze_panes = "A2"

        for row in range(2, row_num):
            if ws_detail.cell(row=row, column=5).value is not None:
                ws_detail.cell(row=row, column=5).number_format = "0.00"
            if ws_detail.cell(row=row, column=6).value is not None:
                ws_detail.cell(row=row, column=6).number_format = "0.00"
            if ws_detail.cell(row=row, column=7).value:
                ws_detail.cell(row=row, column=7).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

    def _create_detail_worksheet(self, wb, request):
        """Create full period detail worksheet."""
        ws_detail = wb.create_sheet("Udziały za cały okres")

        headers = [
            "Autor",
            "Dyscyplina",
            "Kod Dyscypliny",
            "Rodzaj autora",
            "Ilość Udziałów (2022-2025)",
            "Ilość Udziałów - Monografie",
            "Komentarz",
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        udzialy = self._get_filtered_udzialy_calosc_queryset(request)

        row_num = 2
        for udzial in udzialy:
            ws_detail.cell(row=row_num, column=1, value=str(udzial.autor))
            ws_detail.cell(row=row_num, column=2, value=udzial.dyscyplina_naukowa.nazwa)
            ws_detail.cell(row=row_num, column=3, value=udzial.dyscyplina_naukowa.kod)
            ws_detail.cell(
                row=row_num,
                column=4,
                value=(
                    udzial.rodzaj_autora.nazwa
                    if udzial.rodzaj_autora
                    else "Brak danych"
                ),
            )
            ws_detail.cell(row=row_num, column=5, value=float(udzial.ilosc_udzialow))
            ws_detail.cell(
                row=row_num, column=6, value=float(udzial.ilosc_udzialow_monografie)
            )
            komentarz_text = (udzial.komentarz or "").replace("<br>", "\n")
            ws_detail.cell(row=row_num, column=7, value=komentarz_text)
            row_num += 1

        self._format_calosc_detail_worksheet(ws_detail, row_num)


class ExportUdzialyZaCaloscView(GroupRequiredMixin, View):
    """Eksport danych udziałów za cały okres do pliku XLSX"""

    group_required = GR_WPROWADZANIE_DANYCH

    def get(self, request, *args, **kwargs):
        exporter = UdzialyZaCaloscExporter()
        return exporter.export(request)
