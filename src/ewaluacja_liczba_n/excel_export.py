"""Base classes for Excel export functionality in ewaluacja_liczba_n."""

from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from bpp.models import Uczelnia

from .models import LiczbaNDlaUczelni
from .utils import oblicz_liczbe_n_na_koniec_2025


class LiczbaNExcelExporter:
    """Base class for Liczba N Excel exports.

    Provides common methods for creating Excel workbooks with summary,
    non-reported disciplines, and author detail worksheets.

    Subclasses should implement:
    - _create_detail_worksheet(wb, request): Create the main detail worksheet
    - get_filename(): Return the export filename
    """

    def _create_summary_worksheet(self, wb: Workbook, uczelnia: Uczelnia) -> None:
        """Create summary worksheet with Liczba N data."""
        ws_summary = wb.active
        ws_summary.title = "Podsumowanie Liczba N"

        # Headers
        headers = ["Dyscyplina", "Kod", "Liczba N"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        # Data
        liczby_n = (
            LiczbaNDlaUczelni.objects.filter(uczelnia=uczelnia)
            .select_related("dyscyplina_naukowa")
            .order_by("dyscyplina_naukowa__nazwa")
        )

        row_num = 2
        suma = 0
        for liczba_n in liczby_n:
            ws_summary.cell(
                row=row_num, column=1, value=liczba_n.dyscyplina_naukowa.nazwa
            )
            ws_summary.cell(
                row=row_num, column=2, value=liczba_n.dyscyplina_naukowa.kod
            )
            ws_summary.cell(row=row_num, column=3, value=float(liczba_n.liczba_n))
            suma += float(liczba_n.liczba_n)
            row_num += 1

        # Sum row
        ws_summary.cell(row=row_num, column=1, value="SUMA")
        ws_summary.cell(row=row_num, column=3, value=suma)
        ws_summary.cell(row=row_num, column=1).font = Font(bold=True)
        ws_summary.cell(row=row_num, column=3).font = Font(bold=True)

    def _create_nieraportowane_worksheet(
        self, wb: Workbook, uczelnia: Uczelnia
    ) -> None:
        """Create non-reported disciplines worksheet."""
        ws_nieraportowane = wb.create_sheet("Dyscypliny Nieraportowane")

        headers = ["Dyscyplina", "Kod", "Liczba N - średnia", "Liczba N - koniec 2025"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_nieraportowane.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        # Get all disciplines and calculate N values for end of 2025
        wszystkie_liczby_n = (
            LiczbaNDlaUczelni.objects.filter(uczelnia=uczelnia)
            .select_related("dyscyplina_naukowa")
            .order_by("dyscyplina_naukowa__nazwa")
        )
        liczby_n_2025 = oblicz_liczbe_n_na_koniec_2025(uczelnia)

        # Filter only non-reported (N < 12 at end of 2025)
        row_num = 2
        for liczba in wszystkie_liczby_n:
            liczba_n_2025 = liczby_n_2025.get(liczba.dyscyplina_naukowa_id, 0)
            if liczba_n_2025 < 12:
                ws_nieraportowane.cell(
                    row=row_num, column=1, value=liczba.dyscyplina_naukowa.nazwa
                )
                ws_nieraportowane.cell(
                    row=row_num, column=2, value=liczba.dyscyplina_naukowa.kod
                )
                ws_nieraportowane.cell(
                    row=row_num, column=3, value=float(liczba.liczba_n)
                )
                ws_nieraportowane.cell(
                    row=row_num, column=4, value=float(liczba_n_2025)
                )
                row_num += 1

    def _apply_column_widths(self, wb: Workbook) -> None:
        """Adjust column widths for all worksheets."""
        for sheet in wb.worksheets:
            for column_cells in sheet.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = min(length + 2, 50)

    def _create_excel_response(self, wb: Workbook, filename: str) -> HttpResponse:
        """Create HTTP response with Excel file."""
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        # Save to buffer
        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)
        response.write(virtual_workbook.getvalue())

        return response

    def _create_detail_worksheet(self, wb: Workbook, request) -> None:
        """Create the main detail worksheet. Override in subclasses."""
        raise NotImplementedError("Subclasses must implement _create_detail_worksheet")

    def get_filename(self) -> str:
        """Return the export filename. Override in subclasses."""
        raise NotImplementedError("Subclasses must implement get_filename")

    def export(self, request) -> HttpResponse:
        """Generate and return the Excel export response."""
        uczelnia = Uczelnia.objects.get_default()
        wb = Workbook()

        # Sheet 1: Summary of Liczba N for institution
        self._create_summary_worksheet(wb, uczelnia)

        # Sheet 2: Detailed data (implemented by subclass)
        self._create_detail_worksheet(wb, request)

        # Sheet 3: Non-reported disciplines
        self._create_nieraportowane_worksheet(wb, uczelnia)

        # Adjust column widths
        self._apply_column_widths(wb)

        # Return response
        return self._create_excel_response(wb, self.get_filename())
