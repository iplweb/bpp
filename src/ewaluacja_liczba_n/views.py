import sys
import traceback
from io import BytesIO

import rollbar
from braces.views import GroupRequiredMixin
from django.contrib import messages
from django.db.models import Count, Q, Sum
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.views import View
from django.views.generic import ListView, TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from bpp.const import GR_WPROWADZANIE_DANYCH
from bpp.models import Autor_Dyscyplina, Uczelnia

from .models import (
    DyscyplinaNieRaportowana,
    IloscUdzialowDlaAutoraZaCalosc,
    IloscUdzialowDlaAutoraZaRok,
    LiczbaNDlaUczelni,
)
from .utils import oblicz_liczby_n_dla_ewaluacji_2022_2025


class LiczbaNIndexView(GroupRequiredMixin, TemplateView):
    """Główny widok aplikacji liczba N"""

    template_name = "ewaluacja_liczba_n/index.html"
    group_required = GR_WPROWADZANIE_DANYCH

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        uczelnia = Uczelnia.objects.get_default()

        # Pobierz dane liczby N dla uczelni
        context["liczby_n"] = (
            LiczbaNDlaUczelni.objects.filter(uczelnia=uczelnia)
            .select_related("dyscyplina_naukowa")
            .order_by("-liczba_n")
        )

        # Pobierz dyscypliny nieraportowane
        context["dyscypliny_nieraportowane"] = (
            DyscyplinaNieRaportowana.objects.filter(uczelnia=uczelnia)
            .select_related("dyscyplina_naukowa")
            .order_by("dyscyplina_naukowa__nazwa")
        )

        # Oblicz sumę liczby N
        context["suma_liczby_n"] = (
            context["liczby_n"].aggregate(suma=Sum("liczba_n"))["suma"] or 0
        )

        context["uczelnia"] = uczelnia
        return context


class ObliczLiczbeNView(GroupRequiredMixin, View):
    """Widok do obliczania liczby N"""

    group_required = GR_WPROWADZANIE_DANYCH

    def post(self, request, *args, **kwargs):
        uczelnia = Uczelnia.objects.get_default()

        try:
            oblicz_liczby_n_dla_ewaluacji_2022_2025(uczelnia)
            messages.success(
                request, "Pomyślnie obliczono liczbę N dla ewaluacji 2022-2025"
            )
        except Autor_Dyscyplina.DoesNotExist:
            traceback.print_exc()
            rollbar.report_exc_info(sys.exc_info())
            messages.error(
                request,
                "Błąd: Nie znaleziono danych Autor_Dyscyplina dla niektórych "
                "autorów. Upewnij się, że wszystkie dane są poprawnie "
                "wprowadzone w systemie.",
            )
        except NotImplementedError:
            traceback.print_exc()
            rollbar.report_exc_info()
            messages.error(
                request,
                "Błąd: Wykryto niespójność w danych - autor ma przypisane udziały "
                "dla dyscypliny która nie jest ani jego dyscypliną główną ani "
                "subdyscypliną. Odśwież tabelę Autor_Dyscyplina i spróbuj ponownie.",
            )
        except Exception as e:
            traceback.print_exc()
            rollbar.report_exc_info(sys.exc_info())
            messages.error(
                request,
                f"Błąd podczas obliczania liczby N: {str(e)}. "
                "Skontaktuj się z administratorem systemu.",
            )

        return HttpResponseRedirect(reverse("ewaluacja_liczba_n:index"))


class AutorzyLiczbaNListView(GroupRequiredMixin, ListView):
    """Lista autorów z liczbą N dla każdej dyscypliny"""

    template_name = "ewaluacja_liczba_n/autorzy_list.html"
    context_object_name = "autorzy_data"
    paginate_by = 50
    group_required = GR_WPROWADZANIE_DANYCH

    def _filter_by_rodzaj_autora(self, queryset, rodzaj_autora_id, rok):
        """Filter queryset by rodzaj_autora (author type)"""
        # Przygotuj filtr dla Autor_Dyscyplina
        ad_filter = {"rodzaj_autora_id": rodzaj_autora_id}

        # Jeśli filtr roku jest ustawiony, użyj dokładnego roku
        # W przeciwnym razie użyj zakresu lat
        if rok:
            ad_filter["rok"] = rok
        else:
            ad_filter["rok__gte"] = 2022
            ad_filter["rok__lte"] = 2025

        # Pobierz pary (autor_id, rok) z danym rodzajem autora
        autorzy_z_rodzajem = (
            Autor_Dyscyplina.objects.filter(**ad_filter)
            .values_list("autor_id", "rok")
            .distinct()
        )

        # Utwórz listę par (autor_id, rok) do filtrowania
        autor_rok_pairs = list(autorzy_z_rodzajem)

        # Filtruj queryset według par (autor_id, rok)
        if autor_rok_pairs:
            # Stwórz warunek Q dla każdej pary
            q_objects = Q()
            for autor_id, ar_rok in autor_rok_pairs:
                q_objects |= Q(autor_id=autor_id, rok=ar_rok)
            queryset = queryset.filter(q_objects)
        else:
            # Jeśli nie ma żadnych pasujących par, zwróć pusty queryset
            queryset = queryset.none()

        return queryset

    def _apply_memory_sorting_for_rodzaj_autora(self, queryset, sort):
        """Apply in-memory sorting for rodzaj_autora field"""
        # Pobierz wszystkie dane i sortuj w pamięci
        items = list(queryset)

        # Pobierz dane Autor_Dyscyplina dla wszystkich elementów
        autorzy_dyscypliny = {
            (ad.autor_id, ad.rok): ad
            for ad in Autor_Dyscyplina.objects.filter(
                autor__in=[item.autor_id for item in items],
                rok__in=[item.rok for item in items],
            ).select_related("rodzaj_autora")
        }

        def get_rodzaj_autora_key(item):
            ad = autorzy_dyscypliny.get((item.autor_id, item.rok))
            if ad and ad.rodzaj_autora:
                return ad.rodzaj_autora.skrot
            return ""

        items.sort(key=get_rodzaj_autora_key, reverse=(sort == "-rodzaj_autora"))

        # Store the sorted list for later use in get_context_data
        self._sorted_items = items
        return items  # Return list, but handle it in get_context_data

    def _apply_filters(self, queryset):
        """Apply all filters from request GET parameters"""
        # Filtrowanie po nazwisku/imieniu autora
        search = self.request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(autor__nazwisko__icontains=search)
                | Q(autor__imiona__icontains=search)
            )

        # Filtrowanie po dyscyplinie
        dyscyplina_id = self.request.GET.get("dyscyplina")
        if dyscyplina_id:
            queryset = queryset.filter(dyscyplina_naukowa_id=dyscyplina_id)

        # Filtrowanie po roku
        rok = self.request.GET.get("rok")
        if rok:
            queryset = queryset.filter(rok=rok)

        # Filtrowanie po rodzaju autora
        rodzaj_autora_id = self.request.GET.get("rodzaj_autora")
        if rodzaj_autora_id:
            queryset = self._filter_by_rodzaj_autora(queryset, rodzaj_autora_id, rok)

        return queryset

    def _get_sort_fields_mapping(self):
        """Return mapping of sort field names to database field lists"""
        return {
            "autor": ["autor__nazwisko", "autor__imiona", "rok"],
            "-autor": ["-autor__nazwisko", "-autor__imiona", "-rok"],
            "rok": ["rok", "autor__nazwisko", "autor__imiona"],
            "-rok": ["-rok", "autor__nazwisko", "autor__imiona"],
            "dyscyplina": [
                "dyscyplina_naukowa__nazwa",
                "autor__nazwisko",
                "autor__imiona",
            ],
            "-dyscyplina": [
                "-dyscyplina_naukowa__nazwa",
                "autor__nazwisko",
                "autor__imiona",
            ],
            "wymiar_etatu": [
                "rok",
                "autor__nazwisko",
                "autor__imiona",
            ],  # Will be sorted in memory
            "-wymiar_etatu": [
                "-rok",
                "autor__nazwisko",
                "autor__imiona",
            ],  # Will be sorted in memory
            "udzialy": ["ilosc_udzialow", "autor__nazwisko", "autor__imiona"],
            "-udzialy": ["-ilosc_udzialow", "autor__nazwisko", "autor__imiona"],
            "monografie": [
                "ilosc_udzialow_monografie",
                "autor__nazwisko",
                "autor__imiona",
            ],
            "-monografie": [
                "-ilosc_udzialow_monografie",
                "autor__nazwisko",
                "autor__imiona",
            ],
            "rodzaj_autora": [
                "autor__nazwisko",
                "autor__imiona",
                "rok",
            ],
            "-rodzaj_autora": [
                "-autor__nazwisko",
                "-autor__imiona",
                "-rok",
            ],
        }

    def _apply_sorting(self, queryset, sort):
        """Apply sorting to queryset based on sort parameter"""
        sort_mapping = self._get_sort_fields_mapping()
        sort_fields = sort_mapping.get(
            sort, ["autor__nazwisko", "autor__imiona", "rok"]
        )
        return queryset.order_by(*sort_fields)

    def get_queryset(self):
        # Pobierz wszystkie udziały dla autorów
        queryset = IloscUdzialowDlaAutoraZaRok.objects.filter(
            rok__gte=2022, rok__lte=2025
        ).select_related(
            "autor",
            "dyscyplina_naukowa",
            "autor_dyscyplina",
            "autor_dyscyplina__rodzaj_autora",
        )

        # Apply filters
        queryset = self._apply_filters(queryset)

        # Apply sorting
        sort = self.request.GET.get("sort", "autor")
        queryset = self._apply_sorting(queryset, sort)

        # Sortowanie w pamięci dla rodzaj_autora
        if sort in ["rodzaj_autora", "-rodzaj_autora"]:
            return self._apply_memory_sorting_for_rodzaj_autora(queryset, sort)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Dane do filtrowania - tylko dyscypliny które mają dane w bazie
        from decimal import Decimal

        from bpp.models import Dyscyplina_Naukowa
        from ewaluacja_common.models import Rodzaj_Autora

        # Pobierz ID dyscyplin które mają faktyczne dane
        dyscypliny_z_danymi = (
            IloscUdzialowDlaAutoraZaRok.objects.filter(rok__gte=2022, rok__lte=2025)
            .values_list("dyscyplina_naukowa_id", flat=True)
            .distinct()
        )

        # Filtruj tylko te dyscypliny które mają dane
        context["dyscypliny"] = Dyscyplina_Naukowa.objects.filter(
            pk__in=dyscypliny_z_danymi, widoczna=True
        ).order_by("nazwa")

        # Pobierz tylko lata które faktycznie są w bazie
        lata_z_danymi = (
            IloscUdzialowDlaAutoraZaRok.objects.filter(rok__gte=2022, rok__lte=2025)
            .values_list("rok", flat=True)
            .distinct()
            .order_by("rok")
        )

        context["lata"] = list(lata_z_danymi)

        # Pobierz wszystkie rodzaje autorów
        context["rodzaje_autorow"] = Rodzaj_Autora.objects.all().order_by("sort")

        # Aktualne filtry i sortowanie
        context["selected_dyscyplina"] = self.request.GET.get("dyscyplina", "")
        context["selected_rok"] = self.request.GET.get("rok", "")
        context["selected_rodzaj_autora"] = self.request.GET.get("rodzaj_autora", "")
        context["search"] = self.request.GET.get("search", "")
        context["current_sort"] = self.request.GET.get("sort", "autor")

        # Oblicz sumy dla przefiltrowanych danych (przed paginacją)
        queryset_for_sum = self.get_queryset()

        # Handle both queryset and list cases
        if hasattr(self, "_sorted_items") and isinstance(queryset_for_sum, list):
            # We have a sorted list from memory sorting
            context["suma_udzialow"] = sum(
                float(item.ilosc_udzialow) for item in queryset_for_sum
            )
            context["suma_monografie"] = sum(
                float(item.ilosc_udzialow_monografie) for item in queryset_for_sum
            )
            context["total_count"] = len(queryset_for_sum)
        else:
            # Normal queryset case
            sums = queryset_for_sum.aggregate(
                suma_udzialow=Sum("ilosc_udzialow"),
                suma_monografie=Sum("ilosc_udzialow_monografie"),
            )
            context["suma_udzialow"] = sums["suma_udzialow"] or 0
            context["suma_monografie"] = sums["suma_monografie"] or 0
            context["total_count"] = queryset_for_sum.count()

        # Add Autor_Dyscyplina data for each item
        for item in context["object_list"]:
            if item.autor_dyscyplina:
                item.wymiar_etatu = item.autor_dyscyplina.wymiar_etatu

                # Calculate wymiar_etatu_dla_dyscypliny
                if item.autor_dyscyplina.wymiar_etatu:
                    if (
                        item.dyscyplina_naukowa_id
                        == item.autor_dyscyplina.dyscyplina_naukowa_id
                    ):
                        item.wymiar_etatu_dla_dyscypliny = (
                            item.autor_dyscyplina.wymiar_etatu
                            * (item.autor_dyscyplina.procent_dyscypliny or Decimal("0"))
                            / Decimal("100")
                        )
                    elif (
                        item.dyscyplina_naukowa_id
                        == item.autor_dyscyplina.subdyscyplina_naukowa_id
                    ):
                        item.wymiar_etatu_dla_dyscypliny = (
                            item.autor_dyscyplina.wymiar_etatu
                            * (
                                item.autor_dyscyplina.procent_subdyscypliny
                                or Decimal("0")
                            )
                            / Decimal("100")
                        )
                    else:
                        item.wymiar_etatu_dla_dyscypliny = None
                else:
                    item.wymiar_etatu_dla_dyscypliny = None

                # Dodaj rodzaj autora
                item.rodzaj_autora = item.autor_dyscyplina.rodzaj_autora
            else:
                item.wymiar_etatu = None
                item.wymiar_etatu_dla_dyscypliny = None
                item.rodzaj_autora = None

        # Oblicz sumy dla bieżącej strony
        page_suma_udzialow = 0
        page_suma_monografie = 0
        for item in context["object_list"]:
            page_suma_udzialow += float(item.ilosc_udzialow)
            page_suma_monografie += float(item.ilosc_udzialow_monografie)

        context["page_suma_udzialow"] = page_suma_udzialow
        context["page_suma_monografie"] = page_suma_monografie

        return context


class UdzialyZaCaloscListView(GroupRequiredMixin, ListView):
    """Lista udziałów za cały okres ewaluacji"""

    template_name = "ewaluacja_liczba_n/udzialy_za_calosc_list.html"
    context_object_name = "udzialy_data"
    paginate_by = 50
    group_required = GR_WPROWADZANIE_DANYCH

    def get_queryset(self):
        # Pobierz wszystkie udziały za cały okres
        queryset = IloscUdzialowDlaAutoraZaCalosc.objects.all().select_related(
            "autor", "dyscyplina_naukowa", "rodzaj_autora"
        )

        # Filtrowanie po nazwisku/imieniu autora
        search = self.request.GET.get("search")
        if search:
            queryset = queryset.filter(
                Q(autor__nazwisko__icontains=search)
                | Q(autor__imiona__icontains=search)
            )

        # Filtrowanie po dyscyplinie
        dyscyplina_id = self.request.GET.get("dyscyplina")
        if dyscyplina_id:
            queryset = queryset.filter(dyscyplina_naukowa_id=dyscyplina_id)

        # Filtrowanie po rodzaju autora
        # Model IloscUdzialowDlaAutoraZaCalosc ma już pole rodzaj_autora
        rodzaj_autora_id = self.request.GET.get("rodzaj_autora")
        if rodzaj_autora_id:
            queryset = queryset.filter(rodzaj_autora_id=rodzaj_autora_id)

        # Sortowanie
        sort = self.request.GET.get("sort", "autor")

        # Mapowanie pól sortowania
        sort_mapping = {
            "autor": ["autor__nazwisko", "autor__imiona"],
            "-autor": ["-autor__nazwisko", "-autor__imiona"],
            "dyscyplina": ["dyscyplina_naukowa__nazwa"],
            "-dyscyplina": ["-dyscyplina_naukowa__nazwa"],
            "udzialy": ["ilosc_udzialow"],
            "-udzialy": ["-ilosc_udzialow"],
            "monografie": ["ilosc_udzialow_monografie"],
            "-monografie": ["-ilosc_udzialow_monografie"],
        }

        # Zastosuj sortowanie
        sort_fields = sort_mapping.get(sort, ["autor__nazwisko", "autor__imiona"])
        queryset = queryset.order_by(*sort_fields)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Dane do filtrowania - tylko dyscypliny które mają dane w bazie
        from bpp.models import Dyscyplina_Naukowa
        from ewaluacja_common.models import Rodzaj_Autora

        # Pobierz ID dyscyplin które mają faktyczne dane
        dyscypliny_z_danymi = IloscUdzialowDlaAutoraZaCalosc.objects.values_list(
            "dyscyplina_naukowa_id", flat=True
        ).distinct()

        # Filtruj tylko te dyscypliny które mają dane
        context["dyscypliny"] = Dyscyplina_Naukowa.objects.filter(
            pk__in=dyscypliny_z_danymi, widoczna=True
        ).order_by("nazwa")

        # Pobierz wszystkie rodzaje autorów
        context["rodzaje_autorow"] = Rodzaj_Autora.objects.all().order_by("sort")

        # Aktualne filtry i sortowanie
        context["selected_dyscyplina"] = self.request.GET.get("dyscyplina", "")
        context["selected_rodzaj_autora"] = self.request.GET.get("rodzaj_autora", "")
        context["search"] = self.request.GET.get("search", "")
        context["current_sort"] = self.request.GET.get("sort", "autor")

        # Oblicz sumy dla przefiltrowanych danych (przed paginacją)
        queryset_for_sum = self.get_queryset()
        sums = queryset_for_sum.aggregate(
            suma_udzialow=Sum("ilosc_udzialow"),
            suma_monografie=Sum("ilosc_udzialow_monografie"),
        )
        context["suma_udzialow"] = sums["suma_udzialow"] or 0
        context["suma_monografie"] = sums["suma_monografie"] or 0
        context["total_count"] = queryset_for_sum.count()

        # Oblicz sumy dla bieżącej strony
        page_suma_udzialow = 0
        page_suma_monografie = 0
        for item in context["object_list"]:
            page_suma_udzialow += float(item.ilosc_udzialow or 0)
            page_suma_monografie += float(item.ilosc_udzialow_monografie or 0)

        context["page_suma_udzialow"] = page_suma_udzialow
        context["page_suma_monografie"] = page_suma_monografie

        return context


class ExportAutorzyLiczbaNView(GroupRequiredMixin, View):
    """Eksport danych autorów do pliku XLSX"""

    group_required = GR_WPROWADZANIE_DANYCH

    def _create_summary_worksheet(self, wb, uczelnia):
        """Create summary worksheet with Liczba N data"""
        ws_summary = wb.active
        ws_summary.title = "Podsumowanie Liczba N"

        # Nagłówki
        headers = ["Dyscyplina", "Kod", "Liczba N"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        # Dane
        liczby_n = (
            LiczbaNDlaUczelni.objects.filter(uczelnia=uczelnia)
            .select_related("dyscyplina_naukowa")
            .order_by("dyscyplina_naukowa__nazwa")
        )

        row_num = 2
        suma = 0
        for liczba_n in liczby_n:
            ws_summary.cell(
                row=row_num, column=1, value=liczba_n.dyscyplina_naukowa.nazwa
            )
            ws_summary.cell(
                row=row_num, column=2, value=liczba_n.dyscyplina_naukowa.kod
            )
            ws_summary.cell(row=row_num, column=3, value=float(liczba_n.liczba_n))
            suma += float(liczba_n.liczba_n)
            row_num += 1

        # Suma
        ws_summary.cell(row=row_num, column=1, value="SUMA")
        ws_summary.cell(row=row_num, column=3, value=suma)
        ws_summary.cell(row=row_num, column=1).font = Font(bold=True)
        ws_summary.cell(row=row_num, column=3).font = Font(bold=True)

    def _get_filtered_udzialy_queryset(self, request):
        """Get filtered queryset based on request parameters"""
        udzialy = IloscUdzialowDlaAutoraZaRok.objects.filter(
            rok__gte=2022, rok__lte=2025
        )

        # Zastosuj filtry jeśli są w URL
        search = request.GET.get("search")
        if search:
            udzialy = udzialy.filter(
                Q(autor__nazwisko__icontains=search)
                | Q(autor__imiona__icontains=search)
            )

        dyscyplina_id = request.GET.get("dyscyplina")
        if dyscyplina_id:
            udzialy = udzialy.filter(dyscyplina_naukowa_id=dyscyplina_id)

        rok = request.GET.get("rok")
        if rok:
            udzialy = udzialy.filter(rok=rok)

        return udzialy.select_related(
            "autor",
            "dyscyplina_naukowa",
            "autor_dyscyplina",
            "autor_dyscyplina__rodzaj_autora",
        ).order_by(
            "autor__nazwisko", "autor__imiona", "rok", "dyscyplina_naukowa__nazwa"
        )

    def _calculate_wymiar_etatu_dla_dyscypliny(self, udzial, autor_dyscyplina):
        """Calculate wymiar etatu for discipline"""
        from decimal import Decimal

        if not autor_dyscyplina or not autor_dyscyplina.wymiar_etatu:
            return None

        if udzial.dyscyplina_naukowa_id == autor_dyscyplina.dyscyplina_naukowa_id:
            return float(
                autor_dyscyplina.wymiar_etatu
                * (autor_dyscyplina.procent_dyscypliny or Decimal("0"))
                / Decimal("100")
            )
        elif udzial.dyscyplina_naukowa_id == autor_dyscyplina.subdyscyplina_naukowa_id:
            return float(
                autor_dyscyplina.wymiar_etatu
                * (autor_dyscyplina.procent_subdyscypliny or Decimal("0"))
                / Decimal("100")
            )
        return None

    def _format_autorzy_detail_worksheet(self, ws_detail, row_num):
        """Apply formatting to autorzy detail worksheet"""
        if row_num <= 2:
            return

        # Define the table range
        table_range = f"A1:I{row_num - 1}"

        # Create table style
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        # Create the table
        table = Table(
            displayName="AutorzyLiczbaN", ref=table_range, tableStyleInfo=style
        )

        # Add table to worksheet
        ws_detail.add_table(table)

        # Freeze panes to keep headers visible
        ws_detail.freeze_panes = "A2"

        # Apply number formatting for decimal columns
        for row in range(2, row_num):
            # Column 5: Wymiar Etatu (decimal with 2 places)
            if ws_detail.cell(row=row, column=5).value is not None:
                ws_detail.cell(row=row, column=5).number_format = "0.00"

            # Column 6: Wymiar Etatu dla Dyscypliny (decimal with 4 places)
            if ws_detail.cell(row=row, column=6).value is not None:
                ws_detail.cell(row=row, column=6).number_format = "0.0000"

            # Column 7: Ilość Udziałów (decimal with 2 places)
            ws_detail.cell(row=row, column=7).number_format = "0.00"

            # Column 8: Ilość Udziałów - Monografie (decimal with 2 places)
            ws_detail.cell(row=row, column=8).number_format = "0.00"

    def _create_autorzy_detail_worksheet(self, wb, request):
        """Create detailed author data worksheet"""
        ws_detail = wb.create_sheet("Autorzy - Liczba N")

        # Nagłówki
        headers = [
            "Autor",
            "Rok",
            "Dyscyplina",
            "Kod Dyscypliny",
            "Wymiar Etatu",
            "Wymiar Etatu dla Dyscypliny",
            "Ilość Udziałów",
            "Ilość Udziałów - Monografie",
            "Rodzaj Autora",
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        # Dane autorów - z uwzględnieniem filtrów z URL
        udzialy = self._get_filtered_udzialy_queryset(request)

        row_num = 2
        for udzial in udzialy:
            autor_dyscyplina = udzial.autor_dyscyplina

            ws_detail.cell(row=row_num, column=1, value=str(udzial.autor))
            ws_detail.cell(row=row_num, column=2, value=udzial.rok)
            ws_detail.cell(row=row_num, column=3, value=udzial.dyscyplina_naukowa.nazwa)
            ws_detail.cell(row=row_num, column=4, value=udzial.dyscyplina_naukowa.kod)

            if autor_dyscyplina and autor_dyscyplina.wymiar_etatu:
                ws_detail.cell(
                    row=row_num,
                    column=5,
                    value=float(autor_dyscyplina.wymiar_etatu),
                )
                wymiar_dla_dyscypliny = self._calculate_wymiar_etatu_dla_dyscypliny(
                    udzial, autor_dyscyplina
                )
                if wymiar_dla_dyscypliny is not None:
                    ws_detail.cell(
                        row=row_num,
                        column=6,
                        value=wymiar_dla_dyscypliny,
                    )

            if autor_dyscyplina:
                ws_detail.cell(
                    row=row_num,
                    column=9,
                    value=(
                        autor_dyscyplina.rodzaj_autora.nazwa
                        if autor_dyscyplina.rodzaj_autora
                        else ""
                    ),
                )

            # Ilość udziałów
            ws_detail.cell(row=row_num, column=7, value=float(udzial.ilosc_udzialow))
            ws_detail.cell(
                row=row_num, column=8, value=float(udzial.ilosc_udzialow_monografie)
            )

            row_num += 1

        # Create Excel Table and apply formatting
        self._format_autorzy_detail_worksheet(ws_detail, row_num)

    def _create_nieraportowane_worksheet(self, wb, uczelnia):
        """Create non-reported disciplines worksheet"""
        ws_nieraportowane = wb.create_sheet("Dyscypliny Nieraportowane")

        headers = ["Dyscyplina", "Kod"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_nieraportowane.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        dyscypliny_nieraportowane = (
            DyscyplinaNieRaportowana.objects.filter(uczelnia=uczelnia)
            .select_related("dyscyplina_naukowa")
            .order_by("dyscyplina_naukowa__nazwa")
        )

        row_num = 2
        for dn in dyscypliny_nieraportowane:
            ws_nieraportowane.cell(
                row=row_num, column=1, value=dn.dyscyplina_naukowa.nazwa
            )
            ws_nieraportowane.cell(
                row=row_num, column=2, value=dn.dyscyplina_naukowa.kod
            )
            row_num += 1

    def _apply_column_widths(self, wb):
        """Adjust column widths for all worksheets"""
        for sheet in wb.worksheets:
            for column_cells in sheet.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = min(length + 2, 50)

    def _create_excel_response(self, wb, filename):
        """Create HTTP response with Excel file"""
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        # Zapisz do bufora
        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)
        response.write(virtual_workbook.getvalue())

        return response

    def get(self, request, *args, **kwargs):
        # Przygotuj dane
        uczelnia = Uczelnia.objects.get_default()

        # Utworzenie skoroszytu Excel
        wb = Workbook()

        # Arkusz 1: Podsumowanie liczby N dla uczelni
        self._create_summary_worksheet(wb, uczelnia)

        # Arkusz 2: Szczegółowe dane autorów
        self._create_autorzy_detail_worksheet(wb, request)

        # Arkusz 3: Dyscypliny nieraportowane
        self._create_nieraportowane_worksheet(wb, uczelnia)

        # Dostosuj szerokość kolumn
        self._apply_column_widths(wb)

        # Przygotuj odpowiedź
        return self._create_excel_response(wb, "liczba_n_ewaluacja_2022_2025.xlsx")


class ExportUdzialyZaCaloscView(GroupRequiredMixin, View):
    """Eksport danych udziałów za cały okres do pliku XLSX"""

    group_required = GR_WPROWADZANIE_DANYCH

    def _create_summary_worksheet(self, wb, uczelnia):
        """Create summary worksheet with Liczba N data"""
        ws_summary = wb.active
        ws_summary.title = "Podsumowanie Liczba N"

        headers = ["Dyscyplina", "Kod", "Liczba N"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        liczby_n = (
            LiczbaNDlaUczelni.objects.filter(uczelnia=uczelnia)
            .select_related("dyscyplina_naukowa")
            .order_by("dyscyplina_naukowa__nazwa")
        )

        row_num = 2
        suma = 0
        for liczba_n in liczby_n:
            ws_summary.cell(
                row=row_num, column=1, value=liczba_n.dyscyplina_naukowa.nazwa
            )
            ws_summary.cell(
                row=row_num, column=2, value=liczba_n.dyscyplina_naukowa.kod
            )
            ws_summary.cell(row=row_num, column=3, value=float(liczba_n.liczba_n))
            suma += float(liczba_n.liczba_n)
            row_num += 1

        ws_summary.cell(row=row_num, column=1, value="SUMA")
        ws_summary.cell(row=row_num, column=3, value=suma)
        ws_summary.cell(row=row_num, column=1).font = Font(bold=True)
        ws_summary.cell(row=row_num, column=3).font = Font(bold=True)

    def _get_filtered_udzialy_calosc_queryset(self, request):
        """Get filtered udzialy za calosc queryset"""
        udzialy = IloscUdzialowDlaAutoraZaCalosc.objects.all()

        search = request.GET.get("search")
        if search:
            udzialy = udzialy.filter(
                Q(autor__nazwisko__icontains=search)
                | Q(autor__imiona__icontains=search)
            )

        dyscyplina_id = request.GET.get("dyscyplina")
        if dyscyplina_id:
            udzialy = udzialy.filter(dyscyplina_naukowa_id=dyscyplina_id)

        rodzaj_autora_id = request.GET.get("rodzaj_autora")
        if rodzaj_autora_id:
            udzialy = udzialy.filter(rodzaj_autora_id=rodzaj_autora_id)

        return udzialy.select_related(
            "autor", "dyscyplina_naukowa", "rodzaj_autora"
        ).order_by("autor__nazwisko", "autor__imiona", "dyscyplina_naukowa__nazwa")

    def _format_calosc_detail_worksheet(self, ws_detail, row_num):
        """Apply formatting to full period detail worksheet"""
        if row_num <= 2:
            return

        table_range = f"A1:G{row_num - 1}"
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table = Table(
            displayName="UdzialyZaCalosc", ref=table_range, tableStyleInfo=style
        )
        ws_detail.add_table(table)
        ws_detail.freeze_panes = "A2"

        for row in range(2, row_num):
            if ws_detail.cell(row=row, column=5).value is not None:
                ws_detail.cell(row=row, column=5).number_format = "0.00"
            if ws_detail.cell(row=row, column=6).value is not None:
                ws_detail.cell(row=row, column=6).number_format = "0.00"
            if ws_detail.cell(row=row, column=7).value:
                ws_detail.cell(row=row, column=7).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

    def _create_calosc_detail_worksheet(self, wb, request):
        """Create full period detail worksheet"""
        ws_detail = wb.create_sheet("Udziały za cały okres")

        headers = [
            "Autor",
            "Dyscyplina",
            "Kod Dyscypliny",
            "Rodzaj autora",
            "Ilość Udziałów (2022-2025)",
            "Ilość Udziałów - Monografie",
            "Komentarz",
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        udzialy = self._get_filtered_udzialy_calosc_queryset(request)

        row_num = 2
        for udzial in udzialy:
            ws_detail.cell(row=row_num, column=1, value=str(udzial.autor))
            ws_detail.cell(row=row_num, column=2, value=udzial.dyscyplina_naukowa.nazwa)
            ws_detail.cell(row=row_num, column=3, value=udzial.dyscyplina_naukowa.kod)
            ws_detail.cell(
                row=row_num,
                column=4,
                value=(
                    udzial.rodzaj_autora.nazwa
                    if udzial.rodzaj_autora
                    else "Brak danych"
                ),
            )
            ws_detail.cell(row=row_num, column=5, value=float(udzial.ilosc_udzialow))
            ws_detail.cell(
                row=row_num, column=6, value=float(udzial.ilosc_udzialow_monografie)
            )
            komentarz_text = (udzial.komentarz or "").replace("<br>", "\n")
            ws_detail.cell(row=row_num, column=7, value=komentarz_text)
            row_num += 1

        self._format_calosc_detail_worksheet(ws_detail, row_num)

    def _create_nieraportowane_worksheet(self, wb, uczelnia):
        """Create non-reported disciplines worksheet"""
        ws_nieraportowane = wb.create_sheet("Dyscypliny Nieraportowane")

        headers = ["Dyscyplina", "Kod", "Liczba N"]
        for col_num, header in enumerate(headers, 1):
            cell = ws_nieraportowane.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )

        dyscypliny_nieraportowane = (
            DyscyplinaNieRaportowana.objects.filter(uczelnia=uczelnia)
            .select_related("dyscyplina_naukowa")
            .order_by("dyscyplina_naukowa__nazwa")
        )

        row_num = 2
        for dn in dyscypliny_nieraportowane:
            ws_nieraportowane.cell(
                row=row_num, column=1, value=dn.dyscyplina_naukowa.nazwa
            )
            ws_nieraportowane.cell(
                row=row_num, column=2, value=dn.dyscyplina_naukowa.kod
            )
            ws_nieraportowane.cell(row=row_num, column=3, value=float(dn.liczba_n))
            row_num += 1

    def _apply_column_widths(self, wb):
        """Adjust column widths for all worksheets"""
        for sheet in wb.worksheets:
            for column_cells in sheet.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                sheet.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = min(length + 2, 50)

    def _create_excel_response(self, wb, filename):
        """Create HTTP response with Excel file"""
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)
        response.write(virtual_workbook.getvalue())

        return response

    def get(self, request, *args, **kwargs):
        uczelnia = Uczelnia.objects.get_default()
        wb = Workbook()

        self._create_summary_worksheet(wb, uczelnia)
        self._create_calosc_detail_worksheet(wb, request)
        self._create_nieraportowane_worksheet(wb, uczelnia)
        self._apply_column_widths(wb)

        return self._create_excel_response(
            wb, "udzialy_za_calosc_ewaluacja_2022_2025.xlsx"
        )


class WeryfikujBazeView(GroupRequiredMixin, TemplateView):
    """Widok weryfikacji poprawności bazy danych"""

    template_name = "ewaluacja_liczba_n/weryfikuj_baze.html"
    group_required = GR_WPROWADZANIE_DANYCH

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # 1. Total by rodzaj_pracownika for 2022-2025
        from ewaluacja_common.models import Rodzaj_Autora

        context["rodzaje_pracownika"] = (
            Autor_Dyscyplina.objects.filter(rok__gte=2022, rok__lte=2025)
            .values("rodzaj_autora")
            .annotate(liczba=Count("id"))
            .order_by("rodzaj_autora")
        )

        # Pobierz obiekty Rodzaj_Autora dla mapowania
        rodzaje_dict = {r.id: r for r in Rodzaj_Autora.objects.all()}

        for item in context["rodzaje_pracownika"]:
            if item["rodzaj_autora"] is not None:
                rodzaj_obj = rodzaje_dict.get(item["rodzaj_autora"])
                if rodzaj_obj:
                    item["nazwa"] = rodzaj_obj.nazwa
                    # Mapowanie na query_key dla kompatybilności
                    skrot_mapping = {
                        "N": "rodzaj_n",
                        "D": "rodzaj_d",
                        "B": "rodzaj_b",
                        "Z": "rodzaj_z",
                    }
                    item["query_key"] = skrot_mapping.get(
                        rodzaj_obj.skrot, "brak_danych"
                    )
                else:
                    item["nazwa"] = "brak danych"
                    item["query_key"] = "brak_danych"
            else:
                item["nazwa"] = "brak danych"
                item["query_key"] = "brak_danych"

        # Count records without rodzaj_autora (empty/missing employment type)
        # Get IDs of known rodzaj_autora types
        known_rodzaje_ids = [
            r.id for r in rodzaje_dict.values() if r.skrot in ["N", "D", "B", "Z"]
        ]

        context["bez_rodzaju_zatrudnienia"] = (
            Autor_Dyscyplina.objects.filter(
                rok__gte=2022,
                rok__lte=2025,
            )
            .exclude(rodzaj_autora__in=known_rodzaje_ids)
            .count()
        )

        # 2. Records without wymiar_etatu
        context["bez_wymiaru_etatu"] = (
            Autor_Dyscyplina.objects.filter(rok__gte=2022, rok__lte=2025)
            .filter(Q(wymiar_etatu__isnull=True) | Q(wymiar_etatu=0))
            .select_related("autor")
            .count()
        )

        # 3. Records with missing percentage information
        # Missing procent_dyscypliny OR (has subdyscyplina but missing procent_subdyscypliny)
        context["bez_procent"] = (
            Autor_Dyscyplina.objects.filter(rok__gte=2022, rok__lte=2025)
            .filter(
                Q(procent_dyscypliny__isnull=True)  # Missing main discipline percentage
                # Has subdiscipline but no percentage
                | (
                    Q(subdyscyplina_naukowa__isnull=False)
                    & Q(procent_subdyscypliny__isnull=True)
                )
            )
            .count()
        )

        # 4. Check if sum of percentages equals 100
        # We need to check each record individually
        from decimal import Decimal

        problematic_suma = []
        all_records = Autor_Dyscyplina.objects.filter(
            rok__gte=2022, rok__lte=2025
        ).select_related("autor")

        for record in all_records:
            procent_d = record.procent_dyscypliny or Decimal("0")
            procent_s = record.procent_subdyscypliny or Decimal("0")
            suma = procent_d + procent_s

            # Check if sum is not 100 (allowing small rounding differences)
            if suma > Decimal("0") and abs(suma - Decimal("100")) > Decimal("0.01"):
                problematic_suma.append(
                    {
                        "id": record.id,
                        "autor": str(record.autor),
                        "rok": record.rok,
                        "suma": float(suma),
                    }
                )

        context["zla_suma_procent"] = len(problematic_suma)
        context["zla_suma_procent_przykłady"] = problematic_suma[
            :5
        ]  # Show first 5 examples

        # Calculate total count of records
        context["total_count"] = sum(
            item["liczba"] for item in context["rodzaje_pracownika"]
        )

        # Calculate distinct number of authors
        context["distinct_authors_count"] = (
            Autor_Dyscyplina.objects.filter(rok__gte=2022, rok__lte=2025)
            .values("autor")
            .distinct()
            .count()
        )

        # Generate DjangoQL queries
        # DjangoQL needs to reference the related object now
        context["djangoql_queries"] = {
            "bez_wymiaru": "rok >= 2022 and rok <= 2025 and (wymiar_etatu = None or wymiar_etatu = 0)",
            "bez_procent": "rok >= 2022 and rok <= 2025 and (procent_dyscypliny = None or (subdyscyplina_naukowa != "
            "None and procent_subdyscypliny = None))",
            "zla_suma": "rok >= 2022 and rok <= 2025 and procent_dyscypliny != None",  # Can't check sum with DjangoQL
            "rodzaj_n": 'rok >= 2022 and rok <= 2025 and rodzaj_autora.skrot = "N"',
            "rodzaj_d": 'rok >= 2022 and rok <= 2025 and rodzaj_autora.skrot = "D"',
            "rodzaj_b": 'rok >= 2022 and rok <= 2025 and rodzaj_autora.skrot = "B"',
            "rodzaj_z": 'rok >= 2022 and rok <= 2025 and rodzaj_autora.skrot = "Z"',
            "brak_danych": "rok >= 2022 and rok <= 2025 and rodzaj_autora = None",
        }

        return context
