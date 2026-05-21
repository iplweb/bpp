import openpyxl.worksheet.worksheet
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo


def _extract_hyperlink_text(text):
    """Extract display text from hyperlink formula."""
    if text.startswith("=HYPERLINK"):
        try:
            # Wyciągnij z hiperlinku jego faktyczny opis tekstowy na cele
            # liczenia szerokości kolumny
            return text.split('"')[3]
        except IndexError:
            pass
    return text


def _calculate_column_width(col, right_margin, multiplier, max_width):
    """Calculate optimal width for a column based on its content."""
    max_length = 0

    for cell in col:
        if cell.value is None or not str(cell.value):
            continue

        text = str(cell.value)
        text = _extract_hyperlink_text(text)

        max_line_len = max(len(line) for line in text.split("\n"))
        max_length = max(max_length, max_line_len)

    adjusted_width = (max_length + right_margin) * multiplier
    if adjusted_width > max_width:
        adjusted_width = max_width

    return adjusted_width


def worksheet_columns_autosize(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    max_width: int = 55,
    column_widths: dict[str, int] | None = None,
    dont_resize_those_columns: list[int] | None = None,
    right_margin=2,
    multiplier=1.1,
):
    if column_widths is None:
        column_widths = {}

    if dont_resize_those_columns is None:
        dont_resize_those_columns = []

    for ncol, col in enumerate(ws.columns):
        column = col[0].column_letter  # Get the column name

        # Nie ustawiaj szerokosci tym kolumnom, one będą jako auto-size
        if ncol in dont_resize_those_columns:
            continue

        if column in column_widths:
            adjusted_width = column_widths[column]
        else:
            adjusted_width = _calculate_column_width(
                col, right_margin, multiplier, max_width
            )

        ws.column_dimensions[column].width = adjusted_width


# Znaki, którymi zaczynający się tekst Excel/LibreOffice interpretuje
# jako formułę (=, +, -, @) lub jako wstrzyknięcie do innej komórki/komendy
# poprzez separator (Tab, CR, LF). Pełna lista zaleceń OWASP CSV/Formula
# Injection: https://owasp.org/www-community/attacks/CSV_Injection
_XLSX_FORMULA_INJECTION_LEAD = ("=", "+", "-", "@", "\t", "\r", "\n")


def sanitize_xlsx_cell(value):
    """Zwraca wartość bezpieczną do `ws.append()` / `ws.cell().value`.

    Stringi zaczynające się od znaków interpretowanych jako formuła
    (=, +, -, @) lub separator (Tab/CR/LF) są poprzedzane apostrofem,
    co Excel traktuje jako wymuszenie typu „tekst" (apostrof nie jest
    pokazywany w komórce). Pozostałe wartości (None, liczby, daty itd.)
    są zwracane bez zmian.
    """
    if not isinstance(value, str):
        return value
    if value.startswith(_XLSX_FORMULA_INJECTION_LEAD):
        return "'" + value
    return value


def sanitize_xlsx_row(row):
    """Wersja `sanitize_xlsx_cell` dla całego wiersza (`ws.append(row)`)."""
    return [sanitize_xlsx_cell(c) for c in row]


def auto_fit_columns(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    max_width: int = 50,
    padding: int = 2,
):
    """Prosta auto-szerokość kolumn: longest cell value + padding, capped at max_width.

    Wcześniej duplikowane inline w 5 miejscach (eksporty XLSX z auto-fit). Dla
    zaawansowanej wersji z hyperlinkami i multiplier patrz worksheet_columns_autosize.
    """
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                length = len(str(cell.value))
            except (TypeError, ValueError):
                continue
            if length > max_length:
                max_length = length
        ws.column_dimensions[column_letter].width = min(max_length + padding, max_width)


def worksheet_create_table(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    title="Tabela",
    first_table_row=1,
    totals=False,
    table_columns=None,
):
    """
    Formatuje skoroszyt jako tabelę.

    :param first_table_row: pierwszy wiersz tabeli (licząc od nagłówka)

    :param table_columns: określa rodzaj kolumn w tabeli, jeżeli None to tytuły nagłówków zostaną pobrane
    z pierwszego wiersza w arkuszu.
    """
    max_column = ws.max_column
    max_column_letter = get_column_letter(max_column)
    max_row = ws.max_row

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )

    if table_columns is None:
        table_columns = tuple(
            TableColumn(id=h, name=header.value)
            for h, header in enumerate(next(iter(ws.rows), None), start=1)
        )

    tab = Table(
        displayName=title,
        ref=f"A{first_table_row}:{max_column_letter}{max_row}",
        autoFilter=AutoFilter(
            ref=f"A{first_table_row}:{max_column_letter}{max_row - 1}"
        ),
        totalsRowShown=True if totals else False,
        totalsRowCount=1 if totals else False,
        tableStyleInfo=style,
        tableColumns=table_columns,
    )

    ws.add_table(tab)


def worksheet_create_urls(
    ws: openpyxl.worksheet.worksheet.Worksheet, default_link_name: str = "[link]"
):
    """Tworzy adresy URL w postaci klikalnego linku z domyslnym tekstem."""

    for column_cell in ws.iter_cols(1, ws.max_column):  # iterate column cell
        if hasattr(column_cell[0].value, "endswith") and column_cell[0].value.endswith(
            "_url"
        ):
            for data in column_cell[1:]:
                if data.value:
                    data.value = f'=HYPERLINK("{data.value}", "{default_link_name}")'
