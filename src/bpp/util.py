import json
import multiprocessing
import os
import re
import sys
from datetime import datetime, timedelta
from math import ceil, floor
from pathlib import Path
from typing import Dict, List

import bleach
import lxml.html
import openpyxl.worksheet.worksheet
import progressbar
from django.apps import apps
from django.conf import settings
from django.db.models import Max, Min, Value
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo
from psycopg2.extensions import QuotedString
from unidecode import unidecode

from django.contrib.postgres.search import SearchQuery, SearchRank

from django.utils import timezone
from django.utils.html import strip_tags

non_url = re.compile(r"[^\w-]+")


def get_fixture(name):
    p = Path(__file__).parent / "fixtures" / ("%s.json" % name)
    ret = json.load(open(p, "rb"))
    ret = [x["fields"] for x in ret if x["model"] == ("bpp.%s" % name)]
    return {x["skrot"].lower().strip(): x for x in ret}


def fulltext_tokenize(s):
    s = (
        s.replace(":", " ")
        .replace("*", " ")
        .replace('"', " ")
        .replace("|", " ")
        .replace("'", " ")
        .replace("&", " ")
        .replace("\\", " ")
        .replace("(", " ")
        .replace(")", " ")
        .replace("#", " ")
        .replace("@", " ")
        .replace("!", " ")
        .replace("[", " ")
        .replace("]", " ")
        .replace("\t", " ")
        .replace("\n", " ")
        .replace("\r", " ")
        .replace("<", " ")
        .replace(">", " ")
    )
    return [x.strip() for x in s.strip().split(" ") if x.strip()]


class FulltextSearchMixin:
    fts_field = "search"

    def fulltext_filter(self, qstr):
        def quotes(wordlist):
            ret = []
            for elem in wordlist:
                ret.append(str(QuotedString(elem.replace("\\", "").encode("utf-8"))))
            return ret

        def startswith(wordlist):
            return [x + ":*" for x in quotes(wordlist)]

        def negative(wordlist):
            return ["!" + x for x in startswith(wordlist)]

        if qstr is None:
            qstr = ""

        if isinstance(qstr, bytes):
            qstr = qstr.decode("utf-8")

        clean_qstr = strip_tags(qstr)
        words = fulltext_tokenize(clean_qstr)
        if not words:
            return self.none().annotate(**{self.fts_field + "__rank": Value(0)})

        qstr = "(" + " & ".join(startswith(words)) + ") | (" + " & ".join(words) + ")"
        sq = SearchQuery(qstr, search_type="raw", config="bpp_nazwy_wlasne")

        return (
            self.filter(**{self.fts_field: sq})
            .annotate(**{self.fts_field + "__rank": SearchRank(self.fts_field, sq)})
            .order_by(f"-{self.fts_field}__rank")
        )


def strip_html(s):
    if not s:
        return s

    return lxml.html.fromstring(str(s)).text_content()


def slugify_function(s):
    s = unidecode(strip_html(s)).replace(" ", "-")
    while s.find("--") >= 0:
        s = s.replace("--", "-")
    return non_url.sub("", s)


def get_original_object(object_name, object_pk):
    from bpp.models import TABLE_TO_MODEL

    klass = TABLE_TO_MODEL.get(object_name)
    try:
        return klass.objects.get(pk=object_pk)
    except klass.DoesNotExist:
        return


def get_copy_from_db(instance):
    if not instance.pk:
        return None
    return instance.__class__._default_manager.get(pk=instance.pk)


def has_changed(instance, field_or_fields):
    try:
        original = get_copy_from_db(instance)
    except instance.__class__.DoesNotExist:
        return True
        # Jeżeli w bazie danych nie ma tego obiektu, no to bankowo
        # się zmienił...
        return True

    fields = field_or_fields
    if isinstance(field_or_fields, str):
        fields = [field_or_fields]

    for field in fields:
        if not getattr(instance, field) == getattr(original, field):
            return True

    return False


class Getter:
    """Klasa pomocnicza dla takich danych jak Typ_KBN czy
    Charakter_Formalny, umozliwia po zainicjowaniu pobieranie
    tych klas po atrybucie w taki sposob:

    >>> kbn = Getter(Typ_KBN)
    >>> kbn.PO == Typ_KBN.objects.get(skrot='PO')
    True
    """

    def __init__(self, klass, field="skrot"):
        self.field = field
        self.klass = klass

    def __getitem__(self, item):
        kw = {self.field: item}
        return self.klass.objects.get(**kw)

    __getattr__ = __getitem__


class NewGetter(Getter):
    """Zwraca KeyError zamiast DoesNotExist."""

    def __getitem__(self, item):
        kw = {self.field: item}
        try:
            return self.klass.objects.get(**kw)
        except self.klass.DoesNotExist as e:
            raise KeyError(e)

    __getattr__ = __getitem__


def zrob_cache(t):
    zle_znaki = [
        " ",
        ":",
        ";",
        "-",
        ",",
        "-",
        ".",
        "(",
        ")",
        "?",
        "!",
        "ę",
        "ą",
        "ł",
        "ń",
        "ó",
        "ź",
        "ż",
    ]
    for znak in zle_znaki:
        t = t.replace(znak, "")
    return t.lower()


def remove_old_objects(klass, file_field="file", field_name="created_on", days=7):
    since = datetime.now() - timedelta(days=days)

    kwargs = {}
    kwargs["%s__lt" % field_name] = since

    for rec in klass.objects.filter(**kwargs):
        try:
            path = getattr(rec, file_field).path
        except ValueError:
            path = None

        rec.delete()

        if path is not None:
            try:
                os.unlink(path)
            except OSError:
                pass


def rebuild_contenttypes():
    app_config = apps.get_app_config("bpp")
    from django.contrib.contenttypes.management import create_contenttypes

    create_contenttypes(app_config, verbosity=0)


#
# Progress bar
#


def pbar(query, count=None, label="Progres...", disable_progress_bar=False):
    if sys.stdout.isatty() and not disable_progress_bar:
        if count is None:
            if hasattr(query, "count"):
                try:
                    count = query.count()
                except TypeError:
                    count = len(query)
            elif hasattr(query, "__len__"):
                count = len(query)

        return progressbar.progressbar(
            query,
            max_value=count,
            widgets=[
                progressbar.FormatLabel(label),
                " ",
                progressbar.AnimatedMarker(),
                " ",
                progressbar.SimpleProgress(),
                " ",
                progressbar.Timer(),
                " ",
                progressbar.ETA(),
            ],
        )
    else:
        # You're being piped or redirected
        return query


#
# Multiprocessing stuff
#


def partition(min, max, num_proc, fun=ceil):
    s = int(fun((max - min) / num_proc))
    cnt = min
    ret = []
    while cnt < max:
        ret.append((cnt, cnt + s))
        cnt += s
    return ret


def partition_ids(model, num_proc, attr="idt"):
    d = model.objects.aggregate(min=Min(attr), max=Max(attr))
    return partition(d["min"], d["max"], num_proc)


def partition_count(objects, num_proc):
    return partition(0, objects.count(), num_proc, fun=ceil)


def no_threads(multiplier=0.75):
    return max(int(floor(multiprocessing.cpu_count() * multiplier)), 1)


class safe_html_defaults:
    ALLOWED_TAGS = (
        "a",
        "abbr",
        "acronym",
        "b",
        "blockquote",
        "code",
        "em",
        "i",
        "li",
        "ol",
        "strong",
        "ul",
        "font",
        "div",
        "span",
        "br",
        "strike",
        "h2",
        "h3",
        "h4",
        "h5",
        "h6",
        "p",
        "table",
        "tr",
        "td",
        "th",
        "thead",
        "tbody",
        "dl",
        "dd",
        "u",
    )

    ALLOWED_ATTRIBUTES = {
        "*": ["class"],
        "a": ["href", "title", "rel"],
        "abbr": ["title"],
        "acronym": ["title"],
        "font": [
            "face",
            "size",
        ],
        "div": [
            "style",
        ],
        "span": [
            "style",
        ],
        "ul": [
            "style",
        ],
    }


def safe_html(html):
    html = html or ""

    ALLOWED_TAGS = getattr(settings, "ALLOWED_TAGS", safe_html_defaults.ALLOWED_TAGS)
    ALLOWED_ATTRIBUTES = getattr(
        settings, "ALLOWED_ATTRIBUTES", safe_html_defaults.ALLOWED_ATTRIBUTES
    )
    STRIP_TAGS = getattr(settings, "STRIP_TAGS", True)
    return bleach.clean(
        html,
        tags=ALLOWED_TAGS,
        attributes=ALLOWED_ATTRIBUTES,
        strip=STRIP_TAGS,
    )


def set_seq(s):
    if settings.DATABASES["default"]["ENGINE"].find("postgresql") >= 0:
        from django.db import connection

        cursor = connection.cursor()
        cursor.execute(f"SELECT setval('{s}_id_seq', (SELECT MAX(id) FROM {s}))")


def usun_nieuzywany_typ_charakter(klass, field, dry_run):
    from bpp.models import Rekord

    for elem in klass.objects.all():
        kw = {field: elem}
        if not Rekord.objects.filter(**kw).exists():
            print(f"Kasuje {elem}")
            if not dry_run:
                elem.delete()


isbn_regex = re.compile(
    r"^isbn\s*[0-9]*[-| ][0-9]*[-| ][0-9]*[-| ][0-9]*[-| ][0-9]*X?",
    flags=re.IGNORECASE,
)


def wytnij_isbn_z_uwag(uwagi):
    if uwagi is None:
        return

    if uwagi == "":
        return

    if uwagi.lower().find("isbn-10") >= 0 or uwagi.lower().find("isbn-13") >= 0:
        return None

    res = isbn_regex.search(uwagi)
    if res:
        res = res.group()
        isbn = res.replace("ISBN", "").replace("isbn", "").strip()
        reszta = uwagi.replace(res, "").strip()

        while (
            reszta.startswith(".") or reszta.startswith(";") or reszta.startswith(",")
        ):
            reszta = reszta[1:].strip()

        return isbn, reszta


def crispy_form_html(self, key):
    from crispy_forms_foundation.layout import HTML, Column, Row

    from django.utils.functional import lazy

    def _():
        return self.initial.get(key, None) or ""

    return Row(Column(HTML(lazy(_, str)())))


def formdefaults_html_before(form):
    return crispy_form_html(form, "formdefaults_pre_html")


def formdefaults_html_after(form):
    return crispy_form_html(form, "formdefaults_post_html")


def knapsack(W, wt, val, ids, zwracaj_liste_przedmiotow=True):
    """
    :param W: wielkosc plecaka -- maksymalna masa przedmiotów w plecaku (zbierany slot)
    :param wt: masy przedmiotów, które można włożyć do plecaka (sloty prac)
    :param val: ceny przedmiotów, które można włożyc do plecaka (punkty PKdAut prac)
    :param ids: ID prac, które można włożyć do plecaka (rekord.pk)
    :param zwracaj_liste_przedmiotow: gdy True (domyślnie) funkcja zwróci listę z identyfikatorami włożonych
    przedmiotów, gdy False zwrócona lista będzie pusta

    :returns: tuple(mp, lista), gdzie mp to maksymalna możliwa wartość włożonych przedmiotów, a lista to lista
    lub pusta lista gdy parametr `zwracaj_liste_przemiotów` był pusty
    """

    assert len(wt) == len(val) == len(ids), "Listy są różnej długości"

    sum_wt = sum(wt)
    if sum_wt <= W:
        # Jeżeli wszystkie przedmioty zmieszczą się w plecaku, to po co liczyć cokolwiek
        if zwracaj_liste_przedmiotow:
            return sum(val), ids
        return sum(val), []

    n = len(wt)

    K = [[0 for x in range(W + 1)] for x in range(n + 1)]

    for i in range(n + 1):
        for w in range(W + 1):
            if i == 0 or w == 0:
                K[i][w] = 0
            elif wt[i - 1] <= w:
                K[i][w] = max(val[i - 1] + K[i - 1][w - wt[i - 1]], K[i - 1][w])
            else:
                K[i][w] = K[i - 1][w]

    res = maks_punkty = K[n][W]
    lista = []

    if zwracaj_liste_przedmiotow:
        w = W
        for i in range(n, 0, -1):
            if res <= 0:
                break

            if res == K[i - 1][w]:
                continue
            else:
                lista.append(ids[i - 1])

                res = res - val[i - 1]
                w = w - wt[i - 1]

    return maks_punkty, lista


DEC2INT = 10000


def intsack(W, wt, val, ids):
    pkt, ids = knapsack(
        int(W * DEC2INT),
        [int(x * DEC2INT) for x in wt],
        [int(x * DEC2INT) for x in val],
        ids,
    )
    return pkt / DEC2INT, ids


def disable_multithreading_by_monkeypatching_pool(pool):
    def apply(fun, args=()):
        return fun(*args)

    pool.apply = apply

    def starmap(fun, lst):
        for elem in lst:
            fun(*elem)

    pool.starmap = starmap


def year_last_month():
    now = timezone.now().date()
    if now.month >= 2:
        return now.year
    return now.year - 1


#
# Seq Scan check
#


class PerformanceFailure(Exception):
    pass


def fail_if_seq_scan(qset, DEBUG):
    """
    Funkcja weryfikujaca, czy w wyjasnieniu zapytania (EXPLAIN) nie wystapi ciag znakow 'Seq Scan',
    jezeli tak to wyjatek PerformanceFailure z zapytaniem + wyjasnieniem
    """
    if DEBUG:
        explain = qset.explain()
        if explain.find("Seq Scan") >= 0:
            print("\r\n", explain)
            raise PerformanceFailure(str(qset.query), explain)


def rebuild_instances_of_models(modele, *args, **kw):
    from denorm import denorms
    from django.db import transaction

    with transaction.atomic():
        for model in modele:
            denorms.rebuild_instances_of(model, *args, **kw)


def worksheet_columns_autosize(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    max_width: int = 55,
    column_widths: Dict[str, int] | None = None,
    dont_resize_those_columns: List[int] | None = None,
    right_margin=2,
    multiplier=1.1,
):

    if column_widths is None:
        column_widths = {}

    if dont_resize_those_columns is None:
        dont_resize_those_columns = []

    for ncol, col in enumerate(ws.columns):
        max_length = 0
        column = col[0].column_letter  # Get the column name

        # Nie ustawiaj szerokosci tym kolumnom, one będą jako auto-size
        if ncol in dont_resize_those_columns:
            continue

        if column in column_widths:
            adjusted_width = column_widths[column]
        else:
            for cell in col:
                if cell.value is None or not str(cell.value):
                    continue

                text = str(cell.value)

                if text.startswith("=HYPERLINK"):
                    try:
                        # Wyciągnij z hiperlinku jego faktyczny opis tekstowy na cele
                        # liczenia szerokości kolumny
                        text = text.split('"')[3]
                    except IndexError:
                        pass

                max_line_len = max(len(line) for line in text.split("\n"))
                max_length = max(max_length, max_line_len)

            adjusted_width = (max_length + right_margin) * multiplier
            if adjusted_width > max_width:
                adjusted_width = max_width

        ws.column_dimensions[column].width = adjusted_width


def worksheet_create_table(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    title="Tabela",
    first_table_row=1,
    totals=False,
    table_columns=None,
):
    """
    Formatuje skoroszyt jako tabelę.

    :param first_table_row: pierwszy wiersz tabeli (licząc od nagłówka)

    :param table_columns: określa rodzaj kolumn w tabeli, jeżeli None to tytuły nagłówków zostaną pobrane
    z pierwszego wiersza w arkuszu.
    """
    max_column = ws.max_column
    max_column_letter = get_column_letter(max_column)
    max_row = ws.max_row

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )

    if table_columns is None:
        table_columns = tuple(
            TableColumn(id=h, name=header.value)
            for h, header in enumerate(next(iter(ws.rows), None), start=1)
        )

    tab = Table(
        displayName=title,
        ref=f"A{first_table_row}:{max_column_letter}{max_row}",
        autoFilter=AutoFilter(
            ref=f"A{first_table_row}:{max_column_letter}{max_row - 1}"
        ),
        totalsRowShown=True if totals else False,
        totalsRowCount=1 if totals else False,
        tableStyleInfo=style,
        tableColumns=table_columns,
    )

    ws.add_table(tab)


def worksheet_create_urls(
    ws: openpyxl.worksheet.worksheet.Worksheet, default_link_name: str = "[link]"
):
    """Tworzy adresy URL w postaci klikalnego linku z domyslnym tekstem."""

    for column_cell in ws.iter_cols(1, ws.max_column):  # iterate column cell
        if hasattr(column_cell[0].value, "endswith") and column_cell[0].value.endswith(
            "_url"
        ):
            for data in column_cell[1:]:
                if data.value:
                    data.value = '=HYPERLINK("{}", "{}")'.format(
                        data.value, default_link_name
                    )
