"""Serializacja wyników Multiseek do plików eksportu (CSV / XLSX).

Wydzielone z bpp/views/mymultiseek.py — widok trzyma routing i stan sesji,
ten moduł zamienia queryset Rekordów na gotową odpowiedź HTTP z plikiem.
"""

import csv
import html
import io
import re

from django.http import HttpResponse
from django.urls import reverse
from django.utils.html import strip_tags
from django.utils.http import content_disposition_header

from bpp import const
from bpp.models import Uczelnia

MULTISEEK_DEFAULT_REPORT_TITLE = "Rezultat wyszukiwania"
XLSX_WORKSHEET_TITLE_MAX_LENGTH = 31

MULTISEEK_EXPORT_HEADERS = (
    "tytul_oryginalny",
    "autorzy",
    "zrodlo",
    "rok",
    "impact_factor",
    "pk",
    "bpp_id",
    "typ_rekordu",
    "typ_mnisw_mein",
    "id_rekordu",
    "pbn_uid_id",
    "link_do_bpp_url",
    "link_do_bpp_admin_url",
    "link_do_pbn_url",
)

MULTISEEK_EXPORT_XLSX_HEADERS = (
    "Tytuł oryginalny",
    "Autorzy",
    "Źródło",
    "Rok",
    "Impact Factor",
    "PK",
    "BPP ID",
    "Typ rekordu",
    "Typ MNiSW/MEiN",
    "ID rekordu",
    "PBN UID",
    "Link do BPP",
    "Link do edycji w BPP",
    "Link do PBN",
)

MULTISEEK_EXPORT_DANE_FIELDS = (
    "id",
    "tytul_oryginalny",
    "opis_bibliograficzny_zapisani_autorzy_cache",
    "zrodlo__nazwa",
    "rok",
    "impact_factor",
    "punkty_kbn",
    "typ_kbn__nazwa",
    "pbn_uid_id",
)

MULTISEEK_EXPORT_OPIS_XLSX_HEADERS = (
    "Lp.",
    "Opis bibliograficzny",
    "IF",
    "PK",
    "Charakter",
    "Typ MNiSW/MEiN",
)

MULTISEEK_EXPORT_OPIS_FIELDS = (
    "id",
    "opis_bibliograficzny_cache",
    "impact_factor",
    "punkty_kbn",
    "charakter_formalny__nazwa",
    "typ_kbn__nazwa",
)

EXPORT_FILENAME_INVALID_CHARS_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
MULTISEEK_REPORT_TITLE_HTML_BREAK_RE = re.compile(
    r"</?(?:br|hr|p|div|h[1-6])\b[^>]*>",
    re.IGNORECASE,
)
XLSX_WORKSHEET_TITLE_INVALID_CHARS_RE = re.compile(
    r"[\[\]:*?/\\\x00-\x08\x0b\x0c\x0e-\x1f]"
)
SPREADSHEET_FORMULA_INJECTION_LEAD = ("=", "+", "-", "@", "\t", "\r", "\n")


def _export_value(value):
    if value is None:
        return ""
    return str(value)


def _single_line_text(value):
    return re.sub(r"\s+", " ", value).strip()


def plain_multiseek_report_title(value):
    value = _export_value(value)
    value = MULTISEEK_REPORT_TITLE_HTML_BREAK_RE.sub(" ", value)
    value = html.unescape(strip_tags(value))
    return _single_line_text(value) or MULTISEEK_DEFAULT_REPORT_TITLE


def _plain_opis_bibliograficzny(value):
    """opis_bibliograficzny_cache (HTML) -> jednoliniowy czysty tekst.

    Bez fallbacku na tytuł domyślny (w przeciwieństwie do
    plain_multiseek_report_title): puste wejście -> "". Sanityzacja formuł
    dzieje się później, w sanitize_xlsx_row — tu jej NIE wołamy.
    """
    if not value:
        return ""
    value = MULTISEEK_REPORT_TITLE_HTML_BREAK_RE.sub(" ", value)
    value = html.unescape(strip_tags(value))
    return _single_line_text(value)


def _export_filename(export_format, report_title):
    title = EXPORT_FILENAME_INVALID_CHARS_RE.sub(" ", report_title)
    title = _single_line_text(title).strip(". ")
    if not title:
        title = "multiseek"
    return f"eksport-{title}.{export_format}"


def _xlsx_worksheet_title(report_title):
    title = XLSX_WORKSHEET_TITLE_INVALID_CHARS_RE.sub(" ", report_title)
    title = _single_line_text(title).strip("'")
    if not title:
        title = "Multiseek"
    return title[:XLSX_WORKSHEET_TITLE_MAX_LENGTH]


def _sanitize_spreadsheet_cell(value):
    if isinstance(value, str) and value.startswith(SPREADSHEET_FORMULA_INJECTION_LEAD):
        return "'" + value
    return value


def _sanitize_spreadsheet_row(row):
    return tuple(_sanitize_spreadsheet_cell(value) for value in row)


def _pbn_publication_url(pbn_uid_id, pbn_api_root):
    if not pbn_uid_id or not pbn_api_root:
        return ""
    return const.LINK_PBN_DO_PUBLIKACJI.format(
        pbn_api_root=pbn_api_root,
        pbn_uid_id=pbn_uid_id,
    )


def _admin_change_url(rekord, request):
    content_type = rekord.content_type
    url = reverse(
        f"admin:{content_type.app_label}_{content_type.model}_change",
        args=(rekord.object_id,),
    )
    return request.build_absolute_uri(url)


def _iter_export_rows(queryset, request):
    uczelnia = Uczelnia.objects.get_for_request(request)
    pbn_api_root = uczelnia.pbn_api_root if uczelnia is not None else ""

    for rekord in queryset.iterator(chunk_size=1000):
        zrodlo = rekord.zrodlo
        typ_kbn = rekord.typ_kbn
        yield (
            rekord.tytul_oryginalny,
            rekord.opis_bibliograficzny_zapisani_autorzy_cache,
            zrodlo.nazwa if zrodlo is not None else "",
            rekord.rok,
            rekord.impact_factor,
            rekord.punkty_kbn,
            str(tuple(rekord.pk)),
            str(rekord.describe_content_type),
            typ_kbn.nazwa if typ_kbn is not None else "",
            rekord.object_id,
            _export_value(rekord.pbn_uid_id),
            request.build_absolute_uri(rekord.get_absolute_url()),
            _admin_change_url(rekord, request),
            _pbn_publication_url(rekord.pbn_uid_id, pbn_api_root),
        )


def _iter_export_rows_opis(queryset, request):
    for lp, rekord in enumerate(queryset.iterator(chunk_size=1000), start=1):
        charakter = rekord.charakter_formalny
        typ_kbn = rekord.typ_kbn
        yield (
            lp,
            _plain_opis_bibliograficzny(rekord.opis_bibliograficzny_cache),
            rekord.impact_factor,
            rekord.punkty_kbn,
            charakter.nazwa if charakter is not None else "",
            typ_kbn.nazwa if typ_kbn is not None else "",
        )


def _xlsx_columns_where(headers, predicate):
    """1-based indeksy kolumn XLSX, których nagłówek spełnia predykat."""
    return [i for i, h in enumerate(headers, start=1) if predicate(h)]


def _apply_xlsx_number_format(worksheet, columns, number_format):
    for col in columns:
        for row in worksheet.iter_rows(min_row=2, min_col=col, max_col=col):
            row[0].number_format = number_format


def _apply_xlsx_hyperlinks(worksheet, url_cols):
    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in url_cols:
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.value = f'=HYPERLINK("{cell.value}", "[link]")'


# Allowlist dla nh3.clean treści dokumentu eksportu (D8). Unia (nie podzbiór)
# DEFAULT_ALLOWED_TAGS z nowe_raporty.docx_export (m.in. h4/strike/font —
# markup opisu bibliograficznego pochodzi z per-instalacyjnych, DB-konfig.
# szablonów) i tagów strukturalnych listy/tabeli. Atrybuty: td/th → colspan.
EXPORT_HTML_STRUCTURAL_TAGS = frozenset(
    {
        "table",
        "thead",
        "tbody",
        "tfoot",
        "tr",
        "td",
        "th",
        "ol",
        "ul",
        "li",
        "p",
        "div",
        "span",
        "h1",
        "h2",
        "h3",
        "br",
        "hr",
        "pre",
        "code",
    }
)
EXPORT_HTML_ATTRIBUTES = {"td": {"colspan"}, "th": {"colspan"}}


def sanitize_export_html(body_html):
    """Sanityzacja treści dokumentu eksportu przez nh3.clean.

    Plik opuszcza serwis (brak CSP, file://), a body zawiera DB-sourced
    opis_bibliograficzny_cache — czyścimy jak siostrzany nowe_raporty.as_docx.
    """
    import nh3

    from nowe_raporty.docx_export import DEFAULT_ALLOWED_TAGS

    tags = set(DEFAULT_ALLOWED_TAGS) | set(EXPORT_HTML_STRUCTURAL_TAGS)
    return nh3.clean(body_html, tags=tags, attributes=EXPORT_HTML_ATTRIBUTES)


def html_export_response(document_html, report_title):
    response = HttpResponse(document_html, content_type="text/html; charset=utf-8")
    response["Content-Disposition"] = content_disposition_header(
        as_attachment=True,
        filename=_export_filename("html", report_title),
    )
    return response


def docx_export_response(document_html, report_title):
    """Eksport DOCX — konwersja przez nowe_raporty.docx_export.html_to_docx.

    Ścieżka pierwsza (i produkcyjna) to pandoc; przy twardej awarii fallback
    na dockerowy html2docx. DocxConversionError propaguje (500) — nie tłumimy.
    """
    from nowe_raporty.docx_export import html_to_docx

    content = html_to_docx(document_html)
    response = HttpResponse(
        content,
        content_type=(
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        ),
    )
    response["Content-Disposition"] = content_disposition_header(
        as_attachment=True,
        filename=_export_filename("docx", report_title),
    )
    return response


def bibtex_export_response(queryset, report_title):
    """Eksport BibTeX (.bib) — surowy tekst, gdy report_type == 'bibtex'.

    Reuse bpp.export.bibtex.export_to_bibtex (dispatch po model_name);
    rekord.original zwraca konkretną instancję publikacji.
    """
    from bpp.export.bibtex import export_to_bibtex

    content = export_to_bibtex(
        rekord.original for rekord in queryset.iterator(chunk_size=1000)
    )
    response = HttpResponse(content, content_type="application/x-bibtex; charset=utf-8")
    response["Content-Disposition"] = content_disposition_header(
        as_attachment=True,
        filename=_export_filename("bib", report_title),
    )
    return response


def csv_export_response(queryset, request, report_title):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(MULTISEEK_EXPORT_HEADERS)
    writer.writerows(
        _sanitize_spreadsheet_row(row) for row in _iter_export_rows(queryset, request)
    )

    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = content_disposition_header(
        as_attachment=True,
        filename=_export_filename("csv", report_title),
    )
    return response


def xlsx_export_response(queryset, request, report_title, wariant="dane"):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    from bpp.util import (
        sanitize_xlsx_row,
        worksheet_columns_autosize,
        worksheet_create_table,
    )

    if wariant == "opis":
        headers = MULTISEEK_EXPORT_OPIS_XLSX_HEADERS
        rows = _iter_export_rows_opis(queryset, request)
        freeze = "A2"
    else:
        headers = MULTISEEK_EXPORT_XLSX_HEADERS
        rows = _iter_export_rows(queryset, request)
        freeze = "B1"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _xlsx_worksheet_title(report_title)
    worksheet.append(headers)
    for row in rows:
        worksheet.append(sanitize_xlsx_row(row))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    url_cols = _xlsx_columns_where(headers, lambda h: h.startswith("Link"))
    if_cols = _xlsx_columns_where(headers, lambda h: h in ("Impact Factor", "IF"))
    pk_cols = _xlsx_columns_where(headers, lambda h: h == "PK")

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    _apply_xlsx_number_format(worksheet, if_cols, "0.000")
    _apply_xlsx_number_format(worksheet, pk_cols, "0.00")
    _apply_xlsx_hyperlinks(worksheet, url_cols)

    worksheet.freeze_panes = freeze
    worksheet_columns_autosize(worksheet)
    if worksheet.max_row > 1:
        worksheet_create_table(worksheet, title="MultiseekExport")

    output = io.BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = content_disposition_header(
        as_attachment=True,
        filename=_export_filename("xlsx", report_title),
    )
    return response
