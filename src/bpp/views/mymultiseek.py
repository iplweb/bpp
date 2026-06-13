import csv
import hashlib
import html
import io
import json
import logging
import re

from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.cache import cache
from django.db.models import Count, Sum
from django.http import HttpResponse, HttpResponseBadRequest, JsonResponse
from django.urls import reverse
from django.utils.html import strip_tags
from django.utils.http import content_disposition_header
from django.views.decorators.cache import never_cache
from django.views.generic import View
from multiseek.logic import get_registry
from multiseek.views import (
    MULTISEEK_SESSION_KEY_REMOVED,
    MultiseekResults,
    manually_add_or_remove,
)

from bpp import const
from bpp.models import Uczelnia
from bpp.multiseek_registry import registry as multiseek_registry
from bpp.multiseek_registry.djangoql_export import multiseek_form_to_djangoql
from bpp.views.zapytanie import WprowadzanieDanychOrSuperuserMixin

logger = logging.getLogger(__name__)

PKT_WEWN = "pkt_wewn"
PKT_WEWN_BEZ = "pkt_wewn_bez"
TABLE = "table"
MULTISEEK_EXPORT_MAX_ROWS = 5000

# TTL cache agregatów wyników (count + sumy) — patrz get_context_data.
MULTISEEK_AGGREGATE_CACHE_TIMEOUT = 30 * 60
MULTISEEK_REPORT_TITLE_SESSION_KEY = "MULTISEEK_TITLE"
MULTISEEK_DEFAULT_REPORT_TITLE = "Rezultat wyszukiwania"
XLSX_WORKSHEET_TITLE_MAX_LENGTH = 31

MULTISEEK_EXPORT_HEADERS = (
    "tytul_oryginalny",
    "autorzy",
    "rok",
    "impact_factor",
    "pk",
    "bpp_id",
    "typ_rekordu",
    "id_rekordu",
    "pbn_uid_id",
    "link_do_bpp_url",
    "link_do_bpp_admin_url",
    "link_do_pbn_url",
)

MULTISEEK_EXPORT_XLSX_HEADERS = (
    "Tytuł oryginalny",
    "Autorzy",
    "Rok",
    "Impact Factor",
    "PK",
    "BPP ID",
    "Typ rekordu",
    "ID rekordu",
    "PBN UID",
    "Link do BPP",
    "Link do edycji w BPP",
    "Link do PBN",
)

MULTISEEK_EXPORT_FIELDS = (
    "id",
    "tytul_oryginalny",
    "opis_bibliograficzny_zapisani_autorzy_cache",
    "rok",
    "impact_factor",
    "punkty_kbn",
    "pbn_uid_id",
)

MULTISEEK_EXPORT_XLSX_URL_COLUMNS = (10, 11, 12)
EXPORT_FILENAME_INVALID_CHARS_RE = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
MULTISEEK_REPORT_TITLE_HTML_BREAK_RE = re.compile(
    r"</?(?:br|hr|p|div|h[1-6])\b[^>]*>",
    re.IGNORECASE,
)
XLSX_WORKSHEET_TITLE_INVALID_CHARS_RE = re.compile(
    r"[\[\]:*?/\\\x00-\x08\x0b\x0c\x0e-\x1f]"
)
SPREADSHEET_FORMULA_INJECTION_LEAD = ("=", "+", "-", "@", "\t", "\r", "\n")

EXTRA_TYPES = [
    PKT_WEWN,
    PKT_WEWN_BEZ,
    TABLE,
    PKT_WEWN + "_cytowania",
    PKT_WEWN_BEZ + "_cytowania",
    TABLE + "_cytowania",
]


class MyMultiseekResults(MultiseekResults):
    registry = "bpp.multiseek.registry"

    def get_queryset(self, only_those_ids=None):
        registry = get_registry(self.registry)
        if only_those_ids is not None:
            qset = registry.get_query_for_model(self.get_multiseek_data())
            if not only_those_ids:
                return qset.none()
            qset = qset.filter(pk__in=only_those_ids)
        else:
            qset = super().get_queryset()

        if not self.request.user.is_authenticated:
            uczelnia = Uczelnia.objects.get_for_request(self.request)
            if uczelnia is not None:
                ukryte_statusy = uczelnia.ukryte_statusy("multiwyszukiwarka")
                if ukryte_statusy:
                    qset = qset.exclude(status_korekty_id__in=ukryte_statusy)

        flds = ("id", "opis_bibliograficzny_cache")

        # wycięte z multiseek/views.py, get_context_data
        report_type = registry.get_report_type(
            self.get_multiseek_data(), request=self.request
        )

        if report_type in EXTRA_TYPES:
            qset = qset.select_related("charakter_formalny", "typ_kbn")

            flds = flds + (
                "charakter_formalny",
                "typ_kbn",
                "punkty_kbn",
                "impact_factor",
                "adnotacje",
                "uwagi",
                "punktacja_wewnetrzna",
                "charakter_formalny__nazwa",
                "typ_kbn__nazwa",
                # "zrodlo_lub_nadrzedne",
            )

        ret = qset.only(*flds)

        # Add DISTINCT when joining with views that can produce duplicates.
        # This string-based SQL check is not ideal but avoids complex query
        # introspection. The affected tables come from fulltext search joins.
        sql = str(ret.query)
        if "bpp_autorzy_mat" in sql or "bpp_zewnetrzne_bazy_view" in sql:
            ret = ret.distinct()

        return ret

    def get_queryset_for_current_mode(self):
        if not self.request.GET.get("print-removed", False):
            return self.get_queryset()

        return self.get_queryset(
            only_those_ids=self.request.session.get(MULTISEEK_SESSION_KEY_REMOVED, [])
        )

    def _aggregate_cache_key(self):
        """Klucz cache agregatów wyników (count + sumy) — wszystkie wejścia
        wpływające na wynik: formularz multiseek (z report_type
        i ordering), lista wyrzuconych rekordów, tryb print-removed,
        zalogowanie (ukryte statusy) i uczelnia."""
        uczelnia = Uczelnia.objects.get_for_request(self.request)
        payload = json.dumps(
            [
                self.get_multiseek_data(),
                sorted(str(x) for x in self.get_removed_records()),
                bool(self.request.GET.get("print-removed", False)),
                bool(self.request.user.is_authenticated),
                getattr(uczelnia, "pk", None),
            ],
            sort_keys=True,
            default=str,
        )
        return "multiseek_agregaty:" + hashlib.sha256(payload.encode()).hexdigest()

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data()

        qset = self.get_queryset_for_current_mode()
        if self.request.GET.get("print-removed", False):
            ctx["object_list"] = qset
            ctx["print_removed"] = True

        # Agregaty (COUNT + sumy) sa cache'owane, zeby stronicowanie tych
        # samych wynikow nie powtarzalo drogiego skanu (DISTINCT + join do
        # bpp_autorzy_mat) przy kazdej stronie. Staleness ograniczone TTL;
        # kazda zmiana formularza / wyrzucenie rekordu zmienia klucz.
        cache_key = self._aggregate_cache_key()
        agregaty = cache.get(cache_key)
        if agregaty is None:
            if ctx["report_type"] in EXTRA_TYPES:
                # Licznik i sumy jednym skanem — przy DISTINCT + join do
                # bpp_autorzy_mat osobny COUNT podwajalby najdrozsza czesc.
                agregaty = qset.aggregate(
                    Sum("impact_factor"),
                    Sum("liczba_cytowan"),
                    Sum("punkty_kbn"),
                    Sum("index_copernicus"),
                    Sum("punktacja_wewnetrzna"),
                    paginator_count=Count("pk"),
                )
            else:
                agregaty = {"paginator_count": qset.count()}
            cache.set(cache_key, agregaty, MULTISEEK_AGGREGATE_CACHE_TIMEOUT)

        agregaty = dict(agregaty)
        ctx["paginator_count"] = agregaty.pop("paginator_count")
        if ctx["report_type"] in EXTRA_TYPES:
            ctx["sumy"] = agregaty

        ctx["multiseek_export_max_rows"] = MULTISEEK_EXPORT_MAX_ROWS
        object_list = ctx["object_list"]
        object_list.count = lambda *args, **kw: ctx["paginator_count"]

        keys = list(self.request.session.keys())
        if "MULTISEEK_TITLE" not in keys:
            self.request.session["MULTISEEK_TITLE"] = "Rezultat wyszukiwania"
        else:
            if self.request.session["MULTISEEK_TITLE"] == "":
                self.request.session["MULTISEEK_TITLE"] = "Rezultat wyszukiwania"

        return ctx


def _export_value(value):
    if value is None:
        return ""
    return str(value)


def _single_line_text(value):
    return re.sub(r"\s+", " ", value).strip()


def _plain_multiseek_report_title(value):
    value = _export_value(value)
    value = MULTISEEK_REPORT_TITLE_HTML_BREAK_RE.sub(" ", value)
    value = html.unescape(strip_tags(value))
    return _single_line_text(value) or MULTISEEK_DEFAULT_REPORT_TITLE


def _multiseek_report_title(request):
    return _plain_multiseek_report_title(
        request.session.get(
            MULTISEEK_REPORT_TITLE_SESSION_KEY,
            MULTISEEK_DEFAULT_REPORT_TITLE,
        )
    )


def _export_filename(export_format, report_title):
    title = EXPORT_FILENAME_INVALID_CHARS_RE.sub(" ", report_title)
    title = _single_line_text(title).strip(". ")
    if not title:
        title = "multiseek"
    return f"eksport-{title}.{export_format}"


def _xlsx_worksheet_title(report_title):
    title = XLSX_WORKSHEET_TITLE_INVALID_CHARS_RE.sub(" ", report_title)
    title = _single_line_text(title).strip("'")
    if not title:
        title = "Multiseek"
    return title[:XLSX_WORKSHEET_TITLE_MAX_LENGTH]


def _sanitize_spreadsheet_cell(value):
    if isinstance(value, str) and value.startswith(SPREADSHEET_FORMULA_INJECTION_LEAD):
        return "'" + value
    return value


def _sanitize_spreadsheet_row(row):
    return tuple(_sanitize_spreadsheet_cell(value) for value in row)


def _pbn_publication_url(pbn_uid_id, pbn_api_root):
    if not pbn_uid_id or not pbn_api_root:
        return ""
    return const.LINK_PBN_DO_PUBLIKACJI.format(
        pbn_api_root=pbn_api_root,
        pbn_uid_id=pbn_uid_id,
    )


def _admin_change_url(rekord, request):
    content_type = rekord.content_type
    url = reverse(
        f"admin:{content_type.app_label}_{content_type.model}_change",
        args=(rekord.object_id,),
    )
    return request.build_absolute_uri(url)


def _iter_export_rows(queryset, request):
    uczelnia = Uczelnia.objects.get_for_request(request)
    pbn_api_root = uczelnia.pbn_api_root if uczelnia is not None else ""

    for rekord in queryset.iterator(chunk_size=1000):
        yield (
            rekord.tytul_oryginalny,
            rekord.opis_bibliograficzny_zapisani_autorzy_cache,
            rekord.rok,
            rekord.impact_factor,
            rekord.punkty_kbn,
            str(tuple(rekord.pk)),
            str(rekord.describe_content_type),
            rekord.object_id,
            _export_value(rekord.pbn_uid_id),
            request.build_absolute_uri(rekord.get_absolute_url()),
            _admin_change_url(rekord, request),
            _pbn_publication_url(rekord.pbn_uid_id, pbn_api_root),
        )


def _csv_export_response(queryset, request, report_title):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(MULTISEEK_EXPORT_HEADERS)
    writer.writerows(
        _sanitize_spreadsheet_row(row) for row in _iter_export_rows(queryset, request)
    )

    response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = content_disposition_header(
        as_attachment=True,
        filename=_export_filename("csv", report_title),
    )
    return response


def _xlsx_export_response(queryset, request, report_title):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    from bpp.util import (
        sanitize_xlsx_row,
        worksheet_columns_autosize,
        worksheet_create_table,
    )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _xlsx_worksheet_title(report_title)
    worksheet.append(MULTISEEK_EXPORT_XLSX_HEADERS)
    for row in _iter_export_rows(queryset, request):
        worksheet.append(sanitize_xlsx_row(row))

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in worksheet.iter_rows(min_row=2, min_col=4, max_col=5):
        row[0].number_format = "0.000"
        row[1].number_format = "0.00"

    for row_idx in range(2, worksheet.max_row + 1):
        for col_idx in MULTISEEK_EXPORT_XLSX_URL_COLUMNS:
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                cell.value = f'=HYPERLINK("{cell.value}", "[link]")'

    worksheet.freeze_panes = "B1"
    worksheet_columns_autosize(worksheet)
    if worksheet.max_row > 1:
        worksheet_create_table(worksheet, title="MultiseekExport")

    output = io.BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    response["Content-Disposition"] = content_disposition_header(
        as_attachment=True,
        filename=_export_filename("xlsx", report_title),
    )
    return response


class MyMultiseekExport(LoginRequiredMixin, MyMultiseekResults):
    http_method_names = ["get"]

    def get(self, request, export_format, *args, **kwargs):
        if export_format not in {"csv", "xlsx"}:
            return HttpResponseBadRequest("Nieznany format eksportu.")

        queryset = self.get_queryset_for_current_mode()
        count = queryset.count()
        if count > MULTISEEK_EXPORT_MAX_ROWS:
            return HttpResponseBadRequest(
                "Eksport Multiseek jest dostępny dla maksymalnie "
                f"{MULTISEEK_EXPORT_MAX_ROWS} rekordów."
            )

        report_title = _multiseek_report_title(request)
        queryset = queryset.select_related(None).only(*MULTISEEK_EXPORT_FIELDS)
        if export_format == "csv":
            return _csv_export_response(queryset, request, report_title)
        return _xlsx_export_response(queryset, request, report_title)


def _normalize_session_removed(request):
    # JSONSerializer zamienia tuple → list przy zapisie sesji. Upstream
    # manually_add_or_remove() robi `set(data)` na odczycie, co pada na
    # listach (unhashable). Konwertujemy z powrotem do tuple w pamięci —
    # w ramach tego requestu upstream zobaczy tuple; serializacja z
    # powrotem do listy nastąpi przy zapisie sesji.
    data = request.session.get(MULTISEEK_SESSION_KEY_REMOVED)
    if not data:
        return
    normalized = [tuple(x) if isinstance(x, list) else x for x in data]
    if normalized != data:
        request.session[MULTISEEK_SESSION_KEY_REMOVED] = normalized


@never_cache
def bpp_remove_by_hand(request, pk):
    """Add a record's PK to a list of manually removed records.

    User, via the web ui, can add or remove a record to a list of records
    removed "by hand". Those records will be explictly removed
    from the search results in the query function. The list of those
    records is cleaned when there is a form reset.
    """
    pk = tuple(int(x) for x in pk.split("_"))
    _normalize_session_removed(request)
    return manually_add_or_remove(request, pk)


@never_cache
def bpp_remove_from_removed_by_hand(request, pk):
    """Cancel manual record removal."""
    pk = tuple(int(x) for x in pk.split("_"))
    _normalize_session_removed(request)
    return manually_add_or_remove(request, pk, add=False)


class MultiseekToDjangoQLView(WprowadzanieDanychOrSuperuserMixin, View):
    """Tlumaczy biezacy formularz Multiseek (POST 'json') na zapytanie
    DjangoQL nad Rekord. Zwraca {query, warnings, editor_url}."""

    http_method_names = ["post"]

    def post(self, request, *args, **kwargs):
        raw = request.POST.get("json")
        if not raw:
            return HttpResponseBadRequest("Brak parametru 'json'.")
        try:
            form_json = json.loads(raw)
        except (ValueError, TypeError):
            return HttpResponseBadRequest("Niepoprawny JSON formularza.")
        if not isinstance(form_json, dict):
            return HttpResponseBadRequest("Oczekiwano obiektu JSON.")
        try:
            result = multiseek_form_to_djangoql(form_json, multiseek_registry)
        except (KeyError, TypeError, AttributeError):
            logger.exception(
                "Niepoprawna struktura formularza przesłana do MultiseekToDjangoQLView."
            )
            return HttpResponseBadRequest("Niepoprawna struktura formularza.")
        return JsonResponse(
            {
                "query": result.query,
                "warnings": result.warnings,
                "editor_url": result.editor_url,
            }
        )
