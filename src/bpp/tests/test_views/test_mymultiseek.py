"""Regression tests dla bpp.views.mymultiseek."""

import csv
import io
import json
from decimal import Decimal

import pytest
from django.conf import settings
from django.db import connection
from django.test.utils import CaptureQueriesContext
from django.urls import reverse
from django.utils import translation
from model_bakery import baker
from multiseek.logic import STARTS_WITH
from multiseek.views import MULTISEEK_SESSION_KEY, MULTISEEK_SESSION_KEY_REMOVED

from bpp.models import Wydawnictwo_Ciagle
from bpp.models.cache import Rekord
from bpp.tests.util import any_ciagle, any_zwarte
from bpp.views.multiseek_export import (
    MULTISEEK_EXPORT_OPIS_XLSX_HEADERS,
    _plain_opis_bibliograficzny,
)
from bpp.views.mymultiseek import (
    MULTISEEK_DEFAULT_REPORT_TITLE,
    MULTISEEK_EXPORT_HEADERS,
    MULTISEEK_EXPORT_MAX_ROWS,
    MULTISEEK_EXPORT_XLSX_HEADERS,
    MULTISEEK_REPORT_TITLE_SESSION_KEY,
    XLSX_WORKSHEET_TITLE_MAX_LENGTH,
)

EXPORT_TITLE_PREFIX = "Eksport Multiseek"
TABLE_REPORT_TYPE = "1"


def _set_multiseek_title_filter(
    client,
    title_prefix=EXPORT_TITLE_PREFIX,
    report_type="0",
):
    with translation.override(settings.LANGUAGE_CODE):
        operator = str(STARTS_WITH)

    session = client.session
    session[MULTISEEK_SESSION_KEY] = json.dumps(
        {
            "form_data": [
                None,
                {
                    "field": "Tytuł pracy",
                    "operator": operator,
                    "value": title_prefix,
                    "prev_op": None,
                },
            ],
            "ordering": {},
            "report_type": report_type,
        }
    )
    session.save()


def _set_multiseek_report_title(client, title):
    session = client.session
    session[MULTISEEK_REPORT_TITLE_SESSION_KEY] = title
    session.save()


@pytest.fixture
def multiseek_export_rekord(standard_data, denorms, autor_jan_kowalski, jednostka):
    uczelnia = jednostka.uczelnia
    uczelnia.pbn_api_root = "https://pbn.example.org"
    uczelnia.save(update_fields=["pbn_api_root"])
    pbn_publication = baker.make(
        "pbn_api.Publication",
        mongoId="507f1f77bcf86cd799439011",
    )
    wydawnictwo = any_ciagle(
        tytul_oryginalny=f"{EXPORT_TITLE_PREFIX} - alfa",
        rok=2024,
        impact_factor=Decimal("1.230"),
        punkty_kbn=Decimal("42.00"),
        pbn_uid=pbn_publication,
    )
    wydawnictwo.dodaj_autora(
        autor_jan_kowalski,
        jednostka,
        zapisany_jako="Kowalski Jan",
    )
    denorms.flush()
    return Rekord.objects.get_original(wydawnictwo)


@pytest.fixture
def multiseek_export_pair(standard_data, denorms):
    first = any_ciagle(tytul_oryginalny=f"{EXPORT_TITLE_PREFIX} - alfa", rok=2024)
    second = any_ciagle(tytul_oryginalny=f"{EXPORT_TITLE_PREFIX} - beta", rok=2023)
    denorms.flush()
    return (
        Rekord.objects.get_original(first),
        Rekord.objects.get_original(second),
    )


@pytest.mark.django_db
def test_remove_from_removed_after_json_session_roundtrip(client):
    """Drugie wywołanie remove_from_results nie może zwrócić 500.

    JSONSerializer zamienia tuple → list przy zapisie sesji. Upstream
    multiseek.manually_add_or_remove robi set(data) na odczycie, co
    failuje na listach (unhashable). Poprzednio to powodowało HTTP 500
    przy drugim klinięciu "Wyrzuć" (oryginalna regresja Django 5.2 PW).

    Symulacja: wrzucamy do sesji listę [[3, 1]] (post-JSON format),
    wołamy endpoint i oczekujemy 200.
    """
    session = client.session
    session[MULTISEEK_SESSION_KEY_REMOVED] = [[3, 1]]  # post-JSON roundtrip
    session.save()

    response = client.get(reverse("remove_from_removed_results", kwargs={"pk": "3_1"}))
    assert response.status_code == 200

    # Po wywołaniu session ma zostać bez tego id (usunięte)
    session = client.session
    assert session.get(MULTISEEK_SESSION_KEY_REMOVED) == []


@pytest.mark.django_db
def test_remove_by_hand_twice_does_not_500(client):
    """Dwa kolejne wywołania remove_from_results nie 500-ują.

    Po pierwszym wywołaniu sesja ma [[id1, id2]] (JSON-serialized tuple).
    Drugie wywołanie musi poprawnie odczytać i znormalizować.
    """
    url = reverse("remove_from_results", kwargs={"pk": "3_7"})
    assert client.get(url).status_code == 200
    # Drugie wywołanie (to samo pk) — niech nie wybuchnie
    assert client.get(url).status_code == 200


@pytest.mark.django_db
def test_multiseek_export_requires_login(client):
    response = client.get(reverse("multiseek-export", kwargs={"export_format": "csv"}))
    assert response.status_code == 302
    assert settings.LOGIN_URL in response["Location"]


@pytest.mark.django_db
def test_multiseek_export_csv(logged_in_client, multiseek_export_rekord):
    _set_multiseek_title_filter(logged_in_client)

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "csv"})
    )

    assert response.status_code == 200
    assert response["Content-Type"] == "text/csv; charset=utf-8"
    assert (
        f'filename="eksport-{MULTISEEK_DEFAULT_REPORT_TITLE}.csv"'
        in response["Content-Disposition"]
    )

    rows = list(csv.reader(io.StringIO(response.content.decode("utf-8"))))
    assert rows[0] == list(MULTISEEK_EXPORT_HEADERS)
    assert rows[1] == [
        multiseek_export_rekord.tytul_oryginalny,
        "Kowalski Jan",
        multiseek_export_rekord.zrodlo.nazwa,
        "2024",
        "1.230",
        "42.00",
        str(tuple(multiseek_export_rekord.pk)),
        "wydawnictwo ciągłe",
        multiseek_export_rekord.typ_kbn.nazwa,
        str(multiseek_export_rekord.object_id),
        "507f1f77bcf86cd799439011",
        f"http://testserver{multiseek_export_rekord.get_absolute_url()}",
        "http://testserver/admin/bpp/wydawnictwo_ciagle/"
        f"{multiseek_export_rekord.object_id}/change/",
        "https://pbn.example.org/core/#/publication/view/"
        "507f1f77bcf86cd799439011/current",
    ]


@pytest.mark.django_db
def test_multiseek_export_filename_uses_report_title(
    logged_in_client,
    multiseek_export_rekord,
):
    _set_multiseek_title_filter(logged_in_client)
    _set_multiseek_report_title(logged_in_client, "Raport<br/>autorski")

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "csv"})
    )

    assert response.status_code == 200
    assert 'filename="eksport-Raport autorski.csv"' in response["Content-Disposition"]


@pytest.mark.django_db
def test_multiseek_export_csv_sanitizes_formula_like_values(
    logged_in_client,
    standard_data,
    denorms,
):
    title = "=Eksport Multiseek formula"
    any_ciagle(tytul_oryginalny=title)
    denorms.flush()
    _set_multiseek_title_filter(logged_in_client, title_prefix="=Eksport")

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "csv"})
    )

    assert response.status_code == 200
    rows = list(csv.DictReader(io.StringIO(response.content.decode("utf-8"))))
    assert rows[0]["tytul_oryginalny"] == "'" + title


@pytest.mark.django_db
def test_multiseek_export_csv_works_for_table_report_type(
    logged_in_client,
    multiseek_export_rekord,
):
    _set_multiseek_title_filter(logged_in_client, report_type=TABLE_REPORT_TYPE)

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "csv"})
    )

    assert response.status_code == 200


@pytest.mark.django_db
def test_multiseek_export_xlsx(logged_in_client, multiseek_export_rekord):
    _set_multiseek_title_filter(logged_in_client)

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "xlsx"})
    )

    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert (
        f'filename="eksport-{MULTISEEK_DEFAULT_REPORT_TITLE}.xlsx"'
        in response["Content-Disposition"]
    )

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(response.content))
    worksheet = workbook.active
    assert worksheet.title == MULTISEEK_DEFAULT_REPORT_TITLE
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0] == MULTISEEK_EXPORT_XLSX_HEADERS
    assert rows[1] == (
        multiseek_export_rekord.tytul_oryginalny,
        "Kowalski Jan",
        multiseek_export_rekord.zrodlo.nazwa,
        2024,
        1.23,
        42,
        str(tuple(multiseek_export_rekord.pk)),
        "wydawnictwo ciągłe",
        multiseek_export_rekord.typ_kbn.nazwa,
        multiseek_export_rekord.object_id,
        "507f1f77bcf86cd799439011",
        '=HYPERLINK("'
        f"http://testserver{multiseek_export_rekord.get_absolute_url()}"
        '", "[link]")',
        '=HYPERLINK("http://testserver/admin/bpp/wydawnictwo_ciagle/'
        f'{multiseek_export_rekord.object_id}/change/", "[link]")',
        '=HYPERLINK("https://pbn.example.org/core/#/publication/view/'
        '507f1f77bcf86cd799439011/current", "[link]")',
    )
    assert worksheet.freeze_panes == "B1"
    assert worksheet["E2"].data_type == "n"
    assert worksheet["F2"].data_type == "n"
    assert worksheet["E2"].number_format == "0.000"  # Impact Factor (było D)
    assert worksheet["F2"].number_format == "0.00"  # PK (było E)
    # kolumny linków (L, M, N) zawierają =HYPERLINK
    for col in ("L", "M", "N"):
        assert worksheet[f"{col}2"].value.startswith("=HYPERLINK(")


@pytest.mark.django_db
def test_multiseek_export_xlsx_sheet_title_uses_limited_report_title(
    logged_in_client,
    multiseek_export_rekord,
):
    _set_multiseek_title_filter(logged_in_client)
    _set_multiseek_report_title(
        logged_in_client,
        "Raport: A/B*C? \x07[bardzo bardzo bardzo długi]",
    )

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "xlsx"})
    )

    assert response.status_code == 200

    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(response.content))
    worksheet_title = workbook.active.title
    assert worksheet_title.startswith("Raport A B C")
    assert len(worksheet_title) == XLSX_WORKSHEET_TITLE_MAX_LENGTH
    assert not set(r"[]:*?/\\").intersection(worksheet_title)
    assert "\x07" not in worksheet_title


@pytest.mark.django_db
def test_multiseek_export_removed_records(
    logged_in_client,
    multiseek_export_pair,
):
    visible, removed = multiseek_export_pair
    _set_multiseek_title_filter(logged_in_client)
    session = logged_in_client.session
    session[MULTISEEK_SESSION_KEY_REMOVED] = [removed.pk]
    session.save()

    normal_response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "csv"})
    )
    removed_response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "csv"})
        + "?print-removed=1"
    )

    normal_rows = list(
        csv.DictReader(io.StringIO(normal_response.content.decode("utf-8")))
    )
    removed_rows = list(
        csv.DictReader(io.StringIO(removed_response.content.decode("utf-8")))
    )
    assert [row["bpp_id"] for row in normal_rows] == [str(tuple(visible.pk))]
    assert [row["bpp_id"] for row in removed_rows] == [str(tuple(removed.pk))]


@pytest.mark.django_db
def test_multiseek_export_enforces_row_limit(
    logged_in_client,
    multiseek_export_pair,
    monkeypatch,
):
    _set_multiseek_title_filter(logged_in_client)
    monkeypatch.setattr("bpp.views.mymultiseek.MULTISEEK_EXPORT_MAX_ROWS", 1)

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "csv"})
    )

    assert response.status_code == 400
    assert "maksymalnie 1 rekordów" in response.content.decode("utf-8")


@pytest.mark.django_db
def test_multiseek_export_links_hidden_over_limit(
    logged_in_client,
    multiseek_export_pair,
    monkeypatch,
):
    _set_multiseek_title_filter(logged_in_client)
    monkeypatch.setattr("bpp.views.mymultiseek.MULTISEEK_EXPORT_MAX_ROWS", 1)

    response = logged_in_client.get(reverse("live-results"))

    assert response.status_code == 200
    assert (
        reverse("multiseek-export", kwargs={"export_format": "csv"}).encode()
        not in response.content
    )
    assert (
        reverse("multiseek-export", kwargs={"export_format": "xlsx"}).encode()
        not in response.content
    )


def test_multiseek_export_max_rows_is_5000():
    assert MULTISEEK_EXPORT_MAX_ROWS == 5000


@pytest.mark.django_db
def test_multiseek_export_unknown_format(logged_in_client):
    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "pdf"})
    )
    assert response.status_code == 400


@pytest.mark.django_db
def test_multiseek_export_does_not_match_unfiltered_records(
    logged_in_client,
    multiseek_export_rekord,
):
    baker.make(Wydawnictwo_Ciagle, tytul_oryginalny="Poza eksportem")
    _set_multiseek_title_filter(logged_in_client)

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "csv"})
    )

    rows = list(csv.DictReader(io.StringIO(response.content.decode("utf-8"))))
    assert [row["bpp_id"] for row in rows] == [str(tuple(multiseek_export_rekord.pk))]


@pytest.mark.django_db
def test_export_dane_csv_ma_zrodlo_i_typ_mnisw(
    logged_in_client, multiseek_export_rekord
):
    _set_multiseek_title_filter(logged_in_client)

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "csv"})
    )

    assert response.status_code == 200
    text = response.content.decode("utf-8")
    header = text.splitlines()[0]
    cols = header.split(",")
    assert cols[2] == "zrodlo"
    assert cols[8] == "typ_mnisw_mein"
    assert cols[6] == "bpp_id"  # BPP ID nadal przed typ_rekordu (kol. 8)
    assert cols[7] == "typ_rekordu"

    rows = list(csv.DictReader(io.StringIO(text)))
    assert rows[0]["zrodlo"] == multiseek_export_rekord.zrodlo.nazwa
    assert rows[0]["typ_mnisw_mein"] == multiseek_export_rekord.typ_kbn.nazwa


@pytest.mark.django_db
def test_export_dane_csv_ksiazka_bez_zrodla_ma_puste_pole(
    logged_in_client, standard_data, denorms
):
    """Wydawnictwo_Zwarte (książka/monografia) nie ma pola ``zrodlo`` — w
    cache Rekord to ``zrodlo=None``. Każda książka w produkcji trafia w tę
    gałąź _iter_export_rows (``zrodlo.nazwa if zrodlo is not None else
    ""``); żaden dotychczasowy test jej nie ćwiczył."""
    any_zwarte(tytul_oryginalny=f"{EXPORT_TITLE_PREFIX} - książka")
    denorms.flush()
    _set_multiseek_title_filter(logged_in_client)

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "csv"})
    )

    assert response.status_code == 200
    rows = list(csv.DictReader(io.StringIO(response.content.decode("utf-8"))))
    assert len(rows) == 1
    assert rows[0]["zrodlo"] == ""


@pytest.mark.django_db
def test_export_dane_xlsx_ma_zrodlo_i_typ_mnisw(
    logged_in_client, multiseek_export_rekord
):
    from openpyxl import load_workbook

    _set_multiseek_title_filter(logged_in_client)

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "xlsx"})
    )

    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers[2] == "Źródło"
    assert headers[8] == "Typ MNiSW/MEiN"
    assert headers[6] == "BPP ID"
    assert headers[7] == "Typ rekordu"

    rows = list(ws.iter_rows(values_only=True))
    assert rows[1][2] == multiseek_export_rekord.zrodlo.nazwa
    assert rows[1][8] == multiseek_export_rekord.typ_kbn.nazwa


def _make_n_plus_1_rows(prefix, count):
    """``count`` rekordów, każdy z WŁASNYM zrodlo/typ_kbn/charakter_formalny.

    ``any_ciagle()`` bez jawnych kwargs tworzy te FK-i od zera przy każdym
    wywołaniu (model_bakery nie współdzieli/nie cache'uje instancji FK
    między wywołaniami) — dzięki temu ewentualny N+1 na dostępie do tych
    pól faktycznie odpali osobne zapytanie na wiersz, zamiast schować się
    za pamięcią podręczną kluczoną po (współdzielonej) wartości FK.
    """
    for i in range(count):
        any_ciagle(tytul_oryginalny=f"{prefix} - {i}")


def _query_count_for_export(client, url):
    # Rozgrzewka: pierwsze zapytanie w sesji odpala dodatkowy, jednorazowy
    # SELECT z niepowiązanego middleware'u password_policies (sprawdzenie
    # historii haseł raz na sesję) — bez tego call'a porównanie N1 vs N2
    # byłoby fałszywie różne o 1 zapytanie, mimo braku N+1 w eksporcie.
    client.get(url)
    with CaptureQueriesContext(connection) as ctx:
        response = client.get(url)
    assert response.status_code == 200
    return response, len(ctx.captured_queries)


@pytest.mark.django_db
def test_export_dane_xlsx_nie_ma_n_plus_1(logged_in_client, standard_data, denorms):
    """Liczba zapytań SQL dla eksportu 'dane' jest STAŁA względem liczby
    wierszy — to jest definicja braku N+1. Samo sprawdzenie górnego progu
    (max_num_queries) przepuściłoby częściowy N+1 (np. +1 zapytanie na
    wiersz przy dużym marginesie); porównanie N1 vs N2 tego nie przepuści.
    """
    from openpyxl import load_workbook

    prefix = f"{EXPORT_TITLE_PREFIX} - n plus 1"
    url = reverse("multiseek-export", kwargs={"export_format": "xlsx"})

    _make_n_plus_1_rows(prefix, 2)
    denorms.flush()
    _set_multiseek_title_filter(logged_in_client, title_prefix=prefix)

    response, n1_queries = _query_count_for_export(logged_in_client, url)
    load_workbook(io.BytesIO(response.content))

    _make_n_plus_1_rows(prefix, 4)  # razem 6 rekordów
    denorms.flush()

    response, n2_queries = _query_count_for_export(logged_in_client, url)
    load_workbook(io.BytesIO(response.content))

    assert n2_queries == n1_queries, (
        "Liczba zapytań SQL rośnie wraz z liczbą wierszy "
        f"({n1_queries} dla 2 wierszy → {n2_queries} dla 6 wierszy) — N+1."
    )
    assert n1_queries <= 18  # sensowny górny limit, z marginesem


def test_plain_opis_bibliograficzny_czysci_html():
    value = "Tytuł <i>Źródła</i> 2026<br>s. 1-2"

    opis = _plain_opis_bibliograficzny(value)

    assert "<" not in opis and ">" not in opis
    assert "Źródła" in opis
    assert "  " not in opis  # spacje skolapsowane, <br> nie sklejone
    # <br> (tag blokowy) MUSI zostać zamieniony na spację, a nie po prostu
    # zniknąć — inaczej "2026" i "s." skleją się w "2026s." (usunięcie kroku
    # MULTISEEK_REPORT_TITLE_HTML_BREAK_RE.sub(" ", ...) nie zostałoby
    # wykryte przez samo "  " not in opis, bo sklejenie nie tworzy podwójnej
    # spacji).
    assert "2026 s. 1-2" in opis


def test_plain_opis_bibliograficzny_dekoduje_encje_html():
    value = "Kowalski J. &amp; Nowak T. &oacute;wczesny"

    opis = _plain_opis_bibliograficzny(value)

    assert "&amp;" not in opis
    assert "&" in opis
    assert "&oacute;" not in opis
    assert "ó" in opis


def test_plain_opis_bibliograficzny_puste_wejscie_bez_fallbacku():
    assert _plain_opis_bibliograficzny("") == ""
    assert _plain_opis_bibliograficzny(None) == ""


@pytest.mark.django_db
def test_multiseek_export_opis_xlsx_uklad(logged_in_client, multiseek_export_rekord):
    from openpyxl import load_workbook

    _set_multiseek_title_filter(logged_in_client)

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "xlsx"}) + "?wariant=opis"
    )

    assert response.status_code == 200
    wb = load_workbook(io.BytesIO(response.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    assert headers == list(MULTISEEK_EXPORT_OPIS_XLSX_HEADERS)
    assert ws["A2"].value == 1  # numeracja Lp.
    assert ws.freeze_panes == "A2"


@pytest.mark.django_db
def test_multiseek_export_opis_xlsx_wiersz(logged_in_client, multiseek_export_rekord):
    from openpyxl import load_workbook

    _set_multiseek_title_filter(logged_in_client)

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "xlsx"}) + "?wariant=opis"
    )

    assert response.status_code == 200
    ws = load_workbook(io.BytesIO(response.content)).active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[1] == (
        1,
        _plain_opis_bibliograficzny(multiseek_export_rekord.opis_bibliograficzny_cache),
        1.23,
        42,
        multiseek_export_rekord.charakter_formalny.nazwa,
        multiseek_export_rekord.typ_kbn.nazwa,
    )


@pytest.mark.django_db
def test_multiseek_export_opis_csv_degraduje_do_dane(
    logged_in_client, multiseek_export_rekord
):
    _set_multiseek_title_filter(logged_in_client)

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "csv"}) + "?wariant=opis"
    )

    assert response.status_code == 200
    header = response.content.decode("utf-8").splitlines()[0]
    assert header.split(",")[0] == "tytul_oryginalny"  # układ dane, nie opis


@pytest.mark.django_db
def test_multiseek_export_nieznany_wariant_to_dane(
    logged_in_client, multiseek_export_rekord
):
    from openpyxl import load_workbook

    _set_multiseek_title_filter(logged_in_client)

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "xlsx"})
        + "?wariant=cokolwiek"
    )

    assert response.status_code == 200
    ws = load_workbook(io.BytesIO(response.content)).active
    assert ws[1][0].value == "Tytuł oryginalny"  # układ dane


@pytest.mark.django_db
def test_multiseek_export_opis_print_removed_wspolistnieja(
    logged_in_client, multiseek_export_pair
):
    visible, removed = multiseek_export_pair
    _set_multiseek_title_filter(logged_in_client)
    session = logged_in_client.session
    session[MULTISEEK_SESSION_KEY_REMOVED] = [removed.pk]
    session.save()

    response = logged_in_client.get(
        reverse("multiseek-export", kwargs={"export_format": "xlsx"})
        + "?wariant=opis&print-removed=1"
    )

    assert response.status_code == 200


@pytest.mark.django_db
def test_multiseek_export_opis_xlsx_nie_ma_n_plus_1(
    logged_in_client, standard_data, denorms
):
    """Analogicznie do test_export_dane_xlsx_nie_ma_n_plus_1 — liczba zapytań
    SQL musi być stała względem liczby wierszy (nie tylko poniżej progu)."""
    from openpyxl import load_workbook

    prefix = f"{EXPORT_TITLE_PREFIX} - opis n plus 1"
    url = (
        reverse("multiseek-export", kwargs={"export_format": "xlsx"}) + "?wariant=opis"
    )

    _make_n_plus_1_rows(prefix, 2)
    denorms.flush()
    _set_multiseek_title_filter(logged_in_client, title_prefix=prefix)

    response, n1_queries = _query_count_for_export(logged_in_client, url)
    load_workbook(io.BytesIO(response.content))

    _make_n_plus_1_rows(prefix, 4)  # razem 6 rekordów
    denorms.flush()

    response, n2_queries = _query_count_for_export(logged_in_client, url)
    load_workbook(io.BytesIO(response.content))

    assert n2_queries == n1_queries, (
        "Liczba zapytań SQL rośnie wraz z liczbą wierszy "
        f"({n1_queries} dla 2 wierszy → {n2_queries} dla 6 wierszy) — N+1."
    )
    assert n1_queries <= 18  # sensowny górny limit, z marginesem
