from io import BytesIO

import openpyxl
import pytest
from django.urls import reverse
from django_webtest import DjangoTestApp, DjangoWebtestResponse
from model_bakery import baker

from django.contrib.admin import site

from bpp.models import Autor, Wydawnictwo_Ciagle, Wydawnictwo_Zwarte

NAZWA_LINKU_EKSPORTU = "Eksport"


@pytest.mark.parametrize(
    "urlname,klass",
    [
        ("wydawnictwo_ciagle", Wydawnictwo_Ciagle),
        ("wydawnictwo_zwarte", Wydawnictwo_Zwarte),
        ("autor", Autor),
    ],
)
def test_xlsx_export_overflow(urlname, klass, admin_app: DjangoTestApp, settings):
    max_allowed_export_items = 5

    settings.BPP_MAX_ALLOWED_EXPORT_ITEMS = max_allowed_export_items

    modeladmin = site._registry.get(klass)
    if hasattr(modeladmin, "max_allowed_export_items"):
        modeladmin.max_allowed_export_items = max_allowed_export_items

    baker.make(
        klass,
        _quantity=max_allowed_export_items,
        _bulk_create=True,
    )

    page: DjangoWebtestResponse = admin_app.get(
        reverse(f"admin:bpp_{urlname}_changelist")
    )

    with pytest.raises(IndexError):
        page.click(NAZWA_LINKU_EKSPORTU)

    klass.objects.all().first().delete()

    page: DjangoWebtestResponse = admin_app.get(
        reverse(f"admin:bpp_{urlname}_changelist")
    )
    page.click(NAZWA_LINKU_EKSPORTU)


@pytest.mark.parametrize(
    "urlname,klass,cname",
    [
        ("wydawnictwo_ciagle", Wydawnictwo_Ciagle, "id"),
        ("wydawnictwo_zwarte", Wydawnictwo_Zwarte, "id"),
        ("autor", Autor, "nazwisko"),
    ],
)
def test_xlsx_export_data(urlname, klass, cname, admin_app: DjangoTestApp):
    baker.make(klass)

    page: DjangoWebtestResponse = admin_app.get(
        reverse(f"admin:bpp_{urlname}_changelist")
    )

    xlsx_binary_data = page.click(NAZWA_LINKU_EKSPORTU)
    wb = openpyxl.load_workbook(BytesIO(xlsx_binary_data.content))
    assert wb.active["A1"].value == cname


@pytest.mark.django_db
def test_xlsx_export_nazwy_zamiast_numerkow(
    wydawnictwo_ciagle, admin_app: DjangoTestApp
):
    # fikstura wydawnictwo_ciagle tworzy wydawnictwo cialge z jezykiem polskim

    page: DjangoWebtestResponse = admin_app.get(
        reverse("admin:bpp_wydawnictwo_ciagle_changelist")
    )

    xlsx_binary_data = page.click(NAZWA_LINKU_EKSPORTU)
    wb = openpyxl.load_workbook(BytesIO(xlsx_binary_data.content))
    ws = wb.active

    column_name = "jezyk"
    for column_cell in ws.iter_cols(1, ws.max_column):  # iterate column cell
        if column_cell[0].value == column_name:  # check for your column
            for data in column_cell[1:2]:  # iterate your column
                assert data.value == "pol."
