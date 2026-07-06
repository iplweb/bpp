"""Eksport XLSX: kolumny ORCID i poprawny URL do PBN."""

from io import BytesIO

import pytest
from django.utils import timezone
from model_bakery import baker
from openpyxl import load_workbook

from deduplikator_autorow.models import DuplicateCandidate, DuplicateScanRun
from deduplikator_autorow.utils import export_duplicates_to_xlsx


@pytest.fixture
def candidate_with_orcid_and_pbn(db):
    """Para autorów z ORCID i PBN UID, oraz Uczelnia z pbn_api_root."""
    from django.contrib.sites.models import Site

    from bpp.models import Uczelnia

    site, _ = Site.objects.get_or_create(
        domain="testserver", defaults={"name": "testserver"}
    )
    uczelnia, _ = Uczelnia.objects.get_or_create(
        nazwa="Test U",
        defaults={
            "skrot": "TU",
            "slug": "test-u",
            "pbn_api_root": "https://pbn-micro-alpha.opi.org.pl",
            "site": site,
        },
    )
    if not uczelnia.pbn_api_root:
        uczelnia.pbn_api_root = "https://pbn-micro-alpha.opi.org.pl"
        uczelnia.save()

    sci_main = baker.make("pbn_api.Scientist", mongoId="abcd-1234")
    sci_dup = baker.make("pbn_api.Scientist", mongoId="efgh-5678")
    main = baker.make(
        "bpp.Autor",
        nazwisko="Kowalski",
        imiona="Jan",
        orcid="0000-0001-2345-6789",
        pbn_uid=sci_main,
    )
    dup = baker.make(
        "bpp.Autor",
        nazwisko="Kowalski",
        imiona="Jan",
        orcid="0000-0002-3456-7890",
        pbn_uid=sci_dup,
    )
    scan = DuplicateScanRun.objects.create(
        status=DuplicateScanRun.Status.COMPLETED,
        finished_at=timezone.now(),
    )
    DuplicateCandidate.objects.create(
        scan_run=scan,
        main_autor=main,
        duplicate_autor=dup,
        confidence_score=80,
        confidence_percent=0.6,
        main_autor_name="Kowalski Jan",
        duplicate_autor_name="Kowalski Jan",
        scan_mode="pbn",
    )
    return main, dup


def _load_xlsx_first_data_row():
    result = export_duplicates_to_xlsx()
    wb = load_workbook(BytesIO(result))
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    data = [cell.value for cell in ws[2]]
    return headers, data


@pytest.mark.django_db
def test_xlsx_has_orcid_columns(candidate_with_orcid_and_pbn):
    """Eksport XLSX zawiera ORCID głównego autora i ORCID duplikatu."""
    headers, _ = _load_xlsx_first_data_row()
    assert "ORCID głównego autora" in headers, (
        f"Brak kolumny 'ORCID głównego autora' w eksporcie. Nagłówki: {headers}"
    )
    assert "ORCID duplikatu" in headers, (
        f"Brak kolumny 'ORCID duplikatu' w eksporcie. Nagłówki: {headers}"
    )


@pytest.mark.django_db
def test_xlsx_orcid_values_filled(candidate_with_orcid_and_pbn):
    """Wartości ORCID są obecne w wyeksportowanym wierszu."""
    main, dup = candidate_with_orcid_and_pbn
    headers, data = _load_xlsx_first_data_row()
    main_idx = headers.index("ORCID głównego autora")
    dup_idx = headers.index("ORCID duplikatu")
    assert data[main_idx] == main.orcid
    assert data[dup_idx] == dup.orcid


@pytest.mark.django_db
def test_xlsx_pbn_url_uses_pbn_api_root(candidate_with_orcid_and_pbn):
    """PBN URL korzysta z LINK_PBN_DO_AUTORA + pbn_api_root, nie sedno-webapp."""
    headers, data = _load_xlsx_first_data_row()
    main_pbn_idx = headers.index("PBN URL głównego autora")
    pbn_url = data[main_pbn_idx] or ""
    assert "sedno-webapp" not in pbn_url, (
        f"URL nadal wskazuje na stary sedno-webapp: {pbn_url}"
    )
    assert "/core/#/person/view/" in pbn_url, (
        f"URL nie pasuje do wzorca LINK_PBN_DO_AUTORA: {pbn_url}"
    )
    assert "abcd-1234" in pbn_url, f"PBN UID nieobecne w URL: {pbn_url}"
