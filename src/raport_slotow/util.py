import io
from collections.abc import Iterable

import openpyxl
import openpyxl.styles
from django.db import DEFAULT_DB_ALIAS, connections
from django.http import HttpResponseBadRequest
from django_tables2.export import ExportMixin, TableExport
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

# Limit wierszy synchronicznego eksportu XLSX (ochrona przed OOM/DoS). Eksport
# leci w cyklu request/response i openpyxl trzyma cały arkusz w RAM, więc bez
# górnej granicy duży raport (np. cała uczelnia) może wyczerpać pamięć procesu
# web. Wartość dobrana wyżej niż limit Multiseeka (5000), bo raporty slotów są
# podstawowym produktem eksportowym i pojedynczy autor/filtr bywa obszerniejszy,
# ale nadal ograniczona. Przekroczenie → jasny komunikat (bez cichego ucięcia).
RAPORT_SLOTOW_EXPORT_MAX_ROWS = 10000


class ExportRowLimitExceeded(Exception):
    """Eksport przekracza limit wierszy (RAPORT_SLOTOW_EXPORT_MAX_ROWS).

    Podnoszone w trakcie budowy XLSX, przechwytywane na granicy widoku i
    zamieniane na czytelny HTTP 400 — użytkownik dostaje komunikat, a nie
    po cichu ucięty plik ani 500."""


def drop_table(table_name, using=DEFAULT_DB_ALIAS):
    connection = connections[using]
    with connection.cursor() as cursor:
        cursor.execute("DROP TABLE IF EXISTS " + connection.ops.quote_name(table_name))


def _create_table_as(table_name, queryset, using=DEFAULT_DB_ALIAS, temporary=True):
    compiler = queryset.query.get_compiler(using=using)
    sql, params = compiler.as_sql()
    connection = connections[using]
    crt = "CREATE TABLE "
    if temporary:
        crt = "CREATE TEMPORARY TABLE "
    sql = crt + connection.ops.quote_name(table_name) + " AS " + sql
    drop_table(table_name, using=using)
    with connection.cursor() as cursor:
        cursor.execute(sql, params)


def create_temporary_table_as(table_name, queryset, using=DEFAULT_DB_ALIAS):
    return _create_table_as(table_name, queryset, using=using, temporary=True)


def create_table_as(table_name, queryset, using=DEFAULT_DB_ALIAS):
    return _create_table_as(table_name, queryset, using=using, temporary=False)


def insert_into(table_name, queryset, using=DEFAULT_DB_ALIAS):
    compiler = queryset.query.get_compiler(using=using)
    sql, params = compiler.as_sql()
    connection = connections[using]
    sql = "INSERT INTO " + connection.ops.quote_name(table_name) + " " + sql
    with connection.cursor() as cursor:
        cursor.execute(sql, params)


def clone_temporary_table(source_table, target_table, using=DEFAULT_DB_ALIAS):
    connection = connections[using]
    sql = (
        "CREATE TEMPORARY TABLE "
        + connection.ops.quote_name(target_table)
        + " AS SELECT * FROM "
        + connection.ops.quote_name(source_table)
    )
    drop_table(target_table, using=using)
    with connection.cursor() as cursor:
        cursor.execute(sql)


def _write_export_description(ws, export_description):
    """Wypisz wiersze opisu nad tabelą. Dwuelementowe krotki dostają
    pogrubioną, wyrównaną do prawej etykietę i wyrównaną do lewej wartość."""
    for elem in export_description:
        if isinstance(elem, Iterable):
            ws.append(elem)
            if len(elem) == 2:
                ws[ws.max_row][0].font = openpyxl.styles.Font(bold=True)
                ws[ws.max_row][0].alignment = openpyxl.styles.Alignment(
                    horizontal="right"
                )
                ws[ws.max_row][1].alignment = openpyxl.styles.Alignment(
                    horizontal="left"
                )
        else:
            ws.append([elem])

    ws.append([])


def _build_footer_row(columns, table_columns, table_name):
    """Zbuduj wiersz sumy i ustaw funkcje sumujące na kolumnach tabeli.

    Sumowana kolumna po stronie XLSX tworzona jest w taki sposób, że jeżeli
    jakakolwiek kolumna tabeli ma footer, to jest tam wstawiana suma za pomocą
    funkcji =SUBTOTAL(109, ...). Pierwsza kolumna (o indeksie zerowym) używana
    jest dla napisu "Suma".
    """
    footer_row = []
    for no, elem in enumerate(columns):
        if no == 0:
            footer_row.append("Suma")
            table_columns[0].totalsRowLabel = footer_row[0]
            continue

        if elem.has_footer():
            table_columns[no].totalsRowFunction = "sum"
            footer_row.append(f"=SUBTOTAL(109,{table_name}[{elem.header}])")
            continue

        footer_row.append("")
    return footer_row


def _add_table(ws, table_name, first_table_row, table_columns):
    max_column_letter = get_column_letter(ws.max_column)
    max_row = ws.max_row

    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tab = Table(
        displayName=table_name,
        ref=f"A{first_table_row}:{max_column_letter}{max_row}",
        autoFilter=AutoFilter(
            ref=f"A{first_table_row}:{max_column_letter}{max_row - 1}"
        ),
        totalsRowShown=True,
        totalsRowCount=1,
        tableStyleInfo=style,
        tableColumns=table_columns,
    )
    ws.add_table(tab)


def _autofit_columns(ws, max_width=75):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        # Since Openpyxl 2.6, the column name is ".column_letter" as .column
        # became the column number (1-based)
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except (ValueError, TypeError):
                pass
        adjusted_width = (max_length + 2) * 1.1
        if adjusted_width > max_width:
            adjusted_width = max_width
        ws.column_dimensions[column].width = adjusted_width


class MyTableExport(TableExport):
    """
    Fix https://github.com/jazzband/tablib/issues/252
    """

    FORMATS = {
        "xlsx": "application/vnd.ms-excel",
    }

    @classmethod
    def is_valid_format(self, export_format):
        if export_format not in ["xlsx", "pdf"]:
            return False
        return True

    def __init__(
        self,
        export_format,
        table,
        exclude_columns=None,
        export_description=None,
        max_rows=None,
    ):
        # Celowo NIE wołamy super().__init__ — bazowy TableExport zbudowałby od
        # razu pełny tablib.Dataset (materializacja WSZYSTKICH wierszy przez
        # table.as_values()). Strumieniujemy zamiast tego prosto do openpyxl
        # (jedna kopia mniej), więc dataset jest zbędny.
        if not self.is_valid_format(export_format):
            raise TypeError(f'Export format "{export_format}" is not supported.')
        self.format = export_format
        self.table = table
        self.exclude_columns = exclude_columns
        self.export_description = export_description
        self.max_rows = max_rows

    def export(self):
        return getattr(self, f"export_{self.format}")()

    def export_xlsx(self):
        # Limit wierszy sprawdzamy PRZED iteracją — len(table.rows) na tabeli
        # paginowanej to zwykły COUNT (bez materializacji), więc odmawiamy zanim
        # queryset wciągnie cały wynik do RAM. Bez cichego ucięcia: podnosimy
        # wyjątek z czytelnym komunikatem (łapany na granicy widoku → HTTP 400).
        if self.max_rows is not None:
            wiersze = len(self.table.rows)
            if wiersze > self.max_rows:
                raise ExportRowLimitExceeded(
                    f"Eksport XLSX jest dostępny dla maksymalnie {self.max_rows} "
                    f"wierszy, a ten raport ma ich {wiersze}. Zawęź filtry "
                    "(np. rok, jednostkę, dyscyplinę) i spróbuj ponownie."
                )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet 1"

        table_name = "Table1"

        if self.export_description:
            _write_export_description(ws, self.export_description)

        # table.as_values() to generator: pierwszy element = nagłówki, kolejne =
        # wiersze danych. Zamiast materializować pełny tablib.Dataset i kopiować
        # go do openpyxl, dopisujemy wiersze wprost, w miarę jak spływają.
        rows = self.table.as_values(exclude_columns=self.exclude_columns)
        headers = next(rows, [])

        # Write the header row and make cells bold
        ws.append(headers)

        table_columns = tuple(
            TableColumn(id=h, name=header) for h, header in enumerate(headers, start=1)
        )

        footer_row = _build_footer_row(self.table.columns, table_columns, table_name)

        for cell in ws[ws.max_row : ws.max_row]:
            cell.font = openpyxl.styles.Font(bold=True)

        first_table_row = ws.max_row
        data_rows = 0
        for row in rows:
            ws.append(row)
            data_rows += 1
        ws.append(footer_row)

        if data_rows:
            _add_table(ws, table_name, first_table_row, table_columns)

        _autofit_columns(ws)

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


class MyExportMixin(ExportMixin):
    # Limit wierszy synchronicznego eksportu. Widoki, których eksport biegnie
    # asynchronicznie/off-request (albo świadomie dopuszczają duże pliki), mogą
    # nadpisać na None, żeby wyłączyć bramkę.
    export_max_rows = RAPORT_SLOTOW_EXPORT_MAX_ROWS

    def get_export_description(self):
        """Nadpisz tą funkcję, aby wygenerować pola opisowe na potrzeby XLS, np.
        'metkę' z parametrami raportu. Powinna zwrócić listę ciągów znaków, które zostaną
        wstawione przed tabelę, jeden pod drugim."""
        return

    def get_export_max_rows(self):
        return self.export_max_rows

    def create_export(self, export_format):
        exporter = MyTableExport(
            export_format=export_format,
            table=self.get_table(**self.get_table_kwargs()),
            exclude_columns=self.exclude_columns,
            export_description=self.get_export_description(),
            max_rows=self.get_export_max_rows(),
        )

        try:
            return exporter.response(filename=self.get_export_filename(export_format))
        except ExportRowLimitExceeded as e:
            return HttpResponseBadRequest(str(e))


class InitialValuesFromGETMixin:
    def get_initial(self):
        initial = super().get_initial()
        if hasattr(self, "request"):
            for elem in self.get_form_class().base_fields.keys():
                value = self.request.GET.get(elem)
                if value is not None:
                    initial[elem] = value
        return initial
