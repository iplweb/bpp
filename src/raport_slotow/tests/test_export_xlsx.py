"""Charakteryzacyjne testy dla raport_slotow.util.MyTableExport.export_xlsx.

Pinują OBECNE zachowanie generatora XLSX (openpyxl): wiersze opisu,
nagłówek, wiersze danych, wiersz sumy (footer), zdefiniowana tabela
i autofiltr. Umożliwiają bezpieczny refactor (zdjęcie C901).

Nie wymagają bazy danych — django_tables2.Table budowany jest z listy
słowników.
"""

from io import BytesIO

import django_tables2 as tables
import openpyxl
import pytest


class _DemoTable(tables.Table):
    name = tables.Column()
    value = tables.Column(footer=lambda table: sum(int(r["value"]) for r in table.data))


DATA = [{"name": "Alpha", "value": 10}, {"name": "Beta", "value": 20}]


def _load(bytes_):
    return openpyxl.load_workbook(BytesIO(bytes_)).active


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


def test_export_bez_opisu_naglowek_dane_i_suma():
    from raport_slotow.util import MyTableExport

    ws = _load(MyTableExport("xlsx", _DemoTable(DATA)).export_xlsx())
    assert ws.title == "Sheet 1"
    assert _rows(ws) == [
        ("Name", "Value"),
        ("Alpha", 10),
        ("Beta", 20),
        ("Suma", "=SUBTOTAL(109,Table1[Value])"),
    ]


def test_export_naglowek_jest_pogrubiony():
    from raport_slotow.util import MyTableExport

    ws = _load(MyTableExport("xlsx", _DemoTable(DATA)).export_xlsx())
    assert ws["A1"].font.bold is True
    assert ws["B1"].font.bold is True
    # Wiersze danych nie są pogrubione
    assert ws["A2"].font.bold is False


def test_export_definiuje_tabele_z_autofiltrem():
    from raport_slotow.util import MyTableExport

    ws = _load(MyTableExport("xlsx", _DemoTable(DATA)).export_xlsx())
    assert "Table1" in ws.tables
    tab = ws.tables["Table1"]
    # Nagłówek (A1) .. footer (A4) — autofiltr pomija footer (do A3).
    assert tab.ref == "A1:B4"
    assert tab.autoFilter.ref == "A1:B3"
    assert tab.totalsRowShown is True
    assert tab.totalsRowCount == 1
    cols = {c.name: c for c in tab.tableColumns}
    assert cols["Name"].totalsRowLabel == "Suma"
    assert cols["Name"].totalsRowFunction is None
    assert cols["Value"].totalsRowFunction == "sum"
    assert cols["Value"].totalsRowLabel is None


def test_export_z_opisem_dwuelementowe_krotki_pogrubione_i_wyrownane():
    from raport_slotow.util import MyTableExport

    description = [("Nazwa:", "raport"), ("Rok:", 2025)]
    ws = _load(
        MyTableExport(
            "xlsx", _DemoTable(DATA), export_description=description
        ).export_xlsx()
    )
    rows = _rows(ws)
    assert rows[0] == ("Nazwa:", "raport")
    assert rows[1] == ("Rok:", 2025)
    # Pusty wiersz separujący opis od tabeli
    assert rows[2] == (None, None)
    assert rows[3] == ("Name", "Value")
    # Pierwsza kolumna opisu: pogrubiona, wyrównana do prawej; druga do lewej
    assert ws["A1"].font.bold is True
    assert ws["A1"].alignment.horizontal == "right"
    assert ws["B1"].alignment.horizontal == "left"


def test_export_opis_krotka_inna_niz_dwuelementowa_bez_pogrubienia():
    from raport_slotow.util import MyTableExport

    # Krotka o długości != 2 trafia do ws.append, ale bez pogrubienia/align.
    description = [(1, 2, 3)]
    ws = _load(
        MyTableExport(
            "xlsx", _DemoTable(DATA), export_description=description
        ).export_xlsx()
    )
    rows = _rows(ws)
    assert rows[0] == (1, 2, 3)
    assert ws["A1"].font.bold is False


def test_export_opis_element_nieiterowalny_jest_owijany_w_liste():
    from raport_slotow.util import MyTableExport

    # Element nie-Iterable (int) idzie gałęzią else: ws.append([elem]).
    ws = _load(
        MyTableExport("xlsx", _DemoTable(DATA), export_description=[99]).export_xlsx()
    )
    assert _rows(ws)[0] == (99, None)


def test_export_pusty_dataset_nie_tworzy_tabeli():
    from raport_slotow.util import MyTableExport

    ws = _load(MyTableExport("xlsx", _DemoTable([])).export_xlsx())
    # Brak danych → blok "if tablib_dataset" pomijany → brak zdefiniowanej tabeli
    assert len(ws.tables) == 0
    # Nadal jest nagłówek i wiersz sumy
    rows = _rows(ws)
    assert rows[0] == ("Name", "Value")
    assert rows[-1][0] == "Suma"


def test_export_zwraca_bajty():
    from raport_slotow.util import MyTableExport

    out = MyTableExport("xlsx", _DemoTable(DATA)).export_xlsx()
    assert isinstance(out, bytes)
    assert out[:2] == b"PK"  # ZIP/xlsx magic


def test_string_w_opisie_jest_zgloszany_jako_typeerror():
    """Quirk zachowany: zwykły string w opisie wywala TypeError.

    String jest Iterable, więc trafia do ws.append(str), które openpyxl
    odrzuca. To istniejące zachowanie — NIE naprawiamy go w refaktorze.
    """
    from raport_slotow.util import MyTableExport

    with pytest.raises(TypeError):
        MyTableExport(
            "xlsx", _DemoTable(DATA), export_description=["samtekst"]
        ).export_xlsx()
