from io import BytesIO

from django.urls import reverse
from openpyxl import load_workbook


def test_raport_slotow_formularz(admin_client):
    res = admin_client.get(reverse("raport_slotow:index"))
    assert res.status_code == 200


def test_raport_slotow_autor_brak_danych(admin_client, autor_jan_kowalski, rok):
    url = reverse(
        "raport_slotow:raport",
        kwargs={"autor": autor_jan_kowalski.slug, "od_roku": rok, "do_roku": rok,},
    )

    res = admin_client.get(url)
    assert res.status_code == 200
    assert "Brak danych" in res.rendered_content

    res = admin_client.get(url + "?_export=xlsx")
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    assert len(wb.get_sheet_names()) > 0


def test_raport_slotow_autor_sa_dane(
    admin_client, autor_jan_kowalski, rekord_slotu, rok
):
    url = reverse(
        "raport_slotow:raport",
        kwargs={"autor": autor_jan_kowalski.slug, "od_roku": rok, "do_roku": rok,},
    )

    res = admin_client.get(url)
    assert res.status_code == 200
    assert "Brak danych" not in res.rendered_content

    res = admin_client.get(url + "?_export=xlsx")
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    assert len(wb.get_sheet_names()) > 0
