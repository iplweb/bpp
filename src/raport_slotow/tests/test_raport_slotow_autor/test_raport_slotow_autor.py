import sys
from io import BytesIO

import PyPDF2
import pytest
from django.urls import reverse
from openpyxl import load_workbook

from raport_slotow import const
from raport_slotow.forms import AutorRaportSlotowForm
from raport_slotow.views import SESSION_KEY


def test_raport_slotow_formularz(admin_client):
    res = admin_client.get(reverse("raport_slotow:index"))
    assert res.status_code == 200


def test_raport_slotow_autor_brak_danych(admin_client, autor_jan_kowalski, rok):
    url = reverse(
        "raport_slotow:raport",
    )

    dane_raportu = {
        "obiekt": autor_jan_kowalski.pk,
        "od_roku": rok,
        "do_roku": rok,
        "dzialanie": const.DZIALANIE_WSZYSTKO,
        "minimalny_pk": 0,
        "slot": None,
        "_export": "html",
    }

    form = AutorRaportSlotowForm(dane_raportu)
    assert form.is_valid(), form._errors

    s = admin_client.session
    s.update({SESSION_KEY: dane_raportu})
    s.save()

    res = admin_client.get(url)
    assert res.status_code == 200
    assert "Brak danych" in res.rendered_content

    res = admin_client.get(url + "?_export=xlsx")
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    assert len(wb.get_sheet_names()) > 0


def test_raport_slotow_autor_sa_dane_eksport_wszystkiego(
    admin_client,
    autor_jan_kowalski,
    rekord_slotu,
    rok,
):
    url = reverse("raport_slotow:raport")

    dane_raportu = {
        "obiekt": autor_jan_kowalski.pk,
        "od_roku": rok,
        "do_roku": rok,
        "dzialanie": const.DZIALANIE_WSZYSTKO,
        "minimalny_pk": 0,
        "slot": None,
        "_export": "html",
    }
    s = admin_client.session
    s.update({SESSION_KEY: dane_raportu})
    s.save()

    res = admin_client.get(url)
    assert res.status_code == 200
    assert "Brak danych" not in res.rendered_content

    res = admin_client.get(url + "?_export=xlsx")
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    assert len(wb.get_sheet_names()) > 0


@pytest.mark.skipif(sys.platform == "darwin", reason="libpango macports 1.42?")
def test_raport_slotow_autor_sa_dane_eksport_wszystkiego_do_pdf(
    admin_client,
    autor_jan_kowalski,
    rekord_slotu,
    rok,
):
    url = reverse("raport_slotow:raport")

    dane_raportu = {
        "obiekt": autor_jan_kowalski.pk,
        "od_roku": rok,
        "do_roku": rok,
        "dzialanie": const.DZIALANIE_WSZYSTKO,
        "minimalny_pk": 0,
        "slot": None,
        "_export": "html",
    }
    s = admin_client.session
    s.update({SESSION_KEY: dane_raportu})
    s.save()

    res = admin_client.get(url + "?_export=pdf")
    assert res.status_code == 200
    pdfReader = PyPDF2.PdfFileReader(BytesIO(res.content))
    assert pdfReader.numPages >= 1


def test_raport_slotow_autor_zbieraj_slot(
    admin_client, autor_jan_kowalski, rekord_slotu, rok
):
    url = reverse("raport_slotow:raport")

    dane_raportu = {
        "obiekt": autor_jan_kowalski.pk,
        "od_roku": rok,
        "do_roku": rok,
        "dzialanie": const.DZIALANIE_SLOT,
        "minimalny_pk": 0,
        "slot": 20,
        "_export": "html",
    }
    s = admin_client.session
    s.update({SESSION_KEY: dane_raportu})
    s.save()

    res = admin_client.get(url)
    assert res.status_code == 200
    assert "Brak danych" not in res.rendered_content

    res = admin_client.get(url + "?_export=xlsx")
    assert res.status_code == 200
    wb = load_workbook(BytesIO(res.content))
    assert len(wb.get_sheet_names()) > 0


def test_raport_slotow_autor_wartosci_poczatkowe(admin_client):
    url = reverse("raport_slotow:index")
    res = admin_client.get(url, dict(od_roku=5000))
    assert b"5000" in res.content


@pytest.mark.parametrize(
    "dzialanie,slot", [(const.DZIALANIE_WSZYSTKO, None), (const.DZIALANIE_SLOT, 20)]
)
def test_raport_slotow_autor_sa_dane_minimalny_pk(
    admin_client, autor_jan_kowalski, rekord_slotu, rok, dzialanie, slot
):
    w = rekord_slotu.rekord
    w.punkty_pk = 10
    w.save()

    url = reverse("raport_slotow:raport")

    dane_raportu = {
        "obiekt": autor_jan_kowalski.pk,
        "od_roku": rok,
        "do_roku": rok,
        "dzialanie": dzialanie,
        "minimalny_pk": 0,
        "slot": slot,
        "_export": "html",
    }
    s = admin_client.session
    s.update({SESSION_KEY: dane_raportu})
    s.save()

    res = admin_client.get(url)
    assert res.status_code == 200
    assert "Brak danych" not in res.rendered_content

    dane_raportu = {
        "obiekt": autor_jan_kowalski.pk,
        "od_roku": rok,
        "do_roku": rok,
        "dzialanie": dzialanie,
        "minimalny_pk": 200,
        "slot": slot,
        "_export": "html",
    }
    s = admin_client.session
    s.update({SESSION_KEY: dane_raportu})
    s.save()

    res = admin_client.get(url)
    assert res.status_code == 200
    assert "Brak danych" in res.rendered_content
