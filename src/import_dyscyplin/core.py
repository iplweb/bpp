from _decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from bpp.models import Jednostka
from import_common.core import matchuj_autora, matchuj_jednostke, matchuj_wydzial
from import_common.exceptions import (
    BadNoOfSheetsException,
    HeaderNotFoundException,
    ImproperFileException,
)
from import_dyscyplin.models import Import_Dyscyplin_Row


def _matchuj_wydzial_z_cache(wydzial_cache, nazwa_wydzialu):
    """Matchuje wydział z cache lub z bazy."""
    wydzial = wydzial_cache.get(nazwa_wydzialu)
    if not wydzial and nazwa_wydzialu:
        try:
            wydzial = wydzial_cache[nazwa_wydzialu] = matchuj_wydzial(nazwa_wydzialu)
        except KeyError:
            pass
    return wydzial


def _matchuj_jednostke_z_cache(jednostka_cache, nazwa_jednostki):
    """Matchuje jednostkę z cache lub z bazy."""
    jednostka = jednostka_cache.get(nazwa_jednostki)
    if not jednostka and nazwa_jednostki:
        try:
            jednostka = jednostka_cache[nazwa_jednostki] = matchuj_jednostke(
                nazwa_jednostki
            )
        except (KeyError, Jednostka.DoesNotExist, Jednostka.MultipleObjectsReturned):
            jednostka = None
    return jednostka


def _parsuj_procenty(original):
    """Parsuje pola procentowe i zwraca flagę błędu."""
    bledny = False
    for elem in ["procent dyscypliny", "procent subdyscypliny"]:
        v = original.get(elem)
        if v is None:
            continue
        if str(v) == "":
            original[elem] = None
            continue
        try:
            original[elem] = Decimal(original.get(elem))
        except (TypeError, InvalidOperation):
            original[elem] = None
            bledny = True
    return bledny


def _utworz_wiersz_importu(n, parent, original, jednostka, wydzial, autor, bledny):
    """Tworzy i zapisuje wiersz importu."""
    i = Import_Dyscyplin_Row(
        row_no=n,
        parent=parent,
        original=original,
        nazwisko=original["nazwisko"],
        imiona=original["imię"],
        nazwa_jednostki=original.get("nazwa_jednostki") or "",
        jednostka=jednostka,
        nazwa_wydzialu=original.get("wydział") or "",
        wydzial=wydzial,
        autor=autor,
        dyscyplina=original.get("dyscyplina") or "",
        kod_dyscypliny=original.get("kod dyscypliny") or "",
        procent_dyscypliny=original.get("procent dyscypliny"),
        subdyscyplina=original.get("subdyscyplina") or "",
        kod_subdyscypliny=original.get("kod subdyscypliny") or "",
        procent_subdyscypliny=original.get("procent subdyscypliny"),
    )

    if not i.dyscyplina:
        i.procent_dyscypliny = None

    if not i.subdyscyplina:
        i.procent_subdyscypliny = None

    if autor is None:
        i.stan = Import_Dyscyplin_Row.STAN.BLEDNY
        i.info = "Nie można dopasować autora"
    elif not i.dyscyplina and i.subdyscyplina:
        i.stan = Import_Dyscyplin_Row.STAN.BLEDNY
        i.info = "Dyscyplina pusta, subdyscyplina ustawiona."
    elif bledny:
        i.stan = Import_Dyscyplin_Row.STAN.BLEDNY
        i.info = "niepoprawna wartość w polu procent dyscypliny lub subdyscypliny"

    i.save()


def przeanalizuj_plik_xls(sciezka, parent):
    """
    :param sciezka:
    :return: (success:boolean, message)
    """
    try:
        f = openpyxl.load_workbook(sciezka)
    except InvalidFileException as e:
        raise ImproperFileException(e) from e

    if len(f.worksheets) != 1:
        raise BadNoOfSheetsException()

    s = f.worksheets[0]

    if parent.kolumna_set.all().count() == 0:
        raise HeaderNotFoundException()

    naglowek = [k.rodzaj_pola for k in parent.kolumna_set.all()]
    wydzial_cache = {}
    jednostka_cache = {}

    for n, row in enumerate(
        s.iter_rows(
            min_row=parent.wiersz_naglowka + 1, max_row=s.max_row, values_only=True
        ),
        parent.wiersz_naglowka + 1,
    ):
        original = dict(zip(naglowek, [elem for elem in row], strict=False))

        if original["nazwisko"] is None or (not original["nazwisko"].strip()):
            continue

        original["nazwa_jednostki"] = original.get("nazwa jednostki")

        wydzial = _matchuj_wydzial_z_cache(wydzial_cache, original.get("wydział"))
        jednostka = _matchuj_jednostke_z_cache(
            jednostka_cache, original.get("nazwa jednostki")
        )

        autor = matchuj_autora(
            original["imię"],
            original["nazwisko"],
            jednostka=jednostka,
            orcid=original.get("orcid"),
            pbn_id=original.get("pbn_id"),
            tytul_str=original.get("tytuł"),
        )

        bledny = _parsuj_procenty(original)
        _utworz_wiersz_importu(n, parent, original, jednostka, wydzial, autor, bledny)

    return (True, f"przeanalizowano {n - parent.wiersz_naglowka} rekordow")
