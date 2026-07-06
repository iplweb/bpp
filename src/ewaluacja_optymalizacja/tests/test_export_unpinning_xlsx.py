"""Characterization tests dla ``export_unpinning_opportunities_xlsx``.

Te testy pinują OBECNE zachowanie eksportu XLSX (zawartość komórek, filtry,
stopkę) zanim funkcja zostanie uproszczona/zrefaktoryzowana. Mają przechodzić
na nie-zmienionym kodzie; po refaktorze muszą pozostać zielone bez modyfikacji.
"""

import io
from decimal import Decimal

import pytest
from django.urls import reverse
from model_bakery import baker
from openpyxl import load_workbook

from bpp.models import (
    Autor,
    Dyscyplina_Naukowa,
    Jednostka,
    Rekord,
    Uczelnia,
)
from ewaluacja_common.models import Rodzaj_Autora
from ewaluacja_metryki.models import MetrykaAutora

from ..models import UnpinningOpportunity

EXPORT_URL = "ewaluacja_optymalizacja:unpinning-export-xlsx"


@pytest.fixture
def uczelnia():
    # Multi-hosted: mapuj domenę klienta testowego ("testserver") na uczelnię,
    # by uczelnia_dla_odczytu(request) w eksporcie ją rozstrzygnął (po scaleniu
    # dev eksport używa multi-host resolvera, nie Uczelnia.objects.first()).
    from django.contrib.sites.models import Site

    site, _ = Site.objects.get_or_create(
        domain="testserver", defaults={"name": "testserver"}
    )
    return baker.make(Uczelnia, nazwa="Testowa Uczelnia", site=site)


@pytest.fixture
def dyscyplina():
    return baker.make(Dyscyplina_Naukowa, nazwa="Nauki testowe")


@pytest.fixture
def rodzaj_autora_n():
    # update_or_create (nie get_or_create): asercja w teście oczekuje
    # konkretnej nazwy "pracownik naukowy w liczbie N". Baseline bywa
    # niedeterministyczny per-shard (raz ma Rodzaj_Autora(N), raz nie) —
    # get_or_create z krótkim defaultem dawał flakę CI zależną od kolejności.
    obj, _ = Rodzaj_Autora.objects.update_or_create(
        skrot="N",
        defaults=dict(nazwa="pracownik naukowy w liczbie N", sort=1),
    )
    return obj


def _make_metryka(autor, dyscyplina, jednostka, *, procent, nazbierany, maksymalny):
    return baker.make(
        MetrykaAutora,
        autor=autor,
        dyscyplina_naukowa=dyscyplina,
        jednostka=jednostka,
        rodzaj_autora="N",
        procent_wykorzystania_slotow=Decimal(procent),
        slot_nazbierany=Decimal(nazbierany),
        slot_maksymalny=Decimal(maksymalny),
        punkty_nazbierane=Decimal("100"),
    )


def _make_opportunity(
    uczelnia,
    dyscyplina,
    wydawnictwo_zwarte,
    *,
    punkty_kbn="140.00",
    makes_sense=True,
    rekord_tytul="Tytuł testowej pracy",
):
    """Zbuduj kompletny ``UnpinningOpportunity`` z prawdziwym wpisem Rekord."""
    wydawnictwo_zwarte.punkty_kbn = Decimal(punkty_kbn)
    wydawnictwo_zwarte.save()
    Rekord.objects.full_refresh()
    rekord = Rekord.objects.get_for_model(wydawnictwo_zwarte)

    jednostka_a = baker.make(Jednostka, nazwa="Jednostka A")
    jednostka_b = baker.make(Jednostka, nazwa="Jednostka B")
    autor_a = baker.make(
        Autor, nazwisko="Kowalski", imiona="Jan", system_kadrowy_id=123
    )
    autor_b = baker.make(Autor, nazwisko="Nowak", imiona="Anna", system_kadrowy_id=456)
    metryka_a = _make_metryka(
        autor_a,
        dyscyplina,
        jednostka_a,
        procent="50.00",
        nazbierany="2.0000",
        maksymalny="4.0000",
    )
    metryka_b = _make_metryka(
        autor_b,
        dyscyplina,
        jednostka_b,
        procent="75.00",
        nazbierany="3.0000",
        maksymalny="4.0000",
    )

    return baker.make(
        UnpinningOpportunity,
        uczelnia=uczelnia,
        dyscyplina_naukowa=dyscyplina,
        rekord_id=rekord.pk,
        rekord_tytul=rekord_tytul,
        autor_could_benefit=autor_a,
        metryka_could_benefit=metryka_a,
        autor_currently_using=autor_b,
        metryka_currently_using=metryka_b,
        slot_in_work=Decimal("1.0000"),
        slots_missing=Decimal("0.5000"),
        makes_sense=makes_sense,
        punkty_roznica_a=Decimal("10.0000"),
        sloty_roznica_a=Decimal("0.2500"),
        punkty_roznica_b=Decimal("20.0000"),
        sloty_roznica_b=Decimal("0.5000"),
    )


def _load_sheet(response):
    wb = load_workbook(io.BytesIO(response.content))
    return wb.active


EXPECTED_HEADERS = [
    "Lp.",
    "Tytuł pracy",
    "Punktacja źródła",
    "Dyscyplina",
    "Autor A (do odpięcia)",
    "Rodzaj autora A",
    "ID systemu kadrowego A",
    "Jednostka A",
    "% wykorzystania slotów A",
    "Sloty nazbierane A",
    "Sloty maksymalne A",
    "B może wziąć (slotów)",
    "Slot w pracy",
    "Różnica punktów A",
    "Różnica slotów A",
    "Różnica punktów B",
    "Różnica slotów B",
    "Autor B (skorzysta)",
    "Rodzaj autora B",
    "ID systemu kadrowego B",
    "Jednostka B",
    "% wykorzystania slotów B",
    "Sloty nazbierane B",
    "Sloty maksymalne B",
    "Sensowne?",
]


@pytest.mark.django_db
def test_export_no_uczelnia_redirects(client, admin_user):
    """Bez uczelni w systemie eksport przekierowuje na index."""
    Uczelnia.objects.all().delete()
    client.force_login(admin_user)
    response = client.get(reverse(EXPORT_URL))
    assert response.status_code == 302


@pytest.mark.django_db
def test_export_empty_has_headers_and_no_filters(client, admin_user, uczelnia):
    """Pusta lista: poprawne nagłówki, content-type XLSX, 'Brak filtrów'."""
    client.force_login(admin_user)
    response = client.get(reverse(EXPORT_URL))

    assert response.status_code == 200
    assert response["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment; filename=" in response["Content-Disposition"]

    ws = _load_sheet(response)
    headers = [
        ws.cell(row=1, column=c).value for c in range(1, len(EXPECTED_HEADERS) + 1)
    ]
    assert headers == EXPECTED_HEADERS

    # Stopka: brak danych -> "Brak filtrów" gdzieś w arkuszu
    all_values = {
        ws.cell(row=r, column=c).value
        for r in range(1, ws.max_row + 1)
        for c in range(1, 3)
    }
    assert "Brak filtrów" in all_values
    assert "Podsumowanie:" in all_values


@pytest.mark.django_db
def test_export_single_row_cell_values(
    client, admin_user, uczelnia, dyscyplina, rodzaj_autora_n, wydawnictwo_zwarte
):
    """Pojedynczy pełny wiersz: pinujemy zawartość kluczowych kolumn."""
    _make_opportunity(
        uczelnia,
        dyscyplina,
        wydawnictwo_zwarte,
        punkty_kbn="140.00",
        makes_sense=True,
        rekord_tytul="Tytuł testowej pracy",
    )

    client.force_login(admin_user)
    response = client.get(reverse(EXPORT_URL))
    assert response.status_code == 200

    ws = _load_sheet(response)
    row = 2  # pierwszy wiersz danych

    def cell(col):
        return ws.cell(row=row, column=col).value

    assert cell(1) == 1  # Lp.
    assert cell(2) == "Tytuł testowej pracy"  # Tytuł pracy
    assert cell(3) == 140.0  # Punktacja źródła (float)
    assert cell(4) == "Nauki testowe"  # Dyscyplina
    assert "Kowalski" in cell(5)  # Autor A
    assert (
        cell(6) == "pracownik naukowy w liczbie N"
    )  # Rodzaj autora A (lookup skrot N)
    assert cell(7) == 123  # ID systemu kadrowego A
    assert cell(8) == "Jednostka A"  # Jednostka A
    assert cell(9) == pytest.approx(0.5)  # % wykorzystania slotów A (50/100)
    assert cell(10) == pytest.approx(2.0)  # Sloty nazbierane A
    assert cell(11) == pytest.approx(4.0)  # Sloty maksymalne A
    assert cell(12) == pytest.approx(0.5)  # B może wziąć (slots_missing)
    assert cell(13) == pytest.approx(1.0)  # Slot w pracy
    assert cell(14) == pytest.approx(10.0)  # Różnica punktów A
    assert cell(15) == pytest.approx(0.25)  # Różnica slotów A
    assert cell(16) == pytest.approx(20.0)  # Różnica punktów B
    assert cell(17) == pytest.approx(0.5)  # Różnica slotów B
    assert "Nowak" in cell(18)  # Autor B
    assert cell(19) == "pracownik naukowy w liczbie N"  # Rodzaj autora B
    assert cell(20) == 456  # ID systemu kadrowego B
    assert cell(21) == "Jednostka B"  # Jednostka B
    assert cell(22) == pytest.approx(0.75)  # % wykorzystania slotów B
    assert cell(23) == pytest.approx(3.0)  # Sloty nazbierane B
    assert cell(24) == pytest.approx(4.0)  # Sloty maksymalne B
    assert cell(25) == "TAK"  # Sensowne?


@pytest.mark.django_db
def test_export_only_sensible_filter_in_footer(
    client, admin_user, uczelnia, dyscyplina, rodzaj_autora_n, wydawnictwo_zwarte
):
    """only_sensible=1: w stopce pojawia się 'Tylko sensowne: TAK'."""
    _make_opportunity(uczelnia, dyscyplina, wydawnictwo_zwarte, makes_sense=True)
    client.force_login(admin_user)
    response = client.get(reverse(EXPORT_URL), {"only_sensible": "1"})
    assert response.status_code == 200

    ws = _load_sheet(response)
    col1 = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any(v == "Tylko sensowne: TAK" for v in col1)


@pytest.mark.django_db
def test_export_dyscyplina_filter_in_footer(
    client, admin_user, uczelnia, dyscyplina, rodzaj_autora_n, wydawnictwo_zwarte
):
    """Filtr dyscypliny: w stopce 'Dyscyplina: <nazwa>'."""
    _make_opportunity(uczelnia, dyscyplina, wydawnictwo_zwarte)
    client.force_login(admin_user)
    response = client.get(reverse(EXPORT_URL), {"dyscyplina": str(dyscyplina.pk)})
    assert response.status_code == 200

    ws = _load_sheet(response)
    col1 = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any(v == f"Dyscyplina: {dyscyplina.nazwa}" for v in col1)


@pytest.mark.django_db
def test_export_punktacja_zrodla_filter_buckets(
    client, admin_user, uczelnia, dyscyplina, rodzaj_autora_n, wydawnictwo_zwarte
):
    """punktacja_zrodla=140-200 zostawia pracę o punktach 140 i odnotowuje filtr."""
    _make_opportunity(uczelnia, dyscyplina, wydawnictwo_zwarte, punkty_kbn="140.00")
    client.force_login(admin_user)
    response = client.get(reverse(EXPORT_URL), {"punktacja_zrodla": "140-200"})
    assert response.status_code == 200

    ws = _load_sheet(response)
    # wiersz danych obecny (Lp. == 1)
    assert ws.cell(row=2, column=1).value == 1
    col1 = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    assert any(v == "Punktacja źródła: 140-200 punktów" for v in col1)
