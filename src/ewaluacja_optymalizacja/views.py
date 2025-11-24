import logging
from decimal import Decimal

from celery.result import AsyncResult
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q
from django.shortcuts import get_object_or_404, redirect, render

from bpp.models import Uczelnia

from .models import OptimizationRun
from .tasks import solve_all_reported_disciplines

logger = logging.getLogger(__name__)


def _add_run_statistics(run, uczelnia):
    """Add statistics to an optimization run object."""
    from ewaluacja_liczba_n.models import LiczbaNDlaUczelni

    # Użyj _calculate_liczba_n_stats żeby dostać procent slotów poza N
    author_results = run.author_results.select_related("rodzaj_autora").all()
    stats = _calculate_liczba_n_stats(run, author_results)

    # Dodaj procent slotów poza N do obiektu run
    run.outside_n_slots_pct = stats["non_n_slots_pct"]

    # Dodaj statystyki przypięć dla dyscypliny
    run.pin_stats = _get_discipline_pin_stats(run.dyscyplina_naukowa)

    # Pobierz liczbę N dla dyscypliny i oblicz 3*N oraz procent wypełnienia slotów
    try:
        liczba_n_obj = LiczbaNDlaUczelni.objects.get(
            uczelnia=uczelnia, dyscyplina_naukowa=run.dyscyplina_naukowa
        )
        run.liczba_3n = liczba_n_obj.liczba_n * 3
        if run.liczba_3n > 0:
            run.slots_fill_pct = (run.total_slots / run.liczba_3n) * 100
        else:
            run.slots_fill_pct = None
    except LiczbaNDlaUczelni.DoesNotExist:
        run.liczba_3n = None
        run.slots_fill_pct = None

    return run


def _calculate_summary_statistics(recent_runs):
    """Calculate summary statistics from a list of runs."""
    summary = {
        "total_points": Decimal("0"),
        "total_publications": 0,
        "total_slots": Decimal("0"),
        "total_3n": Decimal("0"),
        "total_low_mono_weighted": Decimal("0"),  # suma LOW-MONO % * liczba publikacji
        "total_outside_n_weighted": Decimal("0"),  # suma sloty poza N % * liczba slotów
        "total_slots_fill_weighted": Decimal("0"),  # suma Sloty % * liczba slotów
        "total_publications_for_low_mono": 0,  # suma publikacji (dla średniej ważonej)
        "total_slots_for_outside_n": Decimal("0"),  # suma slotów (dla średniej ważonej)
        "total_slots_for_fill_pct": Decimal(
            "0"
        ),  # suma slotów z danymi 3*N (dla średniej ważonej)
        "total_pinned": 0,
        "total_unpinned": 0,
    }

    for run in recent_runs:
        summary["total_points"] += run.total_points
        summary["total_publications"] += run.total_publications
        summary["total_slots"] += run.total_slots
        summary["total_pinned"] += run.pin_stats["pinned"]
        summary["total_unpinned"] += run.pin_stats["unpinned"]

        # Dodaj 3*N do sumy jeśli dostępne
        if run.liczba_3n is not None:
            summary["total_3n"] += run.liczba_3n

        # Dla średniej ważonej LOW-MONO % (ważona liczbą publikacji)
        if run.total_publications > 0:
            summary["total_low_mono_weighted"] += (
                run.low_mono_percentage * run.total_publications
            )
            summary["total_publications_for_low_mono"] += run.total_publications

        # Dla średniej ważonej slotów poza N % (ważona liczbą slotów)
        if run.total_slots > 0:
            summary["total_outside_n_weighted"] += (
                Decimal(str(run.outside_n_slots_pct)) * run.total_slots
            )
            summary["total_slots_for_outside_n"] += run.total_slots

        # Dla średniej ważonej Sloty % (ważona liczbą slotów)
        if run.total_slots > 0 and run.slots_fill_pct is not None:
            summary["total_slots_fill_weighted"] += (
                Decimal(str(run.slots_fill_pct)) * run.total_slots
            )
            summary["total_slots_for_fill_pct"] += run.total_slots

    # Oblicz średnie ważone
    if summary["total_publications_for_low_mono"] > 0:
        summary["avg_low_mono_pct"] = (
            summary["total_low_mono_weighted"]
            / summary["total_publications_for_low_mono"]
        )
    else:
        summary["avg_low_mono_pct"] = Decimal("0")

    if summary["total_slots_for_outside_n"] > 0:
        summary["avg_outside_n_pct"] = (
            summary["total_outside_n_weighted"] / summary["total_slots_for_outside_n"]
        )
    else:
        summary["avg_outside_n_pct"] = Decimal("0")

    if summary["total_slots_for_fill_pct"] > 0:
        summary["avg_slots_fill_pct"] = (
            summary["total_slots_fill_weighted"] / summary["total_slots_for_fill_pct"]
        )
    else:
        summary["avg_slots_fill_pct"] = Decimal("0")

    # Oblicz procenty przypięte/odpięte
    total_pins = summary["total_pinned"] + summary["total_unpinned"]
    if total_pins > 0:
        summary["pinned_pct"] = (summary["total_pinned"] / total_pins) * 100
        summary["unpinned_pct"] = (summary["total_unpinned"] / total_pins) * 100
    else:
        summary["pinned_pct"] = Decimal("0")
        summary["unpinned_pct"] = Decimal("0")

    return summary


def _check_for_problematic_slots():
    """Check if there are problematic slots < 0.1 for authors with disciplines."""
    from bpp.models import Autor_Dyscyplina, Cache_Punktacja_Autora_Query

    autorzy_z_dyscyplinami = set(
        Autor_Dyscyplina.objects.filter(rok__gte=2022, rok__lte=2025)
        .values_list("autor_id", flat=True)
        .distinct()
    )

    return Cache_Punktacja_Autora_Query.objects.filter(
        slot__lt=Decimal("0.1"),
        rekord__rok__gte=2022,
        rekord__rok__lte=2025,
        autor_id__in=autorzy_z_dyscyplinami,
    ).exists()


@login_required
def index(request):
    """
    Główny widok aplikacji ewaluacja_optymalizacja - lista wszystkich kalkulacji
    """

    from denorm.models import DirtyInstance

    from ewaluacja_liczba_n.models import LiczbaNDlaUczelni

    # Sprawdź czy są rekordy do przeliczenia
    dirty_count = DirtyInstance.objects.count()

    runs = OptimizationRun.objects.select_related(
        "dyscyplina_naukowa", "uczelnia"
    ).all()

    # Pobierz uczelnię dla dalszych obliczeń
    uczelnia = Uczelnia.objects.first()

    # Dla każdego run oblicz procent slotów poza N oraz statystyki przypięć
    runs_with_stats = [_add_run_statistics(run, uczelnia) for run in runs]

    # Sprawdź czy są dane liczby N
    liczba_n_count = 0
    liczba_n_raportowane = 0

    if uczelnia:
        liczba_n_count = LiczbaNDlaUczelni.objects.filter(uczelnia=uczelnia).count()
        liczba_n_raportowane = LiczbaNDlaUczelni.objects.filter(
            uczelnia=uczelnia, liczba_n__gte=12
        ).count()

    # Oblicz agregacje dla ostatnich 10 wierszy (dla podsumowania w tabeli)
    recent_runs = runs_with_stats[:10]
    summary = _calculate_summary_statistics(recent_runs)

    # Sprawdź czy są problematyczne sloty < 0.1 dla przycisku weryfikacji
    has_problematic_slots = _check_for_problematic_slots()

    context = {
        "optimization_runs": runs_with_stats,
        "liczba_n_count": liczba_n_count,
        "liczba_n_raportowane": liczba_n_raportowane,
        "summary": summary,
        "dirty_count": dirty_count,
        "has_problematic_slots": has_problematic_slots,
    }

    return render(request, "ewaluacja_optymalizacja/index.html", context)


@login_required
def run_list(request):
    """
    Lista wszystkich wykonanych optymalizacji
    """
    runs = OptimizationRun.objects.select_related(
        "dyscyplina_naukowa", "uczelnia"
    ).order_by("-started_at")

    # Dla każdego run oblicz procent slotów poza N oraz statystyki przypięć
    runs_with_stats = []
    for run in runs:
        # Użyj _calculate_liczba_n_stats żeby dostać procent slotów poza N
        author_results = run.author_results.select_related("rodzaj_autora").all()
        stats = _calculate_liczba_n_stats(run, author_results)

        # Dodaj procent slotów poza N do obiektu run
        run.outside_n_slots_pct = stats["non_n_slots_pct"]

        # Dodaj statystyki przypięć dla dyscypliny
        run.pin_stats = _get_discipline_pin_stats(run.dyscyplina_naukowa)

        runs_with_stats.append(run)

    return render(
        request, "ewaluacja_optymalizacja/run_list.html", {"runs": runs_with_stats}
    )


@login_required
def run_detail(request, pk):
    """
    Szczegółowy widok wyników optymalizacji z wizualizacją
    """
    run = get_object_or_404(
        OptimizationRun.objects.select_related("dyscyplina_naukowa", "uczelnia"), pk=pk
    )

    # Get author results with annotations
    author_results = (
        run.author_results.select_related("autor", "rodzaj_autora")
        .prefetch_related("publications")
        .annotate(
            pub_count=Count("publications"),
            low_mono_count=Count(
                "publications", filter=Q(publications__is_low_mono=True)
            ),
        )
    )

    # Check which authors have evaluation metrics for this discipline
    from ewaluacja_metryki.models import MetrykaAutora

    autor_ids = [ar.autor_id for ar in author_results]
    existing_metrics = set(
        MetrykaAutora.objects.filter(
            autor_id__in=autor_ids, dyscyplina_naukowa=run.dyscyplina_naukowa
        ).values_list("autor_id", flat=True)
    )

    # Add ma_metryke attribute to each author_result
    for author_result in author_results:
        author_result.ma_metryke = author_result.autor_id in existing_metrics

    # Calculate statistics for liczba N analysis
    stats = _calculate_liczba_n_stats(run, author_results)

    # Prepare data for charts
    chart_data = {
        "labels": ["Autorzy w liczbie N", "Autorzy poza liczbą N"],
        "points": [float(stats["n_points"]), float(stats["non_n_points"])],
        "slots": [float(stats["n_slots"]), float(stats["non_n_slots"])],
        "publications": [stats["n_pubs"], stats["non_n_pubs"]],
        "low_mono_data": {
            "labels": ["LOW-MONO", "Inne publikacje"],
            "values": [
                run.low_mono_count,
                run.total_publications - run.low_mono_count,
            ],
        },
    }

    context = {
        "run": run,
        "author_results": author_results,
        "stats": stats,
        "chart_data": chart_data,
    }

    return render(request, "ewaluacja_optymalizacja/run_detail.html", context)


@login_required
def discipline_comparison(request):
    """
    Porównanie wyników dla różnych dyscyplin
    """
    # Get latest run for each discipline

    latest_runs = {}
    all_runs = OptimizationRun.objects.filter(status="completed")

    for run in all_runs:
        disc_id = run.dyscyplina_naukowa_id
        if (
            disc_id not in latest_runs
            or run.started_at > latest_runs[disc_id].started_at
        ):
            latest_runs[disc_id] = run

    runs = list(latest_runs.values())
    runs.sort(key=lambda x: x.dyscyplina_naukowa.nazwa)

    # Calculate stats for each discipline
    discipline_stats = []
    for run in runs:
        author_results = run.author_results.select_related("rodzaj_autora")
        stats = _calculate_liczba_n_stats(run, author_results)
        stats["discipline"] = run.dyscyplina_naukowa.nazwa
        stats["run_id"] = run.pk
        stats["started_at"] = run.started_at

        # Convert Decimal values to proper float strings for JavaScript
        # This ensures we get dots instead of commas regardless of locale
        stats["n_points_js"] = str(float(stats["n_points"]))
        stats["non_n_points_js"] = str(float(stats["non_n_points"]))
        stats["n_slots_js"] = str(float(stats["n_slots"]))
        stats["non_n_slots_js"] = str(float(stats["non_n_slots"]))
        stats["n_slots_pct_js"] = str(float(stats["n_slots_pct"]))
        stats["non_n_slots_pct_js"] = str(float(stats["non_n_slots_pct"]))
        stats["low_mono_pct_js"] = str(float(stats["low_mono_pct"]))
        stats["non_n_and_low_mono_pct_js"] = str(float(stats["non_n_and_low_mono_pct"]))

        discipline_stats.append(stats)

    context = {
        "discipline_stats": discipline_stats,
    }

    return render(
        request, "ewaluacja_optymalizacja/discipline_comparison.html", context
    )


def _calculate_liczba_n_stats(run, author_results):
    """
    Calculate statistics about liczba N vs non-liczba N authors

    Returns dict with:
    - n_points, n_slots, n_pubs: totals for liczba N authors
    - non_n_points, non_n_slots, non_n_pubs: totals for non-liczba N authors
    - n_percentage, non_n_percentage: percentage splits
    - low_mono_percentage: percentage of LOW-MONO publications
    - non_n_and_low_mono_percentage: combined percentage
    """
    # Initialize counters
    n_points = Decimal("0")
    n_slots = Decimal("0")
    n_pubs = 0

    non_n_points = Decimal("0")
    non_n_slots = Decimal("0")
    non_n_pubs = 0

    # Aggregate by liczba N status
    for author_result in author_results:
        is_in_n = author_result.rodzaj_autora and author_result.rodzaj_autora.jest_w_n

        pub_count = author_result.publications.count()

        if is_in_n:
            n_points += author_result.total_points
            n_slots += author_result.total_slots
            n_pubs += pub_count
        else:
            non_n_points += author_result.total_points
            non_n_slots += author_result.total_slots
            non_n_pubs += pub_count

    # Calculate percentages
    total_points = n_points + non_n_points
    total_slots = n_slots + non_n_slots
    total_pubs = n_pubs + non_n_pubs

    n_points_pct = float(n_points / total_points * 100) if total_points > 0 else 0
    non_n_points_pct = (
        float(non_n_points / total_points * 100) if total_points > 0 else 0
    )

    n_slots_pct = float(n_slots / total_slots * 100) if total_slots > 0 else 0
    non_n_slots_pct = float(non_n_slots / total_slots * 100) if total_slots > 0 else 0

    low_mono_pct = float(run.low_mono_percentage)

    # Combined percentage: LOW-MONO + slots from authors outside liczba N
    # Note: This is an approximation since LOW-MONO is percentage of publications, not slots
    # But it gives a rough indicator of "problematic" content
    non_n_and_low_mono_pct = non_n_slots_pct + low_mono_pct

    return {
        "n_points": n_points,
        "n_slots": n_slots,
        "n_pubs": n_pubs,
        "n_points_pct": n_points_pct,
        "n_slots_pct": n_slots_pct,
        "non_n_points": non_n_points,
        "non_n_slots": non_n_slots,
        "non_n_pubs": non_n_pubs,
        "non_n_points_pct": non_n_points_pct,
        "non_n_slots_pct": non_n_slots_pct,
        "low_mono_pct": low_mono_pct,
        "non_n_and_low_mono_pct": non_n_and_low_mono_pct,
        "total_points": total_points,
        "total_slots": total_slots,
        "total_pubs": total_pubs,
    }


@login_required
def start_bulk_optimization(request):
    """
    Uruchamia zadanie Celery do optymalizacji wszystkich dyscyplin raportowanych.
    """
    from denorm.models import DirtyInstance

    from ewaluacja_liczba_n.models import LiczbaNDlaUczelni

    # Sprawdź czy są rekordy do przeliczenia
    dirty_count = DirtyInstance.objects.count()
    if dirty_count > 0:
        messages.warning(
            request,
            f"Przed optymalizacją poczekaj na przeliczenie punktów. "
            f"Obecnie jest {dirty_count} rekordów do przeliczenia.",
        )
        return redirect("ewaluacja_optymalizacja:index")

    # Pobierz pierwszą uczelnię (zakładamy, że jest tylko jedna)
    uczelnia = Uczelnia.objects.first()

    if not uczelnia:
        messages.error(request, "Nie znaleziono uczelni w systemie.")
        return redirect("ewaluacja_optymalizacja:index")

    # WALIDACJA: Sprawdź czy istnieją dane liczby N dla uczelni
    liczba_n_count = LiczbaNDlaUczelni.objects.filter(uczelnia=uczelnia).count()
    liczba_n_raportowane = LiczbaNDlaUczelni.objects.filter(
        uczelnia=uczelnia, liczba_n__gte=12
    ).count()

    if liczba_n_count == 0:
        messages.error(
            request,
            "Brak danych liczby N dla uczelni! Przed optymalizacją musisz najpierw policzyć liczbę N. "
            'Przejdź do <a href="/ewaluacja_liczba_n/">modułu Liczba N</a> '
            "lub uruchom polecenie: python src/manage.py przelicz_n",
        )
        return redirect("ewaluacja_optymalizacja:index")

    if liczba_n_raportowane == 0:
        messages.warning(
            request,
            f"Znaleziono {liczba_n_count} dyscyplin z danymi liczby N, "
            "ale żadna nie ma liczby N >= 12 (nie są raportowane). "
            "Optymalizacja nie przetworzy żadnych dyscyplin.",
        )
        return redirect("ewaluacja_optymalizacja:index")

    # Skasuj stare rekordy OptimizationRun dla tej uczelni
    from ewaluacja_optymalizacja.models import OptimizationRun

    deleted_count = OptimizationRun.objects.filter(uczelnia=uczelnia).delete()[0]
    logger.info(
        f"Cleared {deleted_count} old OptimizationRun records for uczelnia {uczelnia.pk}"
    )

    # Pobierz wybrany tryb algorytmu z formularza (domyślnie two-phase)
    algorithm_mode = request.POST.get("algorithm_mode", "two-phase")
    if algorithm_mode not in ["two-phase", "single-phase"]:
        algorithm_mode = "two-phase"  # fallback to default

    logger.info(
        f"Starting bulk optimization for {uczelnia}: "
        f"{liczba_n_raportowane} disciplines with liczba_n >= 12, "
        f"algorithm mode: {algorithm_mode}"
    )

    # Uruchom zadanie Celery
    task = solve_all_reported_disciplines.delay(uczelnia.pk, algorithm_mode)

    # Nie pokazuj natychmiastowego komunikatu sukcesu - zadanie dopiero się rozpoczęło
    # Komunikat pojawi się na stronie statusu po faktycznym zakończeniu wszystkich operacji

    # Przekieruj do strony statusu (z uczelnia_id)
    return redirect(
        "ewaluacja_optymalizacja:bulk-status", uczelnia_id=uczelnia.pk, task_id=task.id
    )


def _process_discipline_optimization_status(liczba_n_obj, uczelnia):
    """Process optimization status for a single discipline.

    Returns dict with discipline info including status, points, publications, etc.
    """
    from ewaluacja_optymalizacja.models import OptimizationAuthorResult, OptimizationRun

    dyscyplina = liczba_n_obj.dyscyplina_naukowa

    disc_info = {
        "dyscyplina_id": dyscyplina.pk,
        "dyscyplina_nazwa": dyscyplina.nazwa,
        "liczba_n": float(liczba_n_obj.liczba_n),
    }

    # Znajdź najnowszy OptimizationRun dla tej dyscypliny
    opt_run = (
        OptimizationRun.objects.filter(uczelnia=uczelnia, dyscyplina_naukowa=dyscyplina)
        .order_by("-started_at")
        .first()
    )

    if opt_run:
        if opt_run.status == "completed":
            # Sprawdź czy są zapisane dane autorów
            has_authors = OptimizationAuthorResult.objects.filter(
                optimization_run=opt_run
            ).exists()

            if has_authors:
                disc_info["status"] = "completed"
                disc_info["optimization_run_id"] = opt_run.pk
                disc_info["total_points"] = float(opt_run.total_points)
                disc_info["total_publications"] = opt_run.total_publications
                return disc_info, "completed"

            # OptimizationRun completed ale brak danych - wciąż się zapisuje
            disc_info["status"] = "running"
            return disc_info, "running"

        if opt_run.status == "failed":
            disc_info["status"] = "failed"
            disc_info["error"] = opt_run.notes
            return disc_info, "failed"

        # status == "running"
        disc_info["status"] = "running"
        return disc_info, "running"

    # Brak OptimizationRun - zadanie jeszcze nie utworzyło rekordu
    disc_info["status"] = "pending"
    return disc_info, "pending"


def _build_bulk_progress_context(
    task_id, uczelnia_id, uczelnia, total_disciplines, disciplines_info, status_counts
):
    """Build context dict based on optimization progress status."""
    completed_count = status_counts["completed"]
    failed_count = status_counts["failed"]
    running_count = status_counts["running"]
    pending_count = status_counts["pending"]

    finished_count = completed_count + failed_count
    progress = (
        int((finished_count / total_disciplines) * 100) if total_disciplines > 0 else 0
    )

    all_done = finished_count == total_disciplines

    context = {
        "task_id": task_id,
        "uczelnia_id": uczelnia_id,
    }

    if all_done and completed_count > 0:
        # Zadania NAPRAWDĘ zakończone
        logger.info(
            f"Bulk optimization for {uczelnia} completed: "
            f"{completed_count} successful, {failed_count} failed"
        )
        context["redirect_to_index"] = True
        context["success_message"] = (
            f"Przeliczanie ewaluacji zakończone pomyślnie! "
            f"Przeliczono {completed_count} dyscyplin."
        )
    elif all_done and total_disciplines == 0:
        # Brak dyscyplin do przetworzenia
        context["task_ready"] = True
        context["success"] = False
        context["task_state"] = "FAILURE"
        context["error"] = (
            "Brak dyscyplin do przetworzenia. "
            "Upewnij się, że istnieją dane LiczbaNDlaUczelni z liczbą N >= 12."
        )
        logger.warning(f"Bulk optimization for {uczelnia}: no disciplines to process")
    elif all_done and completed_count == 0 and failed_count == 0:
        # Wszystkie skończone ale nic nie wykonano - zadania nie wystartowały
        context["task_ready"] = False
        context["task_state"] = "PENDING"
        context["info"] = {
            "step": "optimizing",
            "progress": 0,
            "message": "Oczekiwanie na rozpoczęcie przetwarzania...",
            "total_disciplines": total_disciplines,
            "completed": 0,
            "failed": 0,
            "running": 0,
            "pending": total_disciplines,
            "disciplines": disciplines_info,
        }
        logger.info(f"Bulk optimization for {uczelnia}: waiting for tasks to start")
    else:
        # Zadania w trakcie
        context["task_ready"] = False
        context["task_state"] = "PROGRESS"
        context["info"] = {
            "step": "optimizing",
            "progress": progress,
            "message": f"Przetwarzanie dyscyplin: {finished_count}/{total_disciplines}",
            "total_disciplines": total_disciplines,
            "completed": completed_count,
            "failed": failed_count,
            "running": running_count,
            "pending": pending_count,
            "disciplines": disciplines_info,
        }
        logger.debug(
            f"Bulk optimization progress for {uczelnia}: "
            f"{finished_count}/{total_disciplines} ({progress}%)"
        )

    return context


@login_required
def bulk_optimization_status(request, uczelnia_id, task_id):
    """
    Wyświetla status zadania masowej optymalizacji.
    Sprawdza postęp poprzez zapytania do bazy danych zamiast monitorowania Celery.
    Supports HTMX partial updates.
    """
    from bpp.models import Uczelnia
    from ewaluacja_liczba_n.models import LiczbaNDlaUczelni

    try:
        uczelnia = Uczelnia.objects.get(pk=uczelnia_id)
    except Uczelnia.DoesNotExist:
        context = {
            "task_id": task_id,
            "task_ready": True,
            "success": False,
            "error": f"Nie znaleziono uczelni o ID {uczelnia_id}",
        }
        if request.headers.get("HX-Request"):
            return render(
                request,
                "ewaluacja_optymalizacja/_bulk_optimization_progress.html",
                context,
            )
        return render(
            request, "ewaluacja_optymalizacja/bulk_optimization_status.html", context
        )

    # Pobierz listę dyscyplin raportowanych dla uczelni
    raportowane_dyscypliny = LiczbaNDlaUczelni.objects.filter(
        uczelnia=uczelnia, liczba_n__gte=12
    ).select_related("dyscyplina_naukowa")

    total_disciplines = raportowane_dyscypliny.count()

    if total_disciplines == 0:
        context = {
            "task_id": task_id,
            "task_ready": True,
            "success": True,
            "result": {
                "uczelnia_nazwa": str(uczelnia),
                "total_disciplines": 0,
                "successful": 0,
                "failed": 0,
                "disciplines": [],
            },
        }
        if request.headers.get("HX-Request"):
            return render(
                request,
                "ewaluacja_optymalizacja/_bulk_optimization_progress.html",
                context,
            )
        return render(
            request, "ewaluacja_optymalizacja/bulk_optimization_status.html", context
        )

    # Sprawdź status każdej dyscypliny w bazie danych
    status_counts = {"completed": 0, "failed": 0, "running": 0, "pending": 0}
    disciplines_info = []

    for liczba_n_obj in raportowane_dyscypliny:
        disc_info, status = _process_discipline_optimization_status(
            liczba_n_obj, uczelnia
        )
        status_counts[status] += 1
        disciplines_info.append(disc_info)

    # Build context based on progress
    context = _build_bulk_progress_context(
        task_id,
        uczelnia_id,
        uczelnia,
        total_disciplines,
        disciplines_info,
        status_counts,
    )

    # Handle redirect for completed tasks
    if context.get("redirect_to_index"):
        messages.success(request, context["success_message"])

        if request.headers.get("HX-Request"):
            from django.http import HttpResponse
            from django.urls import reverse

            response = HttpResponse(status=200)
            response["HX-Redirect"] = reverse("ewaluacja_optymalizacja:index")
            return response

        return redirect("ewaluacja_optymalizacja:index")

    # If HTMX request, return only the progress partial
    if request.headers.get("HX-Request"):
        return render(
            request, "ewaluacja_optymalizacja/_bulk_optimization_progress.html", context
        )

    return render(
        request, "ewaluacja_optymalizacja/bulk_optimization_status.html", context
    )


@login_required
def optimize_with_unpinning(request):
    """
    Uruchamia zadanie Celery do optymalizacji z odpinaniem niewykazanych slotów.
    """
    from denorm.models import DirtyInstance

    from ewaluacja_liczba_n.models import LiczbaNDlaUczelni

    # Sprawdź czy są rekordy do przeliczenia
    dirty_count = DirtyInstance.objects.count()
    if dirty_count > 0:
        messages.warning(
            request,
            f"Przed optymalizacją z odpinaniem poczekaj na przeliczenie punktów. "
            f"Obecnie jest {dirty_count} rekordów do przeliczenia.",
        )
        return redirect("ewaluacja_optymalizacja:index")

    # Pobierz pierwszą uczelnię (zakładamy, że jest tylko jedna)
    uczelnia = Uczelnia.objects.first()

    if not uczelnia:
        messages.error(request, "Nie znaleziono uczelni w systemie.")
        return redirect("ewaluacja_optymalizacja:index")

    # Sprawdź czy wszystkie dyscypliny raportowane są przeliczone
    liczba_n_raportowane = LiczbaNDlaUczelni.objects.filter(
        uczelnia=uczelnia, liczba_n__gte=12
    )

    if liczba_n_raportowane.count() == 0:
        messages.error(
            request,
            "Brak dyscyplin raportowanych (liczba N >= 12). "
            "Najpierw uruchom optymalizację dla dyscyplin.",
        )
        return redirect("ewaluacja_optymalizacja:index")

    # Sprawdź czy wszystkie dyscypliny raportowane mają wyniki optymalizacji
    for liczba_n_obj in liczba_n_raportowane:
        if not OptimizationRun.objects.filter(
            dyscyplina_naukowa=liczba_n_obj.dyscyplina_naukowa,
            uczelnia=uczelnia,
            status="completed",
        ).exists():
            messages.error(
                request,
                f"Dyscyplina {liczba_n_obj.dyscyplina_naukowa.nazwa} nie ma wykonanej optymalizacji. "
                "Najpierw uruchom 'Policz całą ewaluację'.",
            )
            return redirect("ewaluacja_optymalizacja:index")

    logger.info(f"Starting optimization with unpinning for {uczelnia}")

    # Uruchom zadanie Celery
    # Sprawdź czy nie ma już uruchomionego zadania
    from celery_singleton import DuplicateTaskError

    from .tasks import optimize_and_unpin_task

    try:
        task = optimize_and_unpin_task.delay(uczelnia.pk)
    except DuplicateTaskError:
        messages.error(
            request,
            "Zadanie optymalizacji z odpinaniem jest już uruchomione. "
            "Poczekaj na zakończenie obecnego zadania.",
        )
        return redirect("ewaluacja_optymalizacja:index")

    # Nie pokazuj komunikatu sukcesu - zadanie dopiero się rozpoczyna
    # Przekieruj od razu do strony statusu
    return redirect("ewaluacja_optymalizacja:optimize-unpin-status", task_id=task.id)


def _verify_optimization_data_complete(raportowane_dyscypliny, uczelnia):
    """Verify that all optimization data is fully saved to database."""
    from ewaluacja_optymalizacja.models import (
        OptimizationAuthorResult,
        OptimizationPublication,
        OptimizationRun,
    )

    for liczba_n_obj in raportowane_dyscypliny:
        opt_run = (
            OptimizationRun.objects.filter(
                uczelnia=uczelnia,
                dyscyplina_naukowa=liczba_n_obj.dyscyplina_naukowa,
                status="completed",
            )
            .order_by("-started_at")
            .first()
        )

        if opt_run:
            has_authors = OptimizationAuthorResult.objects.filter(
                optimization_run=opt_run
            ).exists()
            has_publications = OptimizationPublication.objects.filter(
                author_result__optimization_run=opt_run
            ).exists()

            if not (has_authors and has_publications):
                logger.info(
                    f"Data not fully saved for {liczba_n_obj.dyscyplina_naukowa.nazwa}: "
                    f"authors={has_authors}, publications={has_publications}"
                )
                return False
        else:
            logger.warning(
                f"No optimization run found for {liczba_n_obj.dyscyplina_naukowa.nazwa}"
            )
            return False

    return True


def _determine_unpin_task_context(
    task,
    all_data_complete,
    completed_count,
    discipline_count,
    running_count,
    percent_complete,
):
    """Determine the context based on task status and data completion."""
    context_update = {}

    if task.ready():
        if task.failed():
            error_info = str(task.info)
            logger.error(f"Task {task.id} failed: {error_info}")
            context_update.update(
                {
                    "error": error_info,
                    "success": False,
                    "task_ready": True,
                    "task_state": "FAILURE",
                }
            )
        elif task.successful() and all_data_complete:
            result = task.result
            logger.info(f"Task {task.id} successful, all data verified")
            context_update.update(
                {
                    "result": result,
                    "success": True,
                    "task_ready": True,
                    "task_state": "SUCCESS",
                }
            )
        else:
            context_update.update(
                {
                    "task_ready": False,
                    "task_state": "PROGRESS",
                    "info": {
                        "step": "finalizing",
                        "progress": 99,
                        "message": "Finalizowanie zapisów do bazy danych...",
                        "completed_optimizations": completed_count,
                        "total_optimizations": discipline_count,
                    },
                }
            )
    else:
        context_update["task_ready"] = False
        context_update["task_state"] = "PROGRESS"

        if completed_count >= discipline_count and discipline_count > 0:
            context_update["info"] = {
                "step": "finalizing",
                "progress": 99,
                "message": "Finalizowanie zadania...",
                "completed_optimizations": completed_count,
                "total_optimizations": discipline_count,
            }
        else:
            progress = (
                min(65 + int((completed_count / discipline_count) * 30), 95)
                if discipline_count > 0
                else 65
            )
            context_update["info"] = {
                "step": "optimization",
                "progress": progress,
                "message": f"Przeliczono {completed_count} z {discipline_count} dyscyplin ({percent_complete}%)",
                "running_optimizations": running_count,
                "completed_optimizations": completed_count,
                "total_optimizations": discipline_count,
            }

    return context_update


@login_required
def optimize_unpin_status(request, task_id):
    """
    Wyświetla status zadania optymalizacji z odpinaniem.
    - Dla fazy "denorm": pokazuje postęp denormalizacji (używa task.info)
    - Dla fazy "optimization": monitoruje BEZPOŚREDNIO bazę danych OptimizationRun
    Supports HTMX partial updates.
    """
    from denorm.models import DirtyInstance

    from ewaluacja_liczba_n.models import LiczbaNDlaUczelni
    from ewaluacja_optymalizacja.models import OptimizationRun

    task = AsyncResult(task_id)
    uczelnia = Uczelnia.objects.first()

    context = {
        "task_id": task_id,
        "dirty_count": DirtyInstance.objects.count(),
    }

    # KROK 1: Sprawdź fazę zadania z task.info (dla faz przed optymalizacją)
    task_info = task.info if task.info and isinstance(task.info, dict) else {}
    current_step = task_info.get("step", "unknown")

    logger.info(
        f"Checking optimize_unpin task {task_id}: current step={current_step}, "
        f"task.info={task_info}"
    )

    # KROK 2: Jeśli jesteśmy w fazie denormalizacji - użyj task.info
    if current_step == "denorm":
        dirty_count = task_info.get("dirty_count", DirtyInstance.objects.count())
        context["task_state"] = "PROGRESS"
        context["task_ready"] = False
        context["info"] = {
            "step": "denorm",
            "progress": task_info.get("progress", 50),
            "dirty_count": dirty_count,
            "message": f"Przeliczanie punktacji - pozostało {dirty_count} obiektów",
            "unpinned_ciagle": task_info.get("unpinned_ciagle", 0),
            "unpinned_zwarte": task_info.get("unpinned_zwarte", 0),
        }

        logger.info(
            f"Task {task_id} in denorm phase: {dirty_count} dirty objects remaining"
        )

        # If HTMX request, return only the progress partial
        if request.headers.get("HX-Request"):
            return render(
                request,
                "ewaluacja_optymalizacja/_optimize_unpin_progress.html",
                context,
            )
        return render(
            request, "ewaluacja_optymalizacja/optimize_unpin_status.html", context
        )

    # KROK 3: Monitoruj postęp przez bazę danych (niezależnie od stanu task)
    # Pobierz listę dyscyplin raportowanych
    raportowane_dyscypliny = LiczbaNDlaUczelni.objects.filter(
        uczelnia=uczelnia, liczba_n__gte=12
    ).select_related("dyscyplina_naukowa")

    discipline_count = raportowane_dyscypliny.count()
    dyscypliny_ids = list(
        raportowane_dyscypliny.values_list("dyscyplina_naukowa_id", flat=True)
    )

    # Sprawdź ile OptimizationRun jest już w bazie - BEZPOŚREDNIO z bazy, NIE przez Celery
    completed_count = OptimizationRun.objects.filter(
        uczelnia=uczelnia,
        dyscyplina_naukowa_id__in=dyscypliny_ids,
        status="completed",
    ).count()

    running_count = OptimizationRun.objects.filter(
        uczelnia=uczelnia,
        dyscyplina_naukowa_id__in=dyscypliny_ids,
        status="running",
    ).count()

    # Oblicz procent postępu
    if discipline_count > 0:
        percent_complete = int((completed_count / discipline_count) * 100)
    else:
        percent_complete = 0

    logger.info(
        f"Task {task_id} monitoring: task.ready()={task.ready()}, "
        f"{completed_count}/{discipline_count} completed ({percent_complete}%), "
        f"{running_count} running"
    )

    # KROK 4: Sprawdź czy wszystkie dane są zapisane w bazie
    all_data_complete = False
    if completed_count >= discipline_count and discipline_count > 0:
        # Wszystkie OptimizationRun są completed - weryfikuj dane
        logger.info(
            "All OptimizationRun records completed, verifying data integrity..."
        )
        all_data_complete = _verify_optimization_data_complete(
            raportowane_dyscypliny, uczelnia
        )

    # KROK 5: Określ stan na podstawie task.ready() i all_data_complete
    context_update = _determine_unpin_task_context(
        task,
        all_data_complete,
        completed_count,
        discipline_count,
        running_count,
        percent_complete,
    )
    context.update(context_update)

    # KROK 6: Jeśli zadanie się zakończyło pomyślnie i mamy success=True w kontekście
    # to znaczy że wszystko jest OK - przekieruj z flash message
    # WAŻNE: Dla HTMX używamy nagłówka HX-Redirect zamiast zwykłego redirect 302
    if context.get("success"):
        result = context.get("result", {})
        unpinned_total = result.get("unpinned_total", 0)

        messages.success(
            request,
            f"Optymalizacja z odpinaniem zakończona pomyślnie! "
            f"Odpiętych publikacji: {unpinned_total}",
        )

        # Dla HTMX requests używamy nagłówka HX-Redirect aby wykonać full-page redirect
        if request.headers.get("HX-Request"):
            from django.http import HttpResponse
            from django.urls import reverse

            response = HttpResponse(status=200)
            response["HX-Redirect"] = reverse("ewaluacja_optymalizacja:index")
            return response

        # Dla zwykłych requestów normalny redirect
        return redirect("ewaluacja_optymalizacja:index")

    # If HTMX request, return only the progress partial
    if request.headers.get("HX-Request"):
        return render(
            request, "ewaluacja_optymalizacja/_optimize_unpin_progress.html", context
        )

    return render(
        request, "ewaluacja_optymalizacja/optimize_unpin_status.html", context
    )


def _get_discipline_pin_stats(dyscyplina_naukowa):
    """
    Oblicza statystyki przypięć/odpięć dla danej dyscypliny w latach 2022-2025.

    Liczy tylko rekordy gdzie autor ma Autor_Dyscyplina w latach 2022-2025
    dla tej dyscypliny.

    Args:
        dyscyplina_naukowa: Obiekt Dyscyplina_Naukowa

    Returns:
        dict: {'pinned': int, 'unpinned': int, 'total': int,
               'pinned_pct': float, 'unpinned_pct': float}
    """
    from bpp.models import (
        Autor_Dyscyplina,
        Patent_Autor,
        Wydawnictwo_Ciagle_Autor,
        Wydawnictwo_Zwarte_Autor,
    )

    # Pobierz autorów którzy mają Autor_Dyscyplina w latach 2022-2025 dla tej dyscypliny
    autorzy_ids = set(
        Autor_Dyscyplina.objects.filter(
            rok__gte=2022,
            rok__lte=2025,
            dyscyplina_naukowa=dyscyplina_naukowa,
        )
        .values_list("autor_id", flat=True)
        .distinct()
    )

    if not autorzy_ids:
        return {
            "pinned": 0,
            "unpinned": 0,
            "total": 0,
            "pinned_pct": 0.0,
            "unpinned_pct": 0.0,
        }

    # Filtr bazowy dla wszystkich modeli
    base_filter = Q(
        rekord__rok__gte=2022,
        rekord__rok__lte=2025,
        dyscyplina_naukowa=dyscyplina_naukowa,
        autor_id__in=autorzy_ids,
    )

    # Zlicz przypięte i odpięte dla każdego modelu
    pinned_count = 0
    unpinned_count = 0

    for model in [Wydawnictwo_Ciagle_Autor, Wydawnictwo_Zwarte_Autor, Patent_Autor]:
        pinned_count += model.objects.filter(base_filter, przypieta=True).count()
        unpinned_count += model.objects.filter(base_filter, przypieta=False).count()

    total = pinned_count + unpinned_count

    if total == 0:
        return {
            "pinned": 0,
            "unpinned": 0,
            "total": 0,
            "pinned_pct": 0.0,
            "unpinned_pct": 0.0,
        }

    pinned_pct = (pinned_count / total) * 100
    unpinned_pct = (unpinned_count / total) * 100

    return {
        "pinned": pinned_count,
        "unpinned": unpinned_count,
        "total": total,
        "pinned_pct": pinned_pct,
        "unpinned_pct": unpinned_pct,
    }


@login_required
def reset_discipline_pins(request, pk):
    """
    Resetuje przypięcia dla danej dyscypliny (ustawia przypieta=True).
    Tworzy snapshot przed resetowaniem.
    """
    from django.db import transaction

    from bpp.models import (
        Autor_Dyscyplina,
        Dyscyplina_Naukowa,
        Patent_Autor,
        Wydawnictwo_Ciagle_Autor,
        Wydawnictwo_Zwarte_Autor,
    )
    from snapshot_odpiec.models import SnapshotOdpiec

    dyscyplina = get_object_or_404(Dyscyplina_Naukowa, pk=pk)

    # Sprawdź czy są odpięte rekordy
    stats = _get_discipline_pin_stats(dyscyplina)
    if stats["unpinned"] == 0:
        messages.warning(
            request,
            f"Dyscyplina '{dyscyplina.nazwa}' nie ma odpiętych rekordów w latach 2022-2025.",
        )
        return redirect("ewaluacja_optymalizacja:index")

    # Pobierz autorów którzy mają Autor_Dyscyplina w latach 2022-2025 dla tej dyscypliny
    autorzy_ids = set(
        Autor_Dyscyplina.objects.filter(
            rok__gte=2022,
            rok__lte=2025,
            dyscyplina_naukowa=dyscyplina,
        )
        .values_list("autor_id", flat=True)
        .distinct()
    )

    with transaction.atomic():
        # Utwórz snapshot przed resetowaniem
        snapshot = SnapshotOdpiec.objects.create(
            owner=request.user, comment=f"przed resetem przypięć - {dyscyplina.nazwa}"
        )

        logger.info(
            f"Created snapshot {snapshot.pk} before resetting pins for discipline "
            f"'{dyscyplina.nazwa}' (user: {request.user})"
        )

        # Resetuj przypięcia dla wszystkich modeli
        base_filter = Q(
            rekord__rok__gte=2022,
            rekord__rok__lte=2025,
            dyscyplina_naukowa=dyscyplina,
            autor_id__in=autorzy_ids,
        )

        updated_count = 0
        for model in [Wydawnictwo_Ciagle_Autor, Wydawnictwo_Zwarte_Autor, Patent_Autor]:
            count = model.objects.filter(base_filter).update(przypieta=True)
            updated_count += count
            logger.info(
                f"Reset {count} pins in {model.__name__} for discipline '{dyscyplina.nazwa}'"
            )

    # Poczekaj na zakończenie denormalizacji
    logger.info(
        f"Waiting for denormalization to complete after resetting pins for '{dyscyplina.nazwa}'"
    )

    from time import sleep

    from denorm.models import DirtyInstance
    from denorm.tasks import flush_via_queue

    # Trigger denorm processing via Celery queue
    flush_via_queue.delay()
    logger.info("Triggered denorm flush via queue")

    max_wait = 600  # Max 10 minut
    waited = 0
    check_interval = 5

    while DirtyInstance.objects.count() > 0 and waited < max_wait:
        sleep(check_interval)
        waited += check_interval
        dirty_count = DirtyInstance.objects.count()
        logger.info(
            f"Waiting for denormalization after reset... {dirty_count} objects remaining"
        )

    # Uruchom zadanie optymalizacji dla tej dyscypliny
    logger.info(f"Starting optimization for discipline '{dyscyplina.nazwa}'")

    from ewaluacja_liczba_n.models import LiczbaNDlaUczelni

    from .tasks import solve_single_discipline_task

    uczelnia = Uczelnia.objects.first()

    try:
        liczba_n_obj = LiczbaNDlaUczelni.objects.get(
            uczelnia=uczelnia, dyscyplina_naukowa=dyscyplina
        )

        task = solve_single_discipline_task.delay(
            uczelnia.pk, dyscyplina.pk, float(liczba_n_obj.liczba_n)
        )

        logger.info(
            f"Started optimization task {task.id} for discipline '{dyscyplina.nazwa}'"
        )

        messages.success(
            request,
            f"Zresetowano {updated_count} przypięć dla dyscypliny '{dyscyplina.nazwa}' "
            f"(lata 2022-2025). Utworzono snapshot #{snapshot.pk}. "
            f"Trwa przeliczanie optymalizacji...",
        )
    except LiczbaNDlaUczelni.DoesNotExist:
        logger.warning(
            f"No LiczbaNDlaUczelni found for discipline '{dyscyplina.nazwa}', skipping optimization"
        )
        messages.success(
            request,
            f"Zresetowano {updated_count} przypięć dla dyscypliny '{dyscyplina.nazwa}' "
            f"(lata 2022-2025). Utworzono snapshot #{snapshot.pk}.",
        )

    return redirect("ewaluacja_optymalizacja:index")


@login_required
def reset_all_pins(request):
    """
    Uruchamia zadanie Celery do resetowania przypięć dla wszystkich autorów uczelni.
    Resetuje przypięcia dla rekordów 2022-2025 gdzie autor ma dyscyplinę.
    """
    # Pobierz pierwszą uczelnię (zakładamy, że jest tylko jedna)
    uczelnia = Uczelnia.objects.first()

    if not uczelnia:
        messages.error(request, "Nie znaleziono uczelni w systemie.")
        return redirect("ewaluacja_optymalizacja:index")

    logger.info(f"Starting reset all pins for {uczelnia}")

    # Uruchom zadanie Celery
    # Sprawdź czy nie ma już uruchomionego zadania
    from celery_singleton import DuplicateTaskError

    from .tasks import reset_all_pins_task

    try:
        task = reset_all_pins_task.delay(uczelnia.pk)
    except DuplicateTaskError:
        messages.error(
            request,
            "Zadanie resetowania przypięć jest już uruchomione. "
            "Poczekaj na zakończenie obecnego zadania.",
        )
        return redirect("ewaluacja_optymalizacja:index")

    # Przekieruj do strony statusu
    return redirect("ewaluacja_optymalizacja:reset-all-pins-status", task_id=task.id)


@login_required
def reset_all_pins_status(request, task_id):
    """
    Wyświetla status zadania resetowania przypięć dla wszystkich autorów.
    Supports HTMX partial updates.
    """
    from celery.result import AsyncResult
    from denorm.models import DirtyInstance

    task = AsyncResult(task_id)

    context = {
        "task_id": task_id,
        "dirty_count": DirtyInstance.objects.count(),
    }

    # Pobierz informacje o zadaniu
    task_info = task.info if task.info and isinstance(task.info, dict) else {}
    current_step = task_info.get("step", "unknown")

    logger.info(
        f"Checking reset_all_pins task {task_id}: current step={current_step}, "
        f"task.info={task_info}"
    )

    # Jeśli zadanie się nie zakończyło
    if not task.ready():
        context["task_state"] = "PROGRESS"
        context["task_ready"] = False
        context["info"] = task_info
        context["info"]["step"] = current_step

    # Sprawdź czy zadanie zakończyło się błędem
    elif task.failed():
        error_info = str(task.info)
        logger.error(f"Task {task_id} failed: {error_info}")
        context["error"] = error_info
        context["success"] = False
        context["task_ready"] = True

    # Zadanie zakończone pomyślnie
    elif task.successful():
        result = task.result
        logger.info(f"Task {task_id} successful: {result}")
        total_reset = result.get("total_reset", 0)

        messages.success(
            request,
            f"Reset przypięć zakończony pomyślnie! Zresetowano {total_reset} przypięć.",
        )

        # Dla HTMX requests używamy nagłówka HX-Redirect aby wykonać full-page redirect
        if request.headers.get("HX-Request"):
            from django.http import HttpResponse
            from django.urls import reverse

            response = HttpResponse(status=200)
            response["HX-Redirect"] = reverse("ewaluacja_optymalizacja:index")
            return response

        # Dla zwykłych requestów normalny redirect
        return redirect("ewaluacja_optymalizacja:index")

    # If HTMX request, return only the progress partial
    if request.headers.get("HX-Request"):
        return render(
            request,
            "ewaluacja_optymalizacja/_reset_all_pins_progress.html",
            context,
        )

    return render(
        request, "ewaluacja_optymalizacja/reset_all_pins_status.html", context
    )


@login_required
def unpin_all_sensible(request):
    """
    Uruchamia zadanie Celery do odpinania wszystkich sensownych możliwości odpinania.
    Odpina rekordy gdzie makes_sense=True, czeka na denormalizację i przelicza metryki + unpinning.
    """
    # Pobierz pierwszą uczelnię (zakładamy, że jest tylko jedna)
    uczelnia = Uczelnia.objects.first()

    if not uczelnia:
        messages.error(request, "Nie znaleziono uczelni w systemie.")
        return redirect("ewaluacja_optymalizacja:unpinning-list")

    logger.info(f"Starting unpin all sensible for {uczelnia}")

    # Uruchom zadanie Celery
    # Sprawdź czy nie ma już uruchomionego zadania
    from celery_singleton import DuplicateTaskError

    from .tasks import unpin_all_sensible_task

    try:
        task = unpin_all_sensible_task.delay(uczelnia.pk)
    except DuplicateTaskError:
        messages.error(
            request,
            "Zadanie odpinania sensownych możliwości jest już uruchomione. "
            "Poczekaj na zakończenie obecnego zadania.",
        )
        return redirect("ewaluacja_optymalizacja:unpinning-list")

    # Przekieruj do strony statusu
    return redirect(
        "ewaluacja_optymalizacja:unpin-all-sensible-status", task_id=task.id
    )


@login_required
def unpin_all_sensible_status(request, task_id):
    """
    Wyświetla status zadania odpinania wszystkich sensownych możliwości.
    Supports HTMX partial updates.
    """
    from celery.result import AsyncResult
    from denorm.models import DirtyInstance

    task = AsyncResult(task_id)

    context = {
        "task_id": task_id,
        "dirty_count": DirtyInstance.objects.count(),
    }

    # Pobierz informacje o zadaniu
    task_info = task.info if task.info and isinstance(task.info, dict) else {}
    current_step = task_info.get("step", "unknown")

    logger.info(
        f"Checking unpin_all_sensible task {task_id}: current step={current_step}, "
        f"task.info={task_info}"
    )

    # Jeśli zadanie się nie zakończyło
    if not task.ready():
        context["task_state"] = "PROGRESS"
        context["task_ready"] = False
        context["info"] = task_info
        context["info"]["step"] = current_step

    # Sprawdź czy zadanie zakończyło się błędem
    elif task.failed():
        error_info = str(task.info)
        logger.error(f"Task {task_id} failed: {error_info}")
        context["error"] = error_info
        context["success"] = False
        context["task_ready"] = True

    # Zadanie zakończone pomyślnie
    elif task.successful():
        result = task.result
        logger.info(f"Task {task_id} successful: {result}")
        unpinned_count = result.get("unpinned_count", 0)

        messages.success(
            request,
            f"Odpinanie sensownych możliwości zakończone pomyślnie! Odpięto {unpinned_count} rekordów. "
            f"Metryki i analiza unpinning zostały przeliczone.",
        )

        # Dla HTMX requests używamy nagłówka HX-Redirect aby wykonać full-page redirect
        if request.headers.get("HX-Request"):
            from django.http import HttpResponse
            from django.urls import reverse

            response = HttpResponse(status=200)
            response["HX-Redirect"] = reverse("ewaluacja_optymalizacja:unpinning-list")
            return response

        # Dla zwykłych requestów normalny redirect
        return redirect("ewaluacja_optymalizacja:unpinning-list")

    # If HTMX request, return only the progress partial
    if request.headers.get("HX-Request"):
        return render(
            request,
            "ewaluacja_optymalizacja/_unpin_all_sensible_progress.html",
            context,
        )

    return render(
        request, "ewaluacja_optymalizacja/unpin_all_sensible_status.html", context
    )


@login_required
def denorm_progress(request):
    """
    Endpoint HTMX zwracający progress bar dla denormalizacji.
    Śledzi liczbę DirtyInstance i pokazuje postęp przeliczania.

    Parametry GET:
        initial: Początkowa liczba rekordów (opcjonalnie, do obliczenia procentu)
    """
    from denorm.models import DirtyInstance

    dirty_count = DirtyInstance.objects.count()

    # Pobierz initial_count z parametru GET
    initial_count = request.GET.get("initial")
    if initial_count:
        try:
            initial_count = int(initial_count)
        except (ValueError, TypeError):
            initial_count = None

    # Oblicz procent pozostały do przeliczenia i ukończony
    if initial_count and initial_count > 0:
        remaining_pct = (dirty_count / initial_count) * 100
        completed_pct = 100 - remaining_pct
    else:
        remaining_pct = 0
        completed_pct = 0

    # Pokaż komunikat o zakończeniu jeśli było initial_count > 0 a teraz dirty_count == 0
    show_completed = (
        dirty_count == 0 and initial_count is not None and initial_count > 0
    )

    context = {
        "dirty_count": dirty_count,
        "initial_count": initial_count,
        "remaining_pct": remaining_pct,
        "completed_pct": completed_pct,
        "show_completed": show_completed,
    }

    return render(request, "ewaluacja_optymalizacja/_denorm_progress.html", context)


@login_required
def analyze_unpinning_opportunities(request):
    """
    Uruchamia zadanie analizy możliwości odpinania prac wieloautorskich.

    ZAWSZE najpierw przelicza metryki (usuwa stare i generuje nowe),
    a następnie automatycznie uruchamia analizę prac wieloautorskich.

    Używa Celery chain do sekwencyjnego wykonania: metryki -> unpinning.
    """
    from celery import chain

    from ewaluacja_metryki.models import MetrykaAutora, StatusGenerowania
    from ewaluacja_metryki.tasks import generuj_metryki_task_parallel
    from ewaluacja_metryki.utils import get_default_rodzaje_autora
    from ewaluacja_optymalizacja.tasks import run_unpinning_after_metrics_wrapper

    # Pobierz pierwszą uczelnię (zakładamy, że jest tylko jedna)
    uczelnia = Uczelnia.objects.first()

    if not uczelnia:
        messages.error(request, "Nie znaleziono uczelni w systemie.")
        return redirect("ewaluacja_optymalizacja:index")

    # Sprawdź czy przeliczanie metryk już trwa
    status = StatusGenerowania.get_or_create()

    if status.w_trakcie:
        messages.info(
            request,
            "Przeliczanie metryk jest już w trakcie. Poczekaj na zakończenie.",
        )
        return redirect("ewaluacja_optymalizacja:index")

    # Usuń stare metryki
    metryki_count = MetrykaAutora.objects.count()
    if metryki_count > 0:
        logger.info(f"Deleting {metryki_count} old metrics before recalculation...")
        MetrykaAutora.objects.all().delete()
        logger.info("Deleted old metrics")

    # Stwórz Celery chain: metryki -> unpinning
    # Pierwszy task (metryki) zwraca wynik, który jest przekazywany do drugiego (unpinning wrapper)
    workflow = chain(
        generuj_metryki_task_parallel.s(
            rodzaje_autora=get_default_rodzaje_autora(),
            nadpisz=True,
            przelicz_liczbe_n=True,
        ),  # Przelicz metryki z jawnymi parametrami
        run_unpinning_after_metrics_wrapper.s(  # Potem uruchom unpinning
            uczelnia_id=uczelnia.pk, dyscyplina_id=None, min_slot_filled=0.8
        ),
    )

    # Uruchom chain
    result = workflow.apply_async()

    logger.info(
        f"Started metrics->unpinning chain. Chain ID: {result.id}, "
        f"Parent task (metrics): {result.parent.id if result.parent else 'None'}"
    )

    # Store task_id in session for progress tracking
    request.session["unpinning_task_id"] = result.id

    messages.info(
        request,
        "Rozpoczęto przeliczanie metryk i analizę możliwości odpinania. "
        "Zostaniesz przekierowany do strony postępu.",
    )

    # Przekieruj do strony statusu - result.id to ID ostatniego zadania w chain (unpinning)
    return redirect(
        "ewaluacja_optymalizacja:unpinning-combined-status", task_id=result.id
    )


@login_required
def unpinning_combined_status(request, task_id):
    """
    Wyświetla status łączonego zadania: przeliczanie metryk + analiza unpinning.
    Supports HTMX partial updates.
    """
    from ewaluacja_metryki.models import StatusGenerowania

    task = AsyncResult(task_id)

    context = {
        "task_id": task_id,
    }

    # Pobierz informacje o zadaniu
    task_info = task.info if task.info and isinstance(task.info, dict) else {}
    current_stage = task_info.get("stage", "unknown")
    current_step = task_info.get("step", "unknown")

    logger.info(
        f"Checking combined task {task_id}: stage={current_stage}, step={current_step}, "
        f"task.info={task_info}"
    )

    # Jeśli zadanie się nie zakończyło
    if not task.ready():
        context["task_state"] = "PROGRESS"
        context["task_ready"] = False
        context["info"] = task_info
        context["info"]["stage"] = current_stage
        context["info"]["step"] = current_step

        # Check if metrics generation is running and add real progress
        status = StatusGenerowania.get_or_create()
        if status.w_trakcie and current_stage == "unknown":
            # Metrics are still running, update stage info
            context["info"]["stage"] = "metrics"
            context["info"]["step"] = "calculating"
            # Calculate real progress based on StatusGenerowania
            if status.liczba_do_przetworzenia > 0:
                metrics_progress = int(
                    (status.przetworzono_count / status.liczba_do_przetworzenia) * 50
                )
                context["info"]["progress"] = metrics_progress
                context["info"]["analyzed"] = status.przetworzono_count
                context["info"]["total"] = status.liczba_do_przetworzenia
            else:
                context["info"]["progress"] = 0

    # Sprawdź czy zadanie zakończyło się błędem
    elif task.failed():
        error_info = str(task.info)
        logger.error(f"Task {task_id} failed: {error_info}")

        # Clear task_id from session since analysis failed
        if "unpinning_task_id" in request.session:
            del request.session["unpinning_task_id"]

        context["error"] = error_info
        context["success"] = False
        context["task_ready"] = True

    # Zadanie zakończone pomyślnie
    elif task.successful():
        result = task.result
        logger.info(f"Task {task_id} successful: {result}")

        # Clear task_id from session since analysis is complete
        if "unpinning_task_id" in request.session:
            del request.session["unpinning_task_id"]

        messages.success(
            request,
            f"Analiza zakończona pomyślnie! "
            f"Przeliczono {result.get('metrics_count', 0)} metryk. "
            f"Znaleziono {result.get('total_opportunities', 0)} możliwości odpinania, "
            f"w tym {result.get('sensible_opportunities', 0)} sensownych.",
        )

        # Dla HTMX requests używamy nagłówka HX-Redirect aby wykonać full-page redirect
        if request.headers.get("HX-Request"):
            from django.http import HttpResponse
            from django.urls import reverse

            response = HttpResponse(status=200)
            response["HX-Redirect"] = reverse("ewaluacja_optymalizacja:unpinning-list")
            return response

        # Dla zwykłych requestów normalny redirect
        return redirect("ewaluacja_optymalizacja:unpinning-list")

    # If HTMX request, return only the progress partial
    if request.headers.get("HX-Request"):
        return render(
            request,
            "ewaluacja_optymalizacja/_unpinning_combined_progress.html",
            context,
        )

    return render(
        request, "ewaluacja_optymalizacja/unpinning_combined_status.html", context
    )


@login_required
def unpinning_analysis_status(request, task_id):
    """
    Wyświetla status zadania analizy możliwości odpinania.
    Supports HTMX partial updates.
    """
    task = AsyncResult(task_id)

    context = {
        "task_id": task_id,
    }

    # Pobierz informacje o zadaniu
    task_info = task.info if task.info and isinstance(task.info, dict) else {}
    current_step = task_info.get("step", "unknown")

    logger.info(
        f"Checking unpinning analysis task {task_id}: current step={current_step}, "
        f"task.info={task_info}"
    )

    # Jeśli zadanie się nie zakończyło
    if not task.ready():
        context["task_state"] = "PROGRESS"
        context["task_ready"] = False
        context["info"] = task_info
        context["info"]["step"] = current_step

    # Sprawdź czy zadanie zakończyło się błędem
    elif task.failed():
        error_info = str(task.info)
        logger.error(f"Task {task_id} failed: {error_info}")
        context["error"] = error_info
        context["success"] = False
        context["task_ready"] = True

    # Zadanie zakończone pomyślnie
    elif task.successful():
        result = task.result
        logger.info(f"Task {task_id} successful: {result}")

        messages.success(
            request,
            f"Analiza zakończona pomyślnie! "
            f"Znaleziono {result.get('total_opportunities', 0)} możliwości odpinania, "
            f"w tym {result.get('sensible_opportunities', 0)} sensownych.",
        )

        # Dla HTMX requests używamy nagłówka HX-Redirect aby wykonać full-page redirect
        if request.headers.get("HX-Request"):
            from django.http import HttpResponse
            from django.urls import reverse

            response = HttpResponse(status=200)
            response["HX-Redirect"] = reverse("ewaluacja_optymalizacja:unpinning-list")
            return response

        # Dla zwykłych requestów normalny redirect
        return redirect("ewaluacja_optymalizacja:unpinning-list")

    # If HTMX request, return only the progress partial
    if request.headers.get("HX-Request"):
        return render(
            request,
            "ewaluacja_optymalizacja/_unpinning_analysis_progress.html",
            context,
        )

    return render(
        request, "ewaluacja_optymalizacja/unpinning_analysis_status.html", context
    )


def _get_dyscyplina_filter(request):
    """Extract and validate dyscyplina ID from request."""
    dyscyplina_id = request.GET.get("dyscyplina")
    if not dyscyplina_id:
        return None

    try:
        return int(dyscyplina_id)
    except (ValueError, TypeError):
        return None


def _cache_rekord_objects(opportunities_qs):
    """Preload Rekord objects to avoid N+1 queries."""
    from bpp.models import Rekord

    all_opportunities = list(opportunities_qs)
    rekord_ids = [opp.rekord_id for opp in all_opportunities]
    rekordy = {r.pk: r for r in Rekord.objects.filter(pk__in=rekord_ids)}

    # Attach rekord objects to opportunities
    for opp in all_opportunities:
        opp._cached_rekord = rekordy.get(opp.rekord_id)

    return all_opportunities


def _get_punkty_kbn(opportunity):
    """Safely get punkty_kbn from opportunity's rekord."""
    try:
        return opportunity.rekord.original.punkty_kbn or Decimal("0")
    except (AttributeError, TypeError):
        return Decimal("0")


def _filter_by_punktacja(opportunities, punktacja_zrodla):
    """Filter opportunities by punktacja_zrodla range."""
    if not punktacja_zrodla:
        return opportunities

    punktacja_ranges = {
        "0-100": lambda p: p < Decimal("100"),
        "100-140": lambda p: Decimal("100") <= p < Decimal("140"),
        "140-200": lambda p: Decimal("140") <= p < Decimal("200"),
        "200+": lambda p: p >= Decimal("200"),
    }

    filter_func = punktacja_ranges.get(punktacja_zrodla)
    if not filter_func:
        return opportunities

    return [opp for opp in opportunities if filter_func(_get_punkty_kbn(opp))]


def _create_sort_key_function(sort_by):
    """Create a sort key function based on sort_by parameter."""
    # Define sort key extractors
    sort_keys = {
        "tytul": lambda opp: opp.rekord_tytul or "",
        "punktacja": _get_punkty_kbn,
        "dyscyplina": lambda opp: (
            opp.dyscyplina_naukowa.nazwa if opp.dyscyplina_naukowa else ""
        ),
        "autor_a": lambda opp: (
            opp.autor_could_benefit.nazwisko if opp.autor_could_benefit else ""
        ),
        "slots_missing": lambda opp: opp.slots_missing or Decimal("0"),
        "slot_in_work": lambda opp: opp.slot_in_work or Decimal("0"),
        "punkty_a": lambda opp: opp.punkty_roznica_a or Decimal("0"),
        "sloty_a": lambda opp: opp.sloty_roznica_a or Decimal("0"),
        "punkty_b": lambda opp: opp.punkty_roznica_b or Decimal("0"),
        "sloty_b": lambda opp: opp.sloty_roznica_b or Decimal("0"),
        "autor_b": lambda opp: (
            opp.autor_currently_using.nazwisko if opp.autor_currently_using else ""
        ),
        "makes_sense": lambda opp: opp.makes_sense,
    }

    # Return the requested sort key function, defaulting to punkty_b
    return sort_keys.get(sort_by, sort_keys["punkty_b"])


@login_required
def unpinning_opportunities_list(request):
    """
    Wyświetla listę możliwości odpinania prac wieloautorskich.
    """
    from celery.result import AsyncResult

    from bpp.models import Dyscyplina_Naukowa
    from ewaluacja_metryki.models import StatusGenerowania

    from .models import UnpinningOpportunity

    # Pobierz pierwszą uczelnię (zakładamy, że jest tylko jedna)
    uczelnia = Uczelnia.objects.first()

    if not uczelnia:
        messages.error(request, "Nie znaleziono uczelni w systemie.")
        return redirect("ewaluacja_optymalizacja:index")

    # Check if unpinning analysis is already running
    task_id = request.session.get("unpinning_task_id")
    if task_id:
        task = AsyncResult(task_id)
        # Check if task is still running (PENDING, STARTED, or PROGRESS)
        if task.state in ["PENDING", "STARTED", "PROGRESS"]:
            messages.info(
                request,
                "Analiza możliwości odpinania jest w trakcie. "
                "Przekierowuję do strony postępu...",
            )
            return redirect(
                "ewaluacja_optymalizacja:unpinning-combined-status", task_id=task_id
            )
        # Task finished or failed - clear from session
        if task.state in ["SUCCESS", "FAILURE"]:
            del request.session["unpinning_task_id"]

    # Also check if metrics generation is running
    status = StatusGenerowania.get_or_create()
    if status.w_trakcie and task_id:
        # Metrics are running as part of unpinning chain
        messages.info(
            request,
            "Przeliczanie metryk dla analizy odpinania jest w trakcie. "
            "Przekierowuję do strony postępu...",
        )
        return redirect(
            "ewaluacja_optymalizacja:unpinning-combined-status", task_id=task_id
        )

    # Podstawowe filtrowanie
    opportunities_qs = UnpinningOpportunity.objects.filter(uczelnia=uczelnia)

    # Filtr po dyscyplinie
    dyscyplina_id = _get_dyscyplina_filter(request)
    if dyscyplina_id:
        opportunities_qs = opportunities_qs.filter(dyscyplina_naukowa_id=dyscyplina_id)

    # Filtr "tylko sensowne"
    only_sensible = request.GET.get("only_sensible") == "1"
    if only_sensible:
        opportunities_qs = opportunities_qs.filter(makes_sense=True)

    # Select related dla optymalizacji
    opportunities_qs = opportunities_qs.select_related(
        "autor_could_benefit",
        "autor_currently_using",
        "dyscyplina_naukowa",
        "metryka_could_benefit",
        "metryka_currently_using",
        "metryka_could_benefit__autor",
        "metryka_currently_using__autor",
    )

    # Preload Rekord objects
    all_opportunities = _cache_rekord_objects(opportunities_qs)

    # Filtr po punktacji źródła
    punktacja_zrodla = request.GET.get("punktacja_zrodla")
    all_opportunities = _filter_by_punktacja(all_opportunities, punktacja_zrodla)

    # Sortowanie
    sort_by = request.GET.get("sort_by", "punkty_b")
    sort_dir = request.GET.get("sort_dir", "desc")
    sort_key = _create_sort_key_function(sort_by)
    all_opportunities = sorted(
        all_opportunities, key=sort_key, reverse=(sort_dir == "desc")
    )

    # Paginacja po filtrowaniu i sortowaniu
    from django.core.paginator import Paginator

    paginator = Paginator(all_opportunities, 50)
    page_number = request.GET.get("page")
    opportunities = paginator.get_page(page_number)

    # Statystyki
    total_count = UnpinningOpportunity.objects.filter(uczelnia=uczelnia).count()
    sensible_count = UnpinningOpportunity.objects.filter(
        uczelnia=uczelnia, makes_sense=True
    ).count()

    # Lista dyscyplin dla filtra
    dyscypliny = Dyscyplina_Naukowa.objects.filter(
        pk__in=UnpinningOpportunity.objects.filter(uczelnia=uczelnia).values_list(
            "dyscyplina_naukowa_id", flat=True
        )
    ).order_by("nazwa")

    context = {
        "opportunities": opportunities,
        "total_count": total_count,
        "sensible_count": sensible_count,
        "dyscypliny": dyscypliny,
        "selected_dyscyplina": dyscyplina_id,
        "only_sensible": only_sensible,
        "selected_punktacja_zrodla": punktacja_zrodla or "",
        "sort_by": sort_by,
        "sort_dir": sort_dir,
    }

    return render(
        request, "ewaluacja_optymalizacja/unpinning_opportunities_list.html", context
    )


@login_required
def database_verification_view(request):
    """
    Wyświetla listę prac z autorami mającymi sloty poniżej 0.1 w latach 2022-2025.
    Takie sloty należy usunąć przed dalszymi krokami optymalizacji.
    """
    from bpp.models import Autor_Dyscyplina, Cache_Punktacja_Autora_Query

    # Pobierz autorów którzy mają przypisane dyscypliny w latach 2022-2025
    autorzy_z_dyscyplinami = set(
        Autor_Dyscyplina.objects.filter(rok__gte=2022, rok__lte=2025)
        .values_list("autor_id", flat=True)
        .distinct()
    )

    # Zapytanie o prace z problemowymi slotami
    problematic_records = (
        Cache_Punktacja_Autora_Query.objects.filter(
            slot__lt=Decimal("0.1"),
            rekord__rok__gte=2022,
            rekord__rok__lte=2025,
            autor_id__in=autorzy_z_dyscyplinami,
        )
        .select_related("rekord", "autor", "dyscyplina")
        .order_by("slot", "rekord__rok", "autor__nazwisko")
    )

    # Statystyki
    total_count = problematic_records.count()
    unique_works = problematic_records.values("rekord_id").distinct().count()
    unique_authors = problematic_records.values("autor_id").distinct().count()

    context = {
        "problematic_records": problematic_records,
        "total_count": total_count,
        "unique_works": unique_works,
        "unique_authors": unique_authors,
    }

    return render(
        request, "ewaluacja_optymalizacja/database_verification.html", context
    )


@login_required
def export_unpinning_opportunities_xlsx(request):  # noqa: C901
    """Export unpinning opportunities list to XLSX format with filters applied"""
    import datetime
    from decimal import Decimal

    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from bpp.models import Dyscyplina_Naukowa, Rekord
    from ewaluacja_common.models import Rodzaj_Autora

    from .models import UnpinningOpportunity

    # Pobierz pierwszą uczelnię (zakładamy, że jest tylko jedna)
    uczelnia = Uczelnia.objects.first()

    if not uczelnia:
        messages.error(request, "Nie znaleziono uczelni w systemie.")
        return redirect("ewaluacja_optymalizacja:index")

    # Filtrowanie (identyczne jak w unpinning_opportunities_list)
    opportunities_qs = UnpinningOpportunity.objects.filter(uczelnia=uczelnia)

    # Filtr po dyscyplinie (opcjonalnie)
    dyscyplina_id = request.GET.get("dyscyplina")
    if dyscyplina_id:
        try:
            dyscyplina_id = int(dyscyplina_id)
            opportunities_qs = opportunities_qs.filter(
                dyscyplina_naukowa_id=dyscyplina_id
            )
        except (ValueError, TypeError):
            pass

    # Filtr "tylko sensowne"
    only_sensible = request.GET.get("only_sensible")
    if only_sensible == "1":
        opportunities_qs = opportunities_qs.filter(makes_sense=True)

    # Select related dla optymalizacji
    opportunities_qs = opportunities_qs.select_related(
        "autor_could_benefit",
        "autor_currently_using",
        "dyscyplina_naukowa",
        "metryka_could_benefit",
        "metryka_currently_using",
        "metryka_could_benefit__autor",
        "metryka_currently_using__autor",
        "metryka_could_benefit__jednostka",
        "metryka_currently_using__jednostka",
    )

    # Preload Rekord objects to avoid N+1 queries
    all_opportunities = list(opportunities_qs)
    rekord_ids = [opp.rekord_id for opp in all_opportunities]
    rekordy = {r.pk: r for r in Rekord.objects.filter(pk__in=rekord_ids)}

    # Attach rekord objects to opportunities
    for opp in all_opportunities:
        opp._cached_rekord = rekordy.get(opp.rekord_id)

    # Filtr po punktacji źródła (identyczne jak w unpinning_opportunities_list)
    punktacja_zrodla = request.GET.get("punktacja_zrodla")
    if punktacja_zrodla:
        filtered_opportunities = []
        for opp in all_opportunities:
            try:
                punkty = opp.rekord.original.punkty_kbn or Decimal("0")
            except (AttributeError, TypeError):
                punkty = Decimal("0")

            if punktacja_zrodla == "0-100" and punkty < Decimal("100"):
                filtered_opportunities.append(opp)
            elif punktacja_zrodla == "100-140" and Decimal("100") <= punkty < Decimal(
                "140"
            ):
                filtered_opportunities.append(opp)
            elif punktacja_zrodla == "140-200" and Decimal("140") <= punkty < Decimal(
                "200"
            ):
                filtered_opportunities.append(opp)
            elif punktacja_zrodla == "200+" and punkty >= Decimal("200"):
                filtered_opportunities.append(opp)

        all_opportunities = filtered_opportunities

    opportunities_qs = all_opportunities

    # Setup workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Możliwości odpinania"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    even_row_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )

    sensible_row_fill = PatternFill(
        start_color="E7F7E7", end_color="E7F7E7", fill_type="solid"
    )

    # Define headers
    headers = [
        "Lp.",
        "Tytuł pracy",
        "Punktacja źródła",
        "Dyscyplina",
        "Autor A (do odpięcia)",
        "Rodzaj autora A",
        "ID systemu kadrowego A",
        "Jednostka A",
        "% wykorzystania slotów A",
        "Sloty nazbierane A",
        "Sloty maksymalne A",
        "B może wziąć (slotów)",
        "Slot w pracy",
        "Różnica punktów A",
        "Różnica slotów A",
        "Różnica punktów B",
        "Różnica slotów B",
        "Autor B (skorzysta)",
        "Rodzaj autora B",
        "ID systemu kadrowego B",
        "Jednostka B",
        "% wykorzystania slotów B",
        "Sloty nazbierane B",
        "Sloty maksymalne B",
        "Sensowne?",
    ]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Helper function to format rodzaj_autora
    def format_rodzaj_autora(metryka):
        if metryka.rodzaj_autora == " ":
            return "Brak danych"
        try:
            rodzaj = Rodzaj_Autora.objects.get(skrot=metryka.rodzaj_autora)
            return rodzaj.nazwa
        except Rodzaj_Autora.DoesNotExist:
            return metryka.rodzaj_autora

    # Write data rows
    last_data_row = 1
    for row_idx, opp in enumerate(opportunities_qs, 2):
        last_data_row = row_idx

        # Determine row fill (sensible = green, otherwise alternating)
        if opp.makes_sense:
            row_fill = sensible_row_fill
        else:
            row_fill = even_row_fill if row_idx % 2 == 0 else None

        col = 1

        # Lp.
        cell = ws.cell(row=row_idx, column=col, value=row_idx - 1)
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Tytuł pracy
        cell = ws.cell(row=row_idx, column=col, value=opp.rekord_tytul)
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Punktacja źródła
        try:
            punkty_kbn = float(opp.rekord.original.punkty_kbn or 0)
        except (AttributeError, TypeError, ValueError):
            punkty_kbn = 0
        cell = ws.cell(row=row_idx, column=col, value=punkty_kbn)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Dyscyplina
        cell = ws.cell(row=row_idx, column=col, value=opp.dyscyplina_naukowa.nazwa)
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Autor A (do odpięcia)
        cell = ws.cell(row=row_idx, column=col, value=str(opp.autor_could_benefit))
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Rodzaj autora A
        cell = ws.cell(
            row=row_idx,
            column=col,
            value=format_rodzaj_autora(opp.metryka_could_benefit),
        )
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # ID systemu kadrowego A
        cell = ws.cell(
            row=row_idx,
            column=col,
            value=opp.autor_could_benefit.system_kadrowy_id or "",
        )
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Jednostka A
        jednostka_a = (
            opp.metryka_could_benefit.jednostka.nazwa
            if opp.metryka_could_benefit.jednostka
            else "-"
        )
        cell = ws.cell(row=row_idx, column=col, value=jednostka_a)
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # % wykorzystania slotów A
        cell = ws.cell(
            row=row_idx,
            column=col,
            value=float(opp.metryka_could_benefit.procent_wykorzystania_slotow) / 100,
        )
        cell.number_format = "0.00%"
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Sloty nazbierane A
        cell = ws.cell(
            row=row_idx,
            column=col,
            value=float(opp.metryka_could_benefit.slot_nazbierany),
        )
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Sloty maksymalne A
        cell = ws.cell(
            row=row_idx,
            column=col,
            value=float(opp.metryka_could_benefit.slot_maksymalny),
        )
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # B może wziąć (slotów)
        cell = ws.cell(row=row_idx, column=col, value=float(opp.slots_missing))
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Slot w pracy
        cell = ws.cell(row=row_idx, column=col, value=float(opp.slot_in_work))
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Różnica punktów A
        cell = ws.cell(row=row_idx, column=col, value=float(opp.punkty_roznica_a))
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Różnica slotów A
        cell = ws.cell(row=row_idx, column=col, value=float(opp.sloty_roznica_a))
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Różnica punktów B
        cell = ws.cell(row=row_idx, column=col, value=float(opp.punkty_roznica_b))
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Różnica slotów B
        cell = ws.cell(row=row_idx, column=col, value=float(opp.sloty_roznica_b))
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Autor B (skorzysta)
        cell = ws.cell(row=row_idx, column=col, value=str(opp.autor_currently_using))
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Rodzaj autora B
        cell = ws.cell(
            row=row_idx,
            column=col,
            value=format_rodzaj_autora(opp.metryka_currently_using),
        )
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # ID systemu kadrowego B
        cell = ws.cell(
            row=row_idx,
            column=col,
            value=opp.autor_currently_using.system_kadrowy_id or "",
        )
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Jednostka B
        jednostka_b = (
            opp.metryka_currently_using.jednostka.nazwa
            if opp.metryka_currently_using.jednostka
            else "-"
        )
        cell = ws.cell(row=row_idx, column=col, value=jednostka_b)
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # % wykorzystania slotów B
        cell = ws.cell(
            row=row_idx,
            column=col,
            value=float(opp.metryka_currently_using.procent_wykorzystania_slotow) / 100,
        )
        cell.number_format = "0.00%"
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Sloty nazbierane B
        cell = ws.cell(
            row=row_idx,
            column=col,
            value=float(opp.metryka_currently_using.slot_nazbierany),
        )
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Sloty maksymalne B
        cell = ws.cell(
            row=row_idx,
            column=col,
            value=float(opp.metryka_currently_using.slot_maksymalny),
        )
        cell.number_format = "0.00"
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

        # Sensowne?
        cell = ws.cell(
            row=row_idx, column=col, value="TAK" if opp.makes_sense else "NIE"
        )
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        col += 1

    # Setup auto-filter, freeze panes
    if last_data_row > 1:
        last_col_letter = get_column_letter(len(headers))
        filter_range = f"A1:{last_col_letter}{last_data_row}"
        ws.auto_filter.ref = filter_range

    ws.freeze_panes = ws["A2"]

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except BaseException:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add summary
    summary_row = last_data_row + 2 if last_data_row > 1 else 3
    ws.cell(row=summary_row, column=1, value="Podsumowanie:")
    ws.cell(row=summary_row + 1, column=1, value="Liczba wierszy:")
    ws.cell(row=summary_row + 1, column=2, value=len(opportunities_qs))

    # Add filter information
    filter_info = []
    if dyscyplina_id:
        try:
            dyscyplina = Dyscyplina_Naukowa.objects.get(pk=dyscyplina_id)
            filter_info.append(f"Dyscyplina: {dyscyplina.nazwa}")
        except Dyscyplina_Naukowa.DoesNotExist:
            pass

    if punktacja_zrodla:
        punktacja_labels = {
            "0-100": "0-100 punktów",
            "100-140": "100-140 punktów",
            "140-200": "140-200 punktów",
            "200+": "200+ punktów",
        }
        filter_info.append(
            f"Punktacja źródła: {punktacja_labels.get(punktacja_zrodla, punktacja_zrodla)}"
        )

    if only_sensible == "1":
        filter_info.append("Tylko sensowne: TAK")

    filters_row = summary_row + 3
    ws.cell(row=filters_row, column=1, value="Zastosowane filtry:")

    if filter_info:
        ws.cell(row=filters_row + 1, column=1, value="; ".join(filter_info))
    else:
        ws.cell(row=filters_row + 1, column=1, value="Brak filtrów")

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"unpinning_opportunities_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
