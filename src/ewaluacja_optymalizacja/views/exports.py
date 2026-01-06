"""Widoki eksportu danych optymalizacji."""

import logging
import sys

import rollbar
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect

from ..models import OptimizationRun

logger = logging.getLogger(__name__)


def _generate_author_sedn_workbook(author_result, run):
    """Funkcja pomocnicza generująca workbook z pracami autora w formacie SEDN"""
    from openpyxl import Workbook

    from bpp.models.cache import Cache_Punktacja_Autora_Query
    from bpp.util import worksheet_columns_autosize, worksheet_create_table
    from ewaluacja_metryki.models import MetrykaAutora

    autor_pk = author_result.autor_id

    # Pobierz MetrykaAutora
    metryka = MetrykaAutora.objects.filter(
        autor_id=autor_pk, dyscyplina_naukowa=run.dyscyplina_naukowa
    ).first()

    if not metryka or not metryka.prace_wszystkie:
        return None

    # Pobierz wybrane prace (rekord_id) z OptimizationPublication
    selected_rekord_ids = set(
        tuple(pub.rekord_id) for pub in author_result.publications.all()
    )

    # Pobierz wszystkie prace z Cache_Punktacja_Autora_Query
    prace = (
        Cache_Punktacja_Autora_Query.objects.filter(
            rekord_id__in=metryka.prace_wszystkie,
            autor_id=autor_pk,
            dyscyplina_id=run.dyscyplina_naukowa_id,
        )
        .select_related("rekord", "rekord__charakter_formalny")
        .order_by("-pkdaut")
    )

    # Generuj XLSX
    wb = Workbook()
    ws = wb.active
    ws.title = "Prace autora"

    # Nagłówki
    headers = [
        "ID OSIĄGNIĘCIA",
        "TYTUŁ OSIĄGNIĘCIA",
        "RODZAJ",
        "ROK",
        "PUNKTACJA CAŁKOWITA",
        "WARTOŚĆ UDZIAŁÓW",
        "PUNKTACJA DO EWALUACJI",
        "STATUS",
    ]
    ws.append(headers)

    # Mapowanie rodzaj_pbn na wartości tekstowe
    rodzaj_map = {
        1: "artykuł",
        2: "rozdział",
        3: "książka",
        4: "postępowanie",
        None: "brak",
    }

    # Dane
    for praca in prace:
        rekord = praca.rekord
        rekord_id_tuple = tuple(rekord.pk)
        status = "Tak" if rekord_id_tuple in selected_rekord_ids else "Nie"

        # Pobierz PBN UID
        pbn_uid = ""
        try:
            original = rekord.original
            if hasattr(original, "pbn_uid_id") and original.pbn_uid_id:
                pbn_uid = original.pbn_uid_id
        except Exception:
            rollbar.report_exc_info(sys.exc_info())
            logger.debug("Błąd pobierania pbn_uid", exc_info=True)

        # Pobierz RODZAJ z charakter_formalny.rodzaj_pbn
        rodzaj = rodzaj_map.get(rekord.charakter_formalny.rodzaj_pbn, "brak")

        ws.append(
            [
                pbn_uid,
                rekord.tytul_oryginalny,
                rodzaj,
                rekord.rok,
                float(rekord.punkty_kbn) if rekord.punkty_kbn else 0,
                float(praca.slot),
                float(praca.pkdaut),
                status,
            ]
        )

    # Formatowanie
    worksheet_columns_autosize(ws)
    worksheet_create_table(ws, title="PraceAutora")

    return wb


@login_required
def export_author_sedn_xlsx(request, run_pk, autor_pk):
    """Eksport prac autora w formacie SEDN do XLSX"""
    from django.http import HttpResponse

    from ..models import OptimizationAuthorResult

    run = get_object_or_404(OptimizationRun, pk=run_pk)
    author_result = get_object_or_404(
        OptimizationAuthorResult, optimization_run=run, autor_id=autor_pk
    )

    wb = _generate_author_sedn_workbook(author_result, run)

    if wb is None:
        messages.warning(request, "Brak prac do eksportu dla tego autora")
        return redirect(run.get_absolute_url())

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"prace_{author_result.autor.slug}_{run.dyscyplina_naukowa.kod}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_all_authors_zip(request, run_pk):
    """Eksport prac wszystkich autorów w formacie SEDN jako archiwum ZIP"""
    import io
    import zipfile

    from django.http import HttpResponse

    from ..models import OptimizationAuthorResult

    run = get_object_or_404(OptimizationRun, pk=run_pk)

    # Pobierz wszystkie wyniki autorów dla tego uruchomienia
    author_results = OptimizationAuthorResult.objects.filter(
        optimization_run=run
    ).select_related("autor")

    if not author_results.exists():
        messages.warning(request, "Brak autorów do eksportu")
        return redirect(run.get_absolute_url())

    # Utwórz archiwum ZIP w pamięci
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for author_result in author_results:
            wb = _generate_author_sedn_workbook(author_result, run)

            if wb is None:
                # Pomiń autorów bez prac
                continue

            # Zapisz workbook do bufora w pamięci
            xlsx_buffer = io.BytesIO()
            wb.save(xlsx_buffer)
            xlsx_buffer.seek(0)

            # Dodaj plik do archiwum ZIP
            filename = (
                f"prace_{author_result.autor.slug}_{run.dyscyplina_naukowa.kod}.xlsx"
            )
            zip_file.writestr(filename, xlsx_buffer.read())

    # Przygotuj response
    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer.read(), content_type="application/zip")
    filename = f"prace_wszystkie_{run.dyscyplina_naukowa.kod}.zip"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


def generate_all_disciplines_zip_file(status_bulk):
    """Generuje ZIP ze wszystkimi XLS i zapisuje do StatusOptymalizacjiBulk.

    Args:
        status_bulk: Instancja StatusOptymalizacjiBulk do której zapisany zostanie plik.
    """
    import io
    import zipfile

    from django.core.files.base import ContentFile
    from django.utils.text import slugify

    from ..models import OptimizationAuthorResult

    # Pobierz najnowsze ukończone uruchomienia dla każdej dyscypliny
    latest_runs = (
        OptimizationRun.objects.filter(status="completed")
        .order_by("dyscyplina_naukowa", "-started_at")
        .distinct("dyscyplina_naukowa")
    )

    if not latest_runs:
        logger.warning("Brak ukończonych optymalizacji do wygenerowania ZIP")
        return

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for run in latest_runs:
            # Nazwa folderu = slugifikowana nazwa dyscypliny (bez akcentów, _ zamiast spacji)
            folder_name = slugify(run.dyscyplina_naukowa.nazwa).replace("-", "_")

            author_results = OptimizationAuthorResult.objects.filter(
                optimization_run=run
            ).select_related("autor")

            for author_result in author_results:
                wb = _generate_author_sedn_workbook(author_result, run)
                if wb is None:
                    continue

                xlsx_buffer = io.BytesIO()
                wb.save(xlsx_buffer)
                xlsx_buffer.seek(0)

                filename = (
                    f"{folder_name}/prace_{author_result.autor.slug}_{folder_name}.xlsx"
                )
                zip_file.writestr(filename, xlsx_buffer.read())

    # Zapisz do modelu (nadpisuje stary plik)
    zip_buffer.seek(0)

    # Usuń stary plik jeśli istnieje
    if status_bulk.plik_zip_wszystkie_xls:
        status_bulk.plik_zip_wszystkie_xls.delete(save=False)

    status_bulk.plik_zip_wszystkie_xls.save(
        "prace_wszystkie_uczelnia.zip",
        ContentFile(zip_buffer.read()),
        save=True,
    )

    logger.info("Wygenerowano i zapisano plik ZIP ze wszystkimi XLS")


@login_required
def export_all_disciplines_zip(request):
    """Serwuje plik ZIP z cache (StatusOptymalizacjiBulk).

    Jeśli plik nie istnieje, zwraca błąd 404.
    """
    from django.http import Http404
    from django_sendfile import sendfile

    from ..models import StatusOptymalizacjiBulk

    status = StatusOptymalizacjiBulk.get_or_create()

    if not status.plik_zip_wszystkie_xls:
        raise Http404(
            "Plik ZIP nie został jeszcze wygenerowany. Uruchom 'Policz całą ewaluację'."
        )

    return sendfile(
        request,
        status.plik_zip_wszystkie_xls.path,
        attachment=True,
        attachment_filename="prace_wszystkie_uczelnia.zip",
    )


@login_required
def export_sedn_report_1(request):
    """Raport SEDN #1 - eksport wszystkich publikacji ze wszystkich dyscyplin z informacją per-autor.

    Kolumny:
    - ID OSIĄGNIĘCIA (publication PBN UID)
    - TYTUŁ OSIĄGNIĘCIA (title)
    - RODZAJ (type from charakter_formalny.rodzaj_pbn)
    - ROK (year)
    - IMIĘ (first name)
    - NAZWISKO (surname)
    - WARTOŚĆ UDZIAŁÓW JEDNOSTKOWYCH (slot value)
    - PUNKTACJA DO EWALUACJI (pkdaut - points for evaluation)
    - PUNKTACJA CAŁKOWITA PUBLIKACJI (total publication points - punkty_kbn)
    - WSKAZANE PRZEZ ALGORYTM (True/False - if in prace_nazbierane)
    """
    from django.http import HttpResponse
    from openpyxl import Workbook

    from bpp.models.cache import Cache_Punktacja_Autora_Query
    from bpp.util import worksheet_columns_autosize, worksheet_create_table
    from ewaluacja_metryki.models import MetrykaAutora

    # Pobierz najnowsze ukończone uruchomienia dla każdej dyscypliny
    latest_runs = (
        OptimizationRun.objects.filter(status="completed")
        .order_by("dyscyplina_naukowa", "-started_at")
        .distinct("dyscyplina_naukowa")
    )

    if not latest_runs.exists():
        messages.warning(request, "Brak ukończonych optymalizacji do eksportu")
        return redirect("ewaluacja_optymalizacja:index")

    # Mapowanie rodzaj_pbn na wartości tekstowe
    rodzaj_map = {
        1: "artykuł",
        2: "rozdział",
        3: "książka",
        4: "postępowanie",
        None: "brak",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Raport SEDN 1"

    # Nagłówki
    headers = [
        "ID OSIĄGNIĘCIA",
        "TYTUŁ OSIĄGNIĘCIA",
        "RODZAJ",
        "ROK",
        "IMIĘ",
        "NAZWISKO",
        "WARTOŚĆ UDZIAŁÓW JEDNOSTKOWYCH",
        "PUNKTACJA DO EWALUACJI",
        "PUNKTACJA CAŁKOWITA PUBLIKACJI",
        "WSKAZANE PRZEZ ALGORYTM",
    ]
    ws.append(headers)

    # Zbierz ID dyscyplin z najnowszych uruchomień
    discipline_ids = [run.dyscyplina_naukowa_id for run in latest_runs]

    # Pobierz wszystkie metryki dla tych dyscyplin aby zbudować zbiór wybranych prac per autor
    metryki = MetrykaAutora.objects.filter(
        dyscyplina_naukowa_id__in=discipline_ids
    ).values("autor_id", "dyscyplina_naukowa_id", "prace_nazbierane")

    # Buduj lookup: (autor_id, dyscyplina_id) -> set of selected rekord_ids
    selected_works = {}
    for m in metryki:
        key = (m["autor_id"], m["dyscyplina_naukowa_id"])
        selected_works[key] = set(tuple(x) for x in (m["prace_nazbierane"] or []))

    # Pobierz wszystkie Cache_Punktacja_Autora_Query dla tych dyscyplin (lata 2022-2025)
    prace = (
        Cache_Punktacja_Autora_Query.objects.filter(
            dyscyplina_id__in=discipline_ids,
            rekord__rok__gte=2022,
            rekord__rok__lte=2025,
        )
        .select_related(
            "rekord",
            "rekord__charakter_formalny",
            "autor",
        )
        .order_by("dyscyplina_id", "autor__nazwisko", "-pkdaut")
    )

    for praca in prace:
        rekord = praca.rekord
        autor = praca.autor

        # Pobierz PBN UID - bezpośrednio z Rekord cache (bez N+1 query)
        pbn_uid = rekord.pbn_uid_id or ""

        # Pobierz rodzaj
        rodzaj = rodzaj_map.get(rekord.charakter_formalny.rodzaj_pbn, "brak")

        # Sprawdź czy wybrane (w prace_nazbierane dla tego autor+dyscyplina)
        key = (praca.autor_id, praca.dyscyplina_id)
        rekord_id_tuple = tuple(rekord.pk)
        is_selected = rekord_id_tuple in selected_works.get(key, set())

        ws.append(
            [
                pbn_uid,
                rekord.tytul_oryginalny,
                rodzaj,
                rekord.rok,
                autor.imiona,
                autor.nazwisko,
                float(praca.slot),
                float(praca.pkdaut),
                float(rekord.punkty_kbn) if rekord.punkty_kbn else 0,
                "Tak" if is_selected else "Nie",
            ]
        )

    worksheet_columns_autosize(ws)
    worksheet_create_table(ws, title="RaportSEDN1")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="raport_sedn_1.xlsx"'
    wb.save(response)
    return response


@login_required
def export_sedn_report_2(request):
    """Raport SEDN #2 - eksport publikacji z agregacją na poziomie publikacji.

    Kolumny:
    - ROK (year)
    - ID (publication PBN UID)
    - TYTUŁ (title)
    - RODZAJ (type)
    - OSOBY, KTÓRE OŚWIADCZYŁY TO OSIĄGNIĘCIE (people who declared this publication)
    - PUNKTACJA ŹRÓDŁA (source points - punkty_kbn)
    - PUNKTACJA OGÓŁEM (total points - sum of pkdaut for all authors)
    - PUNKTACJA UDZIAŁÓW W EWALUACJI (evaluation share points - sum of pkdaut where selected)
    """
    from collections import defaultdict

    from django.http import HttpResponse
    from openpyxl import Workbook

    from bpp.models.cache import Cache_Punktacja_Autora_Query
    from bpp.util import worksheet_columns_autosize, worksheet_create_table
    from ewaluacja_metryki.models import MetrykaAutora

    # Pobierz najnowsze ukończone uruchomienia dla każdej dyscypliny
    latest_runs = (
        OptimizationRun.objects.filter(status="completed")
        .order_by("dyscyplina_naukowa", "-started_at")
        .distinct("dyscyplina_naukowa")
    )

    if not latest_runs.exists():
        messages.warning(request, "Brak ukończonych optymalizacji do eksportu")
        return redirect("ewaluacja_optymalizacja:index")

    rodzaj_map = {
        1: "artykuł",
        2: "rozdział",
        3: "książka",
        4: "postępowanie",
        None: "brak",
    }

    discipline_ids = [run.dyscyplina_naukowa_id for run in latest_runs]

    # Buduj lookup wybranych prac
    metryki = MetrykaAutora.objects.filter(
        dyscyplina_naukowa_id__in=discipline_ids
    ).values("autor_id", "dyscyplina_naukowa_id", "prace_nazbierane")

    selected_works = {}
    for m in metryki:
        key = (m["autor_id"], m["dyscyplina_naukowa_id"])
        selected_works[key] = set(tuple(x) for x in (m["prace_nazbierane"] or []))

    # Agreguj dane po publikacji (rekord_id)
    # Struktura: rekord_id -> {
    #   'rekord': rekord,
    #   'osoby': list of "Imię Nazwisko",
    #   'punktacja_ogolem': sum of pkdaut,
    #   'punktacja_ewaluacji': sum of pkdaut where selected
    # }
    publication_data = defaultdict(
        lambda: {
            "rekord": None,
            "osoby": [],
            "punktacja_ogolem": 0,
            "punktacja_ewaluacji": 0,
        }
    )

    # Pobierz prace z lat 2022-2025
    prace = (
        Cache_Punktacja_Autora_Query.objects.filter(
            dyscyplina_id__in=discipline_ids,
            rekord__rok__gte=2022,
            rekord__rok__lte=2025,
        )
        .select_related(
            "rekord",
            "rekord__charakter_formalny",
            "autor",
        )
        .order_by("rekord__rok", "rekord__tytul_oryginalny")
    )

    for praca in prace:
        rekord = praca.rekord
        rekord_id = tuple(rekord.pk)
        autor = praca.autor

        data = publication_data[rekord_id]
        data["rekord"] = rekord

        osoba_str = f"{autor.imiona} {autor.nazwisko}"
        if osoba_str not in data["osoby"]:
            data["osoby"].append(osoba_str)

        data["punktacja_ogolem"] += float(praca.pkdaut)

        # Sprawdź czy ten autor wybrał tę pracę
        key = (praca.autor_id, praca.dyscyplina_id)
        if rekord_id in selected_works.get(key, set()):
            data["punktacja_ewaluacji"] += float(praca.pkdaut)

    wb = Workbook()
    ws = wb.active
    ws.title = "Raport SEDN 2"

    headers = [
        "ROK",
        "ID",
        "TYTUŁ",
        "RODZAJ",
        "OSOBY, KTÓRE OŚWIADCZYŁY TO OSIĄGNIĘCIE",
        "PUNKTACJA ŹRÓDŁA",
        "PUNKTACJA OGÓŁEM",
        "PUNKTACJA UDZIAŁÓW W EWALUACJI",
    ]
    ws.append(headers)

    # Sortuj po roku, potem tytule
    sorted_items = sorted(
        publication_data.items(),
        key=lambda x: (x[1]["rekord"].rok or 0, x[1]["rekord"].tytul_oryginalny or ""),
    )

    for _rekord_id, data in sorted_items:
        rekord = data["rekord"]

        # Pobierz PBN UID - bezpośrednio z Rekord cache (bez N+1 query)
        pbn_uid = rekord.pbn_uid_id or ""

        rodzaj = rodzaj_map.get(rekord.charakter_formalny.rodzaj_pbn, "brak")
        osoby_str = "; ".join(data["osoby"])

        ws.append(
            [
                rekord.rok,
                pbn_uid,
                rekord.tytul_oryginalny,
                rodzaj,
                osoby_str,
                float(rekord.punkty_kbn) if rekord.punkty_kbn else 0,
                round(data["punktacja_ogolem"], 4),
                round(data["punktacja_ewaluacji"], 4),
            ]
        )

    worksheet_columns_autosize(ws)
    worksheet_create_table(ws, title="RaportSEDN2")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="raport_sedn_2.xlsx"'
    wb.save(response)
    return response
