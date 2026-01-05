"""Widoki listy możliwości zamiany dyscyplin."""

import datetime
import logging

from celery.result import AsyncResult
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.views.decorators.http import require_POST

from bpp.models import Dyscyplina_Naukowa, Uczelnia

logger = logging.getLogger(__name__)


def _cache_rekord_objects(opportunities_qs):
    """Preload Rekord objects to avoid N+1 queries."""
    from bpp.models import Rekord

    all_opportunities = list(opportunities_qs)
    rekord_ids = [opp.rekord_id for opp in all_opportunities]
    rekordy = {r.pk: r for r in Rekord.objects.filter(pk__in=rekord_ids)}

    # Attach rekord objects to opportunities
    for opp in all_opportunities:
        opp._cached_rekord = rekordy.get(opp.rekord_id)

    return all_opportunities


def _parse_int_filter(value):
    """Parse integer filter value, return None if invalid."""
    if not value:
        return None
    try:
        return int(value)
    except (ValueError, TypeError):
        return None


def _apply_filters(opportunities_qs, request):
    """Apply common filters to the opportunities queryset."""
    # Filtr po obecnej dyscyplinie
    current_disc_id = _parse_int_filter(request.GET.get("current_discipline"))
    if current_disc_id:
        opportunities_qs = opportunities_qs.filter(
            current_discipline_id=current_disc_id
        )

    # Filtr po docelowej dyscyplinie
    target_disc_id = _parse_int_filter(request.GET.get("target_discipline"))
    if target_disc_id:
        opportunities_qs = opportunities_qs.filter(target_discipline_id=target_disc_id)

    # Filtr "tylko sensowne"
    only_sensible = request.GET.get("only_sensible") == "1"
    if only_sensible:
        opportunities_qs = opportunities_qs.filter(makes_sense=True)

    # Filtr po typie publikacji
    pub_type = request.GET.get("pub_type")
    if pub_type in ["Ciagle", "Zwarte"]:
        opportunities_qs = opportunities_qs.filter(rekord_typ=pub_type)

    # Filtr po roku
    rok = _parse_int_filter(request.GET.get("rok"))
    if rok:
        opportunities_qs = opportunities_qs.filter(rekord_rok=rok)

    # Filtr po zgodności ze źródłem
    zrodlo_match = request.GET.get("zrodlo_match")
    if zrodlo_match == "1":
        opportunities_qs = opportunities_qs.filter(zrodlo_discipline_match=True)
    elif zrodlo_match == "0":
        opportunities_qs = opportunities_qs.filter(zrodlo_discipline_match=False)

    return opportunities_qs


def _check_running_analysis(status):
    """Check if analysis task is running and return task if active."""
    if not (status.w_trakcie and status.task_id):
        return None

    task = AsyncResult(status.task_id)
    if task.state in ["PENDING", "STARTED", "PROGRESS"]:
        return task
    if task.state in ["SUCCESS", "FAILURE"]:
        status.zakoncz("Zadanie zakończone")
    return None


@login_required
def discipline_swap_opportunities_list(request):
    """
    Wyświetla listę możliwości zamiany dyscyplin.
    """
    from ..models import DisciplineSwapOpportunity, StatusDisciplineSwapAnalysis

    uczelnia = Uczelnia.objects.first()

    if not uczelnia:
        messages.error(request, "Nie znaleziono uczelni w systemie.")
        return redirect("ewaluacja_optymalizacja:index")

    # Sprawdź czy analiza działa
    status = StatusDisciplineSwapAnalysis.get_or_create()
    if _check_running_analysis(status):
        messages.info(
            request,
            "Analiza możliwości zamiany dyscyplin jest w trakcie. "
            "Przekierowuję do strony postępu...",
        )
        return redirect(
            "ewaluacja_optymalizacja:discipline-swap-status",
            task_id=status.task_id,
        )

    # Podstawowe filtrowanie
    opportunities_qs = DisciplineSwapOpportunity.objects.filter(uczelnia=uczelnia)

    # Zastosuj filtry
    opportunities_qs = _apply_filters(opportunities_qs, request)

    # Select related dla optymalizacji
    opportunities_qs = opportunities_qs.select_related(
        "autor",
        "current_discipline",
        "target_discipline",
    )

    # Sortowanie
    sort_by = request.GET.get("sort_by", "point_improvement")
    sort_dir = request.GET.get("sort_dir", "desc")

    sort_fields = {
        "tytul": "rekord_tytul",
        "rok": "rekord_rok",
        "autor": "autor__nazwisko",
        "point_improvement": "point_improvement",
        "zrodlo_match": "zrodlo_discipline_match",
        "makes_sense": "makes_sense",
    }

    order_field = sort_fields.get(sort_by, "point_improvement")
    if sort_dir == "desc":
        order_field = f"-{order_field}"
    opportunities_qs = opportunities_qs.order_by(order_field)

    # Paginacja
    paginator = Paginator(opportunities_qs, 50)
    page_number = request.GET.get("page")
    opportunities = paginator.get_page(page_number)

    # Statystyki
    total_count = DisciplineSwapOpportunity.objects.filter(uczelnia=uczelnia).count()
    sensible_count = DisciplineSwapOpportunity.objects.filter(
        uczelnia=uczelnia, makes_sense=True
    ).count()
    sensible_percentage = (
        round((sensible_count / total_count) * 100, 1) if total_count > 0 else 0
    )

    # Lista dyscyplin dla filtrów
    dyscypliny = Dyscyplina_Naukowa.objects.filter(widoczna=True).order_by("nazwa")

    # Lista lat dostępnych w danych
    years = (
        DisciplineSwapOpportunity.objects.filter(uczelnia=uczelnia)
        .values_list("rekord_rok", flat=True)
        .distinct()
        .order_by("rekord_rok")
    )

    # Data ostatniej analizy
    last_analysis = (
        DisciplineSwapOpportunity.objects.filter(uczelnia=uczelnia)
        .order_by("-created_at")
        .first()
    )

    context = {
        "opportunities": opportunities,
        "total_count": total_count,
        "sensible_count": sensible_count,
        "sensible_percentage": sensible_percentage,
        "dyscypliny": dyscypliny,
        "years": years,
        "selected_current_discipline": request.GET.get("current_discipline") or "",
        "selected_target_discipline": request.GET.get("target_discipline") or "",
        "only_sensible": request.GET.get("only_sensible") == "1",
        "selected_pub_type": request.GET.get("pub_type") or "",
        "selected_rok": request.GET.get("rok") or "",
        "selected_zrodlo_match": request.GET.get("zrodlo_match") or "",
        "sort_by": sort_by,
        "sort_dir": sort_dir,
        "last_analysis": last_analysis,
        "status": status,
    }

    return render(
        request,
        "ewaluacja_optymalizacja/discipline_swap_list.html",
        context,
    )


def _create_xlsx_styles():
    """Create and return Excel styles for export."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    return {
        "header_font": Font(bold=True, color="FFFFFF"),
        "header_fill": PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        ),
        "header_alignment": Alignment(
            horizontal="center", vertical="center", wrap_text=True
        ),
        "thin_border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
        "even_row_fill": PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        ),
        "sensible_row_fill": PatternFill(
            start_color="E7F7E7", end_color="E7F7E7", fill_type="solid"
        ),
        "center_alignment": Alignment(horizontal="center"),
        "right_alignment": Alignment(horizontal="right"),
    }


def _write_xlsx_header(ws, headers, styles):
    """Write header row to worksheet."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.alignment = styles["header_alignment"]
        cell.border = styles["thin_border"]


def _get_row_fill(opp, row_idx, styles):
    """Determine the fill color for a row."""
    if opp.makes_sense:
        return styles["sensible_row_fill"]
    return styles["even_row_fill"] if row_idx % 2 == 0 else None


def _write_xlsx_cell(ws, row, col, value, styles, row_fill, alignment=None, fmt=None):
    """Write a cell with styling."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = styles["thin_border"]
    if row_fill:
        cell.fill = row_fill
    if alignment:
        cell.alignment = alignment
    if fmt:
        cell.number_format = fmt
    return cell


def _write_opportunity_row(ws, row_idx, opp, styles):
    """Write a single opportunity row to the worksheet."""
    row_fill = _get_row_fill(opp, row_idx, styles)
    col = 1

    # Lp.
    _write_xlsx_cell(ws, row_idx, col, row_idx - 1, styles, row_fill)
    col += 1

    # Tytuł
    _write_xlsx_cell(ws, row_idx, col, opp.rekord_tytul, styles, row_fill)
    col += 1

    # Rok
    _write_xlsx_cell(
        ws,
        row_idx,
        col,
        opp.rekord_rok,
        styles,
        row_fill,
        alignment=styles["center_alignment"],
    )
    col += 1

    # Typ
    _write_xlsx_cell(ws, row_idx, col, opp.rekord_typ, styles, row_fill)
    col += 1

    # Autor
    _write_xlsx_cell(ws, row_idx, col, str(opp.autor), styles, row_fill)
    col += 1

    # Obecna dyscyplina
    _write_xlsx_cell(ws, row_idx, col, opp.current_discipline.nazwa, styles, row_fill)
    col += 1

    # Docelowa dyscyplina
    _write_xlsx_cell(ws, row_idx, col, opp.target_discipline.nazwa, styles, row_fill)
    col += 1

    # Punkty przed
    _write_xlsx_cell(
        ws,
        row_idx,
        col,
        float(opp.points_before),
        styles,
        row_fill,
        alignment=styles["right_alignment"],
        fmt="0.00",
    )
    col += 1

    # Punkty po
    _write_xlsx_cell(
        ws,
        row_idx,
        col,
        float(opp.points_after),
        styles,
        row_fill,
        alignment=styles["right_alignment"],
        fmt="0.00",
    )
    col += 1

    # Poprawa
    _write_xlsx_cell(
        ws,
        row_idx,
        col,
        float(opp.point_improvement),
        styles,
        row_fill,
        alignment=styles["right_alignment"],
        fmt="0.00",
    )
    col += 1

    # Dyscyplina pasuje do źródła
    _write_xlsx_cell(
        ws,
        row_idx,
        col,
        "TAK" if opp.zrodlo_discipline_match else "NIE",
        styles,
        row_fill,
        alignment=styles["center_alignment"],
    )
    col += 1

    # Sensowne?
    _write_xlsx_cell(
        ws,
        row_idx,
        col,
        "TAK" if opp.makes_sense else "NIE",
        styles,
        row_fill,
        alignment=styles["center_alignment"],
    )


def _adjust_column_widths(ws):
    """Auto-adjust column widths based on content."""
    from openpyxl.utils import get_column_letter

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, ValueError):
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


@login_required
def export_discipline_swap_xlsx(request):
    """Export discipline swap opportunities to XLSX format with filters applied."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    from ..models import DisciplineSwapOpportunity

    uczelnia = Uczelnia.objects.first()

    if not uczelnia:
        messages.error(request, "Nie znaleziono uczelni.")
        return redirect("ewaluacja_optymalizacja:discipline-swap-list")

    # Filtrowanie
    opportunities_qs = DisciplineSwapOpportunity.objects.filter(uczelnia=uczelnia)
    opportunities_qs = _apply_filters(opportunities_qs, request)

    opportunities_qs = opportunities_qs.select_related(
        "autor", "current_discipline", "target_discipline"
    ).order_by("-point_improvement")

    # Utwórz workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Możliwości zamiany dyscyplin"

    # Style
    styles = _create_xlsx_styles()

    # Nagłówki
    headers = [
        "Lp.",
        "Tytuł pracy",
        "Rok",
        "Typ",
        "Autor",
        "Obecna dyscyplina",
        "Docelowa dyscyplina",
        "Punkty przed",
        "Punkty po",
        "Poprawa",
        "Dyscyplina pasuje do źródła",
        "Sensowne?",
    ]

    _write_xlsx_header(ws, headers, styles)

    # Dane
    last_data_row = 1
    for row_idx, opp in enumerate(opportunities_qs, 2):
        last_data_row = row_idx
        _write_opportunity_row(ws, row_idx, opp, styles)

    # Auto-filter i freeze
    if last_data_row > 1:
        last_col_letter = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_data_row}"

    ws.freeze_panes = ws["A2"]

    # Auto-adjust column widths
    _adjust_column_widths(ws)

    # Utwórz odpowiedź
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = (
        f"discipline_swap_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
@require_POST
def cancel_discipline_swap_task(request, task_id):
    """
    Anuluje zadanie analizy zamiany dyscyplin.
    """
    from django_bpp.celery_tasks import app

    from ..models import StatusDisciplineSwapAnalysis

    task = AsyncResult(task_id)  # noqa: F841

    status = StatusDisciplineSwapAnalysis.get_or_create()
    if not status.w_trakcie or status.task_id != task_id:
        messages.error(
            request,
            "Nie możesz anulować tego zadania - nie jest aktualnie uruchomione.",
        )
        return redirect("ewaluacja_optymalizacja:index")

    try:
        app.control.revoke(task_id, terminate=True)
        AsyncResult(task_id).forget()
        status.zakoncz("Zadanie anulowane przez użytkownika")

        logger.info(f"Task {task_id} cancelled by user {request.user}")
        messages.success(request, "Zadanie zostało anulowane.")

    except Exception as e:
        logger.error(f"Failed to cancel task {task_id}: {e}")
        messages.error(request, f"Nie udało się anulować zadania: {e}")

    return redirect("ewaluacja_optymalizacja:index")
