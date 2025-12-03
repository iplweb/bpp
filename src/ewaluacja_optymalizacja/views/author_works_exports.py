"""Eksport prac autora w kontekście optymalizacji."""

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

from bpp.models import Autor
from bpp.models.cache import Cache_Punktacja_Autora_Query
from bpp.models.patent import Patent_Autor
from bpp.models.wydawnictwo_ciagle import Wydawnictwo_Ciagle_Autor
from bpp.models.wydawnictwo_zwarte import Wydawnictwo_Zwarte_Autor
from bpp.util import worksheet_columns_autosize, worksheet_create_table
from ewaluacja_metryki.models import MetrykaAutora

from ..models import OptimizationRun


def _get_author_works_data(run_pk, autor_pk):
    """Helper to get author works data for exports."""
    run = get_object_or_404(
        OptimizationRun.objects.select_related("dyscyplina_naukowa"),
        pk=run_pk,
    )
    autor = get_object_or_404(Autor, pk=autor_pk)
    metryka = get_object_or_404(
        MetrykaAutora,
        autor=autor,
        dyscyplina_naukowa=run.dyscyplina_naukowa,
    )

    prace_nazbierane_ids = metryka.prace_nazbierane or []
    prace_wszystkie_ids = set(metryka.prace_wszystkie or [])
    prace_nienazbierane_ids = prace_wszystkie_ids - set(prace_nazbierane_ids)

    prace_nazbierane = (
        Cache_Punktacja_Autora_Query.objects.filter(
            rekord_id__in=prace_nazbierane_ids,
            autor=autor,
            dyscyplina=run.dyscyplina_naukowa,
        )
        .select_related("rekord", "jednostka")
        .order_by("-pkdaut")
    )

    prace_nienazbierane = (
        Cache_Punktacja_Autora_Query.objects.filter(
            rekord_id__in=prace_nienazbierane_ids,
            autor=autor,
            dyscyplina=run.dyscyplina_naukowa,
        )
        .select_related("rekord", "jednostka")
        .order_by("-pkdaut")
    )

    # Unpinned works
    prace_odpiete = []
    for model, typ in [
        (Wydawnictwo_Ciagle_Autor, "Wydawnictwo ciagłe"),
        (Wydawnictwo_Zwarte_Autor, "Wydawnictwo zwarte"),
        (Patent_Autor, "Patent"),
    ]:
        odpiete = model.objects.filter(
            autor=autor,
            dyscyplina_naukowa=run.dyscyplina_naukowa,
            przypieta=False,
        ).select_related("rekord")
        for wa in odpiete:
            prace_odpiete.append({"rekord": wa.rekord, "typ": typ})

    prace_odpiete.sort(
        key=lambda x: (-(x["rekord"].rok or 0), x["rekord"].tytul_oryginalny)
    )

    return run, autor, metryka, prace_nazbierane, prace_nienazbierane, prace_odpiete


def _create_works_worksheet(ws, title, prace, include_points=True):
    """Helper to populate a worksheet with works data."""
    ws.title = title

    if include_points:
        headers = ["Tytul", "Rok", "Punkty", "PKDaut", "Slot"]
        ws.append(headers)
        suma_pkdaut = 0
        suma_slot = 0
        for praca in prace:
            punkty_kbn = praca.rekord.punkty_kbn
            ws.append(
                [
                    praca.rekord.tytul_oryginalny,
                    praca.rekord.rok,
                    float(punkty_kbn) if punkty_kbn else None,
                    float(praca.pkdaut),
                    float(praca.slot),
                ]
            )
            suma_pkdaut += float(praca.pkdaut)
            suma_slot += float(praca.slot)
        # Add summary row
        ws.append(["SUMA:", "", "", suma_pkdaut, suma_slot])
    else:
        # For unpinned works (dict format)
        headers = ["Tytul", "Rok", "Punkty", "Typ"]
        ws.append(headers)
        for praca in prace:
            punkty_kbn = praca["rekord"].punkty_kbn
            ws.append(
                [
                    praca["rekord"].tytul_oryginalny,
                    praca["rekord"].rok,
                    float(punkty_kbn) if punkty_kbn else None,
                    praca["typ"],
                ]
            )

    worksheet_columns_autosize(ws)
    # Create table excluding summary row for include_points=True
    row_count = len(list(ws.rows))
    if row_count > 1:
        if include_points and row_count > 2:
            # Exclude summary row from table
            from openpyxl.worksheet.table import Table, TableStyleInfo

            table_range = f"A1:E{row_count - 1}"
            table = Table(displayName=title.replace(" ", ""), ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            table.tableStyleInfo = style
            ws.add_table(table)
        else:
            worksheet_create_table(ws, title=title.replace(" ", ""))


@login_required
def export_prace_nazbierane_xlsx(request, run_pk, autor_pk):
    """Export collected works to XLSX."""
    from openpyxl import Workbook

    run, autor, metryka, prace_nazbierane, _, _ = _get_author_works_data(
        run_pk, autor_pk
    )

    wb = Workbook()
    ws = wb.active
    _create_works_worksheet(
        ws, "Prace nazbierane", prace_nazbierane, include_points=True
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"prace_nazbierane_{autor.slug}_{run.dyscyplina_naukowa.kod}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_prace_nienazbierane_xlsx(request, run_pk, autor_pk):
    """Export not collected works to XLSX."""
    from openpyxl import Workbook

    run, autor, metryka, _, prace_nienazbierane, _ = _get_author_works_data(
        run_pk, autor_pk
    )

    wb = Workbook()
    ws = wb.active
    _create_works_worksheet(
        ws, "Prace nienazbierane", prace_nienazbierane, include_points=True
    )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"prace_nienazbierane_{autor.slug}_{run.dyscyplina_naukowa.kod}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_prace_odpiete_xlsx(request, run_pk, autor_pk):
    """Export unpinned works to XLSX."""
    from openpyxl import Workbook

    run, autor, metryka, _, _, prace_odpiete = _get_author_works_data(run_pk, autor_pk)

    wb = Workbook()
    ws = wb.active
    _create_works_worksheet(ws, "Prace odpiete", prace_odpiete, include_points=False)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"prace_odpiete_{autor.slug}_{run.dyscyplina_naukowa.kod}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_all_works_xlsx(request, run_pk, autor_pk):
    """Export all works (collected, not collected, unpinned) to XLSX with separate sheets."""
    from openpyxl import Workbook

    (
        run,
        autor,
        metryka,
        prace_nazbierane,
        prace_nienazbierane,
        prace_odpiete,
    ) = _get_author_works_data(run_pk, autor_pk)

    wb = Workbook()

    # Sheet 1: Prace nazbierane
    ws1 = wb.active
    _create_works_worksheet(
        ws1, "Prace nazbierane", prace_nazbierane, include_points=True
    )

    # Sheet 2: Prace nienazbierane
    ws2 = wb.create_sheet()
    _create_works_worksheet(
        ws2, "Prace nienazbierane", prace_nienazbierane, include_points=True
    )

    # Sheet 3: Prace odpiete
    ws3 = wb.create_sheet()
    _create_works_worksheet(ws3, "Prace odpiete", prace_odpiete, include_points=False)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = f"prace_wszystkie_{autor.slug}_{run.dyscyplina_naukowa.kod}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
