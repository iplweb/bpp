"""Widoki listy możliwości odpinania."""

import logging
from collections import namedtuple
from decimal import Decimal

from celery.result import AsyncResult
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render
from django.views.decorators.http import require_POST
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from bpp.models import Uczelnia

logger = logging.getLogger(__name__)


def _get_dyscyplina_filter(request):
    """Extract and validate dyscyplina ID from request."""
    dyscyplina_id = request.GET.get("dyscyplina")
    if not dyscyplina_id:
        return None

    try:
        return int(dyscyplina_id)
    except (ValueError, TypeError):
        return None


def _cache_rekord_objects(opportunities_qs):
    """Preload Rekord objects to avoid N+1 queries."""
    from bpp.models import Rekord

    all_opportunities = list(opportunities_qs)
    rekord_ids = [opp.rekord_id for opp in all_opportunities]
    rekordy = {r.pk: r for r in Rekord.objects.filter(pk__in=rekord_ids)}

    # Attach rekord objects to opportunities
    for opp in all_opportunities:
        opp._cached_rekord = rekordy.get(opp.rekord_id)

    return all_opportunities


def _get_punkty_kbn(opportunity):
    """Safely get punkty_kbn from opportunity's rekord."""
    try:
        return opportunity.rekord.original.punkty_kbn or Decimal("0")
    except (AttributeError, TypeError):
        return Decimal("0")


def _filter_by_punktacja(opportunities, punktacja_zrodla):
    """Filter opportunities by punktacja_zrodla range."""
    if not punktacja_zrodla:
        return opportunities

    punktacja_ranges = {
        "0-100": lambda p: p < Decimal("100"),
        "100-140": lambda p: Decimal("100") <= p < Decimal("140"),
        "140-200": lambda p: Decimal("140") <= p < Decimal("200"),
        "200+": lambda p: p >= Decimal("200"),
    }

    filter_func = punktacja_ranges.get(punktacja_zrodla)
    if not filter_func:
        return opportunities

    return [opp for opp in opportunities if filter_func(_get_punkty_kbn(opp))]


def _create_sort_key_function(sort_by):
    """Create a sort key function based on sort_by parameter."""
    # Define sort key extractors
    sort_keys = {
        "tytul": lambda opp: opp.rekord_tytul or "",
        "punktacja": _get_punkty_kbn,
        "dyscyplina": lambda opp: (
            opp.dyscyplina_naukowa.nazwa if opp.dyscyplina_naukowa else ""
        ),
        "autor_a": lambda opp: (
            opp.autor_could_benefit.nazwisko if opp.autor_could_benefit else ""
        ),
        "slots_missing": lambda opp: opp.slots_missing or Decimal("0"),
        "slot_in_work": lambda opp: opp.slot_in_work or Decimal("0"),
        "punkty_a": lambda opp: opp.punkty_roznica_a or Decimal("0"),
        "sloty_a": lambda opp: opp.sloty_roznica_a or Decimal("0"),
        "punkty_b": lambda opp: opp.punkty_roznica_b or Decimal("0"),
        "sloty_b": lambda opp: opp.sloty_roznica_b or Decimal("0"),
        "autor_b": lambda opp: (
            opp.autor_currently_using.nazwisko if opp.autor_currently_using else ""
        ),
        "makes_sense": lambda opp: opp.makes_sense,
    }

    # Return the requested sort key function, defaulting to punkty_b
    return sort_keys.get(sort_by, sort_keys["punkty_b"])


@login_required
def unpinning_opportunities_list(request):
    """
    Wyświetla listę możliwości odpinania prac wieloautorskich.
    """
    from bpp.models import Dyscyplina_Naukowa
    from ewaluacja_metryki.models import StatusGenerowania

    from ..models import StatusUnpinningAnalyzy, UnpinningOpportunity

    # Pobierz pierwszą uczelnię (zakładamy, że jest tylko jedna)
    uczelnia = Uczelnia.objects.first()

    if not uczelnia:
        messages.error(request, "Nie znaleziono uczelni w systemie.")
        return redirect("ewaluacja_optymalizacja:index")

    # Check if unpinning analysis is already running - używa modelu singleton zamiast sesji
    status_unpinning = StatusUnpinningAnalyzy.get_or_create()
    task_id = status_unpinning.task_id if status_unpinning.w_trakcie else None

    if task_id:
        task = AsyncResult(task_id)
        # Check if task is still running (PENDING, STARTED, or PROGRESS)
        if task.state in ["PENDING", "STARTED", "PROGRESS"]:
            messages.info(
                request,
                "Analiza możliwości odpinania jest w trakcie. "
                "Przekierowuję do strony postępu...",
            )
            return redirect(
                "ewaluacja_optymalizacja:unpinning-combined-status", task_id=task_id
            )
        # Task finished or failed - clear from database
        if task.state in ["SUCCESS", "FAILURE"]:
            status_unpinning.zakoncz("Zadanie zakończone")

    # Also check if metrics generation is running
    status = StatusGenerowania.get_or_create()
    if status.w_trakcie and task_id:
        # Metrics are running as part of unpinning chain
        messages.info(
            request,
            "Przeliczanie metryk dla analizy odpinania jest w trakcie. "
            "Przekierowuję do strony postępu...",
        )
        return redirect(
            "ewaluacja_optymalizacja:unpinning-combined-status", task_id=task_id
        )

    # Podstawowe filtrowanie
    opportunities_qs = UnpinningOpportunity.objects.filter(uczelnia=uczelnia)

    # Filtr po dyscyplinie
    dyscyplina_id = _get_dyscyplina_filter(request)
    if dyscyplina_id:
        opportunities_qs = opportunities_qs.filter(dyscyplina_naukowa_id=dyscyplina_id)

    # Filtr "tylko sensowne"
    only_sensible = request.GET.get("only_sensible") == "1"
    if only_sensible:
        opportunities_qs = opportunities_qs.filter(makes_sense=True)

    # Select related dla optymalizacji
    opportunities_qs = opportunities_qs.select_related(
        "autor_could_benefit",
        "autor_currently_using",
        "dyscyplina_naukowa",
        "metryka_could_benefit",
        "metryka_currently_using",
        "metryka_could_benefit__autor",
        "metryka_currently_using__autor",
    )

    # Preload Rekord objects
    all_opportunities = _cache_rekord_objects(opportunities_qs)

    # Filtr po punktacji źródła
    punktacja_zrodla = request.GET.get("punktacja_zrodla")
    all_opportunities = _filter_by_punktacja(all_opportunities, punktacja_zrodla)

    # Sortowanie
    sort_by = request.GET.get("sort_by", "punkty_b")
    sort_dir = request.GET.get("sort_dir", "desc")
    sort_key = _create_sort_key_function(sort_by)
    all_opportunities = sorted(
        all_opportunities, key=sort_key, reverse=(sort_dir == "desc")
    )

    # Paginacja po filtrowaniu i sortowaniu
    from django.core.paginator import Paginator

    paginator = Paginator(all_opportunities, 50)
    page_number = request.GET.get("page")
    opportunities = paginator.get_page(page_number)

    # Statystyki
    total_count = UnpinningOpportunity.objects.filter(uczelnia=uczelnia).count()
    sensible_count = UnpinningOpportunity.objects.filter(
        uczelnia=uczelnia, makes_sense=True
    ).count()
    sensible_percentage = (
        round((sensible_count / total_count) * 100, 1) if total_count > 0 else 0
    )

    # Lista dyscyplin dla filtra
    dyscypliny = Dyscyplina_Naukowa.objects.filter(
        pk__in=UnpinningOpportunity.objects.filter(uczelnia=uczelnia).values_list(
            "dyscyplina_naukowa_id", flat=True
        )
    ).order_by("nazwa")

    context = {
        "opportunities": opportunities,
        "total_count": total_count,
        "sensible_count": sensible_count,
        "sensible_percentage": sensible_percentage,
        "dyscypliny": dyscypliny,
        "selected_dyscyplina": dyscyplina_id,
        "only_sensible": only_sensible,
        "selected_punktacja_zrodla": punktacja_zrodla or "",
        "sort_by": sort_by,
        "sort_dir": sort_dir,
        "status_unpinning": status_unpinning,
    }

    return render(
        request, "ewaluacja_optymalizacja/unpinning_opportunities_list.html", context
    )


def _format_rodzaj_autora(metryka):
    """Zwróć czytelną nazwę rodzaju autora na podstawie skrótu w metryce."""
    from ewaluacja_common.models import Rodzaj_Autora

    if metryka.rodzaj_autora == " ":
        return "Brak danych"
    try:
        return Rodzaj_Autora.objects.get(skrot=metryka.rodzaj_autora).nazwa
    except Rodzaj_Autora.DoesNotExist:
        return metryka.rodzaj_autora


def _punktacja_zrodla_cell(opp):
    """Punktacja źródła jako float (komórka XLSX), 0 gdy brak danych."""
    try:
        return float(opp.rekord.original.punkty_kbn or 0)
    except (AttributeError, TypeError, ValueError):
        return 0


def _jednostka_nazwa(metryka):
    return metryka.jednostka.nazwa if metryka.jednostka else "-"


# Opis kolumn eksportu XLSX. Każdy wpis to (nagłówek, getter, format, align),
# gdzie getter dostaje (opp, lp) i zwraca wartość komórki. Pętla pisząca wiersz
# jest jedna — dzięki temu znika 25 powtórzonych bloków ``if row_fill``.
_ExportColumn = namedtuple("_ExportColumn", "header getter number_format alignment")

_FMT_DECIMAL = "0.00"
_FMT_PERCENT = "0.00%"
_ALIGN_CENTER = Alignment(horizontal="center")
_ALIGN_RIGHT = Alignment(horizontal="right")


def _build_export_columns():
    """Zbuduj listę deskryptorów kolumn (lazy — używa Alignment z openpyxl)."""

    def metryka_a(opp):
        return opp.metryka_could_benefit

    def metryka_b(opp):
        return opp.metryka_currently_using

    return [
        _ExportColumn("Lp.", lambda opp, lp: lp, None, None),
        _ExportColumn("Tytuł pracy", lambda opp, lp: opp.rekord_tytul, None, None),
        _ExportColumn(
            "Punktacja źródła",
            lambda opp, lp: _punktacja_zrodla_cell(opp),
            None,
            _ALIGN_CENTER,
        ),
        _ExportColumn(
            "Dyscyplina",
            lambda opp, lp: opp.dyscyplina_naukowa.nazwa,
            None,
            None,
        ),
        _ExportColumn(
            "Autor A (do odpięcia)",
            lambda opp, lp: str(opp.autor_could_benefit),
            None,
            None,
        ),
        _ExportColumn(
            "Rodzaj autora A",
            lambda opp, lp: _format_rodzaj_autora(metryka_a(opp)),
            None,
            None,
        ),
        _ExportColumn(
            "ID systemu kadrowego A",
            lambda opp, lp: opp.autor_could_benefit.system_kadrowy_id or "",
            None,
            None,
        ),
        _ExportColumn(
            "Jednostka A",
            lambda opp, lp: _jednostka_nazwa(metryka_a(opp)),
            None,
            None,
        ),
        _ExportColumn(
            "% wykorzystania slotów A",
            lambda opp, lp: float(metryka_a(opp).procent_wykorzystania_slotow) / 100,
            _FMT_PERCENT,
            _ALIGN_RIGHT,
        ),
        _ExportColumn(
            "Sloty nazbierane A",
            lambda opp, lp: float(metryka_a(opp).slot_nazbierany),
            _FMT_DECIMAL,
            _ALIGN_RIGHT,
        ),
        _ExportColumn(
            "Sloty maksymalne A",
            lambda opp, lp: float(metryka_a(opp).slot_maksymalny),
            _FMT_DECIMAL,
            _ALIGN_RIGHT,
        ),
        _ExportColumn(
            "B może wziąć (slotów)",
            lambda opp, lp: float(opp.slots_missing),
            _FMT_DECIMAL,
            _ALIGN_RIGHT,
        ),
        _ExportColumn(
            "Slot w pracy",
            lambda opp, lp: float(opp.slot_in_work),
            _FMT_DECIMAL,
            _ALIGN_RIGHT,
        ),
        _ExportColumn(
            "Różnica punktów A",
            lambda opp, lp: float(opp.punkty_roznica_a),
            _FMT_DECIMAL,
            _ALIGN_RIGHT,
        ),
        _ExportColumn(
            "Różnica slotów A",
            lambda opp, lp: float(opp.sloty_roznica_a),
            _FMT_DECIMAL,
            _ALIGN_RIGHT,
        ),
        _ExportColumn(
            "Różnica punktów B",
            lambda opp, lp: float(opp.punkty_roznica_b),
            _FMT_DECIMAL,
            _ALIGN_RIGHT,
        ),
        _ExportColumn(
            "Różnica slotów B",
            lambda opp, lp: float(opp.sloty_roznica_b),
            _FMT_DECIMAL,
            _ALIGN_RIGHT,
        ),
        _ExportColumn(
            "Autor B (skorzysta)",
            lambda opp, lp: str(opp.autor_currently_using),
            None,
            None,
        ),
        _ExportColumn(
            "Rodzaj autora B",
            lambda opp, lp: _format_rodzaj_autora(metryka_b(opp)),
            None,
            None,
        ),
        _ExportColumn(
            "ID systemu kadrowego B",
            lambda opp, lp: opp.autor_currently_using.system_kadrowy_id or "",
            None,
            None,
        ),
        _ExportColumn(
            "Jednostka B",
            lambda opp, lp: _jednostka_nazwa(metryka_b(opp)),
            None,
            None,
        ),
        _ExportColumn(
            "% wykorzystania slotów B",
            lambda opp, lp: float(metryka_b(opp).procent_wykorzystania_slotow) / 100,
            _FMT_PERCENT,
            _ALIGN_RIGHT,
        ),
        _ExportColumn(
            "Sloty nazbierane B",
            lambda opp, lp: float(metryka_b(opp).slot_nazbierany),
            _FMT_DECIMAL,
            _ALIGN_RIGHT,
        ),
        _ExportColumn(
            "Sloty maksymalne B",
            lambda opp, lp: float(metryka_b(opp).slot_maksymalny),
            _FMT_DECIMAL,
            _ALIGN_RIGHT,
        ),
        _ExportColumn(
            "Sensowne?",
            lambda opp, lp: "TAK" if opp.makes_sense else "NIE",
            None,
            None,
        ),
    ]


def _row_fill(opp, row_idx, sensible_fill, even_fill):
    """Kolor wiersza: zielony dla sensownych, naprzemienny dla reszty."""
    if opp.makes_sense:
        return sensible_fill
    return even_fill if row_idx % 2 == 0 else None


def _build_filter_info(dyscyplina_id, punktacja_zrodla, only_sensible):
    """Złóż listę opisów zastosowanych filtrów do stopki arkusza."""
    from bpp.models import Dyscyplina_Naukowa

    filter_info = []

    if dyscyplina_id:
        try:
            dyscyplina = Dyscyplina_Naukowa.objects.get(pk=dyscyplina_id)
            filter_info.append(f"Dyscyplina: {dyscyplina.nazwa}")
        except Dyscyplina_Naukowa.DoesNotExist:
            pass

    if punktacja_zrodla:
        punktacja_labels = {
            "0-100": "0-100 punktów",
            "100-140": "100-140 punktów",
            "140-200": "140-200 punktów",
            "200+": "200+ punktów",
        }
        filter_info.append(
            "Punktacja źródła: "
            f"{punktacja_labels.get(punktacja_zrodla, punktacja_zrodla)}"
        )

    if only_sensible == "1":
        filter_info.append("Tylko sensowne: TAK")

    return filter_info


def _get_export_opportunities(request, uczelnia):
    """Pobierz i przefiltruj możliwości odpinania zgodnie z parametrami GET."""
    from ..models import UnpinningOpportunity

    opportunities_qs = UnpinningOpportunity.objects.filter(uczelnia=uczelnia)

    dyscyplina_id = _get_dyscyplina_filter(request)
    if dyscyplina_id:
        opportunities_qs = opportunities_qs.filter(dyscyplina_naukowa_id=dyscyplina_id)

    only_sensible = request.GET.get("only_sensible")
    if only_sensible == "1":
        opportunities_qs = opportunities_qs.filter(makes_sense=True)

    opportunities_qs = opportunities_qs.select_related(
        "autor_could_benefit",
        "autor_currently_using",
        "dyscyplina_naukowa",
        "metryka_could_benefit",
        "metryka_currently_using",
        "metryka_could_benefit__autor",
        "metryka_currently_using__autor",
        "metryka_could_benefit__jednostka",
        "metryka_currently_using__jednostka",
    )

    all_opportunities = _cache_rekord_objects(opportunities_qs)

    punktacja_zrodla = request.GET.get("punktacja_zrodla")
    all_opportunities = _filter_by_punktacja(all_opportunities, punktacja_zrodla)

    return all_opportunities, dyscyplina_id, punktacja_zrodla, only_sensible


def _write_export_sheet(ws, opportunities):
    """Zapisz nagłówki i wiersze danych do arkusza; zwróć numer ostatniego wiersza."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    even_row_fill = PatternFill(
        start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
    )
    sensible_row_fill = PatternFill(
        start_color="E7F7E7", end_color="E7F7E7", fill_type="solid"
    )

    columns = _build_export_columns()

    for col, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=column.header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    last_data_row = 1
    for row_idx, opp in enumerate(opportunities, 2):
        last_data_row = row_idx
        row_fill = _row_fill(opp, row_idx, sensible_row_fill, even_row_fill)
        lp = row_idx - 1

        for col, column in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col, value=column.getter(opp, lp))
            cell.border = thin_border
            if column.number_format:
                cell.number_format = column.number_format
            if column.alignment:
                cell.alignment = column.alignment
            if row_fill:
                cell.fill = row_fill

    return last_data_row, len(columns)


def _write_export_footer(ws, last_data_row, n_columns, count, filter_info):
    """Auto-filtr, zamrożenie nagłówka, podsumowanie i opis filtrów."""
    from openpyxl.utils import get_column_letter

    from bpp.util import auto_fit_columns

    if last_data_row > 1:
        last_col_letter = get_column_letter(n_columns)
        ws.auto_filter.ref = f"A1:{last_col_letter}{last_data_row}"

    ws.freeze_panes = ws["A2"]
    auto_fit_columns(ws)

    summary_row = last_data_row + 2 if last_data_row > 1 else 3
    ws.cell(row=summary_row, column=1, value="Podsumowanie:")
    ws.cell(row=summary_row + 1, column=1, value="Liczba wierszy:")
    ws.cell(row=summary_row + 1, column=2, value=count)

    filters_row = summary_row + 3
    ws.cell(row=filters_row, column=1, value="Zastosowane filtry:")
    ws.cell(
        row=filters_row + 1,
        column=1,
        value="; ".join(filter_info) if filter_info else "Brak filtrów",
    )


@login_required
def export_unpinning_opportunities_xlsx(request):
    """Export unpinning opportunities list to XLSX format with filters applied"""
    import datetime

    from django.http import HttpResponse
    from openpyxl import Workbook

    # Pobierz pierwszą uczelnię (zakładamy, że jest tylko jedna)
    uczelnia = Uczelnia.objects.first()
    if not uczelnia:
        messages.error(request, "Nie znaleziono uczelni w systemie.")
        return redirect("ewaluacja_optymalizacja:index")

    (
        opportunities,
        dyscyplina_id,
        punktacja_zrodla,
        only_sensible,
    ) = _get_export_opportunities(request, uczelnia)

    wb = Workbook()
    ws = wb.active
    ws.title = "Możliwości odpinania"

    last_data_row, n_columns = _write_export_sheet(ws, opportunities)

    filter_info = _build_filter_info(dyscyplina_id, punktacja_zrodla, only_sensible)
    _write_export_footer(ws, last_data_row, n_columns, len(opportunities), filter_info)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    filename = (
        "unpinning_opportunities_"
        f"{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
@require_POST
def cancel_unpinning_task(request, task_id):
    """
    Anuluje zadanie analizy odpinania (revoke + forget z Redis).
    """
    from django_bpp.celery_tasks import app

    from ..models import StatusUnpinningAnalyzy

    task = AsyncResult(task_id)

    # Sprawdź czy to jest aktualnie działające zadanie w bazie danych
    status_unpinning = StatusUnpinningAnalyzy.get_or_create()
    if not status_unpinning.w_trakcie or status_unpinning.task_id != task_id:
        messages.error(
            request,
            "Nie możesz anulować tego zadania - nie jest aktualnie uruchomione.",
        )
        return redirect("ewaluacja_optymalizacja:index")

    try:
        # Revoke the task (terminate if running)
        app.control.revoke(task_id, terminate=True)

        # Forget the result from backend (Redis)
        task.forget()

        # Clear from database
        status_unpinning.zakoncz("Zadanie anulowane przez użytkownika")

        logger.info(f"Task {task_id} cancelled by user {request.user}")
        messages.success(request, "Zadanie zostało anulowane.")

    except Exception as e:
        logger.error(f"Failed to cancel task {task_id}: {e}")
        messages.error(request, f"Nie udało się anulować zadania: {e}")

    return redirect("ewaluacja_optymalizacja:index")
