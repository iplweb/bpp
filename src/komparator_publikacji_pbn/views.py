import io

from django.contrib.admin.views.decorators import staff_member_required
from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse
from django.utils.decorators import method_decorator
from django.utils.html import strip_tags
from django.views.generic import ListView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from bpp.models import Wydawnictwo_Ciagle, Wydawnictwo_Zwarte
from import_common.normalization import normalize_isbn, normalize_issn
from pbn_downloader_app.freshness import is_pbn_publications_data_fresh

# Default year range for evaluation period
DEFAULT_YEAR_MIN = 2022
DEFAULT_YEAR_MAX = 2025

# All available comparison fields
COMPARISON_FIELDS = {
    "title": "Tytuł",
    "year": "Rok",
    "doi": "DOI",
    "isbn": "ISBN",
    "issn": "ISSN",
    "url": "URL/WWW",
    "authors": "Liczba autorów",
    "volume": "Tom",
    "pages": "Strony",
    "publisher": "Wydawca",
    "issue": "Numer zeszytu",
    "conference": "Konferencja",
    "keywords": "Słowa kluczowe",
    "abstract": "Streszczenie",
    "language": "Język",
    "publication_place": "Miejsce wydania",
    "series": "Seria wydawnicza",
    "points": "Punkty",
}


DEFAULT_FIELDS = ["title", "year", "doi", "isbn", "issn", "url", "authors"]
SESSION_KEY = "komparator_publikacji_filters"


def _get_default_filters():
    """Return default filter values."""
    return {
        "query": "",
        "publication_type": "all",
        "year_min": str(DEFAULT_YEAR_MIN),
        "year_max": str(DEFAULT_YEAR_MAX),
        "enabled_fields": DEFAULT_FIELDS.copy(),
    }


def _parse_year(value, default):
    """Parse year string to int, returning default on failure."""
    try:
        return int(value) if value else default
    except ValueError:
        return default


@method_decorator(staff_member_required, name="dispatch")
class PublicationComparisonView(ListView):
    """View for comparing BPP publications with PBN API data."""

    template_name = "komparator_publikacji_pbn/comparison_list.html"
    context_object_name = "comparisons"
    paginate_by = 5

    def _get_filters(self):
        """Get filters from GET params or session."""
        has_filter_params = any(
            key in self.request.GET
            for key in ["q", "type", "year_min", "year_max", "fields"]
        )

        if has_filter_params:
            filters = {
                "query": self.request.GET.get("q", ""),
                "publication_type": self.request.GET.get("type", "all"),
                "year_min": self.request.GET.get("year_min", str(DEFAULT_YEAR_MIN)),
                "year_max": self.request.GET.get("year_max", str(DEFAULT_YEAR_MAX)),
                "enabled_fields": self.request.GET.getlist("fields")
                or DEFAULT_FIELDS.copy(),
            }
            self.request.session[SESSION_KEY] = filters
        else:
            filters = self.request.session.get(SESSION_KEY, _get_default_filters())

        return filters

    def _process_publications(self, queryset, pub_type, enabled_fields, comparisons):
        """Process publications and add comparisons with differences."""
        for pub in queryset[:100]:
            comparison = self._compare_publication(pub, pub_type, enabled_fields)
            if comparison["has_differences"]:
                comparisons.append(comparison)

    def get_queryset(self):
        """Get all publications with pbn_uid_id and compare them."""
        comparisons = []
        filters = self._get_filters()

        query = filters["query"]
        publication_type = filters["publication_type"]
        year_min = _parse_year(filters["year_min"], DEFAULT_YEAR_MIN)
        year_max = _parse_year(filters["year_max"], DEFAULT_YEAR_MAX)
        enabled_fields = filters["enabled_fields"]

        if publication_type in ["all", "ciagle"]:
            ciagle_qs = Wydawnictwo_Ciagle.objects.filter(
                pbn_uid_id__isnull=False, rok__gte=year_min, rok__lte=year_max
            ).select_related("pbn_uid", "zrodlo")
            if query:
                ciagle_qs = ciagle_qs.filter(
                    Q(tytul_oryginalny__icontains=query)
                    | Q(pbn_uid__title__icontains=query)
                    | Q(doi__icontains=query)
                )
            self._process_publications(ciagle_qs, "ciagle", enabled_fields, comparisons)

        if publication_type in ["all", "zwarte"]:
            zwarte_qs = Wydawnictwo_Zwarte.objects.filter(
                pbn_uid_id__isnull=False, rok__gte=year_min, rok__lte=year_max
            ).select_related("pbn_uid", "wydawca")
            if query:
                zwarte_qs = zwarte_qs.filter(
                    Q(tytul_oryginalny__icontains=query)
                    | Q(pbn_uid__title__icontains=query)
                    | Q(doi__icontains=query)
                    | Q(isbn__icontains=query)
                )
            self._process_publications(zwarte_qs, "zwarte", enabled_fields, comparisons)

        return comparisons

    def _compare_publication(self, bpp_pub, pub_type, enabled_fields):
        """Compare a BPP publication with its PBN counterpart."""
        pbn_pub = bpp_pub.pbn_uid
        differences = []

        comparers = [
            ("title", self._compare_title),
            ("year", self._compare_year),
            ("doi", self._compare_doi),
            ("isbn", lambda b, p, d: self._compare_isbn(b, p, pub_type, d)),
            ("url", self._compare_url),
            ("authors", self._compare_authors),
            ("volume", lambda b, p, d: self._compare_volume(b, p, pub_type, d)),
            ("pages", lambda b, p, d: self._compare_pages(b, p, pub_type, d)),
            ("publisher", lambda b, p, d: self._compare_publisher(b, p, pub_type, d)),
            ("issn", self._compare_issn),
            ("issue", lambda b, p, d: self._compare_issue(b, p, pub_type, d)),
            ("conference", self._compare_conference),
            ("keywords", self._compare_keywords),
            ("abstract", self._compare_abstract),
            ("language", self._compare_language),
            (
                "publication_place",
                lambda b, p, d: self._compare_publication_place(b, p, pub_type, d),
            ),
            ("series", self._compare_series),
            ("points", self._compare_points),
        ]

        for field_name, comparer in comparers:
            if field_name in enabled_fields:
                comparer(bpp_pub, pbn_pub, differences)

        return {
            "bpp_publication": bpp_pub,
            "pbn_publication": pbn_pub,
            "type": pub_type,
            "differences": differences,
            "has_differences": len(differences) > 0,
            "pbn_id": pbn_pub.mongoId,
            "bpp_id": bpp_pub.pk,
        }

    def _compare_title(self, bpp_pub, pbn_pub, differences):
        bpp_title = bpp_pub.tytul_oryginalny or ""
        pbn_title = pbn_pub.title or pbn_pub.value_or_none("object", "title") or ""
        bpp_title_clean = strip_tags(bpp_title)
        if bpp_title_clean.strip() != pbn_title.strip():
            differences.append(
                {
                    "field": "Tytuł",
                    "bpp_value": bpp_title[:100]
                    + ("..." if len(bpp_title) > 100 else ""),
                    "pbn_value": pbn_title[:100]
                    + ("..." if len(pbn_title) > 100 else ""),
                }
            )

    def _compare_year(self, bpp_pub, pbn_pub, differences):
        bpp_year = bpp_pub.rok
        pbn_year = pbn_pub.year or pbn_pub.value_or_none("object", "year")
        if bpp_year and pbn_year and str(bpp_year) != str(pbn_year):
            differences.append(
                {
                    "field": "Rok",
                    "bpp_value": str(bpp_year),
                    "pbn_value": str(pbn_year),
                }
            )

    def _compare_doi(self, bpp_pub, pbn_pub, differences):
        bpp_doi = bpp_pub.doi or ""
        pbn_doi = pbn_pub.doi or pbn_pub.value_or_none("object", "doi") or ""
        if bpp_doi.strip().lower() != pbn_doi.strip().lower():
            differences.append(
                {"field": "DOI", "bpp_value": bpp_doi, "pbn_value": pbn_doi}
            )

    def _compare_isbn(self, bpp_pub, pbn_pub, pub_type, differences):
        if pub_type != "zwarte" or not hasattr(bpp_pub, "isbn"):
            return
        bpp_isbn_raw = bpp_pub.isbn or ""
        pbn_isbn_raw = pbn_pub.isbn or pbn_pub.value_or_none("object", "isbn") or ""
        bpp_isbn_normalized = normalize_isbn(bpp_isbn_raw) or ""
        pbn_isbn_normalized = normalize_isbn(pbn_isbn_raw) or ""

        if bpp_isbn_normalized == pbn_isbn_normalized:
            return

        if hasattr(bpp_pub, "e_isbn") and bpp_pub.e_isbn:
            bpp_e_isbn_normalized = normalize_isbn(bpp_pub.e_isbn) or ""
            if bpp_e_isbn_normalized == pbn_isbn_normalized:
                return

        differences.append(
            {"field": "ISBN", "bpp_value": bpp_isbn_raw, "pbn_value": pbn_isbn_raw}
        )

    def _compare_url(self, bpp_pub, pbn_pub, differences):
        bpp_www = (
            getattr(bpp_pub, "public_www", None) or getattr(bpp_pub, "www", "") or ""
        )
        pbn_uri = (
            pbn_pub.publicUri or pbn_pub.value_or_none("object", "publicUri") or ""
        )
        if bpp_www.strip() != pbn_uri.strip():
            differences.append(
                {
                    "field": "URL/WWW",
                    "bpp_value": bpp_www[:80] + ("..." if len(bpp_www) > 80 else ""),
                    "pbn_value": pbn_uri[:80] + ("..." if len(pbn_uri) > 80 else ""),
                }
            )

    def _compare_authors(self, bpp_pub, pbn_pub, differences):
        bpp_authors_count = bpp_pub.autorzy_set.count()
        pbn_authors_count = (
            pbn_pub.policz_autorow() if hasattr(pbn_pub, "policz_autorow") else 0
        )
        if bpp_authors_count != pbn_authors_count:
            differences.append(
                {
                    "field": "Liczba autorów",
                    "bpp_value": str(bpp_authors_count),
                    "pbn_value": str(pbn_authors_count),
                }
            )

    def _compare_volume(self, bpp_pub, pbn_pub, pub_type, differences):
        if pub_type != "ciagle":
            return
        bpp_tom = bpp_pub.tom if hasattr(bpp_pub, "tom") else ""
        pbn_volume = pbn_pub.volume() if hasattr(pbn_pub, "volume") else ""
        if bpp_tom and pbn_volume and str(bpp_tom).strip() != str(pbn_volume).strip():
            differences.append(
                {
                    "field": "Tom",
                    "bpp_value": str(bpp_tom),
                    "pbn_value": str(pbn_volume),
                }
            )

    def _compare_pages(self, bpp_pub, pbn_pub, pub_type, differences):
        if pub_type != "ciagle" or not hasattr(bpp_pub, "strony"):
            return
        bpp_pages = bpp_pub.strony or ""
        pbn_pages = pbn_pub.value_or_none("object", "pages") or ""
        if bpp_pages.strip() != pbn_pages.strip():
            differences.append(
                {"field": "Strony", "bpp_value": bpp_pages, "pbn_value": pbn_pages}
            )

    def _compare_publisher(self, bpp_pub, pbn_pub, pub_type, differences):
        if pub_type != "zwarte":
            return
        bpp_publisher = ""
        if hasattr(bpp_pub, "wydawca") and bpp_pub.wydawca:
            bpp_publisher = bpp_pub.wydawca.nazwa
        elif hasattr(bpp_pub, "wydawca_opis"):
            bpp_publisher = bpp_pub.wydawca_opis or ""
        pbn_publisher = pbn_pub.value_or_none("object", "publisher", "name") or ""
        if bpp_publisher.strip() != pbn_publisher.strip():
            differences.append(
                {
                    "field": "Wydawca",
                    "bpp_value": bpp_publisher[:50]
                    + ("..." if len(bpp_publisher) > 50 else ""),
                    "pbn_value": pbn_publisher[:50]
                    + ("..." if len(pbn_publisher) > 50 else ""),
                }
            )

    def _compare_issn(self, bpp_pub, pbn_pub, differences):
        if not hasattr(bpp_pub, "issn"):
            return
        bpp_issn_raw = bpp_pub.issn or ""
        pbn_issn_raw = (
            pbn_pub.value_or_none("object", "issn")
            or pbn_pub.value_or_none("object", "journal", "issn")
            or ""
        )
        bpp_issn_normalized = normalize_issn(bpp_issn_raw) or ""
        pbn_issn_normalized = normalize_issn(pbn_issn_raw) or ""

        if bpp_issn_normalized == pbn_issn_normalized:
            return

        pbn_eissn = pbn_pub.value_or_none("object", "journal", "eissn") or ""
        pbn_eissn_normalized = normalize_issn(pbn_eissn) or ""

        if bpp_issn_normalized == pbn_eissn_normalized:
            return

        if hasattr(bpp_pub, "e_issn") and bpp_pub.e_issn:
            bpp_e_issn_normalized = normalize_issn(bpp_pub.e_issn) or ""
            if (
                bpp_e_issn_normalized == pbn_issn_normalized
                or bpp_e_issn_normalized == pbn_eissn_normalized
            ):
                return

        differences.append(
            {"field": "ISSN", "bpp_value": bpp_issn_raw, "pbn_value": pbn_issn_raw}
        )

    def _compare_issue(self, bpp_pub, pbn_pub, pub_type, differences):
        if pub_type != "ciagle":
            return
        bpp_issue = bpp_pub.nr_zeszytu if hasattr(bpp_pub, "nr_zeszytu") else ""
        bpp_issue = bpp_issue or ""
        pbn_issue = pbn_pub.value_or_none("object", "issue") or ""
        if bpp_issue.strip() != pbn_issue.strip():
            differences.append(
                {
                    "field": "Numer zeszytu",
                    "bpp_value": bpp_issue,
                    "pbn_value": pbn_issue,
                }
            )

    def _compare_conference(self, bpp_pub, pbn_pub, differences):
        if not hasattr(bpp_pub, "konferencja"):
            return
        bpp_conference = bpp_pub.konferencja.nazwa if bpp_pub.konferencja else ""
        pbn_conference = pbn_pub.value_or_none("object", "conferenceName") or ""
        if bpp_conference.strip() != pbn_conference.strip():
            differences.append(
                {
                    "field": "Konferencja",
                    "bpp_value": bpp_conference[:100]
                    + ("..." if len(bpp_conference) > 100 else ""),
                    "pbn_value": pbn_conference[:100]
                    + ("..." if len(pbn_conference) > 100 else ""),
                }
            )

    def _compare_keywords(self, bpp_pub, pbn_pub, differences):
        if not hasattr(bpp_pub, "slowa_kluczowe"):
            return
        bpp_keywords = bpp_pub.slowa_kluczowe or ""
        pbn_keywords = pbn_pub.value_or_none("object", "keywords") or ""
        if isinstance(pbn_keywords, list):
            pbn_keywords = ", ".join(pbn_keywords)
        if bpp_keywords.strip() != pbn_keywords.strip():
            differences.append(
                {
                    "field": "Słowa kluczowe",
                    "bpp_value": bpp_keywords[:200]
                    + ("..." if len(bpp_keywords) > 200 else ""),
                    "pbn_value": pbn_keywords[:200]
                    + ("..." if len(pbn_keywords) > 200 else ""),
                }
            )

    def _compare_abstract(self, bpp_pub, pbn_pub, differences):
        if not hasattr(bpp_pub, "streszczenie"):
            return
        bpp_abstract = bpp_pub.streszczenie or ""
        pbn_abstract = pbn_pub.value_or_none("object", "abstract") or ""
        bpp_has_abstract = len(bpp_abstract.strip()) > 10
        pbn_has_abstract = len(pbn_abstract.strip()) > 10

        if bpp_has_abstract != pbn_has_abstract:
            differences.append(
                {
                    "field": "Streszczenie",
                    "bpp_value": "Tak" if bpp_has_abstract else "Brak",
                    "pbn_value": "Tak" if pbn_has_abstract else "Brak",
                }
            )
        elif bpp_has_abstract and pbn_has_abstract:
            len_diff = abs(len(bpp_abstract) - len(pbn_abstract))
            if len_diff > max(len(bpp_abstract), len(pbn_abstract)) * 0.3:
                differences.append(
                    {
                        "field": "Streszczenie (długość)",
                        "bpp_value": f"{len(bpp_abstract)} znaków",
                        "pbn_value": f"{len(pbn_abstract)} znaków",
                    }
                )

    def _compare_language(self, bpp_pub, pbn_pub, differences):
        if not hasattr(bpp_pub, "jezyk"):
            return
        bpp_language = bpp_pub.jezyk.nazwa if bpp_pub.jezyk else ""
        pbn_language = pbn_pub.value_or_none("object", "language") or ""
        if bpp_language.strip().lower() != pbn_language.strip().lower():
            differences.append(
                {
                    "field": "Język",
                    "bpp_value": bpp_language,
                    "pbn_value": pbn_language,
                }
            )

    def _compare_publication_place(self, bpp_pub, pbn_pub, pub_type, differences):
        import re

        if pub_type != "zwarte":
            return
        bpp_place = ""
        if hasattr(bpp_pub, "miejsce_i_rok"):
            bpp_place = bpp_pub.miejsce_i_rok or ""
            place_match = re.match(r"^([^0-9]+)", bpp_place.strip())
            if place_match:
                bpp_place = place_match.group(1).strip()
        pbn_place = pbn_pub.value_or_none("object", "publicationPlace") or ""
        if bpp_place.strip() != pbn_place.strip():
            differences.append(
                {
                    "field": "Miejsce wydania",
                    "bpp_value": bpp_place,
                    "pbn_value": pbn_place,
                }
            )

    def _compare_series(self, bpp_pub, pbn_pub, differences):
        if not hasattr(bpp_pub, "seria_wydawnicza"):
            return
        bpp_series = bpp_pub.seria_wydawnicza.nazwa if bpp_pub.seria_wydawnicza else ""
        pbn_series = pbn_pub.value_or_none("object", "series") or ""
        if bpp_series.strip() != pbn_series.strip():
            differences.append(
                {
                    "field": "Seria wydawnicza",
                    "bpp_value": bpp_series[:100]
                    + ("..." if len(bpp_series) > 100 else ""),
                    "pbn_value": pbn_series[:100]
                    + ("..." if len(pbn_series) > 100 else ""),
                }
            )

    def _compare_points(self, bpp_pub, pbn_pub, differences):
        if not hasattr(bpp_pub, "punkty_kbn"):
            return
        bpp_points = bpp_pub.punkty_kbn or 0
        pbn_points = pbn_pub.value_or_none("object", "points") or 0
        try:
            bpp_points = float(bpp_points)
            pbn_points = float(pbn_points)
            if abs(bpp_points - pbn_points) > 0.01:
                differences.append(
                    {
                        "field": "Punkty",
                        "bpp_value": str(bpp_points),
                        "pbn_value": str(pbn_points),
                    }
                )
        except (ValueError, TypeError):
            if str(bpp_points) != str(pbn_points):
                differences.append(
                    {
                        "field": "Punkty",
                        "bpp_value": str(bpp_points),
                        "pbn_value": str(pbn_points),
                    }
                )

    def get(self, request, *args, **kwargs):
        """Handle GET request, including XLSX export and reset."""
        if request.GET.get("export") == "xlsx":
            return self.export_xlsx()

        # Handle reset action
        if request.GET.get("reset") == "1":
            if SESSION_KEY in request.session:
                del request.session[SESSION_KEY]
            # Redirect to clean URL
            from django.shortcuts import redirect

            return redirect("komparator_publikacji_pbn:comparison_list")

        return super().get(request, *args, **kwargs)

    def export_xlsx(self):
        """Export comparisons to XLSX file."""
        comparisons = self.get_queryset()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Porównanie publikacji BPP-PBN"

        # Headers
        headers = [
            "Typ",
            "ID BPP",
            "ID PBN",
            "Tytuł BPP",
            "Rok",
            "Pole",
            "Wartość BPP",
            "Wartość PBN",
        ]
        ws.append(headers)

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Add data
        row_num = 2
        for comparison in comparisons:
            for diff in comparison["differences"]:
                ws.append(
                    [
                        (
                            "Wydawnictwo ciągłe"
                            if comparison["type"] == "ciagle"
                            else "Wydawnictwo zwarte"
                        ),
                        comparison["bpp_id"],
                        comparison["pbn_id"],
                        comparison["bpp_publication"].tytul_oryginalny[:100],
                        comparison["bpp_publication"].rok,
                        diff["field"],
                        diff["bpp_value"],
                        diff["pbn_value"],
                    ]
                )
                row_num += 1

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except BaseException:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Create response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            'attachment; filename="komparator_publikacji_bpp_pbn.xlsx"'
        )
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Manual pagination since we're working with a list
        queryset = self.get_queryset()
        paginator = Paginator(queryset, self.paginate_by)
        page_number = self.request.GET.get("page")
        page_obj = paginator.get_page(page_number)

        context["comparisons"] = page_obj.object_list
        context["page_obj"] = page_obj
        context["is_paginated"] = page_obj.has_other_pages()

        # Get filters from session
        filters = self.request.session.get(SESSION_KEY, _get_default_filters())

        context["query"] = filters["query"]
        context["publication_type"] = filters["publication_type"]
        context["year_min"] = filters["year_min"]
        context["year_max"] = filters["year_max"]
        context["enabled_fields"] = filters["enabled_fields"]
        context["all_fields"] = COMPARISON_FIELDS

        # Build export URL with current filters
        export_params = {
            "q": filters["query"],
            "type": filters["publication_type"],
            "year_min": filters["year_min"],
            "year_max": filters["year_max"],
            "export": "xlsx",
        }
        # Add fields to export params
        for field in filters["enabled_fields"]:
            if "fields" not in export_params:
                export_params["fields"] = []
            export_params["fields"].append(field)

        # Build URL string
        export_url_parts = []
        for key, value in export_params.items():
            if key == "fields":
                for field_val in value:
                    export_url_parts.append(f"fields={field_val}")
            else:
                export_url_parts.append(f"{key}={value}")
        context["export_url"] = f"?{'&'.join(export_url_parts)}"

        # Add PBN data freshness check
        pbn_data_fresh, pbn_stale_message, pbn_last_download = (
            is_pbn_publications_data_fresh()
        )
        context["pbn_data_fresh"] = pbn_data_fresh
        context["pbn_stale_message"] = pbn_stale_message
        context["pbn_last_download"] = pbn_last_download

        return context
