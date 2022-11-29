import math
import os
import pprint
import re
import sys
import warnings
from collections import defaultdict

import openpyxl
from dbfread import DBF, FieldParser
from django.db import IntegrityError, transaction
from django.db.models import Count, Q
from django.forms.models import model_to_dict

from import_dbf import models as dbf
from miniblog.models import Article
from .codecs import custom_search_function  # noqa

from django.contrib.contenttypes.models import ContentType

from django.utils import timezone

from bpp import const
from bpp import models as bpp
from bpp.const import CHARAKTER_OGOLNY_KSIAZKA, RODZAJ_PBN_KSIAZKA
from bpp.models import (
    cache,
    parse_informacje,
    parse_informacje_as_dict,
    wez_zakres_stron,
)
from bpp.system import User
from bpp.util import pbar, set_seq

custom_search_function  # noqa


def addslashes(v):
    if not v:
        return v
    if not hasattr(v, "replace"):
        return v
    return v.replace("'", "''")


exp_split_poz_regex = re.compile("(\\#\\d\\d\\d\\$)")


def exp_split_poz_str(s):
    s = s.replace("\r\n", "").replace("\r", "").replace("\n", "")
    spl = exp_split_poz_regex.split(s)[1:]
    for n in range(0, len(spl), 2):
        yield spl[n] + spl[n + 1]


def exp_combine(a, b, sep=", "):
    ret = ""
    if a:
        ret = a

    if b:
        if ret:
            ret += sep
        ret += b

    while ret.find("  ") >= 0:
        ret = ret.replace("  ", " ")
    return ret


class MyFieldParser(FieldParser):
    """Nie ucinaj spacji na końcu pól znakowych"""

    dont_strip = ["tresc", "title", "tytul_or"]

    def parseC(self, field, data):
        """Parse char field and return unicode string"""
        if field.name.lower() not in self.dont_strip:
            data = data.rstrip(b"\0 ")

        return self.decode_text(data)


def dbf2sql(filename, appname="import_dbf"):
    tablename = appname + "_" + os.path.basename(filename.split(".")[0]).lower()
    dbf = DBF(filename, encoding="my_cp1250", parserclass=MyFieldParser)

    output = open(filename + ".sql", "w")
    output.write("BEGIN;\n")
    output.write("DROP TABLE IF EXISTS %s CASCADE;\n" % tablename)
    output.write("CREATE TABLE %s(\n" % tablename)
    for field in dbf.fields:
        output.write("\t%s text,\n" % field.name.lower())
    output.write("\t_ignore_me text);\n")

    for record in dbf:
        output.write(
            "INSERT INTO %s(%s) VALUES(%s);\n"
            % (
                tablename,
                ", ".join([f.lower() for f in record]),
                ", ".join(["'%s'" % addslashes(v or "") for v in record.values()]),
            )
        )

    output.write("COMMIT;\n")
    output.close()


def exp_parse_str(input):
    """
    :param s: tekst z pola tekstowego bazy danych Expertus, np "#102$ #a$ Tytul #b$ #c$ ...
    :return: słownik zawierający pole 'id' oraz kolejno pola 'a', 'b', 'c' itd
    """
    s = input

    assert len(s) >= 5
    if s[0] != "#":
        if not s.strip():  # To moze byc pusty ciąg znaków
            raise ValueError
        raise AssertionError(s)  # Jeżeli jest niepusty i nie zaczyna się od #, błąd

    assert s[4] == "$"

    ret = {}

    ret["id"] = int(s[1:4])

    s = s[5:]  # Nie stripuj, bo Expertus ma spacje na koncu niekiedy: .strip()

    if s[0] == " ":
        s = s[1:]

    if s[0] != "#":
        raise ValueError(input)

    literki = "abcdefghij"
    cnt = 0

    while True:
        try:
            literka = literki[cnt]
        except IndexError:
            break

        sep = f"#{literka}$"
        pos = s.find(sep)
        if pos < 0:
            # Nie ma takiego separatora w tym ciągu znakw
            continue
        assert pos == 0, "Entry string: %r" % input
        s = s[3:]

        while True:
            try:
                nastepna = literki[cnt + 1]
            except IndexError:
                ret[literka] = s.strip()
                break

            next_pos = s.find(f"#{nastepna}$")
            if next_pos < 0:
                # Nie ma następnego separatora, szukaj kolejnego
                cnt += 1
                continue

            ret[literka] = s[:next_pos].strip()
            break

        s = s[next_pos:]
        cnt += 1

    return ret


def exp_add_spacing(s):
    # s = s.replace(".", ". ")
    while s.find("  ") >= 0:
        s = s.replace("  ", " ")
    s = s.replace(". )", ".)")
    s = s.replace(". -", ".-")
    s = s.replace(". ,", ".,")
    return s  # nie stripuj: .strip()


def integruj_uczelnia(nazwa="Domyślna Uczelnia", skrot="DU"):
    """Jeżeli istnieje jakikolwiek obiekt uczelnia w systemie, zwróć go.

    Jeżeli nie ma żadnych obiektw Uczelnia w systemie, utwórz obiekt z
    ustawieniami domyślnymi i zwróć go."""

    if bpp.Uczelnia.objects.exists():
        return bpp.Uczelnia.objects.first()

    uczelnia, created = bpp.Uczelnia.objects.get_or_create(nazwa=nazwa, skrot=skrot)
    if created:
        User.objects.create_superuser(username="admin", password="admin", email=None)
    return uczelnia


@transaction.atomic
def integruj_wydzialy():
    """Utwórz wydziały na podstawie importu DBF"""
    uczelnia = bpp.Uczelnia.objects.first()
    for wydzial in dbf.Wyd.objects.all():
        bpp.Wydzial.objects.get_or_create(
            uczelnia=uczelnia, nazwa=wydzial.nazwa, skrot=wydzial.skrot
        )


@transaction.atomic
def integruj_jednostki():
    uczelnia = bpp.Uczelnia.objects.first()
    for jednostka in dbf.Jed.objects.all():
        # print(f"JEDNOSTKA: [{jednostka.nazwa}] [{jednostka.skrot}]")
        if bpp.Jednostka.objects.filter(
            nazwa=jednostka.nazwa, skrot=jednostka.skrot
        ).exists():
            # Jeżeli istnieje jednostka o dokładnie takiej nazwie, to przejdź dalej
            continue

        widoczna = True
        if jednostka.nazwa == "":
            # Jeżeli nazwa jest pusta, nadaj identyfikator ale i ukryj później
            jednostka.nazwa = "Jednostka %s" % jednostka.idt_jed
            widoczna = False

        while bpp.Jednostka.objects.filter(nazwa=jednostka.nazwa).exists():
            jednostka.nazwa += "*"

        while bpp.Jednostka.objects.filter(skrot=jednostka.skrot).exists():
            jednostka.skrot += "*"

        bpp_jednostka, _ign = bpp.Jednostka.objects.get_or_create(
            pk=jednostka.pk,  # ustaw to samo ID co po stronie DBF
            nazwa=jednostka.nazwa,
            skrot=jednostka.skrot,
            email=jednostka.email,
            www=jednostka.www,
            wydzial=bpp.Wydzial.objects.get(skrot=jednostka.wyd_skrot),
            uczelnia=uczelnia,
            widoczna=widoczna,
        )
        jednostka.bpp_jednostka = bpp_jednostka
        jednostka.save()

    set_seq("bpp_jednostka")


def data_or_null(s):
    if s == "15011961":
        return
    if not s:
        return None
    return f"{s[0:4]}-{s[4:6]}-{s[6:8]}"


def get_dict(model, attname):
    return {getattr(i, attname): i for i in model.objects.all()}


@transaction.atomic
def integruj_tytuly_autorow():
    for (tytul,) in (
        dbf.Aut.objects.all().order_by().values_list("tytul").distinct("tytul")
    ):
        tytul = tytul.strip()
        if not tytul:
            continue

        tytul = exp_add_spacing(tytul)
        try:
            bpp.Tytul.objects.get(skrot=tytul)
        except bpp.Tytul.DoesNotExist:
            bpp.Tytul.objects.create(nazwa=tytul, skrot=tytul)


@transaction.atomic
def integruj_funkcje_autorow():
    for (funkcja,) in (
        dbf.Aut.objects.all()
        .order_by()
        .values_list("stanowisko")
        .distinct("stanowisko")
    ):
        funkcja = funkcja.strip()
        if not funkcja:
            continue

        funkcja = exp_add_spacing(funkcja)
        try:
            bpp.Funkcja_Autora.objects.get(nazwa=funkcja)
        except bpp.Funkcja_Autora.DoesNotExist:
            bpp.Funkcja_Autora.objects.create(nazwa=funkcja, skrot=funkcja)


def exp_autor(base_aut, already_seen=None):
    if already_seen is None:
        already_seen = set()

    already_seen.add(base_aut)

    if base_aut.ref_id is not None and base_aut.ref_id != base_aut.idt_aut:
        if base_aut.ref not in already_seen:
            already_seen.add(base_aut.ref)
            already_seen.update(exp_autor(base_aut.ref, already_seen))

    if base_aut.exp_id_id is not None and base_aut.exp_id_id != base_aut.idt_aut:
        base_aut_exp_id = None
        try:
            base_aut_exp_id = base_aut.exp_id
        except dbf.Aut.DoesNotExist:
            pass
        if base_aut_exp_id is not None and base_aut_exp_id not in already_seen:
            already_seen.add(base_aut.exp_id)
            already_seen.update(exp_autor(base_aut.exp_id, already_seen))

    return already_seen


@transaction.atomic
def integruj_autorow(silent=False):
    tytuly = get_dict(bpp.Tytul, "skrot")
    funkcje = get_dict(bpp.Funkcja_Autora, "skrot")

    base_query = (
        dbf.Aut.objects.filter(bpp_autor_id=None)
        .values_list("idt_aut")
        .order_by("idt_aut")
    )

    if not silent:
        base_query = pbar(base_query)

    for (autor_id,) in base_query:
        a = dbf.Aut.objects.get(pk=autor_id)
        if a.bpp_autor_id is not None:
            continue

        autorzy = list(exp_autor(a))
        autorzy.sort(key=lambda x: x.idt_aut)

        autor = None
        forma_glowna = [autor for autor in autorzy if autor.fg == "*"]
        if forma_glowna and len(forma_glowna) == 1:
            autor = forma_glowna[0]
        else:
            autor = autorzy[0]

        if len(forma_glowna) > 1:
            print("*** Autor ma dwie lub więcej formy główne: %s" % autor)

        autorzy.remove(autor)
        bpp_autor = None

        if autor.bpp_autor_id is not None:
            # ten autor byl juz widziany; sprawdz czy jest ustalony taki sam dla calej grupy
            bpp_autor = autor.bpp_autor

        if bpp_autor is None:
            if autor.pbn_id or autor.orcid_id:
                if autor.pbn_id:
                    try:
                        bpp_autor = bpp.Autor.objects.get(pbn_id=autor.pbn_id)
                    except bpp.Autor.DoesNotExist:
                        pass

                if bpp_autor is None and autor.orcid_id:
                    try:
                        bpp_autor = bpp.Autor.objects.get(orcid=autor.orcid_id)
                    except bpp.Autor.DoesNotExist:
                        pass

            if bpp_autor is None:
                tytul = None
                if autor.tytul:
                    tytul = tytuly.get(autor.tytul)
                bpp_autor = bpp.Autor.objects.create(
                    nazwisko=autor.nazwisk_bz,
                    imiona=autor.imiona_bz,
                    tytul=tytul,
                    pbn_id=autor.pbn_id or None,
                    orcid=autor.orcid_id or None,
                    expertus_id=autor.exp_id_id or None,
                )

            autor.bpp_autor = bpp_autor
            autor.save()

        for elem in autorzy:
            if (elem.bpp_autor_id is None) or (
                elem.bpp_autor_id is not None and elem.bpp_autor_id != bpp_autor.pk
            ):
                elem.bpp_autor = bpp_autor
                elem.save()

        # Jeżeli nie ma tego przypisania do jednostki, to utworz:
        if not autor.idt_jed_id:
            continue

        try:
            jednostka = autor.idt_jed
        except BaseException:
            print(
                f"XXX autor {autor.pk} ma nieistniejace ID jednostki {autor.idt_jed_id}, NIE TWORZE AUTOR_JEDNOSTKA!"
            )
            continue
        try:
            bpp.Autor_Jednostka.objects.get(
                autor=bpp_autor, jednostka=jednostka.bpp_jednostka
            )
        except bpp.Autor_Jednostka.DoesNotExist:
            funkcja_autora = None
            if autor.stanowisko:
                funkcja_autora = funkcje.get(autor.stanowisko)

            bpp.Autor_Jednostka.objects.create(
                autor=bpp_autor,
                jednostka=jednostka.bpp_jednostka,
                funkcja=funkcja_autora,
                rozpoczal_prace=data_or_null(autor.prac_od),
                zakonczyl_prace=data_or_null(autor.dat_zwol),
            )


def sprawdz_zamapowanie_autorow():
    assert dbf.Aut.objects.filter(bpp_autor=None).count() == 0


def integruj_charaktery():
    for rec in dbf.Pub.objects.all():
        try:
            charakter = bpp.Charakter_Formalny.objects.get(skrot=rec.skrot)
        except bpp.Charakter_Formalny.DoesNotExist:
            charakter = bpp.Charakter_Formalny.objects.create(
                nazwa=rec.nazwa,
                skrot=rec.skrot,
            )
        rec.bpp_id = charakter
        rec.save()


def wzbogacaj_charaktery(fp):
    for elem in xls2dict(fp):
        ch = bpp.Charakter_Formalny.objects.get(skrot=elem["skrot"])
        save = False

        if elem["charakter_pbn"]:
            ch.charakter_pbn = bpp.Charakter_PBN.objects.get(opis=elem["charakter_pbn"])
            save = True

        if elem["rodzaj_dla_pbn"]:
            if elem["rodzaj_dla_pbn"] == "Książka":
                ch.rodzaj_pbn = const.RODZAJ_PBN_KSIAZKA
            elif elem["rodzaj_dla_pbn"] == "Artykuł":
                ch.rodzaj_pbn = const.RODZAJ_PBN_ARTYKUL
            elif elem["rodzaj_dla_pbn"] == "Rozdział":
                ch.rodzaj_pbn = const.RODZAJ_PBN_ROZDZIAL
            else:
                raise ValueError("Niezdefiniowany rodzaj: %r" % elem["rodzaj_dla_pbn"])
            save = True

        if elem["charakter_dla_slotów"]:
            if elem["charakter_dla_slotów"] == "Książka":
                ch.charakter_sloty = const.CHARAKTER_SLOTY_KSIAZKA
            elif elem["charakter_dla_slotów"] == "Rozdział":
                ch.charakter_sloty = const.CHARAKTER_SLOTY_ROZDZIAL
            else:
                raise ValueError(
                    "Niezdefiniowany rozdaj: %r" % elem["charakter_dla_slotow"]
                )
            save = True

        if save:
            ch.save()


@transaction.atomic
def integruj_kbn():
    for rec in dbf.Kbn.objects.all():
        try:
            kbn = bpp.Typ_KBN.objects.get(skrot=rec.skrot)
        except bpp.Typ_KBN.DoesNotExist:
            kbn = bpp.Typ_KBN.objects.create(nazwa=rec.nazwa, skrot=rec.skrot)
        rec.bpp_id = kbn
        rec.save()


@transaction.atomic
def integruj_jezyki():
    for rec in dbf.Jez.objects.all():
        try:
            jez = bpp.Jezyk.objects.get(Q(nazwa=rec.nazwa) | Q(skrot=rec.skrot))
        except bpp.Jezyk.DoesNotExist:
            jez = bpp.Jezyk.objects.create(nazwa=rec.nazwa, skrot=rec.skrot)
        rec.bpp_id = jez
        rec.save()


@transaction.atomic
def integruj_zrodla():
    # pole usm_f:
    # - 988 -> numer ISBN,
    # - 985 -> jakiś identyfikator (dwa rekordy),
    # - 969 -> kilka rekordow nazwanych od okreslen openaccess
    # - 154 -> monografia
    # - 107 -> 10 rekordw (ISSN, numerki, nazwa pracy(?))
    # - 100 -> źrdło indeksowane

    rodzaj = bpp.Rodzaj_Zrodla.objects.get(nazwa="periodyk")

    for rec in dbf.Usi.objects.filter(usm_f__in=["100", "102", "206"]):
        rec.nazwa = exp_add_spacing(rec.nazwa)

        try:
            zrodlo = bpp.Zrodlo.objects.get(nazwa=rec.nazwa)
        except bpp.Zrodlo.DoesNotExist:
            zrodlo = bpp.Zrodlo.objects.create(
                nazwa=rec.nazwa, skrot=rec.skrot or rec.nazwa, rodzaj=rodzaj
            )
        rec.bpp_id = zrodlo
        rec.save()

    for rec in dbf.Usi.objects.filter(Q(usm_f="152") | Q(usm_f="153", usm_sf="b")):
        rec.nazwa = exp_add_spacing(rec.nazwa)
        try:
            wydawca = bpp.Wydawca.objects.get(nazwa=rec.nazwa)
        except bpp.Wydawca.DoesNotExist:
            wydawca = bpp.Wydawca.objects.create(
                nazwa=rec.nazwa,
            )
        rec.bpp_wydawca_id = wydawca
        rec.save()

    for rec in dbf.Usi.objects.filter(usm_f__in=["154", "206"]):
        rec.nazwa = exp_add_spacing(rec.nazwa)
        try:
            seria = bpp.Seria_Wydawnicza.objects.get(nazwa=rec.nazwa)
        except bpp.Seria_Wydawnicza.DoesNotExist:
            seria = bpp.Seria_Wydawnicza.objects.create(
                nazwa=rec.nazwa,
            )
        rec.bpp_seria_wydawnicza_id = seria
        rec.save()


@transaction.atomic
def mapuj_elementy_publikacji(offset, limit):
    # Zbiera elementy 'zrodlo', 'poz_', 'b_u' do jednej tabeli

    ids_list = (
        dbf.Bib.objects.all()
        .exclude(analyzed=True)
        .order_by("pk")
        .values_list("pk")[offset:limit]
    )

    base_query = dbf.Bib.objects.filter(pk__in=ids_list).select_for_update()

    iter = base_query
    if offset == 0:
        iter = pbar(base_query, base_query.count())

    for rec in iter:

        poz_a = dbf.Poz.objects.get_for_model(rec.idt, "A")
        if poz_a:
            poz_a_pos = rec.tytul_or.find(poz_a)
            append_poz_a = True
            if poz_a_pos > 5 and len(poz_a) > 1:
                if rec.idt in [78911]:
                    append_poz_a = False
                else:
                    raise ValueError(
                        "rec.tytul_or zawiera juz poz_a?", rec.tytul_or, poz_a, rec.idt
                    )

            if append_poz_a:
                rec.tytul_or += poz_a

        # Jeżeli pole 'zrodlo' jest dosc dlugie, reszta tego pola moze znalezc
        # sie w tabeli "Poz" z oznaczeniem literowym 'C'
        zrodlo_cd = dbf.Poz.objects.get_for_model(rec.idt, "C")
        if zrodlo_cd:
            poz_c_pos = rec.zrodlo.find(zrodlo_cd)
            if poz_c_pos > 4 and len(zrodlo_cd) > 1:
                raise ValueError(
                    "rec.zrodlo zawiera juz poz_c?", rec.zrodlo, zrodlo_cd, rec.idt
                )
            rec.zrodlo += zrodlo_cd

        # Dodaj elementy z B_U do POZ_G tak, aby były na koncu czyli
        # były zaimportowane jako najaktualniejsze
        # bpp=# select distinct(substr(comm, 1, 3)) from import_dbf_b_u;
        #  substr
        # --------
        #  985 -> jeden rekord z numerkiem 28940458 -> pubmed ID?
        #  988 -> ISBN, niekiedy punktacja, niekiedy DWA ISBN
        #  202 -> b#Polskie Towrazystwo Walki z Kalectwem; Akad. Med; Wydaw. Continuo etc
        #  154 -> "Rozprawy Habilitacyjne", "Zdrowe Życie",
        #  100 -> Źródło indeksowane
        #  203 -> Onkologia Kliniczna, Biblioteka Polskiego Przeglądu Chir
        #  107 -> ISSN
        #  969 -> OpenAccess
        #  152 -> Wydawnictwo (Volumed, GEOPOL, Akad. Med)
        # (9 rows)

        for source, data in [
            ("poz_g", dbf.Poz.objects.get_for_model(rec.idt, "G")),
            ("poz_n", dbf.Poz.objects.get_for_model(rec.idt, "N")),
        ]:
            for element in exp_split_poz_str(data):
                parsed = exp_parse_str(element)
                id = parsed["id"]
                del parsed["id"]

                dbf.Bib_Desc.objects.create(
                    idt=rec, elem_id=id, value=parsed, source=source
                )

        bu = defaultdict(dict)
        for elem in rec.b_u_set.all():
            id, literka, wartosc = elem.comm.split("#")
            id = int(id)
            bu[id]["id"] = int(id)
            bu[id][literka] = "|" + "%.10i" % elem.idt_usi_id
            bu[id]["source"] = "b_u"

        if bu:
            for id, value in bu.items():
                # id = element['id']
                del value["id"]
                dbf.Bib_Desc.objects.create(
                    idt=rec, elem_id=id, value=value, source="b_u"
                )

        if rec.zrodlo.strip():
            zrodlo = exp_parse_str(rec.zrodlo.strip())
            id = zrodlo["id"]
            del zrodlo["id"]
            dbf.Bib_Desc.objects.create(idt=rec, elem_id=id, value=zrodlo, source="b_u")

        rec.analyzed = True
        rec.save()


def to_pubmed_id(ident):
    if not ident.strip():
        return

    try:
        return int(ident)
    except (TypeError, ValueError):
        return


@transaction.atomic
def integruj_publikacje(offset=None, limit=None):
    if cache.enabled():
        cache.disable()

    # zagadnienie nr 1 BPP: prace mają rok i rok_punkt inny: select rok, rok_punkt from
    # import_dbf_bib where rok_punkt != rok and rok_punkt != '' order by rok;

    # zagadnienie nr 2: kody w polach tekstowych:
    #
    # postgres=# select distinct substr(tytul_or, 0, 6) from import_dbf_bib;
    #  substr
    # --------
    #  #102$
    #  #150$
    #  #204$

    # bpp=# select distinct substr(title, 1, 4) from import_dbf_bib;
    #  substr
    # --------
    #  #105
    #  #157
    #  #207

    # postgres=# select distinct substr(zrodlo, 0, 6) from import_dbf_bib;
    #  substr
    # --------
    #  #152$ -> (a: miejsce, b: numer id lub brak, c: rok)
    #  #100$ -> (a: numer id -> idt_usi -> Usi, B_U)
    #  #200$ -> (a: tytuł, b: podtytuł, c: pod red)

    statusy_korekt = {a.nazwa: a for a in bpp.Status_Korekty.objects.all()}
    mapping_korekt = {
        "!": "przed korektą",
        "*": "w trakcie korekty",
        "ű": "po korekcie",
    }

    id_list = (
        dbf.Bib.objects.filter(object_id=None, analyzed=True)
        .order_by("pk")[offset:limit]
        .values_list("pk")
    )

    base_query = dbf.Bib.objects.filter(pk__in=id_list)

    base_query = base_query.select_for_update(nowait=True, of=("self",))

    base_query = base_query.select_related()

    typy_kbn = {x.skrot: x.bpp_id for x in dbf.Kbn.objects.all()}
    typ_kbn_pusty = bpp.Typ_KBN.objects.get(skrot="000")

    try:
        jezyk_pusty = bpp.Jezyk.objects.get(skrot="000")
    except bpp.Jezyk.DoesNotExist:
        jezyk_pusty = bpp.Jezyk.objects.get(skrot="in.")
    jezyki = {x.skrot: x.bpp_id for x in dbf.Jez.objects.all()}

    charaktery_formalne = {x.skrot: x.bpp_id for x in dbf.Pub.objects.all()}

    iter = base_query

    # pbar wyłączony kompletnie, bo ostrzeżenia na konsoli są w zamian
    # i ma wszystko pójść do logu

    # if offset == 0 or offset is None:
    #    iter = pbar(base_query, base_query.count())

    for rec in iter:
        delayed_creation = []
        wos = False
        kw = {}

        tytul = exp_parse_str(rec.tytul_or)
        if rec.title.strip():
            try:
                title = exp_parse_str(rec.title)
            except ValueError:
                continue

            kw["tytul"] = exp_combine(title.get("a"), title.get("b"), sep=": ")
            for literka in "cdefgh":
                if literka in title.keys():
                    raise NotImplementedError("co mam z tym zrobic %r" % title)

        try:
            kw["charakter_formalny"] = charaktery_formalne[rec.charakter]
        except KeyError:
            kw["charakter_formalny"] = bpp.Charakter_Formalny.objects.get(skrot="000")

        try:
            kw["typ_kbn"] = typy_kbn[rec.kbn]
        except KeyError:
            kw["typ_kbn"] = typ_kbn_pusty

        if not rec.jezyk:
            jez = jezyk_pusty
        else:
            jez = jezyki[rec.jezyk]

        kw["jezyk"] = jez

        if rec.jezyk2:
            try:
                kw["jezyk_alt"] = jezyki[rec.jezyk2]
            except KeyError:
                assert not kw.get("adnotacje")
                kw["adnotacje"] = "Praca deklaruje jezyk2 jako " + rec.jezyk2

        kw["rok"] = rec.rok
        kw["recenzowana"] = rec.recenzowan.strip() == "1"

        # afiliowana -- nie ma takiego pola
        # kw['afiliowana'] = rec.afiliowana.strip() == '1'

        # kbr -- co to za pole?
        # bpp=# select distinct kbr, count(*) from import_dbf_bib group by kbr;
        #  kbr | count
        # -----+-------
        #      | 86354
        #  000 |    41
        #  OA  |    21
        # (3 rows)

        # lf -- co to za pole?
        # bpp=# select distinct lf, count(*) from import_dbf_bib group by lf;
        #  lf | count
        # ----+-------
        #     |  9605
        #  0  | 66617
        #  1  | 10194
        # (3 rows)

        # study_gr -- co to za pole?
        # bpp=# select distinct study_gr, count(*) from import_dbf_bib group by study_gr;
        #  study_gr | count
        # ----------+-------
        #  0        | 86026
        #  1        |   390
        # (2 rows)

        # idt2 - 5 rekordw

        # pun_max - 1 rekord

        # kwartyl -- co to za pole?
        # bpp=# select distinct kwartyl, count(*) from import_dbf_bib group by kwartyl;
        #  kwartyl | count
        # ---------+-------
        #          | 85236
        #  A       |    37
        #  A1      |   831
        #  B       |    13
        #  B1      |   296
        #  CC      |     3
        # (6 rows)

        # lis_numer -- co to za pole?
        # bpp=# select distinct lis_numer, count(*) from import_dbf_bib group by lis_numer;
        #  lis_numer | count
        # -----------+-------
        #            | 86365
        #  10226     |     1
        #  116       |     8
        #  118       |     1
        #  1436      |     2
        #  157       |     1
        #  162       |     1
        #  165       |     1
        #  167       |     4
        #  182       |     1
        #  205       |     1
        #  22        |     8
        #  2700      |     1
        #  3180      |     1
        #  3903      |     1
        #  4930      |     2
        #  5206      |     1
        #  54        |     3
        #  6728      |     1
        #  75        |     3
        #  756       |     1
        #  7742      |     1
        #  8         |     1
        #  888       |     2
        #  89        |     3
        #  972       |     1
        # (26 rows)

        kw["status_korekty"] = statusy_korekt[mapping_korekt[rec.status]]

        if rec.punkty_kbn:
            kw["punkty_kbn"] = rec.punkty_kbn

        if rec.impact:
            kw["impact_factor"] = rec.impact

        if rec.link:
            if rec.link.startswith("dx.doi.org/"):
                kw["doi"] = rec.link[len("dx.doi.org/") :].strip()
            else:
                kw["www"] = rec.link

        if rec.uwagi2:
            kw["adnotacje"] = rec.uwagi2

        if rec.pun_wl:
            kw["punktacja_wewnetrzna"] = rec.pun_wl

        if rec.issn:
            kw["issn"] = rec.issn

        if rec.eissn:
            kw["e_issn"] = rec.eissn

        klass = None

        if tytul["id"] == 204:
            klass = bpp.Wydawnictwo_Zwarte
            klass_ext = bpp.Wydawnictwo_Zwarte_Zewnetrzna_Baza_Danych
            kw["tytul_oryginalny"] = exp_combine(tytul["a"], tytul.get("b"), sep=": ")

            kw["strony"] = exp_add_spacing(tytul["e"])
            if tytul.get("e"):
                kw["szczegoly"] = exp_combine(kw.get("szczegoly"), tytul.get("e"))

            if tytul.get("f"):
                kw["szczegoly"] = exp_combine(kw.get("szczegoly"), tytul["f"])

            if tytul.get("c"):
                kw["tytul_oryginalny"] = exp_combine(
                    kw["tytul_oryginalny"], tytul.get("c")
                )

            if tytul.get("d"):
                kw["tytul_oryginalny"] = exp_combine(
                    kw["tytul_oryginalny"], tytul.get("d")
                )

            if kw.get("informacje") and kw.get("strony") is None:
                kw["strony"] = wez_zakres_stron(kw["informacje"])

        elif tytul["id"] == 100:
            klass = bpp.Wydawnictwo_Ciagle
            klass_ext = bpp.Wydawnictwo_Ciagle_Zewnetrzna_Baza_Danych
            kw["tytul_oryginalny"] = tytul["a"]

            for literka in "bc":
                if tytul.get(literka):
                    kw["tytul_oryginalny"] = exp_combine(
                        kw["tytul_oryginalny"], tytul.get(literka)
                    )

            for literka in "de":
                if tytul.get(literka):
                    warnings.warn(
                        f"Co mam zrobic z literka {literka} dla rekordu {rec} tytul {tytul}"
                    )

        elif tytul["id"] in [200, 250]:
            klass = bpp.Wydawnictwo_Zwarte
            klass_ext = bpp.Wydawnictwo_Zwarte_Zewnetrzna_Baza_Danych
            kw["tytul_oryginalny"] = tytul["a"]

            for literka in "bc":
                if tytul.get(literka):
                    kw["tytul_oryginalny"] = exp_combine(
                        kw["tytul_oryginalny"], tytul.get(literka)
                    )
            for literka in "de":
                if tytul.get(literka):
                    warnings.warn(
                        f"Niezaimportowane dane: {tytul.get(literka)} tytul {tytul} rekord {rec} o ID {rec.idt}"
                    )

        elif tytul["id"] == 102:
            klass = bpp.Wydawnictwo_Ciagle
            klass_ext = bpp.Wydawnictwo_Ciagle_Zewnetrzna_Baza_Danych
            kw["tytul_oryginalny"] = tytul["a"]

            if tytul.get("b"):
                kw["tytul_oryginalny"] += ": " + tytul.get("b")

            if tytul.get("c"):
                assert not kw.get("tekst_po_ostatnim_autorze"), (kw, tytul, rec)
                kw["tekst_po_ostatnim_autorze"] = tytul.get("c")

            if (
                tytul.get("d")
                and tytul["d"] != "`"
                and (
                    tytul["a"].find(tytul["d"].split("#b$ #c$ #d$")[0].strip()) < 0
                    and tytul["c"].find(tytul["d"].split("#d$")[0].strip()) < 0
                )
            ):
                assert not kw.get("tytul"), (kw, tytul, rec)
                kw["tytul"] = tytul.get("d")

            for literka in "ef":
                assert not tytul.get(literka), "co mam robic z %r" % tytul

        elif tytul["id"] == 150:
            klass = bpp.Wydawnictwo_Zwarte
            klass_ext = bpp.Wydawnictwo_Zwarte_Zewnetrzna_Baza_Danych
            kw["tytul_oryginalny"] = exp_combine(
                tytul["a"], exp_combine(tytul.get("b"), tytul.get("d")), sep=" "
            )
            if tytul.get("c"):
                print(
                    "*** Pole 150C do adnotacji: ", tytul.get("c"), "dla rekordu", rec
                )
                kw["adnotacje"] = exp_combine(kw.get("adnotacje"), tytul.get("c"))

        else:
            raise NotImplementedError(tytul)

        for bdesc in rec.bib_desc_set.all():

            elem = bdesc.value
            elem["id"] = bdesc.elem_id

            if elem["id"] == 202:
                wydawca_id = None
                try:
                    wydawca_id = int(elem["b"][1:])
                except (TypeError, ValueError, IndexError):
                    pass
                if wydawca_id is not None:
                    kw["wydawca"] = dbf.Usi.objects.get(
                        idt_usi=wydawca_id
                    ).bpp_wydawca_id

                if elem.get("a"):
                    kw["miejsce_i_rok"] = f"{elem['a']} {elem['c']}"

            elif elem["id"] in [203, 154]:
                # Seria wydawnicza

                try:
                    seria_wydawnicza_id = int(elem["a"][1:])
                except ValueError:
                    warnings.warn(
                        f"Nie można skonwertowac numeru serii wydawniczej: {elem} "
                        f"dla rekordu {rec} dodaje do pola 'uwagi'"
                    )
                    kw["uwagi"] = exp_combine(kw.get("uwagi"), elem.get("a"))
                    continue

                kw["seria_wydawnicza"] = dbf.Usi.objects.get(
                    idt_usi=seria_wydawnicza_id
                ).bpp_seria_wydawnicza_id

                if elem.get("d") and elem["d"].startswith("ISSN"):
                    kw["issn"] = elem["d"].split("ISSN")[1].strip()
                    del elem["d"]

                # Pola C lub D idą za chwilę do numeru serii:
                # if elem.get("c"):
                #     kw["informacje"] = exp_combine(kw.get("informacje"), elem.get("c"))
                #
                # if elem.get("d"):
                #     kw["informacje"] = exp_combine(kw.get("informacje"), elem.get("d"))

                kw["numer_w_serii"] = ""

                for literka in "bcd":
                    if elem.get(literka):
                        kw["numer_w_serii"] = exp_combine(
                            kw["numer_w_serii"], elem.get(literka)
                        )

                if elem.get("e", "").find("ISSN") > -1:
                    kw["issn"] = elem["e"][4:].strip()
                    del elem["e"]

                for literka in "ef":
                    assert not elem.get(literka), (elem, rec)

            elif elem["id"] == 201:
                # odpowiedzialni za wydanie dla rozdziału
                # 201 'a' informacje o wydaniu, 'b' odpoweidzialni za wydanie
                assert not kw.get("oznaczenie_wydania")
                kw["oznaczenie_wydania"] = elem.get("a")

                # elem.get("b") ignorujemy
                # https://mantis.iplweb.pl/view.php?id=843
                # Obecnie testuję wersję 202005.37 w której planuję wprowadzić poprawkę na ten problem; pojawia się
                # ]problem z podwójnymi autorami jak w załączniku, pozwolę sobie wyrzucić import pola "B" z w/wym
                # elementów
                # EKG : 150 przypadków.Wyd.3 pol. red. nauk. Jacek Smereka. [AUT.] JOHN HAMPTON, DAVID ADLAM, JOANNA
                # HAMPTON, [RED.] JACEK SMEREKA. XII, 308 s.ryc, ISBN 978-83-66548-04-6. Wrocław 2020, Edra Urban &
                # Partner, 978-83-66548-04-6.

                for literka in "cde":
                    assert not elem.get(literka), (elem, rec, rec.idt)

            elif elem["id"] == 206:

                if elem.get("a").startswith("|"):

                    snazwa = dbf.Usi.objects.get(idt_usi=elem.get("a")[1:]).nazwa
                    seria_wydawnicza = bpp.Seria_Wydawnicza.objects.get_or_create(
                        nazwa=snazwa,
                    )[0]

                    if (
                        kw.get("seria_wydawnicza")
                        and kw.get("seria_wydawnicza") != seria_wydawnicza
                    ):
                        raise NotImplementedError
                    kw["seria_wydawnicza"] = seria_wydawnicza
                else:
                    # wydanie
                    kw["uwagi"] = exp_combine(
                        kw.get("uwagi", ""), elem.get("a"), sep=" "
                    )  # Wyd 1 pol

                kw["uwagi"] = exp_combine(
                    kw.get("uwagi", ""), elem.get("b"), sep=" "
                )  # pod red A. Kowalski

                kw["uwagi"] = exp_combine(
                    kw.get("uwagi", ""), elem.get("c"), sep=" "
                )  # Tom 20

                for literka in "de":
                    if elem.get(literka):
                        warnings.warn(
                            f"Nieobslugiwane literki przy polu 206 {elem} rekord {rec} o ID {rec.idt}"
                        )

            elif elem["id"] in [205, 255, 153]:
                if elem.get("a") and not elem.get("a").startswith("|"):
                    kw["uwagi"] = exp_combine(kw.get("uwagi"), elem.get("a"), sep=". ")
                else:
                    if elem.get("a") and elem.get("c"):
                        miejsce = dbf.Usi.objects.get(idt_usi=elem["a"][1:]).nazwa
                        rok = elem.get("c")
                        kw["miejsce_i_rok"] = f"{miejsce} {rok}"

                    if elem.get("b"):
                        wydawca = dbf.Usi.objects.get(idt_usi=elem["b"][1:]).nazwa
                        bpp_wydawca = bpp.Wydawca.objects.get_or_create(nazwa=wydawca)[
                            0
                        ]
                        kw["wydawca"] = bpp_wydawca

                for literka in "de":
                    assert not elem.get(literka), elem

            elif elem["id"] == 101:
                # 101A to informacje, 101D i 101E to szczegoly

                # A: rok,
                # B: tom
                # C: numer
                # D: strony
                # E: bibliogr. poz

                kw["informacje"] = elem.get("a")

                # A: moze byc to 'rok tom' lub 'rok numer' lub 'rok numer tom'
                pi = parse_informacje_as_dict(elem.get("a"))
                if pi.get("rok") and (pi.get("numer") or pi.get("tom")):
                    assert not kw.get("tom")
                    assert not kw.get("nr_zeszytu")

                    kw["tom"] = pi.get("tom")
                    kw["nr_zeszytu"] = pi.get("numer")

                kw["szczegoly"] = exp_combine(kw.get("szczegoly"), elem.get("b"))

                if elem.get("b"):
                    if elem.get("b").startswith("nr "):
                        assert not kw.get("nr_zeszytu")
                        kw["nr_zeszytu"] = elem.get("b").replace("nr ", "")
                    else:
                        assert not kw.get("tom")
                        kw["tom"] = elem.get("b")

                kw["szczegoly"] = exp_combine(kw["szczegoly"], elem.get("c"))

                if elem.get("d"):
                    kw["strony"] = exp_combine(kw.get("strony"), elem.get("d"))
                    kw["szczegoly"] = exp_combine(kw["szczegoly"], elem.get("d"))

                if elem.get("e"):
                    kw["szczegoly"] = exp_combine(kw.get("szczegoly"), elem.get("e"))

            elif elem["id"] == 103:
                # Uwagi (nie: konferencja)
                # Zgłoszenie #794 Przy imporcie wszystkie uwagi w opisach artykułów z czasopism
                # (pole 103) stały informacjami o konferencjach, a spora część dotyczy czegoś
                # innego. Pole to należy przenieść przy imporcie do pola "szczegoly"
                # (lub innego odpowiadającego uwadze).
                if kw.get("uwagi") is None:
                    kw["uwagi"] = elem["a"]
                else:
                    kw["uwagi"] += elem["a"]

            # Dla UMW to miało sens
            # elif elem["id"] == 153:
            #     # assert not kw.get("informacje")
            #     if elem.get("a"):
            #         kw["szczegoly"] = exp_combine(kw.get("szczegoly"), elem["a"])
            #         if kw.get("strony"):
            #             warnings.warn(
            #                 f"Pole strony juz ma wartosc {kw.get('strony')}, nadpisuje! Rekord {rec} o
            #                 ID {rec.idt}; element {elem}"
            #             )
            #         kw["strony"] = elem["a"]
            #     else:
            #         warnings.warn(
            #             f"Pole nie do konca obslugiwane: {elem} rekoer {rec} o id {rec.idt}"
            #         )
            #     kw["szczegoly"] = exp_combine(kw.get("szczegoly"), elem.get("b"))
            #     kw["szczegoly"] = exp_combine(kw.get("szczegoly"), elem.get("c"))

            elif elem["id"] == 104:
                # assert not kw.get("uwagi"), (kw["uwagi"], elem, rec, rec.idt)
                if kw.get("uwagi") is None:
                    kw["uwagi"] = elem["a"]
                else:
                    print(
                        "Łączę pola UWAGI: %s | %s | %s "
                        % (rec.idt, kw.get("uwagi"), elem["a"])
                    )
                    kw["uwagi"] = exp_combine(kw.get("uwagi"), elem["a"], sep=". ")
                    kw["adnotacje"] = exp_combine(
                        kw.get("adnotacje"), "połączone pola uwag", sep=". "
                    )

                for literka in "bcd":
                    assert not elem.get(literka), (elem, rec, rec.idt)

            elif elem["id"] == 151:
                assert not kw.get("oznaczenie_wydania")
                kw["oznaczenie_wydania"] = elem.get("a")

                # elem.get("b") ignorujemy
                # https://mantis.iplweb.pl/view.php?id=843
                # Obecnie testuję wersję 202005.37 w której planuję wprowadzić poprawkę na ten problem; pojawia się
                # ]problem z podwójnymi autorami jak w załączniku, pozwolę sobie wyrzucić import pola "B" z w/wym
                # elementów
                # EKG : 150 przypadków.Wyd.3 pol. red. nauk. Jacek Smereka. [AUT.] JOHN HAMPTON, DAVID ADLAM, JOANNA
                # HAMPTON, [RED.] JACEK SMEREKA. XII, 308 s.ryc, ISBN 978-83-66548-04-6. Wrocław 2020, Edra Urban &
                # Partner, 978-83-66548-04-6.

                for literka in "cde":
                    assert not elem.get(literka), (elem, rec, rec.idt)

            elif elem["id"] in [155, 156]:
                # 155 "Komunikat tegoż w ... / 156 "toż w wersji polskiej"
                kw["uwagi"] = exp_combine(kw.get("uwagi"), elem["a"], sep=". ")
                for literka in "bcde":
                    if elem.get(literka):
                        warnings.warn(
                            f"Nieobslugiwana literka {literka} dla elementu {elem} rekord {rec} ID {rec.idt}"
                        )

            elif elem["id"] == 995:
                if kw.get("www"):
                    if "http://" + kw["www"] == elem["a"] or kw["www"] == elem["a"]:
                        del kw["www"]
                    else:
                        kw["adnotacje"] = exp_combine(
                            kw.get("adnotacje", ""),
                            "Drugi adres WWW? " + kw["www"],
                            sep="\n",
                        )
                        del kw["www"]

                assert not kw.get("www"), (elem, rec, kw)
                kw["www"] = elem["a"]
                for literka in "bcd":
                    assert not elem.get(literka), (elem, rec, rec.idt)

            elif elem["id"] == 991:
                # DOI
                if kw.get("doi"):
                    if kw["doi"] != elem["a"]:
                        kw["adnotacje"] = exp_combine(
                            kw.get("adnotacje"),
                            "Drugi nr DOI: %s" % elem["a"],
                            sep=". ",
                        )
                else:
                    kw["doi"] = elem["a"]

            elif elem["id"] == 969:
                # ({'id': 969, 'a': '|0000005187', 'b': '|0000005493',
                # 'c': '|0000005494', 'd': '', 'e': '|0000005496'},
                # <Bib: Mediators of pruritus in psoriasis>, 38163)
                # open-access-text-version: FINAL_PUBLISHED
                # open-access-licence: CC BY
                # open-access-release-time: AT_PUBLICATION
                # open-access-article-mode: OPEN_JOURNAL

                if klass == bpp.Wydawnictwo_Ciagle:
                    oa_klass = bpp.Tryb_OpenAccess_Wydawnictwo_Ciagle
                elif klass == bpp.Wydawnictwo_Zwarte:
                    oa_klass = bpp.Tryb_OpenAccess_Wydawnictwo_Zwarte
                else:
                    raise NotImplementedError(klass)

                del elem["id"]

                if elem == {
                    "a": "|0000005187",
                    "b": "|0000005188",
                    "c": "",
                    "d": "|0000005189",
                    "e": "",
                }:
                    # Rekordy z takim schematem informacji pojawiaja sie co jakis czas
                    # i nie bardzo widze ich powiazanie z informacja wyswietlana na stornie
                    # Najprawdopodobniej te informacje OpenAccess beda zaimportowane w sposob
                    # nieprawidlowy. Na ten moment (31.10.2019) zostaje jak-jest:
                    kw[
                        "openaccess_wersja_tekstu"
                    ] = bpp.Wersja_Tekstu_OpenAccess.objects.get(
                        skrot="FINAL_PUBLISHED"
                    )
                    kw["openaccess_licencja"] = bpp.Licencja_OpenAccess.objects.get(
                        skrot="CC-BY-NC-SA"
                    )
                    kw[
                        "openaccess_czas_publikacji"
                    ] = bpp.Czas_Udostepnienia_OpenAccess.objects.get(
                        skrot="AT_PUBLICATION"
                    )
                    kw["openaccess_tryb_dostepu"] = oa_klass.objects.get(
                        skrot="OPEN_JOURNAL"
                    )
                else:

                    if elem.get("a"):
                        s = dbf.Usi.objects.get(idt_usi=elem["a"][1:])
                        try:
                            o = bpp.Wersja_Tekstu_OpenAccess.objects.get(skrot=s.nazwa)
                        except bpp.Wersja_Tekstu_OpenAccess.DoesNotExist:
                            raise ValueError(
                                f"Nie ma takiej wersji tekstu openaccess: {s.nazwa}, rekord {kw}"
                            )
                        kw["openaccess_wersja_tekstu"] = o

                    if elem.get("b"):
                        s = dbf.Usi.objects.get(idt_usi=elem["b"][1:])
                        if s.nazwa == "CC":
                            s.nazwa = "CC-ZERO"

                        try:
                            o = bpp.Licencja_OpenAccess.objects.get(
                                skrot=s.nazwa.replace(" ", "-")
                            )
                        except bpp.Licencja_OpenAccess.DoesNotExist:
                            raise ValueError("Niepoprawny skrot licencji: %s" % s.nazwa)

                        kw["openaccess_licencja"] = o

                    if elem.get("c"):
                        s = dbf.Usi.objects.get(idt_usi=elem["c"][1:])
                        o = bpp.Czas_Udostepnienia_OpenAccess.objects.get(skrot=s.nazwa)
                        kw["openaccess_czas_publikacji"] = o

                    if elem.get("d"):
                        # Zazwyczaj prowadzi do pustego wpisu
                        s = dbf.Usi.objects.get(idt_usi=elem["d"][1:])

                        if s.nazwa != "0":
                            warnings.warn(
                                f"Niezaimportowana nazwa: {s.nazwa}, ({elem}, {kw}, {rec})"
                            )

                    if elem.get("e"):
                        s = dbf.Usi.objects.get(idt_usi=elem["e"][1:])
                        try:
                            o = oa_klass.objects.get(skrot=s.nazwa)
                        except oa_klass.DoesNotExist:
                            warnings.warn(
                                f"Nieistniejacy tryb dostepu openaccesS: {s.nazwa} dla {elem} {rec} {rec.idt}"
                            )
                            o = None

                        if o is not None:
                            kw["openaccess_tryb_dostepu"] = o

            elif elem["id"] == 107:
                s = dbf.Usi.objects.get(idt_usi=elem["a"][1:])
                if s.nazwa.startswith("ISSN"):
                    kw["issn"] = s.nazwa[4:].strip()
                elif s.nazwa.find("-") in [4, 5]:  # sam "goły" nr ISSN
                    kw["issn"] = s.nazwa
                else:
                    raise NotImplementedError(elem, kw, rec)

                for literka in "bcd":
                    assert not elem.get(literka), (elem, kw, rec)

            elif elem["id"] == 997:
                if elem["a"] == "3786" or elem["a"] == "":
                    assert not kw.get("adnotacje"), (elem, kw, rec)
                    kw["adnotacje"] = (
                        "Import bazy danych: niejasny element POZ_G: %s" % elem
                    )
                elif elem["a"] == "Publikacja uwzględniona w Web of Science":
                    wos = True
                elif elem["a"] == "Publikacja w wydawnictwie spoza listy MNiSW":
                    kw["uwagi"] = exp_combine(kw.get("uwagi", ""), elem.get("a"))
                elif elem["a"].startswith("http"):
                    if kw.get("www"):
                        assert not kw.get("public_www"), (elem, kw, rec)
                        kw["public_www"] = elem["a"]
                    else:
                        kw["www"] = elem["a"]
                else:
                    raise NotImplementedError(elem, rec)

            elif elem["id"] == 884:
                # PubMedID
                if elem.get("a"):
                    kw["pubmed_id"] = to_pubmed_id(elem["a"])
                kw["pmc_id"] = elem["b"]

                if (
                    kw["pmc_id"]
                    == "https://www-1ncbi-1nlm-1nih-1gov-10031ebjc023e.han.bg.umed.wroc.pl/pmc/articles/PMC4773206/"
                ):
                    kw["pmc_id"] = "PMC4773206"

                for literka in "cde":
                    assert not elem.get(literka)

            elif elem["id"] == 983:
                kw["adnotacje"] = exp_combine(
                    kw.get("adnotacje", ""),
                    "Import bazy danych. Niejasny element POZ_G: %s" % elem,
                    sep="\n",
                )

            elif elem["id"] == 988:
                if rec.idt == 77270:
                    # dwa numery ISBN, do tego z problemem w zapisie
                    kw["isbn"] = dbf.Usi.objects.get(idt_usi=elem["a"][1:]).nazwa
                    kw["e_isbn"] = dbf.Usi.objects.get(idt_usi=elem["b"][1:11]).nazwa
                else:
                    if elem.get("a"):
                        kw["isbn"] = dbf.Usi.objects.get(idt_usi=elem["a"][1:]).nazwa

                    if elem.get("b"):
                        if elem["b"].startswith("#"):
                            kw["adnotacje"] = exp_combine(
                                kw.get("adnotacje"),
                                "Niejasny element importu: %r" % elem["b"],
                            )
                        else:
                            kw["e_isbn"] = dbf.Usi.objects.get(
                                idt_usi=elem["b"][1:11]
                            ).nazwa

                    for literka in "cde":
                        assert not elem.get(literka), (elem, kw, rec)

            elif elem["id"] == 985:
                # elem['a'] prowadzi do IDT_USI ktorego nazwa to
                # PubMed ID
                pmid = dbf.Usi.objects.get(idt_usi=elem["a"][1:11]).nazwa
                if kw.get("pubmed_id"):
                    assert kw["pubmed_id"] == pmid, (elem, kw, rec)
                else:
                    kw["pubmed_id"] = pmid

                for literka in "bcde":
                    assert not elem.get(literka), (elem, kw, rec)

            #
            # rec.zrodlo
            #

            elif elem["id"] == 200:
                # Wydawnictwo zwarte
                assert klass == bpp.Wydawnictwo_Zwarte

                for literka in "efg":
                    assert not elem.get(
                        literka
                    ), f"co mam z tym zrobic literka {literka} w {elem!r}"

                assert not kw.get("informacje"), (
                    kw["informacje"],
                    rec,
                    rec.idt,
                    rec.__dict__,
                )
                kw["informacje"] = elem["a"]
                if elem.get("b"):
                    kw["informacje"] += ": " + elem.get("b")

                if elem.get("c"):
                    z200c = elem.get("c")

                    if True or z200c.strip().startswith("pod red."):
                        # Redaktorzy zostaną dodani jako Wydawnictwo_..._Autor, typ odpowiedzialnosci
                        # to będzie redaktor
                        pass
                    # else:
                    #     print(
                    #         f"*** Zrodlo pole 200C niepuste i zaczyna sie od: {z200c[:10]}"
                    #     )
                    #     kw["szczegoly"] += "; " + z200c

                if elem.get("d"):
                    kw["informacje"] += "; " + elem.get("d")

                if kw["informacje"]:
                    res = parse_informacje(kw["informacje"], "tom")
                    if res:
                        assert not kw.get("tom")
                        kw["tom"] = res

                    # Zwarte NIE ma numeru zeszytu
                    # kw['nr_zeszytu'] = parse_informacje(kw['informacje'], "numer")

            elif elem["id"] == 100:
                # Wydawnictwo_Ciagle
                assert klass == bpp.Wydawnictwo_Ciagle
                for literka in "bcde":
                    if literka in elem.keys():
                        raise NotImplementedError("co mam z tym zrobic %r" % elem)

                kw["zrodlo"] = dbf.Usi.objects.get(idt_usi=elem["a"][1:]).bpp_id

            elif elem["id"] == 152:
                # Wydawca indeksowany
                assert klass == bpp.Wydawnictwo_Zwarte

                for literka in "def":
                    assert literka not in elem.keys()

                if elem.get("b"):
                    try:
                        int(elem["b"][1:])
                    except ValueError:
                        warnings.warn(
                            f"Mial byc wydawca indeksowany, jest jakis tekst {elem} dla rec {rec} o ID {rec.idt}"
                        )
                        kw["wydawca_opis"] = elem.get("b")
                        continue

                    wydawca = dbf.Usi.objects.get(idt_usi=elem["b"][1:]).bpp_wydawca_id
                    if kw.get("wydawca") and kw["wydawca"] != wydawca:
                        raise NotImplementedError(
                            "Juz jest wydawca, prawdopodobnie z tabeli Poz"
                        )
                    kw["wydawca"] = wydawca

                if elem.get("a"):
                    kw[
                        "miejsce_i_rok"
                    ] = f"{elem.get('a', '')} {elem.get('c', '')}".strip()

            # Koniec rec.zrodlo

            #
            # poz_n
            #

            elif elem["id"] == 206:
                if kw.get("uwagi"):
                    kw["adnotacje"] = exp_combine(
                        kw.get("adnotacje", ""), elem["a"], ". "
                    )
                else:
                    kw["uwagi"] = elem["a"]
                for literka in "bcd":
                    assert not elem.get(literka), (elem, rec, kw)
            elif elem["id"] == 205:
                if elem["a"].startswith("ISBN "):
                    kw["isbn"] = elem["a"][5:].strip()
                elif elem["a"].startswith("Toż w:"):
                    kw["uwagi"] = exp_combine(kw.get("uwagi", ""), elem.get("a"))
                elif elem["a"].startswith("https"):
                    kw["adnotacje"] = exp_combine(
                        kw.get("adnotacje", ""),
                        "Drugi adres strony WWW: " + elem.get("a"),
                    )
                else:
                    raise NotImplementedError(elem, rec, kw)

                for literka in "bcd":
                    assert not elem.get(literka), (elem, rec, kw)

            elif elem["id"] == 995:
                if kw.get("www"):
                    if "http://" + kw["www"] == elem["a"] or kw["www"] == elem["a"]:
                        del kw["www"]
                    else:
                        kw["adnotacje"] = exp_combine(
                            kw.get("adnotacje", ""),
                            "Drugi adres WWW? " + kw["www"],
                            sep="\n",
                        )
                        del kw["www"]

                assert not kw.get("www"), (elem, rec, kw)
                kw["www"] = elem["a"]
                for literka in "bcd":
                    assert not elem.get(literka), (elem, rec, rec.idt)

            #
            # Koniec Poz_n
            #

            elif elem["id"] == 102:
                zrodlo = dbf.Usi.objects.get(
                    usm_f="102", idt_usi=int(elem["a"][1:])
                ).bpp_id
                if kw.get("zrodlo"):
                    if kw["zrodlo"] != zrodlo:
                        warnings.warn(
                            f"DWA ROZNE ZRODLA? rekord {rec} o ID {rec.idt} "
                            f"ma już ustalone {kw['zrodlo']} ale mialby miec "
                            f"tez zrodlo: {zrodlo}?!"
                        )
                else:
                    # assert not kw.get("zrodlo")
                    kw["zrodlo"] = zrodlo

            elif elem["id"] in [106, 260, 159, 209]:
                kw["numer_odbitki"] = elem["a"]

            elif elem["id"] == 253:
                # jeżeil w 'a' jest tekst, to konferencja
                if elem["a"].startswith("|"):
                    raise NotImplementedError("Tu bylo wywolanie pdb")
                else:
                    nazwa = elem["a"]
                    if elem.get("b"):
                        nazwa += "; " + elem.get("b")
                    konferencja = bpp.Konferencja.objects.get_or_create(nazwa=nazwa)[0]
                kw["konferencja"] = konferencja

            elif elem["id"] == 996:
                if elem.get("a"):
                    kw["issn"] = dbf.Usi.objects.get(idt_usi=elem["a"][1:]).nazwa

                if elem.get("b"):
                    kw["e_issn"] = dbf.Usi.objects.get(idt_usi=elem["b"][1:]).nazwa

                for literka in "cde":
                    assert not elem.get(literka), (elem, rec, rec.idt)

            elif elem["id"] == 252:
                assert not kw.get("informacje")
                kw["informacje"] = elem.get("a")

                for literka in "bcde":
                    if elem.get(literka):
                        warnings.warn(
                            f"Nieobslugiwane literki przy polu 252 {elem} rekord {rec} o ID {rec.idt}"
                        )

            elif elem["id"] == 982:
                delayed_creation.append(
                    (
                        bpp.Element_Repozytorium.objects.create,
                        dict(
                            rodzaj=elem.get("a"),
                            nazwa_pliku=elem.get("b"),
                            tryb_dostepu=elem.get("c"),
                        ),
                    )
                )

            elif elem["id"] == 889:
                if elem.get("c"):
                    grant = bpp.Grant.objects.get_or_create(
                        numer_projektu=dbf.Usi.objects.get(
                            idt_usi=elem.get("c")[1:]
                        ).nazwa,
                    )[0]

                    entry = (bpp.Grant_Rekordu.objects.create, dict(grant_id=grant.pk))
                    add = True
                    for elem in delayed_creation:
                        if elem == entry:
                            warnings.warn(
                                f"__ grant podany dwukrotnie! grant {grant} rekord {rec}"
                            )
                            add = False

                    if add:
                        delayed_creation.append(entry)

                else:
                    warnings.warn(
                        f"__ grant bez numeru grantu! Rekord {rec} o ID {rec.idt}, elementy grantu (pole 889): {elem}"
                    )

            elif elem["id"] == 250:
                warnings.warn(
                    f"Nieobsługiwane ID pola {elem['id']} (wartosci: {elem}) w kontekście rekordu {rec} o ID {rec.idt}"
                )
                # CMKP TODO

            elif elem["id"] == 210:
                assert not kw.get("isbn")
                kw["isbn"] = elem["a"]
                for literka in "bcd":
                    assert not elem.get(literka), (elem, rec, rec.idt)

            elif elem["id"] == 254:
                warnings.warn("Ignoruje pole nr 254")

            else:
                raise NotImplementedError(elem, rec, rec.idt)

        # Zaimportuj z ID takim samym jak po stronie pliku DBF
        kw["id"] = rec.idt

        try:
            for i, v in kw.items():
                if (
                    v is not None
                    and hasattr(v, "__len__")
                    and len(v) >= 512
                    and i
                    not in [
                        "tytul_oryginalny",
                        "tekst_po_ostatnim_autorze",
                        "informacje",
                    ]
                ):
                    print(
                        "*** UWAGA: przycinam pole %s do 512 znakow, wczesniej wartosc: %s"
                        % (i, v)
                    )
                    kw[i] = v[:512]
            res = klass.objects.create(**kw)
        except Exception as e:
            pprint.pprint(kw)
            raise e

        rec.object = res
        rec.save()

        if delayed_creation:
            for create_function, kw in delayed_creation:
                kw["rekord"] = res
                create_function(**kw)

        if wos:
            klass_ext.objects.create(
                rekord=res, baza=bpp.Zewnetrzna_Baza_Danych.objects.get(skrot="WOS")
            )


@transaction.atomic
def przypisz_grupy_punktowe():
    for rec in dbf.Bib.objects.exclude(idt2=None):
        # print(rec.tytul_or)
        if rec.idt2.object is None:
            print(
                f"GP blad: Praca {rec.idt} ({rec.tytul_or[:50]}) ma IDT2 {rec.idt2.pk} ({rec.idt2.tytul_or[:50]}) "
                f"jednak obiekt w BPP od IDT2 to NULL!"
            )
        else:
            if rec.idt2.object.__class__ == bpp.Wydawnictwo_Ciagle:
                print(
                    f"GP blad: Praca {rec.idt} ({rec.tytul_or[:50]}) ma IDT2 {rec.idt2.pk} "
                    f"({rec.idt2.object.tytul_oryginalny[:50]}) jednak obiekt w BPP to Wydawnictwo_Ciagle!"
                )
            else:
                o = rec.object
                if o.__class__ == bpp.Wydawnictwo_Zwarte:
                    if o.wydawnictwo_nadrzedne is None:
                        # print("Saved")
                        o.wydawnictwo_nadrzedne = rec.idt2.object
                        o.save()

                else:
                    print(
                        f"*** Mam wydawnictwo nadrzedne dla wydawnictwa ciągłego..? Nie tworze. "
                        f"{rec.idt} {rec.tytul_or[:50]} ma nadrzedne {rec.idt2.pk} "
                        f"{rec.idt2.object.tytul_oryginalny[:50]})"
                    )


def set_sequences():
    for elem in [
        "bpp_wydawnictwo_ciagle",
        "bpp_wydawnictwo_zwarte",
        "bpp_patent",
        "bpp_praca_doktorska",
        "bpp_praca_habilitacyjna",
    ]:
        set_seq(elem)


def wyswietl_prace_bez_dopasowania(logger):
    bez = dbf.Bib.objects.filter(object_id=None)
    if bez.exists():
        logger.info("*** Prace bez dopasowania: %i rekordow. " % bez.count())
        for rec in bez[:30]:
            logger.info((rec.idt, rec.rok, rec.tytul_or_s))
    else:
        logger.info("+++ Wszystkie prace mają dopasowanie.")


@transaction.atomic
def zatwierdz_podwojne_przypisania(logger):
    """Zakładamy, że podwójne przypisania w bazie danych są POPRAWNE. Expertus
    nie umożliwia sytuacji, gdzie dwóch Janów Kowalskich ma dwa rekordy
    (zdaniem BG UMW), stąd w przypadku podwójnych przypisań autorów, utwórz
    nowy rekord dla drugiego(czy na pewno?) autora

    Genialnie byłoby tu upewnić się, że każdy tworzony dodatkowy rekord
    to rekord w obcej jednostce
    """
    from django.conf import settings

    setattr(settings, "ENABLE_DATA_AKT_PBN_UPDATE", False)

    for elem in (
        dbf.B_A.objects.order_by()
        .values(
            "idt_id",
            "idt_aut_id",
        )
        .annotate(cnt=Count("*"))
        .filter(cnt__gt=1)
    ):
        elem["cnt"]
        for melem in dbf.B_A.objects.filter(
            idt_id=elem["idt_id"], idt_aut_id=elem["idt_aut_id"]
        )[1:]:
            if melem.idt_jed is None:
                warnings.warn(
                    f"Rekord {melem.idt} ma przypisanego autora {melem.idt_aut} bez przypisanej jednostki"
                )
                continue

            logger.info(
                (
                    "Tworzę dodatkowe (podwójne) przypisanie",
                    melem.idt.tytul_or_s,
                    "(",
                    melem.idt.rok,
                    ") - ",
                    melem.idt_aut,
                    ", jedn. ",
                    melem.idt_jed,
                )
            )

            kw = model_to_dict(melem.idt_aut)
            kw["idt_jed_id"] = kw.pop("idt_jed")
            kw["exp_id"] = None
            kw["bpp_autor"] = None
            kw["bpp_autor_id"] = None
            kw["orcid_id"] = None
            kw["pbn_id"] = None
            kw["ref"] = None
            from django.db.models import Max

            kw["idt_aut"] = dbf.Aut.objects.all().aggregate(c=Max("idt_aut"))["c"] + 1
            idt_aut = dbf.Aut.objects.create(**kw)

            melem.idt_aut = idt_aut
            melem.save()

    integruj_autorow()


@transaction.atomic
def usun_podwojne_przypisania_b_a(logger):
    from django.conf import settings

    setattr(settings, "ENABLE_DATA_AKT_PBN_UPDATE", False)

    for elem in (
        dbf.B_A.objects.order_by()
        .values(
            "idt_id",
            "idt_aut_id",
        )
        .annotate(cnt=Count("*"))
        .filter(cnt__gt=1)
    ):
        cnt = elem["cnt"]
        for melem in dbf.B_A.objects.filter(
            idt_id=elem["idt_id"], idt_aut_id=elem["idt_aut_id"]
        ):
            logger.info(
                (
                    "1 Podwojne przypisanie",
                    melem.idt.tytul_or_s,
                    "(",
                    melem.idt.rok,
                    ") - ",
                    melem.idt_aut,
                )
            )
            melem.delete()
            cnt -= 1
            if cnt == 1:
                break

    query = (
        dbf.B_A.objects.order_by()
        .values("idt_id", "idt_aut__bpp_autor_id")
        .order_by()
        .annotate(cnt=Count("*"))
        .filter(cnt__gt=1)
    )
    for elem in query:
        print(elem)
        cnt = elem["cnt"]
        for melem in dbf.B_A.objects.filter(
            idt_id=elem["idt_id"], idt_aut__bpp_autor_id=elem["idt_aut__bpp_autor_id"]
        ):
            logger.info(
                (
                    "2 Podwojne przypisanie",
                    melem.idt.tytul_or_s,
                    "(",
                    melem.idt.rok,
                    ") - ",
                    melem.idt_aut,
                )
            )
            melem.delete()
            cnt -= 1
            if cnt == 1:
                break


@transaction.atomic
def integruj_b_a(offset=None, limit=None):
    from django.conf import settings

    setattr(settings, "ENABLE_DATA_AKT_PBN_UPDATE", False)
    setattr(settings, "BPP_DODAWAJ_JEDNOSTKE_PRZY_ZAPISIE_PRACY", False)

    if cache.enabled():
        cache.disable()

    ctype_to_klass_map = {
        ContentType.objects.get_by_natural_key(
            "bpp", "wydawnictwo_ciagle"
        ).pk: bpp.Wydawnictwo_Ciagle_Autor,
        ContentType.objects.get_by_natural_key(
            "bpp", "wydawnictwo_zwarte"
        ).pk: bpp.Wydawnictwo_Zwarte_Autor,
    }

    # klass_to_ctype_map = {
    #     bpp.Wydawnictwo_Ciagle_Autor: ContentType.objects.get_by_natural_key(
    #         "bpp", "wydawnictwo_ciagle_autor"
    #     ).pk,
    #     bpp.Wydawnictwo_Zwarte_Autor: ContentType.objects.get_by_natural_key(
    #         "bpp", "wydawnictwo_zwarte_autor"
    #     ).pk,
    # }

    ta_id = bpp.Typ_Odpowiedzialnosci.objects.get(skrot="aut.").pk
    tr_id = bpp.Typ_Odpowiedzialnosci.objects.get(skrot="red.").pk

    from django.db import connection, reset_queries

    base_query = dbf.B_A.objects.filter(object_id=None).exclude(idt__object_id=None)

    if offset is not None and limit is not None:
        base_query = base_query[offset:limit]

    count_queries = True

    base_query = base_query.select_related("idt", "idt_aut", "idt_jed").only(
        "idt__content_type_id",
        "idt__object_id",
        "idt_aut__bpp_autor_id",
        "idt_aut__imiona",
        "idt_aut__nazwisko",
        "idt_jed__bpp_jednostka_id",
        "afiliacja",
        "lp",
        "idt__idt",
        "idt__tytul_or_s",
    )

    iter = base_query
    if offset is not None and offset == 0:
        iter = pbar(base_query, base_query.count())

    for rec in iter:
        if count_queries:
            reset_queries()

        bpp_ctype_id = rec.idt.content_type_id
        bpp_rec_id = rec.idt.object_id

        klass = ctype_to_klass_map.get(bpp_ctype_id)

        bpp_autor_id = rec.idt_aut.bpp_autor_id
        bpp_jednostka_id = rec.idt_jed.bpp_jednostka_id

        lp = 0
        if rec.lp:
            if len(rec.lp) == 1:
                lp = ord(rec.lp)
            else:
                lp = int(rec.lp)

        typ_odp = ta_id
        zj = f"{rec.idt_aut.imiona} {rec.idt_aut.nazwisko}"

        red1 = rec.idt_aut.imiona.strip().startswith("<")
        red2 = rec.idt_aut.nazwisko.strip().startswith("<")
        if red1 and red2:
            typ_odp = tr_id
            zj = zj.replace("<", "").replace(">", "")

        try:
            k = klass(
                rekord_id=bpp_rec_id,
                autor_id=bpp_autor_id,
                jednostka_id=bpp_jednostka_id,
                zapisany_jako=zj,
                afiliuje=rec.afiliacja == "*",
                kolejnosc=lp,
                typ_odpowiedzialnosci_id=typ_odp,
            )
            k.save(__disable_bmoa_clean_method=True)
        except IntegrityError as e:
            print(
                "Rekord: %s, %s, %s, %s -> IntegrityError"
                % (rec.idt.tytul_or_s, rec.idt.idt, rec.idt_aut.nazwisko, lp)
            )
            raise e

        if count_queries:
            #
            # Jeżeli wewnątrz pętli będzie więcej, niż jedno zapytanie SQL, to zbuntuj się.
            # W ten sposb sprawdzamy regresję + wyłączenie mechanizmw cache
            #
            if len(connection.queries) > 1:
                for elem in connection.queries:
                    print(elem)
                sys.exit(-1)
            count_queries = False

        # rec.object_id = wxa.pk
        # rec.content_type_id = klass_to_ctype_map.get(klass)
        # rec.save(update_fields=['object_id', 'content_type_id'])


@transaction.atomic
def przypisz_jednostki():
    query = (
        dbf.B_A.objects.values_list(
            "idt_aut__bpp_autor_id", "idt_jed__bpp_jednostka_id"
        )
        .order_by()
        .distinct()
    )
    for elem in pbar(query, query.count()):
        bpp.Autor_Jednostka.objects.get_or_create(
            autor_id=elem[0], jednostka_id=elem[1]
        )


def xls2dict(fp):
    """Wczytuje plik XLS do słownika. Tylko pierwszy skoroszyt.
    Pierwszy wiersz w pliku XLS to nazwy kolumn, które będą użyte
    w zwracanych słownikach."""

    book = openpyxl.load_workbook(fp)
    first_sheet = book.worksheets[0]
    headers = [
        x.lower().replace(" ", "_")
        for x in list(first_sheet.iter_rows(1, 1, values_only=True))[0]
    ]

    for row in first_sheet.iter_rows(2, first_sheet.max_row, values_only=True):
        yield dict(zip(headers, row))


def dodaj_aktualnosc():
    """Dodaje do 'Aktualności' wpis z datą i czasem importu."""
    a = Article.objects.get_or_create(
        status=Article.STATUS.published,
        title="Informacja o imporcie bazy danych",
    )[0]
    a.article_body = "Bazę danych zaimportowano: %s" % timezone.make_naive(
        timezone.now()
    )
    a.save()


@transaction.atomic
def utworz_szkielety_ksiazek(logger):
    """
    https://mantis.iplweb.pl/file_download.php?file_id=85&type=bug

    Na podstawie pola 200 A dla rekordów, które nie posiadają wydawnictw nadrzędnych,
    utwórz "szkielety" książek
    """

    if cache.enabled():
        cache.disable()

    for bib in dbf.Bib.objects.filter(
        Q(zrodlo__icontains="#200$")
        | Q(zrodlo__icontains="#202$")
        | Q(zrodlo__icontains="#205$")
    ).order_by("pk"):
        try:
            wz = bpp.Wydawnictwo_Zwarte.objects.get(
                pk=bib.object_id, wydawnictwo_nadrzedne=None
            )
        except bpp.Wydawnictwo_Zwarte.DoesNotExist:
            continue

        fld = exp_parse_str(bib.zrodlo)

        tz = fld.get("a")
        try:
            wn = bpp.Wydawnictwo_Zwarte.objects.get(tytul_oryginalny=tz, rok=wz.rok)
            logger.info(
                f"WNWJ Przypisano wydawnictwo nadrzedne\t"
                f"{wn.tytul_oryginalny} ID={wn.pk}\t"
                f"dla rekordu\t"
                f"{wz.tytul_oryginalny} ID={wz.pk}\t"
            )
        except bpp.Wydawnictwo_Zwarte.MultipleObjectsReturned:
            logger.info(
                f"WNWX Otrzymano liczne rekordy dla zapytania o tytul\t{tz}\t dla rekordu\t"
                f"{wz.tytul_oryginalny}\tID={wz.pk}\t"
                f", nie zmieniam nic"
            )
            continue
        except bpp.Wydawnictwo_Zwarte.DoesNotExist:
            logger.info(
                f"WNWN Tworze wydawnictwo zwarte:\t{tz}\trok\t{wz.rok}\tdla rekordu\t{wz.tytul_oryginalny}\tID={wz.pk}"
            )

            # Spróbuj znaleźć wydanwictwo o POŁOWIE długości tytułu, jeżeli jest to napisz
            # na ten temat (ale nadal utwórz)
            try:
                x = bpp.Wydawnictwo_Zwarte.objects.get(
                    tytul_oryginalny__istartswith=tz[: int(math.ceil(len(tz) / 2))]
                )
                logger.info(
                    f"WNWN ... jednakże znalazłem prace zaczynającą się podobnie:\t"
                    f"{x.tytul_oryginalny}\tID={x.pk}\t"
                    f"ale tworze nowy rekord i tak!"
                )
            except bpp.Wydawnictwo_Zwarte.DoesNotExist:
                pass
            except bpp.Wydawnictwo_Zwarte.MultipleObjectsReturned:
                logger.info(
                    "WNWN ... znaleziono liczne prace dla połowy tytułu, tworze nową"
                )

            # ISBN to pole 205
            skelet_isbn = current_isbn_value = None
            try:
                current_isbn_value = dbf.Bib_Desc.objects.get(
                    idt=bib, elem_id=205, source="poz_g"
                ).value
            except dbf.Bib_Desc.DoesNotExist:
                pass

            if current_isbn_value is not None:
                skelet_isbn = current_isbn_value.get("a").replace("ISBN ", "")

            # Miejsce wydania to pole 202a
            skelet_place = pole202 = None
            try:
                pole202 = dbf.Bib_Desc.objects.get(
                    idt=bib, elem_id=202, source="poz_g"
                ).value
            except dbf.Bib_Desc.DoesNotExist:
                pass

            if pole202 is not None:
                skelet_place = pole202.get("a")

                if skelet_place:
                    skelet_place += " " + str(wz.rok)

            # kopiowanie wydawcy - podpole 202 B
            skelet_wydawca = None
            if pole202 is not None and (pole202.get("b") or "").strip():
                skelet_wydawca = bpp.Wydawca.objects.get_or_create(
                    nazwa=pole202.get("b").strip()
                )[0]

            skelet_adnotacje = "Automatycznie utworzone wydawnictwo szkieletowe. " + (
                fld.get("c") or ""
            )  # "red. Jan Kowalski"

            if skelet_isbn is not None:
                if len(skelet_isbn) > 64:
                    logger.info(f"WNWN ISBN za dlugi {skelet_isbn}, idzie do adnotacji")
                    skelet_adnotacje += " " + skelet_isbn
                    skelet_adnotacje = skelet_adnotacje.strip()
                    skelet_isbn = None

            wn = bpp.Wydawnictwo_Zwarte.objects.create(
                tytul_oryginalny=tz,
                # kopiowanie informacji z podpola 200 c (redaktorzy) do pola "adnotacje"
                adnotacje=skelet_adnotacje,
                # ISBN
                isbn=skelet_isbn,
                # miejsce i rok
                miejsce_i_rok=skelet_place,
                # wydawca
                wydawca=skelet_wydawca,
                # rok
                rok=wz.rok,
                charakter_formalny=bpp.Charakter_Formalny.objects.get_or_create(
                    skrot="SWN",
                    nazwa="Szkieletowe wydawnictwo nadrzędne",
                    charakter_ogolny=CHARAKTER_OGOLNY_KSIAZKA,
                    charakter_pbn=bpp.Charakter_PBN.objects.get(
                        identyfikator="scholarly-textbook"
                    ),
                    rodzaj_pbn=RODZAJ_PBN_KSIAZKA,
                )[0],
                typ_kbn=bpp.Typ_KBN.objects.get(skrot="000"),
                # kopiowanie informacji o języku
                jezyk=wz.jezyk,
                status_korekty=bpp.Status_Korekty.objects.get(nazwa="przed korektą"),
            )

        wz.wydawnictwo_nadrzedne = wn
        wz.save()


@transaction.atomic
def integruj_dyscypliny():
    for rec in dbf.Dys.objects.all():
        aut = dbf.Aut.objects.filter(orcid_id=rec.orcid_id).first()
        if aut is None:
            warnings.warn(f"ORCID {rec.orcid_id} nie ma przypisanego autora w bazie!")
            continue

        if aut.bpp_autor_id is None:
            warnings.warn(
                f"Autor z expertusowym ORICDem {rec.orcid_id} nie ma autora po stronie BPP"
            )
            continue

        aut = bpp.Autor.objects.get(id=aut.bpp_autor_id)

        for literka, rok in [
            ("a", 2017),
            ("b", 2018),
            ("c", 2019),
            ("d", 2020),
        ]:
            try:
                dys = getattr(rec, f"{literka}_dysc_1")
            except dbf.Ldy.DoesNotExist:
                continue
            pro_dys = getattr(rec, f"{literka}_dysc_1_e")

            try:
                dys = bpp.Dyscyplina_Naukowa.objects.get(
                    kod=dys.id.replace("0", "").strip()
                )
            except bpp.Dyscyplina_Naukowa.DoesNotExist:
                warnings.warn(f"Nie mam kodu na dyscypline f{dys}, tworze")
                dys = bpp.Dyscyplina_Naukowa.objects.create(
                    kod=dys.id.replace("0", "").strip(),
                    widoczna=True,
                    nazwa=dys.dyscyplina,
                )

            try:
                subdys = getattr(rec, f"{literka}_dysc_2")
                pro_subdys = getattr(rec, f"{literka}_dysc_2_e")
            except dbf.Ldy.DoesNotExist:
                subdys = None
                pro_subdys = None

            if subdys is not None:
                try:
                    subdys = bpp.Dyscyplina_Naukowa.objects.get(
                        kod=subdys.id.replace("0", "").strip()
                    )
                except bpp.Dyscyplina_Naukowa.DoesNotExist:
                    warnings.warn(f"Nie mam kodu na dyscypline f{subdys}, tworze")
                    subdys = bpp.Dyscyplina_Naukowa.objects.create(
                        kod=subdys.id.replace("0", "").strip(),
                        widoczna=True,
                        nazwa=subdys.dyscyplina,
                    )

            kw = dict(
                rok=rok,
                autor=aut,
                rodzaj_autora=getattr(rec, f"{literka}_n"),
                dyscyplina_naukowa=dys,
                procent_dyscypliny=pro_dys,
                subdyscyplina_naukowa=subdys,
                procent_subdyscypliny=pro_subdys,
            )
            try:
                juz = bpp.Autor_Dyscyplina.objects.get(rok=rok, autor=aut)

                save = False
                for key, value in kw.items():
                    if getattr(juz, key) != value:
                        warnings.warn(
                            f"Autor {aut} ma powyzej 1 wpisu w tabeli dyscyplin "
                            f"i rozni sie on dla klucza {key} wartoscia {value}, "
                            f"stara wartosc {getattr(juz, key)}"
                        )
                        setattr(juz, key, value)
                        save = True

                if save:
                    juz.save()
                continue
            except bpp.Autor_Dyscyplina.DoesNotExist:
                bpp.Autor_Dyscyplina.objects.create(**kw)
