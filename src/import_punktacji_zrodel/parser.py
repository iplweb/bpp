"""Czysty parser pliku JCR (Clarivate Journal Citation Reports).

Bez zależności od Django ORM — łatwy do testów jednostkowych.
Obsługuje XLSX (openpyxl) oraz CSV (stdlib csv).
"""

import csv
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

KWARTYL_MAP = {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}

_ROK_JIF_RE = re.compile(r"^\s*(\d{4})\s+JIF\s*$")
_ROK_META_RE = re.compile(r"Selected JCR Year:\s*(\d{4})")
_FOOTER_MARKERS = ("Clarivate", "Terms of Use")

_COL_NAZWA = "Journal name"
_COL_ISSN = "ISSN"
_COL_EISSN = "eISSN"
_COL_KATEGORIA = "Category"
_COL_KWARTYL = "JIF Quartile"


@dataclass
class CzasopismoJCR:
    nazwa: str
    issn: str | None
    e_issn: str | None
    impact_factor: Decimal | None
    kwartyl_wos: int | None
    kategorie: list[tuple[str, int | None]] = field(default_factory=list)


@dataclass
class ParsedJCR:
    rok: int | None
    czasopisma: list[CzasopismoJCR]


def _clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    if s == "" or s.upper() == "N/A":
        return None
    return s


def _parse_if(v) -> Decimal | None:
    s = _clean(v)
    if s is None:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        # Niepoprawna/nienumeryczna wartość IF traktowana jak brak danych (None).
        return None


def _parse_kwartyl(v) -> int | None:
    s = _clean(v)
    if s is None:
        return None
    return KWARTYL_MAP.get(s.upper())


def _iter_rows_xlsx(path: str):
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            yield list(row)
    finally:
        wb.close()


def _iter_rows_csv(path: str):
    with open(path, newline="", encoding="utf-8-sig") as f:
        yield from csv.reader(f)


def _find_header_index(rows: list[list]) -> int:
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if _COL_NAZWA in cells and _COL_KWARTYL in cells:
            return i
    raise ValueError("Nie znaleziono wiersza nagłówka w pliku JCR")


def _detect_rok(header: list[str], rows_before: list[list]) -> int | None:
    for cell in header:
        m = _ROK_JIF_RE.match(str(cell or ""))
        if m:
            return int(m.group(1))
    for row in rows_before:
        for cell in row:
            m = _ROK_META_RE.search(str(cell or ""))
            if m:
                return int(m.group(1))
    return None


def wczytaj_plik_jcr(path: str) -> ParsedJCR:
    if path.lower().endswith(".csv"):
        rows = list(_iter_rows_csv(path))
    else:
        rows = list(_iter_rows_xlsx(path))

    hidx = _find_header_index(rows)
    header = [str(c).strip() if c is not None else "" for c in rows[hidx]]
    rok = _detect_rok(header, rows[:hidx])

    idx = {name: i for i, name in enumerate(header)}
    i_nazwa = idx[_COL_NAZWA]
    i_issn = idx.get(_COL_ISSN)
    i_eissn = idx.get(_COL_EISSN)
    i_kat = idx.get(_COL_KATEGORIA)
    i_kw = idx[_COL_KWARTYL]
    i_if = next((i for i, c in enumerate(header) if _ROK_JIF_RE.match(c)), None)

    def _cell(row, i):
        if i is None or i >= len(row):
            return None
        return row[i]

    grupy: dict[tuple, CzasopismoJCR] = {}
    for row in rows[hidx + 1 :]:
        nazwa = _clean(_cell(row, i_nazwa))
        if nazwa is None or any(m in str(nazwa) for m in _FOOTER_MARKERS):
            continue
        issn = _clean(_cell(row, i_issn))
        e_issn = _clean(_cell(row, i_eissn))
        impact = _parse_if(_cell(row, i_if))
        kwartyl = _parse_kwartyl(_cell(row, i_kw))
        kategoria = _clean(_cell(row, i_kat))

        key = (issn, e_issn, nazwa)
        cz = grupy.get(key)
        if cz is None:
            cz = CzasopismoJCR(
                nazwa=nazwa,
                issn=issn,
                e_issn=e_issn,
                impact_factor=impact,
                kwartyl_wos=kwartyl,
                kategorie=[],
            )
            grupy[key] = cz
        else:
            if cz.impact_factor is None and impact is not None:
                cz.impact_factor = impact
            if kwartyl is not None and (
                cz.kwartyl_wos is None or kwartyl < cz.kwartyl_wos
            ):
                cz.kwartyl_wos = kwartyl  # najlepszy kwartyl (min)
        cz.kategorie.append((kategoria, kwartyl))

    return ParsedJCR(rok=rok, czasopisma=list(grupy.values()))
