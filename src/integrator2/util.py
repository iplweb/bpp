import openpyxl


def find_header_row(sheet, first_column_value, max_row_range=10, max_col_range=10):
    """
    Procedura poszukuje wiersza nagłówkowego.

    :param sheet: :class:`xlrd.sheet.Sheet`
    :param first_column_value: pierwsza poszukiwana wartość
    :type first_column_value: str
    :param max_row_range: maksymalny zakres "w dół" gdzie szukamy,
    :param max_col_range: maksymalny zakres "w prawo" gdzie szukamy.
    :return: zwraca numer wiersza, w którym znajduje się nagłówek.
    """
    for n, row in enumerate(
        sheet.iter_rows(1, max_row=min(max_row_range, sheet.max_row)), start=1
    ):
        for b in range(min(max_col_range, sheet.max_column)):
            f = row[b].value
            if hasattr(f, "upper") and f.upper() == first_column_value.upper():
                return n


def build_mapping(xls_columns, wanted_columns):
    """
    Buduje mapowanie pomiędzy kolumnami dostarczonymi w XLS a kolumnami wymaganymi.

    Przykładowo, przy wejściu:

    xls_columns=["c", "b", "a"]
    wanted_columns={"A": "foo", "B": "bar", "c": "baz"}

    zwraca listę o kolejności elementów:
    ["baz", "bar", "foo"]

    :type xls_columns: list
    :type wanted_columns: dict
    :rtype: list
    """
    ret = []
    upper_wanted_columns = {
        x.upper().strip(): wanted_columns[x] for x in list(wanted_columns.keys())
    }
    for elem in xls_columns:
        if elem.value:
            val = elem.value.upper().strip()
            if val in upper_wanted_columns:
                ret.append(upper_wanted_columns[val])
                continue
        ret.append(None)

    return ret


class ReadDataException(Exception):
    pass


def read_xls_data(
    filename,
    column_mapping,
    header_row_name,
    ignored_sheet_names=None,
    limit=None,
    limit_sheets=None,
    transformations=None,
):
    """
    Importuje dane z pliku XLS - ze wszystkich kolejnych zeszytów.

    :param filename: nazwa pliku XLS do odczytu
    :param column_mapping: mapowanie nazwa kolumny z XLS do atrybutu danych
    :param header_row_name: poszukiwany ciąg znaków z pierwszej kolumny wiersza nagłówka
    :param ignored_sheet_names: nazwy zeszytów ignorowanych - te nie zostaną ujęte w imporcie.
    :param transformations: słownik funkcji, transformujących wartość z arkusza przed zwróceniem jej
    :return:
    """
    book = openpyxl.load_workbook(filename=filename)
    sheets = book.worksheets
    if limit_sheets:
        sheets = sheets[:limit_sheets]
    for sheet in sheets:
        if ignored_sheet_names is not None:
            if sheet.name in ignored_sheet_names:
                continue
        header_row_no = find_header_row(sheet, header_row_name)
        if header_row_no is None:
            raise ReadDataException(
                "Brak wiersza nagłówka, poszukiwano %r" % header_row_name
            )
        header_row_values = list(sheet.iter_rows(header_row_no, header_row_no + 1))[0]
        read_data_mapping = list(
            enumerate(build_mapping(header_row_values, column_mapping))
        )

        max_row = sheet.max_row
        if limit is not None:
            max_row = min(header_row_no + 1 + limit, max_row)

        for row in sheet.iter_rows(header_row_no + 1, max_row):
            dct = {"__sheet__": sheet.title}
            for no, elem in read_data_mapping:
                value = row[no].value

                if transformations is not None:
                    if elem in transformations:
                        value = transformations[elem](value)

                dct[elem] = value
            yield dct
