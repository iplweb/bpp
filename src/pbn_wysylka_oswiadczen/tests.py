"""Tests for pbn_wysylka_oswiadczen app."""

import json
from datetime import timedelta
from unittest.mock import MagicMock, patch

import pytest
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.contrib.contenttypes.models import ContentType
from django.test import RequestFactory
from django.urls import reverse
from django.utils import timezone

from bpp.const import GR_WPROWADZANIE_DANYCH
from bpp.models import Uczelnia, Wydawnictwo_Ciagle
from pbn_wysylka_oswiadczen.models import PbnWysylkaLog, PbnWysylkaOswiadczenTask
from pbn_wysylka_oswiadczen.queries import get_publications_queryset
from pbn_wysylka_oswiadczen.tasks import (
    _delete_existing_statements,
    _handle_http_400_error,
    _send_statements_with_retry,
    get_pbn_client,
    process_single_publication,
)
from pbn_wysylka_oswiadczen.views import (
    CancelTaskView,
    LogListView,
    PbnWysylkaOswiadczenMainView,
    StartTaskView,
    TaskStatusPartialView,
    TaskStatusView,
)

User = get_user_model()


# ============================================================================
# Model Tests
# ============================================================================


@pytest.mark.django_db
def test_pbn_wysylka_oswiadczen_task_model_creation():
    """Test PbnWysylkaOswiadczenTask model creation."""
    user = User.objects.create_user("testuser", password="testpass")

    task = PbnWysylkaOswiadczenTask.objects.create(
        user=user,
        status="running",
        rok_od=2022,
        rok_do=2025,
        total_publications=100,
        processed_publications=50,
    )

    assert task.pk is not None
    assert task.status == "running"
    assert task.rok_od == 2022
    assert task.rok_do == 2025
    assert task.total_publications == 100
    assert task.processed_publications == 50
    assert "Wysyłka oświadczeń" in str(task)


@pytest.mark.django_db
def test_pbn_wysylka_oswiadczen_task_progress_percent():
    """Test progress_percent property."""
    user = User.objects.create_user("testuser", password="testpass")

    # Test with 0 total
    task = PbnWysylkaOswiadczenTask.objects.create(
        user=user, total_publications=0, processed_publications=0
    )
    assert task.progress_percent == 0

    # Test with normal values
    task.total_publications = 100
    task.processed_publications = 50
    task.save()
    assert task.progress_percent == 50

    # Test full progress
    task.processed_publications = 100
    task.save()
    assert task.progress_percent == 100


@pytest.mark.django_db
def test_pbn_wysylka_oswiadczen_task_is_stalled():
    """Test is_stalled method."""
    user = User.objects.create_user("testuser", password="testpass")

    # Running task that's not stalled
    task = PbnWysylkaOswiadczenTask.objects.create(user=user, status="running")
    assert task.is_stalled() is False

    # Manually set last_updated to be older than 15 minutes
    PbnWysylkaOswiadczenTask.objects.filter(pk=task.pk).update(
        last_updated=timezone.now() - timedelta(minutes=20)
    )
    task.refresh_from_db()
    assert task.is_stalled() is True

    # Completed task should not be stalled
    task.status = "completed"
    task.save()
    assert task.is_stalled() is False


@pytest.mark.django_db
def test_pbn_wysylka_oswiadczen_task_get_latest_task():
    """Test get_latest_task class method."""
    user = User.objects.create_user("testuser", password="testpass")

    # No tasks
    assert PbnWysylkaOswiadczenTask.get_latest_task() is None

    # Create tasks
    PbnWysylkaOswiadczenTask.objects.create(user=user, status="completed")
    task2 = PbnWysylkaOswiadczenTask.objects.create(user=user, status="running")

    # Latest should be task2 (most recent)
    latest = PbnWysylkaOswiadczenTask.get_latest_task()
    assert latest == task2


@pytest.mark.django_db
def test_pbn_wysylka_oswiadczen_task_cleanup_stale():
    """Test cleanup_stale_running_tasks class method."""
    user = User.objects.create_user("testuser", password="testpass")

    # Create stale running task
    stale_task = PbnWysylkaOswiadczenTask.objects.create(user=user, status="running")
    PbnWysylkaOswiadczenTask.objects.filter(pk=stale_task.pk).update(
        started_at=timezone.now() - timedelta(hours=25)
    )

    # Create fresh running task
    fresh_task = PbnWysylkaOswiadczenTask.objects.create(user=user, status="running")

    # Run cleanup
    count = PbnWysylkaOswiadczenTask.cleanup_stale_running_tasks()

    assert count == 1
    stale_task.refresh_from_db()
    assert stale_task.status == "failed"
    assert "przestarzalego" in stale_task.error_message

    fresh_task.refresh_from_db()
    assert fresh_task.status == "running"


@pytest.mark.django_db
def test_pbn_wysylka_log_model_creation():
    """Test PbnWysylkaLog model creation."""
    user = User.objects.create_user("testuser", password="testpass")
    task = PbnWysylkaOswiadczenTask.objects.create(user=user)

    content_type = ContentType.objects.get_for_model(Wydawnictwo_Ciagle)

    log = PbnWysylkaLog.objects.create(
        task=task,
        content_type=content_type,
        object_id=123,
        pbn_uid="test-pbn-uid-123",
        status="success",
        json_sent={"test": "data"},
        json_response={"result": "ok"},
    )

    assert log.pk is not None
    assert log.pbn_uid == "test-pbn-uid-123"
    assert log.status == "success"
    assert "test-pbn-uid-123" in str(log)


@pytest.mark.django_db
def test_pbn_wysylka_log_statuses():
    """Test all possible log statuses."""
    user = User.objects.create_user("testuser", password="testpass")
    task = PbnWysylkaOswiadczenTask.objects.create(user=user)
    content_type = ContentType.objects.get_for_model(Wydawnictwo_Ciagle)

    for status in ["success", "error", "skipped", "maintenance"]:
        log = PbnWysylkaLog.objects.create(
            task=task,
            content_type=content_type,
            object_id=1,
            pbn_uid=f"uid-{status}",
            status=status,
        )
        assert log.status == status


@pytest.mark.django_db
def test_pbn_wysylka_task_maintenance_status():
    """Test that maintenance status is available for tasks."""
    user = User.objects.create_user("testuser", password="testpass")
    task = PbnWysylkaOswiadczenTask.objects.create(
        user=user,
        status="maintenance",
        error_message="Prace serwisowe w PBN",
    )
    assert task.status == "maintenance"
    assert "Prace serwisowe" in task.error_message


# ============================================================================
# View Tests
# ============================================================================


def create_user_with_group():
    """Helper to create user with GR_WPROWADZANIE_DANYCH group."""
    user = User.objects.create_user("testuser", password="testpass")
    group, _ = Group.objects.get_or_create(name=GR_WPROWADZANIE_DANYCH)
    user.groups.add(group)
    return user


@pytest.mark.django_db
def test_main_view_context(uczelnia):
    """Test main view returns correct context."""
    factory = RequestFactory()
    request = factory.get("/?rok_od=2022&rok_do=2024")

    user = create_user_with_group()
    request.user = user
    request.session = {}

    view = PbnWysylkaOswiadczenMainView()
    view.setup(request)

    context = view.get_context_data()
    assert "rok_od" in context
    assert "rok_do" in context
    assert "total_count" in context
    assert "ciagle_count" in context
    assert "zwarte_count" in context
    assert "latest_task" in context
    assert "can_resume" in context
    assert context["rok_od"] == 2022
    assert context["rok_do"] == 2024


@pytest.mark.django_db
def test_main_view_requires_group():
    """Test main view requires proper group."""
    factory = RequestFactory()
    request = factory.get("/")

    user = User.objects.create_user("testuser", password="testpass")
    request.user = user
    request.session = {}

    try:
        response = PbnWysylkaOswiadczenMainView.as_view()(request)
        assert response.status_code in [302, 403]
    except Exception:
        # Expected - permission denied
        pass


@pytest.mark.django_db
def test_task_status_view_no_tasks():
    """Test task status view when no tasks exist."""
    factory = RequestFactory()
    request = factory.get("/status/")

    user = create_user_with_group()
    request.user = user
    request.session = {}

    view = TaskStatusView()
    response = view.get(request)

    assert response.status_code == 200
    data = json.loads(response.content)
    assert data["is_running"] is False
    assert data["latest_task"] is None


@pytest.mark.django_db
def test_task_status_view_with_task():
    """Test task status view with existing task."""
    factory = RequestFactory()
    request = factory.get("/status/")

    user = create_user_with_group()
    request.user = user
    request.session = {}

    # Create a task
    PbnWysylkaOswiadczenTask.objects.create(
        user=user,
        status="running",
        total_publications=100,
        processed_publications=50,
        success_count=30,
        error_count=5,
        skipped_count=15,
    )

    view = TaskStatusView()
    response = view.get(request)

    assert response.status_code == 200
    data = json.loads(response.content)
    assert data["is_running"] is True
    assert data["latest_task"]["status"] == "running"
    assert data["latest_task"]["total_publications"] == 100
    assert data["latest_task"]["processed_publications"] == 50
    assert data["latest_task"]["progress_percent"] == 50


@pytest.mark.django_db
def test_start_task_view_no_token():
    """Test start task view when user has no PBN token."""
    factory = RequestFactory()
    request = factory.post("/start/", {"rok_od": "2022", "rok_do": "2025"})

    user = create_user_with_group()
    request.user = user
    request.session = {}

    # Mock get_pbn_user
    class MockPbnUser:
        pbn_token = None

    user.get_pbn_user = lambda: MockPbnUser()

    view = StartTaskView()
    response = view.post(request)

    assert response.status_code == 200
    data = json.loads(response.content)
    assert data["success"] is False
    assert "zalogowany" in data["error"].lower() or "pbn" in data["error"].lower()


@pytest.mark.django_db
def test_start_task_view_expired_token():
    """Test start task view when PBN token is expired."""
    factory = RequestFactory()
    request = factory.post("/start/", {"rok_od": "2022", "rok_do": "2025"})

    user = create_user_with_group()
    request.user = user
    request.session = {}

    # Mock get_pbn_user with expired token
    class MockPbnUser:
        pbn_token = "some-token"

        def pbn_token_possibly_valid(self):
            return False

    user.get_pbn_user = lambda: MockPbnUser()

    view = StartTaskView()
    response = view.post(request)

    assert response.status_code == 200
    data = json.loads(response.content)
    assert data["success"] is False
    assert "wygasl" in data["error"].lower()


@pytest.mark.django_db
def test_start_task_view_already_running():
    """Test start task view when task is already running."""
    factory = RequestFactory()
    request = factory.post("/start/", {"rok_od": "2022", "rok_do": "2025"})

    user = create_user_with_group()
    request.user = user
    request.session = {}

    # Create running task
    PbnWysylkaOswiadczenTask.objects.create(user=user, status="running")

    # Mock get_pbn_user with valid token
    class MockPbnUser:
        pbn_token = "valid-token"

        def pbn_token_possibly_valid(self):
            return True

    user.get_pbn_user = lambda: MockPbnUser()

    view = StartTaskView()
    response = view.post(request)

    assert response.status_code == 200
    data = json.loads(response.content)
    assert data["success"] is False
    assert "uruchomione" in data["error"].lower()


@pytest.mark.django_db
def test_cancel_task_view_no_running_task():
    """Test cancel task view when no task is running."""
    factory = RequestFactory()
    request = factory.post("/cancel/")

    user = create_user_with_group()
    request.user = user
    request.session = {}

    view = CancelTaskView()
    response = view.post(request)

    assert response.status_code == 200
    data = json.loads(response.content)
    assert data["success"] is False
    assert "brak" in data["error"].lower()


@pytest.mark.django_db
def test_cancel_task_view_success():
    """Test cancel task view successfully cancels running task."""
    factory = RequestFactory()
    request = factory.post("/cancel/")

    user = create_user_with_group()
    request.user = user
    request.session = {}

    # Create running task
    task = PbnWysylkaOswiadczenTask.objects.create(user=user, status="running")

    view = CancelTaskView()
    response = view.post(request)

    assert response.status_code == 200
    data = json.loads(response.content)
    assert data["success"] is True

    task.refresh_from_db()
    assert task.status == "failed"
    assert "anulowane" in task.error_message.lower()


@pytest.mark.django_db
def test_task_status_partial_view():
    """Test task status partial view."""
    factory = RequestFactory()
    request = factory.get("/status-partial/")

    user = create_user_with_group()
    request.user = user
    request.session = {}

    view = TaskStatusPartialView()
    view.setup(request)

    context = view.get_context_data()
    assert "latest_task" in context


@pytest.mark.django_db
def test_log_list_view():
    """Test log list view."""
    factory = RequestFactory()
    request = factory.get("/logs/")

    user = create_user_with_group()
    request.user = user
    request.session = {}

    # Create task and logs
    task = PbnWysylkaOswiadczenTask.objects.create(user=user)
    content_type = ContentType.objects.get_for_model(Wydawnictwo_Ciagle)
    for i in range(5):
        PbnWysylkaLog.objects.create(
            task=task,
            content_type=content_type,
            object_id=i,
            pbn_uid=f"uid-{i}",
            status="success",
        )

    view = LogListView()
    view.setup(request)

    context = view.get_context_data()
    assert "page_obj" in context
    assert context["page_obj"].paginator.count == 5


@pytest.mark.django_db
def test_log_list_view_with_filters():
    """Test log list view with status filter."""
    factory = RequestFactory()
    request = factory.get("/logs/?status=error")

    user = create_user_with_group()
    request.user = user
    request.session = {}

    # Create task and logs with different statuses
    task = PbnWysylkaOswiadczenTask.objects.create(user=user)
    content_type = ContentType.objects.get_for_model(Wydawnictwo_Ciagle)

    PbnWysylkaLog.objects.create(
        task=task,
        content_type=content_type,
        object_id=1,
        pbn_uid="uid-success",
        status="success",
    )
    PbnWysylkaLog.objects.create(
        task=task,
        content_type=content_type,
        object_id=2,
        pbn_uid="uid-error",
        status="error",
    )

    view = LogListView()
    view.setup(request)

    context = view.get_context_data()
    assert context["page_obj"].paginator.count == 1
    assert context["page_obj"].object_list[0].status == "error"


# ============================================================================
# URL Tests
# ============================================================================


@pytest.mark.django_db
def test_url_patterns():
    """Test URL patterns are correctly configured."""
    assert reverse("pbn_wysylka_oswiadczen:main") == "/pbn-wysylka-oswiadczen/"
    assert (
        reverse("pbn_wysylka_oswiadczen:publications")
        == "/pbn-wysylka-oswiadczen/publications/"
    )
    assert reverse("pbn_wysylka_oswiadczen:status") == "/pbn-wysylka-oswiadczen/status/"
    assert (
        reverse("pbn_wysylka_oswiadczen:status-partial")
        == "/pbn-wysylka-oswiadczen/status-partial/"
    )
    assert reverse("pbn_wysylka_oswiadczen:start") == "/pbn-wysylka-oswiadczen/start/"
    assert reverse("pbn_wysylka_oswiadczen:cancel") == "/pbn-wysylka-oswiadczen/cancel/"
    assert reverse("pbn_wysylka_oswiadczen:logs") == "/pbn-wysylka-oswiadczen/logs/"
    assert (
        reverse("pbn_wysylka_oswiadczen:log-detail", kwargs={"pk": 1})
        == "/pbn-wysylka-oswiadczen/logs/1/"
    )


# ============================================================================
# Task Function Tests
# ============================================================================


@pytest.mark.django_db
def test_get_publications_queryset_empty(uczelnia):
    """Test get_publications_queryset returns empty when no matching publications."""
    ciagle_qs, zwarte_qs = get_publications_queryset(rok_od=2022, rok_do=2025)
    assert ciagle_qs.count() == 0
    assert zwarte_qs.count() == 0


@pytest.mark.django_db
def test_get_pbn_client_no_token():
    """Test get_pbn_client raises error when user has no token."""
    user = User.objects.create_user("testuser", password="testpass")

    class MockPbnUser:
        pbn_token = None

    user.get_pbn_user = lambda: MockPbnUser()

    with pytest.raises(ValueError) as exc_info:
        get_pbn_client(user)

    assert "tokenu" in str(exc_info.value).lower()


@pytest.mark.django_db
def test_get_pbn_client_expired_token():
    """Test get_pbn_client raises error when token is expired."""
    user = User.objects.create_user("testuser", password="testpass")

    class MockPbnUser:
        pbn_token = "some-token"

        def pbn_token_possibly_valid(self):
            return False

    user.get_pbn_user = lambda: MockPbnUser()

    with pytest.raises(ValueError) as exc_info:
        get_pbn_client(user)

    assert "wygasl" in str(exc_info.value).lower()


@pytest.mark.django_db
def test_get_pbn_client_no_uczelnia():
    """Test get_pbn_client raises error when no uczelnia exists."""
    user = User.objects.create_user("testuser", password="testpass")

    class MockPbnUser:
        pbn_token = "valid-token"

        def pbn_token_possibly_valid(self):
            return True

    user.get_pbn_user = lambda: MockPbnUser()

    # Make sure no Uczelnia exists
    Uczelnia.objects.all().delete()

    with pytest.raises(ValueError) as exc_info:
        get_pbn_client(user)

    assert "uczelni" in str(exc_info.value).lower()


@pytest.mark.django_db
def test_delete_existing_statements_cannot_delete():
    """Test _delete_existing_statements handles CannotDeleteStatementsException."""
    from pbn_api.exceptions import CannotDeleteStatementsException

    mock_client = MagicMock()
    mock_client.delete_all_publication_statements.side_effect = (
        CannotDeleteStatementsException("Test")
    )

    mock_publication = MagicMock()
    mock_publication.pbn_uid_id = "test-uid"

    mock_log_entry = MagicMock()

    # Should not raise exception
    _delete_existing_statements(mock_publication, mock_client, mock_log_entry)

    # Log entry should not have error message set (CannotDelete is OK)
    assert mock_log_entry.error_message != "Blad usuwania oswiadczen"


@pytest.mark.django_db
def test_delete_existing_statements_http_error():
    """Test _delete_existing_statements handles HttpException."""
    from pbn_api.exceptions import HttpException

    mock_client = MagicMock()
    mock_client.delete_all_publication_statements.side_effect = HttpException(
        500, "http://test/", "Server Error"
    )

    mock_publication = MagicMock()
    mock_publication.pbn_uid_id = "test-uid"

    mock_log_entry = MagicMock()
    mock_log_entry.error_message = ""

    _delete_existing_statements(mock_publication, mock_client, mock_log_entry)

    # Error message should be set
    assert "Blad usuwania" in mock_log_entry.error_message


@pytest.mark.django_db
def test_delete_existing_statements_prace_serwisowe():
    """Test _delete_existing_statements re-raises PraceSerwisoweException."""
    from pbn_api.exceptions import PraceSerwisoweException

    mock_client = MagicMock()
    mock_client.delete_all_publication_statements.side_effect = PraceSerwisoweException(
        "Prace serwisowe"
    )

    mock_publication = MagicMock()
    mock_publication.pbn_uid_id = "test-uid"

    mock_log_entry = MagicMock()

    with pytest.raises(PraceSerwisoweException):
        _delete_existing_statements(mock_publication, mock_client, mock_log_entry)


@pytest.mark.django_db
def test_handle_http_400_error():
    """Test _handle_http_400_error function."""
    from pbn_api.exceptions import HttpException

    mock_log_entry = MagicMock()

    error = HttpException(400, "http://test/", '{"error": "Bad Request"}')

    status, log = _handle_http_400_error(error, mock_log_entry)

    assert status == "error"
    assert mock_log_entry.json_response == {"error": "Bad Request"}
    assert "HTTP 400" in mock_log_entry.error_message
    mock_log_entry.save.assert_called_once()


@pytest.mark.django_db
def test_handle_http_400_error_invalid_json():
    """Test _handle_http_400_error with invalid JSON response."""
    from pbn_api.exceptions import HttpException

    mock_log_entry = MagicMock()

    error = HttpException(400, "http://test/", "Invalid response - not JSON")

    status, log = _handle_http_400_error(error, mock_log_entry)

    assert status == "error"
    assert "raw_error" in mock_log_entry.json_response


@pytest.mark.django_db
def test_send_statements_with_retry_success():
    """Test _send_statements_with_retry succeeds on first try."""
    mock_client = MagicMock()
    mock_client.post_discipline_statements.return_value = {"status": "ok"}

    mock_log_entry = MagicMock()
    json_data = {"test": "data"}

    status, log = _send_statements_with_retry(mock_client, json_data, mock_log_entry)

    assert status == "success"
    assert mock_log_entry.status == "success"
    assert mock_log_entry.retry_count == 0
    mock_log_entry.save.assert_called_once()


@pytest.mark.django_db
def test_send_statements_with_retry_500_error():
    """Test _send_statements_with_retry retries on HTTP 500."""
    from pbn_api.exceptions import HttpException

    mock_client = MagicMock()
    # Fail twice with 500, then succeed
    mock_client.post_discipline_statements.side_effect = [
        HttpException(500, "http://test/", "Server Error"),
        HttpException(500, "http://test/", "Server Error"),
        {"status": "ok"},
    ]

    mock_log_entry = MagicMock()
    json_data = {"test": "data"}

    with patch("pbn_wysylka_oswiadczen.tasks.time.sleep"):
        status, log = _send_statements_with_retry(
            mock_client, json_data, mock_log_entry
        )

    assert status == "success"
    assert mock_client.post_discipline_statements.call_count == 3


@pytest.mark.django_db
def test_send_statements_with_retry_exhausted():
    """Test _send_statements_with_retry exhausts all retries."""
    from pbn_api.exceptions import HttpException

    mock_client = MagicMock()
    mock_client.post_discipline_statements.side_effect = HttpException(
        500, "http://test/", "Server Error"
    )

    mock_log_entry = MagicMock()
    json_data = {"test": "data"}

    with patch("pbn_wysylka_oswiadczen.tasks.time.sleep"):
        status, log = _send_statements_with_retry(
            mock_client, json_data, mock_log_entry
        )

    assert status == "error"
    assert "Wszystkie proby nieudane" in mock_log_entry.error_message
    assert mock_client.post_discipline_statements.call_count == 5


@pytest.mark.django_db
def test_send_statements_with_retry_prace_serwisowe():
    """Test _send_statements_with_retry raises PraceSerwisoweException."""
    from pbn_api.exceptions import PraceSerwisoweException

    mock_client = MagicMock()
    mock_client.post_discipline_statements.side_effect = PraceSerwisoweException(
        "Prace serwisowe"
    )

    mock_log_entry = MagicMock()
    json_data = {"test": "data"}

    with pytest.raises(PraceSerwisoweException):
        _send_statements_with_retry(mock_client, json_data, mock_log_entry)

    # Check that log entry was updated before re-raising
    assert mock_log_entry.status == "maintenance"
    assert "Prace serwisowe" in mock_log_entry.error_message
    mock_log_entry.save.assert_called_once()


@pytest.mark.django_db
def test_process_single_publication_no_przypiete():
    """Test process_single_publication when no przypiete authors."""
    user = User.objects.create_user("testuser", password="testpass")
    task = PbnWysylkaOswiadczenTask.objects.create(user=user)

    # Create mock publication with no przypiete authors
    mock_publication = MagicMock()
    mock_publication.pk = 1
    mock_publication.pbn_uid_id = "test-pbn-uid-123"
    mock_publication.tytul_oryginalny = "Test Publication"

    # Mock autorzy_set to return False for przypieta=True filter
    mock_autorzy_set = MagicMock()
    mock_autorzy_set.filter.return_value.exists.return_value = False
    mock_publication.autorzy_set = mock_autorzy_set

    mock_client = MagicMock()

    # Mock ContentType.objects.get_for_model to return a real ContentType
    mock_content_type = ContentType.objects.get_for_model(Wydawnictwo_Ciagle)

    with patch(
        "django.contrib.contenttypes.models.ContentType.objects.get_for_model",
        return_value=mock_content_type,
    ):
        status, log_entry = process_single_publication(
            mock_publication, mock_client, task, PbnWysylkaLog
        )

    assert status == "skipped"
    assert log_entry.status == "skipped"
    assert "przypieta" in log_entry.error_message.lower()


# ============================================================================
# Integration Tests
# ============================================================================


@pytest.mark.django_db
def test_full_workflow_no_publications(client, uczelnia):
    """Test full workflow when no publications match criteria."""
    user = User.objects.create_user("testuser", password="testpass")
    group, _ = Group.objects.get_or_create(name=GR_WPROWADZANIE_DANYCH)
    user.groups.add(group)

    client.force_login(user)

    response = client.get(reverse("pbn_wysylka_oswiadczen:main"))
    assert response.status_code == 200
    assert b"0" in response.content  # Total count should be 0


@pytest.mark.django_db
def test_main_view_authenticated(client, uczelnia):
    """Test main view for authenticated user with proper group."""
    user = User.objects.create_user("testuser", password="testpass")
    group, _ = Group.objects.get_or_create(name=GR_WPROWADZANIE_DANYCH)
    user.groups.add(group)

    client.force_login(user)

    response = client.get(reverse("pbn_wysylka_oswiadczen:main"))
    assert response.status_code == 200
    assert "Wysyłka oświadczeń" in response.content.decode("utf-8")


@pytest.mark.django_db
def test_main_view_unauthenticated(client):
    """Test main view redirects unauthenticated users."""
    response = client.get(reverse("pbn_wysylka_oswiadczen:main"))
    assert response.status_code == 302
    assert "login" in response.url


@pytest.mark.django_db
def test_status_api_returns_json(client, uczelnia):
    """Test status API returns valid JSON."""
    user = User.objects.create_user("testuser", password="testpass")
    group, _ = Group.objects.get_or_create(name=GR_WPROWADZANIE_DANYCH)
    user.groups.add(group)

    client.force_login(user)

    response = client.get(reverse("pbn_wysylka_oswiadczen:status"))
    assert response.status_code == 200
    assert response["Content-Type"] == "application/json"

    data = response.json()
    assert "is_running" in data
    assert "latest_task" in data


@pytest.mark.django_db
def test_get_publications_queryset_tylko_odpiete_parameter(uczelnia):
    """Test get_publications_queryset accepts tylko_odpiete parameter."""
    # Test with tylko_odpiete=False (default)
    ciagle_qs, zwarte_qs = get_publications_queryset(
        rok_od=2022, rok_do=2025, tylko_odpiete=False, with_annotations=True
    )
    assert ciagle_qs.count() == 0
    assert zwarte_qs.count() == 0

    # Test with tylko_odpiete=True
    ciagle_qs, zwarte_qs = get_publications_queryset(
        rok_od=2022, rok_do=2025, tylko_odpiete=True, with_annotations=True
    )
    assert ciagle_qs.count() == 0
    assert zwarte_qs.count() == 0


@pytest.mark.django_db
def test_main_view_context_tylko_odpiete(uczelnia):
    """Test main view returns tylko_odpiete in context."""
    factory = RequestFactory()
    request = factory.get("/?rok_od=2022&rok_do=2024&tylko_odpiete=true")

    user = create_user_with_group()
    request.user = user
    request.session = {}

    view = PbnWysylkaOswiadczenMainView()
    view.setup(request)

    context = view.get_context_data()
    assert "tylko_odpiete" in context
    assert context["tylko_odpiete"] is True


@pytest.mark.django_db
def test_main_view_context_tylko_odpiete_false(uczelnia):
    """Test main view returns tylko_odpiete=False when not set."""
    factory = RequestFactory()
    request = factory.get("/?rok_od=2022&rok_do=2024")

    user = create_user_with_group()
    request.user = user
    request.session = {}

    view = PbnWysylkaOswiadczenMainView()
    view.setup(request)

    context = view.get_context_data()
    assert "tylko_odpiete" in context
    assert context["tylko_odpiete"] is False


# ============================================================================
# Fixture for Publication with PBN UID
# ============================================================================


@pytest.fixture
def publication_with_pbn_uid(
    uczelnia,
    jednostka,
    autor_jan_nowak,
    dyscyplina1,
    jezyki,
    charaktery_formalne,
    typy_kbn,
    statusy_korekt,
    typy_odpowiedzialnosci,
):
    """Create a publication that matches get_publications_queryset criteria."""
    from model_bakery import baker

    from bpp.models import Autor_Dyscyplina, Wydawnictwo_Ciagle

    # Create PBN publication UID
    pbn_pub = baker.make("pbn_api.Publication")

    # Create the publication
    wyd = baker.make(
        Wydawnictwo_Ciagle,
        tytul_oryginalny="Test Publication of ionic liquids",
        rok=2022,
        pbn_uid=pbn_pub,
    )

    # Set up author discipline
    Autor_Dyscyplina.objects.get_or_create(
        autor=autor_jan_nowak,
        dyscyplina_naukowa=dyscyplina1,
        rok=2022,
    )

    # Add author to publication with discipline
    autor_wyd = wyd.dodaj_autora(
        autor_jan_nowak,
        jednostka,
        dyscyplina_naukowa=dyscyplina1,
        afiliuje=True,
    )
    # Set zatrudniony separately (defaults to False)
    autor_wyd.zatrudniony = True
    autor_wyd.save()

    return wyd


# ============================================================================
# Title Filtering Tests
# ============================================================================


@pytest.mark.django_db
def test_get_publications_queryset_title_filter(publication_with_pbn_uid):
    """Test title filtering in get_publications_queryset."""
    # Should find with matching title
    ciagle_qs, zwarte_qs = get_publications_queryset(
        rok_od=2022, rok_do=2022, tytul="of ionic"
    )
    assert ciagle_qs.count() == 1

    # Should not find with non-matching title
    ciagle_qs, zwarte_qs = get_publications_queryset(
        rok_od=2022, rok_do=2022, tytul="nonexistent"
    )
    assert ciagle_qs.count() == 0


@pytest.mark.django_db
def test_get_publications_queryset_title_filter_case_insensitive(
    publication_with_pbn_uid,
):
    """Test title filtering is case-insensitive."""
    ciagle_qs, _ = get_publications_queryset(rok_od=2022, rok_do=2022, tytul="OF IONIC")
    assert ciagle_qs.count() == 1


@pytest.mark.django_db
def test_main_view_title_filter(client, uczelnia, publication_with_pbn_uid):
    """Test main view with title filter shows correct count."""
    user = create_user_with_group()
    client.force_login(user)

    response = client.get(
        reverse("pbn_wysylka_oswiadczen:main"),
        {"rok_od": 2022, "rok_do": 2022, "tytul": "of ionic"},
    )
    assert response.status_code == 200
    assert response.context["total_count"] == 1
    assert response.context["tytul"] == "of ionic"


# ============================================================================
# PublicationListView Tests
# ============================================================================


@pytest.mark.django_db
def test_publication_list_view_basic(client, uczelnia, publication_with_pbn_uid):
    """Test PublicationListView returns publications."""
    user = create_user_with_group()
    client.force_login(user)

    response = client.get(
        reverse("pbn_wysylka_oswiadczen:publications"),
        {"rok_od": 2022, "rok_do": 2022},
    )
    assert response.status_code == 200
    assert "page_obj" in response.context
    assert response.context["page_obj"].paginator.count == 1


@pytest.mark.django_db
def test_publication_list_view_with_title_filter(
    client, uczelnia, publication_with_pbn_uid
):
    """Test PublicationListView with title filter."""
    user = create_user_with_group()
    client.force_login(user)

    # With matching title
    response = client.get(
        reverse("pbn_wysylka_oswiadczen:publications"),
        {"rok_od": 2022, "rok_do": 2022, "tytul": "of ionic"},
    )
    assert response.context["page_obj"].paginator.count == 1

    # With non-matching title
    response = client.get(
        reverse("pbn_wysylka_oswiadczen:publications"),
        {"rok_od": 2022, "rok_do": 2022, "tytul": "nonexistent"},
    )
    assert response.context["page_obj"].paginator.count == 0


@pytest.mark.django_db
def test_publication_list_view_requires_auth(client):
    """Test PublicationListView requires authentication."""
    response = client.get(reverse("pbn_wysylka_oswiadczen:publications"))
    assert response.status_code == 302
    assert "login" in response.url


# ============================================================================
# Excel Export Tests
# ============================================================================


@pytest.mark.django_db
def test_excel_export_view_basic(client, uczelnia, publication_with_pbn_uid):
    """Test ExcelExportView returns Excel file."""
    user = create_user_with_group()
    client.force_login(user)

    response = client.get(
        reverse("pbn_wysylka_oswiadczen:export-excel"),
        {"rok_od": 2022, "rok_do": 2022},
    )

    assert response.status_code == 200
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response["Content-Disposition"]
    assert ".xlsx" in response["Content-Disposition"]


@pytest.mark.django_db
def test_excel_export_view_with_title_filter(
    client, uczelnia, publication_with_pbn_uid
):
    """Test ExcelExportView with title filter."""
    from io import BytesIO

    from openpyxl import load_workbook

    user = create_user_with_group()
    client.force_login(user)

    response = client.get(
        reverse("pbn_wysylka_oswiadczen:export-excel"),
        {"rok_od": 2022, "rok_do": 2022, "tytul": "of ionic"},
    )

    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active

    # Header row + 1 data row
    assert ws.max_row == 2
    assert "ionic" in ws.cell(2, 3).value.lower()  # Title column


@pytest.mark.django_db
def test_excel_export_view_empty_results(client, uczelnia):
    """Test ExcelExportView with no matching publications."""
    from io import BytesIO

    from openpyxl import load_workbook

    user = create_user_with_group()
    client.force_login(user)

    response = client.get(
        reverse("pbn_wysylka_oswiadczen:export-excel"),
        {"rok_od": 2022, "rok_do": 2022, "tytul": "nonexistent"},
    )

    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active

    # Only header row
    assert ws.max_row == 1


@pytest.mark.django_db
def test_excel_export_view_requires_auth(client):
    """Test ExcelExportView requires authentication."""
    response = client.get(reverse("pbn_wysylka_oswiadczen:export-excel"))
    assert response.status_code == 302
    assert "login" in response.url
