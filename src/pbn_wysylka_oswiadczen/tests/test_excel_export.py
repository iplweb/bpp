"""Excel export tests for pbn_wysylka_oswiadczen app."""

import pytest
from django.urls import reverse

from ._helpers import create_user_with_group


@pytest.mark.django_db
def test_excel_export_view_basic(client, uczelnia, publication_with_pbn_uid):
    """Test ExcelExportView returns Excel file."""
    user = create_user_with_group()
    client.force_login(user)

    response = client.get(
        reverse("pbn_wysylka_oswiadczen:export-excel"),
        {"rok_od": 2022, "rok_do": 2022},
    )

    assert response.status_code == 200
    assert (
        response["Content-Type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response["Content-Disposition"]
    assert ".xlsx" in response["Content-Disposition"]


@pytest.mark.django_db
def test_excel_export_view_with_title_filter(
    client, uczelnia, publication_with_pbn_uid
):
    """Test ExcelExportView with title filter."""
    from io import BytesIO

    from openpyxl import load_workbook

    user = create_user_with_group()
    client.force_login(user)

    response = client.get(
        reverse("pbn_wysylka_oswiadczen:export-excel"),
        {"rok_od": 2022, "rok_do": 2022, "tytul": "of ionic"},
    )

    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active

    # Header row + 1 data row
    assert ws.max_row == 2
    assert "ionic" in ws.cell(2, 3).value.lower()  # Title column


@pytest.mark.django_db
def test_excel_export_view_empty_results(client, uczelnia):
    """Test ExcelExportView with no matching publications."""
    from io import BytesIO

    from openpyxl import load_workbook

    user = create_user_with_group()
    client.force_login(user)

    response = client.get(
        reverse("pbn_wysylka_oswiadczen:export-excel"),
        {"rok_od": 2022, "rok_do": 2022, "tytul": "nonexistent"},
    )

    assert response.status_code == 200

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active

    # Only header row
    assert ws.max_row == 1


@pytest.mark.django_db
def test_excel_export_view_requires_auth(client):
    """Test ExcelExportView requires authentication."""
    response = client.get(reverse("pbn_wysylka_oswiadczen:export-excel"))
    assert response.status_code == 302
    assert "login" in response.url
