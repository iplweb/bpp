"""Zawężanie rozbieżności do uczelni oglądającego (multi-hosted, read-side).

Rekord (``Wydawnictwo_Ciagle``) należy do uczelni, gdy którakolwiek jednostka
zapisana na autorstwie należy do tej uczelni — reguła wspólna ze stroną główną
(``scope_rekord_do_uczelni``). W single-host (jedna uczelnia) zawężenie MUSI
być no-opem (guard ``tylko_jedna_uczelnia``).
"""

from io import BytesIO

import pytest
from model_bakery import baker
from openpyxl import load_workbook

from fixtures.conftest_multisite import make_request_for_site
from rozbieznosci.views import RozbieznosciExportView, RozbieznosciView


def _make_rozbieznosc_if(jednostka, tytul, rok=2023):
    """Wydawnictwo_Ciagle będące rozbieżnością IF, przypisane do ``jednostka``.

    Praca ma ``impact_factor`` różny od źródła (``Punktacja_Zrodla``) za ten
    sam rok → trafia do bazowego querysetu metryki ``if`` w trybie standardowym.
    """
    zrodlo = baker.make("bpp.Zrodlo")
    baker.make("bpp.Punktacja_Zrodla", zrodlo=zrodlo, rok=rok, impact_factor="2.500")
    wc = baker.make(
        "bpp.Wydawnictwo_Ciagle",
        zrodlo=zrodlo,
        rok=rok,
        impact_factor="1.500",
        tytul_oryginalny=tytul,
    )
    autor = baker.make("bpp.Autor")
    baker.make(
        "bpp.Wydawnictwo_Ciagle_Autor",
        rekord=wc,
        autor=autor,
        jednostka=jednostka,
    )
    return wc


def _view_queryset(view_cls, request, metryka="if"):
    view = view_cls()
    view.setup(request, metryka=metryka)
    return view


@pytest.mark.django_db
def test_multihost_view_zaweza_do_uczelni_ogladajacego(
    settings, site1, jednostka_uczelnia1, jednostka_uczelnia2
):
    """Multi-host: widok dla uczelni A pokazuje rekord A, ukrywa rekord B."""
    settings.ALLOWED_HOSTS = ["*"]
    wc_a = _make_rozbieznosc_if(jednostka_uczelnia1, "Praca uczelni A")
    wc_b = _make_rozbieznosc_if(jednostka_uczelnia2, "Praca uczelni B")

    request = make_request_for_site(site1, path="/rozbieznosci/if/")
    view = _view_queryset(RozbieznosciView, request)
    pks = set(view.get_queryset().values_list("pk", flat=True))

    assert wc_a.pk in pks
    assert wc_b.pk not in pks


@pytest.mark.django_db
def test_singlehost_view_no_op(uczelnia, jednostka):
    """Single-host: jedna uczelnia → filtr no-op, rozbieżność widoczna.

    ``jednostka`` fixture należy do jedynej ``uczelnia`` — brak drugiej uczelni,
    więc ``tylko_jedna_uczelnia()`` = True i zawężenie jest pomijane.
    """
    from bpp.models import Uczelnia

    assert Uczelnia.objects.count() == 1
    wc = _make_rozbieznosc_if(jednostka, "Praca single-host")

    request = make_request_for_site(uczelnia.site, path="/rozbieznosci/if/")
    view = _view_queryset(RozbieznosciView, request)
    pks = set(view.get_queryset().values_list("pk", flat=True))

    assert wc.pk in pks


@pytest.mark.django_db
def test_multihost_view_bez_uczelni_nie_rzuca(jednostka_uczelnia1, jednostka_uczelnia2):
    """Brak mapowania Site→Uczelnia (uczelnia=None) → fail-open, bez wyjątku."""
    from django.test import RequestFactory

    wc_a = _make_rozbieznosc_if(jednostka_uczelnia1, "Praca A")
    wc_b = _make_rozbieznosc_if(jednostka_uczelnia2, "Praca B")

    request = RequestFactory().get("/rozbieznosci/if/")
    request._uczelnia = None
    view = _view_queryset(RozbieznosciView, request)
    pks = set(view.get_queryset().values_list("pk", flat=True))

    # Fail-open: brak zawężenia, oba rekordy widoczne.
    assert wc_a.pk in pks
    assert wc_b.pk in pks


@pytest.mark.django_db
def test_multihost_export_zaweza_do_uczelni(
    settings, site1, jednostka_uczelnia1, jednostka_uczelnia2
):
    """Multi-host: eksport XLSX dla uczelni A zawiera pracę A, nie zawiera B."""
    settings.ALLOWED_HOSTS = ["*"]
    _make_rozbieznosc_if(jednostka_uczelnia1, "Eksport praca A")
    _make_rozbieznosc_if(jednostka_uczelnia2, "Eksport praca B")

    request = make_request_for_site(site1, path="/rozbieznosci/if/export/")
    view = RozbieznosciExportView()
    view.setup(request, metryka="if")
    response = view.get(request, metryka="if")

    wb = load_workbook(BytesIO(response.content))
    ws = wb.active
    tytuly = {row[0] for row in ws.iter_rows(min_row=2, values_only=True)}

    assert "Eksport praca A" in tytuly
    assert "Eksport praca B" not in tytuly
