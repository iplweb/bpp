import itertools
from collections import OrderedDict
from decimal import Decimal
from enum import Enum
from typing import Any, Dict, List, Optional, Union

import openpyxl.worksheet.worksheet
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import TableColumn
from unidecode import unidecode

from django.contrib.sites.models import Site

from django.utils.functional import cached_property

from bpp.util import worksheet_columns_autosize, worksheet_create_table


def chunker(n, iterable):
    iterable = iter(iterable)
    while True:
        x = tuple(itertools.islice(iterable, n))
        if not x:
            return
        yield x


class SHUFFLE_TYPE(Enum):
    BEGIN = 1
    MIDDLE = 2
    END = 3
    RANDOM = 4


import random


def shuffle_array(
    array, start, length, no_shuffles=1, shuffle_type=SHUFFLE_TYPE.MIDDLE
):

    first = array[:start]
    second = array[start : start + length]
    third = array[start + length :]

    i = shuffle_type
    if shuffle_type == SHUFFLE_TYPE.RANDOM:
        i = random.randint(1, 3)

    if i == SHUFFLE_TYPE.BEGIN:
        for a in range(no_shuffles):
            random.shuffle(first)
    elif i == SHUFFLE_TYPE.MIDDLE:
        for a in range(no_shuffles):
            random.shuffle(second)
    elif i == SHUFFLE_TYPE.END:
        for a in range(no_shuffles):
            random.shuffle(third)

    return first + second + third


def output_table_to_xlsx(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    title: str,
    headers: List[str],
    dataset: List[List[Any]],
    totals: List[str] = None,
    first_column_url: str = "https://{site_name}/bpp/rekord/",
    autor_column_url: Optional[int] = 4,
    autor_column_link: Optional[int] = 1,
    autosize_columns: List[str] = None,
    column_widths: Dict[int, int] = None,
):
    ws.append(headers)

    table_columns = tuple(
        TableColumn(id=h, name=header) for h, header in enumerate(headers, start=1)
    )

    footer_row = []
    if totals:
        for no, elem in enumerate(table_columns):
            if no == 0:
                footer_row.append("Suma")
                total_column = table_columns[0]
                total_column.totalsRowLabel = footer_row[0]
                continue

            if elem.name in totals:
                count_column = table_columns[no]
                count_column.totalsRowFunction = "sum"
                footer_row.append(f"=SUBTOTAL(109,{title}[{elem.name}])")
                continue

            footer_row.append("")

    for cell in ws[ws.max_row : ws.max_row]:
        cell.font = openpyxl.styles.Font(bold=True)

    first_table_row = ws.max_row

    site_name = Site.objects.first().domain
    url = first_column_url.format(site_name=site_name)
    autor_url = f"https://{site_name}/bpp/autor/"
    for row in dataset:
        ws.append(row)

        # URL dla pierwszej kolumny -- odnośnik do BPP
        id_for_url = row[0].replace(" ", "")
        if id_for_url.startswith("("):
            id_for_url = id_for_url[1:-1]
        ws.cell(row=ws.max_row, column=1).value = '=HYPERLINK("{}", "{}")'.format(
            url + id_for_url + "/", row[0]
        )

        if autor_column_link is not None:
            # Druga kolumna z ID autora -> bpp
            ws.cell(
                row=ws.max_row, column=autor_column_link + 1
            ).value = '=HYPERLINK("{}", "{}")'.format(
                autor_url + str(row[autor_column_link]), row[autor_column_link]
            )

        if autor_column_url is not None:
            # URL dla czwartej kolumny -- odnośnik do pliku autora
            ws.cell(
                row=ws.max_row, column=autor_column_url + 1
            ).value = '=HYPERLINK("{}", "{}")'.format(
                string2fn(row[autor_column_url]) + ".xlsx", row[autor_column_url]
            )

    if footer_row:
        ws.append(footer_row)

    if dataset:
        worksheet_create_table(
            ws,
            title=title,
            first_table_row=first_table_row,
            totals=totals,
            table_columns=table_columns,
        )

    if totals is None:
        totals = []
    for elem in totals:
        letter = get_column_letter(headers.index(elem) + 1)
        for row in range(2, ws.max_row):
            ws[f"{letter}{row}"].number_format = "#,####0.0000"

        # Ustaw automatyczny rozmiar
        ws.column_dimensions[letter].bestFit = True

    dont_resize_those_columns = []
    for ncol, col in enumerate(ws.columns):
        if headers[ncol] in totals:
            dont_resize_those_columns.append(ncol)

    worksheet_columns_autosize(ws, dont_resize_those_columns=dont_resize_those_columns)


def string2fn(s):
    return s.strip().replace(" ", "_").replace("*", "x").replace("-", "_")


def autor2fn(autor):
    return string2fn(f"{autor.nazwisko}_{autor.imiona}".lower())


def normalize_xlsx_header_column_name(s):
    if s is None:
        return
    s = str(s).replace(".", " ").replace("-", " ").replace("/", " ").strip()
    if not s:
        return

    while s.find("  ") >= 0:
        s = s.replace("  ", " ")

    return unidecode(s.strip()).replace(" ", "_").lower()


def find_header_row(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    header_row: List[str],
    max_header_row=100,
) -> Union[None, int]:
    """
    Poszukuje wierwsza nagłówka w skoroszycie ``worksheet``.

    :param max_header_row: maksymalny wiersz, w którym poszukujemy nagłówka
    :return: wiersza nagłówkowego - jeżeli znaleziony
    """

    normalized_header_row = [normalize_xlsx_header_column_name(v) for v in header_row]
    max_header_row = min(worksheet.max_row, max_header_row)

    for nrow, row in enumerate(worksheet[1:max_header_row], start=1):
        normalized_cell_values = [
            normalize_xlsx_header_column_name(cell.value)
            for cell in row[: len(header_row)]
        ]
        if normalized_cell_values == normalized_header_row:
            return nrow


class InputXLSX:
    def __init__(self, fn, header_cols):
        self.fn = fn
        self.header_cols = header_cols

    @cached_property
    def workbook(self):
        return openpyxl.load_workbook(self.fn)

    @cached_property
    def worksheet(self):
        return self.workbook.worksheets[0]

    @cached_property
    def header_normalized(self):
        return [normalize_xlsx_header_column_name(col) for col in self.header_cols]

    @cached_property
    def header_row(self):
        return find_header_row(self.worksheet, self.header_cols)

    def rows_as_list(self):
        if self.header_row is None:
            raise ValueError("Brak wiersza nagłówka w pliku XLS.")

        min_row = self.header_row + 1
        max_row = self.worksheet.max_row
        for row in self.worksheet[min_row:max_row]:
            yield [cell.value for cell in row]

    def rows_as_dict(self):
        keys = self.header_normalized[:]
        keys.append("__nrow__")

        for nrow, row in enumerate(self.rows_as_list(), start=self.header_row + 1):
            row = OrderedDict(zip(keys, row))
            row["__nrow__"] = nrow
            yield row


def float_or_string_or_int_or_none_to_decimal(i, decimal_places=4):
    if i is None:
        return i
    if type(i) in [int, Decimal, str]:
        return Decimal(i)
    if isinstance(i, float):
        return Decimal(f"%.{decimal_places}f" % round(i, decimal_places))
    raise NotImplementedError(f"Type {type(i)} not supported.")
