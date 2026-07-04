"""TDD widoków deduplikatora źródeł: lista par + start skanu + akcje statusu."""

from datetime import timedelta

import pytest
from django.urls import reverse
from django.utils import timezone

from deduplikator_zrodel.models import (
    IgnoredSource,
    NotADuplicate,
    ScanZrodelForDuplicates,
    SourceDuplicateCandidate,
)

LIST_URL = reverse("deduplikator_zrodel:duplicate_sources")


# --------------------------------------------------------------------------- #
# Lista par                                                                    #
# --------------------------------------------------------------------------- #


@pytest.mark.django_db
def test_list_shows_candidates_of_latest_completed_scan(
    admin_client, admin_user, make_zrodlo, completed_scan, make_candidate
):
    old = completed_scan()
    a, b = make_zrodlo(nazwa="STARE-A"), make_zrodlo(nazwa="STARE-B")
    make_candidate(old, a, b)

    new = completed_scan()
    c, d = make_zrodlo(nazwa="NOWE-C"), make_zrodlo(nazwa="NOWE-D")
    make_candidate(new, c, d)

    resp = admin_client.get(LIST_URL)
    assert resp.status_code == 200
    content = resp.content.decode()
    assert "NOWE-C" in content
    assert "STARE-A" not in content


@pytest.mark.django_db
def test_list_excludes_ignored_and_notaduplicate(
    admin_client, make_zrodlo, completed_scan, make_candidate
):
    scan = completed_scan()
    a, b = make_zrodlo(nazwa="PARA-IGNORED-A"), make_zrodlo(nazwa="PARA-IGNORED-B")
    c, d = make_zrodlo(nazwa="PARA-ND-C"), make_zrodlo(nazwa="PARA-ND-D")
    e, f = make_zrodlo(nazwa="PARA-OK-E"), make_zrodlo(nazwa="PARA-OK-F")
    make_candidate(scan, a, b)
    make_candidate(scan, c, d)
    make_candidate(scan, e, f)

    IgnoredSource.objects.create(zrodlo=a)
    NotADuplicate.objects.create(zrodlo=c, duplikat=d)
    NotADuplicate.objects.create(zrodlo=d, duplikat=c)

    content = admin_client.get(LIST_URL).content.decode()
    assert "PARA-OK-E" in content
    assert "PARA-IGNORED-A" not in content
    assert "PARA-ND-C" not in content


@pytest.mark.django_db
def test_list_status_filter_hides_skipped_by_default(
    admin_client, make_zrodlo, completed_scan, make_candidate
):
    scan = completed_scan()
    a, b = make_zrodlo(nazwa="PENDING-A"), make_zrodlo(nazwa="PENDING-B")
    c, d = make_zrodlo(nazwa="SKIPPED-C"), make_zrodlo(nazwa="SKIPPED-D")
    make_candidate(scan, a, b)
    make_candidate(scan, c, d, status=SourceDuplicateCandidate.Status.SKIPPED)

    default = admin_client.get(LIST_URL).content.decode()
    assert "PENDING-A" in default
    assert "SKIPPED-C" not in default

    skipped = admin_client.get(LIST_URL, {"status": "skipped"}).content.decode()
    assert "SKIPPED-C" in skipped
    assert "PENDING-A" not in skipped


# --------------------------------------------------------------------------- #
# Start skanu                                                                  #
# --------------------------------------------------------------------------- #


@pytest.mark.django_db
def test_start_scan_anonymous_is_rejected(client):
    url = reverse("deduplikator_zrodel:start_scan")
    resp = client.post(url)
    assert resp.status_code in (301, 302)
    assert ScanZrodelForDuplicates.objects.count() == 0


@pytest.mark.django_db
def test_start_scan_creates_op_and_redirects_to_live(admin_client):
    url = reverse("deduplikator_zrodel:start_scan")
    resp = admin_client.post(url)
    assert ScanZrodelForDuplicates.objects.count() == 1
    op = ScanZrodelForDuplicates.objects.get()
    assert resp.status_code == 302
    assert resp.url == op.get_absolute_url()


@pytest.mark.django_db
def test_start_scan_guard_blocks_second_fresh_scan(admin_client, admin_user):
    ScanZrodelForDuplicates.objects.create(owner=admin_user, started_on=timezone.now())
    resp = admin_client.post(reverse("deduplikator_zrodel:start_scan"))
    # Nie tworzy drugiego skanu; wraca na listę.
    assert ScanZrodelForDuplicates.objects.count() == 1
    assert resp.status_code == 302
    assert resp.url == LIST_URL


@pytest.mark.django_db
def test_start_scan_stale_scan_does_not_block(admin_client, admin_user):
    stale = ScanZrodelForDuplicates.objects.create(
        owner=admin_user, started_on=timezone.now() - timedelta(hours=5)
    )
    # created_on jest auto_now_add — przesuwamy w przeszłość przez update().
    ScanZrodelForDuplicates.objects.filter(pk=stale.pk).update(
        created_on=timezone.now() - timedelta(hours=5)
    )
    resp = admin_client.post(reverse("deduplikator_zrodel:start_scan"))
    assert ScanZrodelForDuplicates.objects.count() == 2
    assert resp.status_code == 302


# --------------------------------------------------------------------------- #
# Akcje statusu kandydata                                                      #
# --------------------------------------------------------------------------- #


@pytest.mark.django_db
def test_skip_candidate_sets_status(
    admin_client, make_zrodlo, completed_scan, make_candidate
):
    scan = completed_scan()
    cand = make_candidate(scan, make_zrodlo(), make_zrodlo())
    resp = admin_client.post(
        reverse("deduplikator_zrodel:skip_candidate"), {"candidate_id": cand.pk}
    )
    assert resp.status_code == 302
    cand.refresh_from_db()
    assert cand.status == SourceDuplicateCandidate.Status.SKIPPED


@pytest.mark.django_db
def test_unskip_candidate_restores_pending(
    admin_client, make_zrodlo, completed_scan, make_candidate
):
    scan = completed_scan()
    cand = make_candidate(
        scan,
        make_zrodlo(),
        make_zrodlo(),
        status=SourceDuplicateCandidate.Status.SKIPPED,
    )
    admin_client.post(
        reverse("deduplikator_zrodel:unskip_candidate"), {"candidate_id": cand.pk}
    )
    cand.refresh_from_db()
    assert cand.status == SourceDuplicateCandidate.Status.PENDING


@pytest.mark.django_db
def test_reset_skipped_unskips_all(
    admin_client, make_zrodlo, completed_scan, make_candidate
):
    scan = completed_scan()
    c1 = make_candidate(
        scan,
        make_zrodlo(),
        make_zrodlo(),
        status=SourceDuplicateCandidate.Status.SKIPPED,
    )
    c2 = make_candidate(
        scan,
        make_zrodlo(),
        make_zrodlo(),
        status=SourceDuplicateCandidate.Status.SKIPPED,
    )
    admin_client.post(reverse("deduplikator_zrodel:reset_skipped"))
    c1.refresh_from_db()
    c2.refresh_from_db()
    assert c1.status == SourceDuplicateCandidate.Status.PENDING
    assert c2.status == SourceDuplicateCandidate.Status.PENDING


@pytest.mark.django_db
def test_download_xlsx_contains_candidates(
    admin_client, make_zrodlo, completed_scan, make_candidate
):
    scan = completed_scan()
    a = make_zrodlo(nazwa="XLSX-MAIN", issn="1111-2222")
    b = make_zrodlo(nazwa="XLSX-DUP", issn="1111-2222")
    make_candidate(scan, a, b, confidence_score=175)

    resp = admin_client.get(reverse("deduplikator_zrodel:download_duplicates_xlsx"))
    assert resp.status_code == 200
    assert "spreadsheetml" in resp["Content-Type"]

    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(resp.content))
    cells = [cell.value for row in wb.active.iter_rows() for cell in row]
    assert "XLSX-MAIN" in cells
    assert 175 in cells


@pytest.mark.django_db
def test_start_scan_eager_runs_and_live_page_renders(admin_client, make_zrodlo):
    """E2E pod RUNNER=eager: POST uruchamia skan synchronicznie, strona live
    (host-template + {% live_operation %}) renderuje się, a para trafia na listę."""
    from model_bakery import baker

    from pbn_api.models import Journal

    make_zrodlo(
        nazwa="Acta E2E",
        skrot="AE",
        issn="9999-8888",
        pbn_uid=baker.make(Journal, mniswId=111),
    )
    make_zrodlo(nazwa="Acta E2E", skrot="AE", issn="9999-8888")

    admin_client.post(reverse("deduplikator_zrodel:start_scan"))
    op = ScanZrodelForDuplicates.objects.get()
    op.refresh_from_db()
    assert op.get_state() == "FINISHED_OK"
    assert op.duplicates_found == 1

    live = admin_client.get(op.get_absolute_url())
    assert live.status_code == 200

    listing = admin_client.get(LIST_URL).content.decode()
    assert "Acta E2E" in listing


@pytest.mark.django_db
def test_list_excludes_candidate_after_remap(
    admin_client, make_zrodlo, completed_scan, make_candidate
):
    """Po przemapowaniu duplikat ma 0 publikacji → para znika z listy
    (przemapuj_zrodlo przenosi publikacje, nie kasuje źródła)."""
    scan = completed_scan()
    main = make_zrodlo(nazwa="REMAP-MAIN")  # zachowuje publikacje
    dup = make_zrodlo(nazwa="REMAP-DUP", z_publikacja=False)  # 0 publikacji
    make_candidate(scan, main, dup)

    content = admin_client.get(LIST_URL).content.decode()
    assert "REMAP-DUP" not in content


@pytest.mark.django_db
def test_live_page_shows_cancel_form_while_running(admin_client, admin_user):
    """Strona live pokazuje pewny (plain POST) formularz Anuluj, gdy skan trwa."""
    op = ScanZrodelForDuplicates.objects.create(
        owner=admin_user, started_on=timezone.now()
    )
    resp = admin_client.get(op.get_absolute_url())
    assert resp.status_code == 200
    content = resp.content.decode()
    assert "Anuluj skanowanie" in content
    cancel_url = reverse(
        "liveops:cancel", kwargs={"op_type": op.op_type_key(), "pk": op.pk}
    )
    assert cancel_url in content
