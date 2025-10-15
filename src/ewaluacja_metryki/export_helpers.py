"""Helper functions for XLSX export views in ewaluacja_metryki."""


def export_globalne_stats(ws, header_font, header_fill, header_alignment):
    """Export global statistics table."""
    from django.db.models import Avg, Count, Sum

    from .models import MetrykaAutora

    ws.title = "Statystyki globalne"
    wszystkie = MetrykaAutora.objects.all()
    stats = wszystkie.aggregate(
        liczba_wierszy=Count("id"),
        liczba_autorow=Count("autor", distinct=True),
        srednia_wykorzystania=Avg("procent_wykorzystania_slotow"),
        srednia_pkd_slot=Avg("srednia_za_slot_nazbierana"),
        suma_punktow=Sum("punkty_nazbierane"),
        suma_slotow=Sum("slot_nazbierany"),
    )

    headers = ["Metryka", "Wartość"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    data = [
        ("Liczba wierszy", stats["liczba_wierszy"] or 0),
        ("Liczba autorów", stats["liczba_autorow"] or 0),
        (
            "Średnie wykorzystanie slotów (%)",
            f"{stats['srednia_wykorzystania'] or 0:.1f}",
        ),
        ("Średnia PKDaut/slot", f"{stats['srednia_pkd_slot'] or 0:.2f}"),
        ("Suma punktów", f"{stats['suma_punktow'] or 0:.0f}"),
        ("Suma slotów", f"{stats['suma_slotow'] or 0:.0f}"),
    ]

    for row_idx, (label, value) in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)


def get_rodzaj_autora_display(rodzaj_autora):
    """Get display value for rodzaj_autora (author type)."""
    from ewaluacja_common.models import Rodzaj_Autora

    if rodzaj_autora == " ":
        return "Brak danych"

    try:
        rodzaj = Rodzaj_Autora.objects.get(skrot=rodzaj_autora)
        return rodzaj.nazwa
    except Rodzaj_Autora.DoesNotExist:
        return rodzaj_autora or "-"


def write_author_metric_row(ws, row_idx, metryka, lp):
    """Write a single author metric row with common fields."""
    ws.cell(row=row_idx, column=1, value=lp)
    ws.cell(row=row_idx, column=2, value=str(metryka.autor))
    ws.cell(
        row=row_idx, column=3, value=get_rodzaj_autora_display(metryka.rodzaj_autora)
    )
    ws.cell(row=row_idx, column=4, value=metryka.autor.system_kadrowy_id or "")
    ws.cell(row=row_idx, column=5, value=metryka.autor.orcid or "")
    ws.cell(row=row_idx, column=6, value=metryka.autor.pbn_uid_id or "")
    ws.cell(
        row=row_idx,
        column=7,
        value=metryka.jednostka.nazwa if metryka.jednostka else "-",
    )
    ws.cell(
        row=row_idx,
        column=8,
        value=metryka.dyscyplina_naukowa.nazwa if metryka.dyscyplina_naukowa else "-",
    )
    ws.cell(row=row_idx, column=9, value=float(metryka.slot_nazbierany))
    ws.cell(row=row_idx, column=10, value=float(metryka.procent_wykorzystania_slotow))
    ws.cell(row=row_idx, column=11, value=float(metryka.srednia_za_slot_nazbierana))


def get_author_metrics_headers():
    """Return standard author metrics headers."""
    return [
        "Lp.",
        "Autor",
        "Rodzaj autora",
        "ID systemu kadrowego",
        "ORCID",
        "PBN UID ID",
        "Jednostka",
        "Dyscyplina",
        "Sloty wypełnione",
        "% wykorzystania",
        "PKDaut/slot",
    ]


def write_headers(ws, headers, header_font, header_fill, header_alignment, thin_border):
    """Write headers to worksheet."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def add_autofilter_and_freeze(ws, headers, last_data_row):
    """Add autofilter and freeze panes to worksheet."""
    from openpyxl.utils import get_column_letter

    if last_data_row > 1:
        last_col_letter = get_column_letter(len(headers))
        filter_range = f"A1:{last_col_letter}{last_data_row}"
        ws.auto_filter.ref = filter_range
    ws.freeze_panes = ws["A2"]


def export_top_autorzy(ws, header_font, header_fill, header_alignment, thin_border):
    """Export top 20 authors by PKDaut/slot."""
    from .models import MetrykaAutora

    ws.title = "Top 20 autorów PKDaut-slot"
    queryset = MetrykaAutora.objects.select_related(
        "autor", "dyscyplina_naukowa", "jednostka"
    ).order_by("-srednia_za_slot_nazbierana")[:20]

    headers = get_author_metrics_headers()
    write_headers(ws, headers, header_font, header_fill, header_alignment, thin_border)

    last_data_row = 1
    for row_idx, metryka in enumerate(queryset, 2):
        last_data_row = row_idx
        write_author_metric_row(ws, row_idx, metryka, row_idx - 1)

    add_autofilter_and_freeze(ws, headers, last_data_row)


def export_top_sloty(ws, header_font, header_fill, header_alignment, thin_border):
    """Export top 20 authors by slots filled."""
    from .models import MetrykaAutora

    ws.title = "Top 20 autorów sloty wypełnione"
    queryset = (
        MetrykaAutora.objects.select_related("autor", "dyscyplina_naukowa", "jednostka")
        .filter(slot_nazbierany__gt=0)
        .order_by("-slot_nazbierany", "-srednia_za_slot_nazbierana")[:20]
    )

    headers = get_author_metrics_headers()
    write_headers(ws, headers, header_font, header_fill, header_alignment, thin_border)

    for row_idx, metryka in enumerate(queryset, 2):
        write_author_metric_row(ws, row_idx, metryka, row_idx - 1)


def export_bottom_pkd(ws, header_font, header_fill, header_alignment, thin_border):
    """Export bottom 20 authors by PKDaut/slot."""
    from .models import MetrykaAutora

    ws.title = "Najniższe 20 PKDaut-slot"
    queryset = (
        MetrykaAutora.objects.select_related("autor", "dyscyplina_naukowa", "jednostka")
        .filter(srednia_za_slot_nazbierana__gt=0)
        .order_by("srednia_za_slot_nazbierana")[:20]
    )

    headers = get_author_metrics_headers()
    write_headers(ws, headers, header_font, header_fill, header_alignment, thin_border)

    for row_idx, metryka in enumerate(queryset, 2):
        write_author_metric_row(ws, row_idx, metryka, row_idx - 1)


def export_bottom_sloty(ws, header_font, header_fill, header_alignment, thin_border):
    """Export bottom 20 authors by slots filled."""
    from .models import MetrykaAutora

    ws.title = "Najniższe 20 sloty wypełnione"
    queryset = (
        MetrykaAutora.objects.select_related("autor", "dyscyplina_naukowa", "jednostka")
        .filter(slot_nazbierany__gt=0)
        .order_by("slot_nazbierany")[:20]
    )

    headers = get_author_metrics_headers()
    write_headers(ws, headers, header_font, header_fill, header_alignment, thin_border)

    for row_idx, metryka in enumerate(queryset, 2):
        write_author_metric_row(ws, row_idx, metryka, row_idx - 1)


def export_zerowi(ws, header_font, header_fill, header_alignment, thin_border):
    """Export authors with zero metrics.

    UWAGA: Eksportowane są tylko autorzy i lata, za które mieli
    Autor_Dyscyplina.rodzaj_autora.jest_w_n = True
    """
    from bpp.models import Autor_Dyscyplina

    from .models import MetrykaAutora

    ws.title = "Autorzy zerowi"

    queryset_raw = (
        MetrykaAutora.objects.select_related("autor", "dyscyplina_naukowa", "jednostka")
        .filter(srednia_za_slot_nazbierana=0)
        .order_by("autor__nazwisko", "autor__imiona")
    )

    headers = [
        "Lp.",
        "Autor",
        "ID systemu kadrowego",
        "ORCID",
        "PBN UID ID",
        "Jednostka",
        "Dyscyplina",
        "Lata",
    ]
    write_headers(ws, headers, header_font, header_fill, header_alignment, thin_border)

    row_idx = 2
    lp = 1
    for metryka in queryset_raw:
        # Pobierz lata TYLKO dla rodzajów autorów z jest_w_n = True
        lata_dyscypliny = list(
            Autor_Dyscyplina.objects.filter(
                autor=metryka.autor,
                dyscyplina_naukowa=metryka.dyscyplina_naukowa,
                rok__gte=metryka.rok_min,
                rok__lte=metryka.rok_max,
                rodzaj_autora__jest_w_n=True,
            )
            .values_list("rok", flat=True)
            .order_by("rok")
        )

        # Pomiń autorów, którzy nie mają żadnych lat z jest_w_n = True
        if not lata_dyscypliny:
            continue

        ws.cell(row=row_idx, column=1, value=lp)
        ws.cell(row=row_idx, column=2, value=str(metryka.autor))
        ws.cell(row=row_idx, column=3, value=metryka.autor.system_kadrowy_id or "")
        ws.cell(row=row_idx, column=4, value=metryka.autor.orcid or "")
        ws.cell(row=row_idx, column=5, value=metryka.autor.pbn_uid_id or "")
        ws.cell(
            row=row_idx,
            column=6,
            value=metryka.jednostka.nazwa if metryka.jednostka else "-",
        )
        ws.cell(
            row=row_idx,
            column=7,
            value=metryka.dyscyplina_naukowa.nazwa
            if metryka.dyscyplina_naukowa
            else "-",
        )

        lata_str = ", ".join(str(rok) for rok in lata_dyscypliny)
        ws.cell(row=row_idx, column=8, value=lata_str)

        row_idx += 1
        lp += 1


def export_jednostki(ws, header_font, header_fill, header_alignment, thin_border):
    """Export unit statistics."""
    from django.db.models import Avg, Count, Sum

    from .models import MetrykaAutora

    ws.title = "Statystyki jednostek"
    stats = (
        MetrykaAutora.objects.values("jednostka__nazwa", "jednostka__skrot")
        .annotate(
            liczba_autorow=Count("id"),
            srednia_wykorzystania=Avg("procent_wykorzystania_slotow"),
            srednia_pkd_slot=Avg("srednia_za_slot_nazbierana"),
            suma_punktow=Sum("punkty_nazbierane"),
        )
        .order_by("-srednia_pkd_slot")[:10]
    )

    headers = [
        "Jednostka",
        "Liczba autorów",
        "Śr. wykorzystanie (%)",
        "Śr. PKDaut/slot",
        "Suma punktów",
    ]
    write_headers(ws, headers, header_font, header_fill, header_alignment, thin_border)

    for row_idx, stat in enumerate(stats, 2):
        nazwa = stat["jednostka__nazwa"] or "-"
        if stat["jednostka__skrot"]:
            nazwa = f"{nazwa} ({stat['jednostka__skrot']})"
        ws.cell(row=row_idx, column=1, value=nazwa)
        ws.cell(row=row_idx, column=2, value=stat["liczba_autorow"])
        ws.cell(
            row=row_idx, column=3, value=f"{stat['srednia_wykorzystania'] or 0:.1f}"
        )
        ws.cell(row=row_idx, column=4, value=f"{stat['srednia_pkd_slot'] or 0:.2f}")
        ws.cell(row=row_idx, column=5, value=f"{stat['suma_punktow'] or 0:.0f}")


def export_dyscypliny(ws, header_font, header_fill, header_alignment, thin_border):
    """Export discipline statistics."""
    from django.db.models import Avg, Count, Sum

    from .models import MetrykaAutora

    ws.title = "Statystyki dyscyplin"
    stats = (
        MetrykaAutora.objects.values(
            "dyscyplina_naukowa__nazwa", "dyscyplina_naukowa__kod"
        )
        .annotate(
            liczba_autorow=Count("id"),
            srednia_wykorzystania=Avg("procent_wykorzystania_slotow"),
            srednia_pkd_slot=Avg("srednia_za_slot_nazbierana"),
            suma_punktow=Sum("punkty_nazbierane"),
        )
        .order_by("-srednia_pkd_slot")
    )

    headers = [
        "Dyscyplina",
        "Kod",
        "Liczba autorów",
        "Śr. wykorzystanie (%)",
        "Śr. PKDaut/slot",
        "Suma punktów",
    ]
    write_headers(ws, headers, header_font, header_fill, header_alignment, thin_border)

    for row_idx, stat in enumerate(stats, 2):
        ws.cell(row=row_idx, column=1, value=stat["dyscyplina_naukowa__nazwa"] or "-")
        ws.cell(row=row_idx, column=2, value=stat["dyscyplina_naukowa__kod"] or "-")
        ws.cell(row=row_idx, column=3, value=stat["liczba_autorow"])
        ws.cell(
            row=row_idx, column=4, value=f"{stat['srednia_wykorzystania'] or 0:.1f}"
        )
        ws.cell(row=row_idx, column=5, value=f"{stat['srednia_pkd_slot'] or 0:.2f}")
        ws.cell(row=row_idx, column=6, value=f"{stat['suma_punktow'] or 0:.0f}")


def export_wykorzystanie(ws, header_font, header_fill, header_alignment, thin_border):
    """Export slot utilization distribution."""
    from .models import MetrykaAutora

    ws.title = "Rozkład wykorzystania slotów"
    wszystkie = MetrykaAutora.objects.all()

    headers = ["Przedział", "Liczba wierszy", "Procent"]
    write_headers(ws, headers, header_font, header_fill, header_alignment, thin_border)

    total = wszystkie.count()
    ranges = [
        ("0-25%", wszystkie.filter(procent_wykorzystania_slotow__lt=25).count()),
        (
            "25-50%",
            wszystkie.filter(
                procent_wykorzystania_slotow__gte=25,
                procent_wykorzystania_slotow__lt=50,
            ).count(),
        ),
        (
            "50-75%",
            wszystkie.filter(
                procent_wykorzystania_slotow__gte=50,
                procent_wykorzystania_slotow__lt=75,
            ).count(),
        ),
        (
            "75-99%",
            wszystkie.filter(
                procent_wykorzystania_slotow__gte=75,
                procent_wykorzystania_slotow__lt=99,
            ).count(),
        ),
        ("99-100%", wszystkie.filter(procent_wykorzystania_slotow__gte=99).count()),
    ]

    for row_idx, (range_name, count) in enumerate(ranges, 2):
        ws.cell(row=row_idx, column=1, value=range_name)
        ws.cell(row=row_idx, column=2, value=count)
        ws.cell(
            row=row_idx,
            column=3,
            value=f"{(count / total * 100) if total > 0 else 0:.1f}%",
        )


def auto_adjust_column_widths(ws):
    """Auto-adjust column widths based on content."""
    from openpyxl.utils import get_column_letter

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except BaseException:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
