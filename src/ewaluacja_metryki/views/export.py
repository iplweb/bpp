from django.contrib.auth.decorators import user_passes_test
from django.db.models import Q
from django.utils.decorators import method_decorator
from django.views import View

from bpp.models import Jednostka, Wydzial
from bpp.models.uczelnia import Uczelnia
from ewaluacja_common.models import Rodzaj_Autora

from ..models import MetrykaAutora
from .mixins import ma_uprawnienia_ewaluacji


@method_decorator(user_passes_test(ma_uprawnienia_ewaluacji), name="dispatch")
class ExportStatystykiXLSX(View):
    """Export statistics tables to XLSX format"""

    def _setup_workbook_and_styles(self):
        """Create workbook and define styles."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        return wb, ws, header_font, header_fill, header_alignment, thin_border

    def _create_response(self, wb, table_type):
        """Create HTTP response with workbook."""
        import datetime

        from django.http import HttpResponse

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = (
            f"metryki_statystyki_{table_type}_"
            f"{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def get(self, request, table_type):
        from django.http import HttpResponse

        from ..export_helpers import (
            auto_adjust_column_widths,
            export_bottom_pkd,
            export_bottom_sloty,
            export_dyscypliny,
            export_globalne_stats,
            export_jednostki,
            export_top_autorzy,
            export_top_sloty,
            export_wykorzystanie,
            export_zerowi,
        )

        # Dispatch table to appropriate export handler
        table_handlers = {
            "globalne": export_globalne_stats,
            "top-autorzy": export_top_autorzy,
            "top-sloty": export_top_sloty,
            "bottom-pkd": export_bottom_pkd,
            "bottom-sloty": export_bottom_sloty,
            "zerowi": export_zerowi,
            "jednostki": export_jednostki,
            "dyscypliny": export_dyscypliny,
            "wykorzystanie": export_wykorzystanie,
        }

        if table_type not in table_handlers:
            return HttpResponse("Nieznany typ tabeli", status=400)

        wb, ws, header_font, header_fill, header_alignment, thin_border = (
            self._setup_workbook_and_styles()
        )

        # Call the appropriate handler
        handler = table_handlers[table_type]
        if table_type == "globalne":
            handler(ws, header_font, header_fill, header_alignment)
        else:
            handler(ws, header_font, header_fill, header_alignment, thin_border)

        auto_adjust_column_widths(ws)

        return self._create_response(wb, table_type)


@method_decorator(user_passes_test(ma_uprawnienia_ewaluacji), name="dispatch")
class ExportListaXLSX(View):
    """Export the main metrics list to XLSX format with filters applied"""

    def _setup_workbook_styles(self):
        """Setup workbook, worksheet and define all styles."""
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            NamedStyle,
            PatternFill,
            Side,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Metryki ewaluacyjne"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        even_row_fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )

        # Create named styles for numbers
        try:
            percent_style = NamedStyle(name="percent_style")
            percent_style.number_format = "0.00%"
            percent_style.alignment = Alignment(horizontal="right")
            wb.add_named_style(percent_style)
        except ValueError:
            percent_style = "percent_style"

        try:
            decimal_style = NamedStyle(name="decimal_style")
            decimal_style.number_format = "0.00"
            decimal_style.alignment = Alignment(horizontal="right")
            wb.add_named_style(decimal_style)
        except ValueError:
            decimal_style = "decimal_style"

        return {
            "wb": wb,
            "ws": ws,
            "header_font": header_font,
            "header_fill": header_fill,
            "header_alignment": header_alignment,
            "thin_border": thin_border,
            "even_row_fill": even_row_fill,
            "percent_style": percent_style,
            "decimal_style": decimal_style,
        }

    def _apply_filters_to_queryset(self, queryset, request):
        """Apply all filters from request to queryset."""
        autor_id = request.GET.get("autor_id")
        if autor_id:
            queryset = queryset.filter(autor_id=autor_id)

        nazwisko = request.GET.get("nazwisko")
        if nazwisko:
            queryset = queryset.filter(
                Q(autor__nazwisko__icontains=nazwisko)
                | Q(autor__imiona__icontains=nazwisko)
            )

        jednostka_id = request.GET.get("jednostka")
        if jednostka_id:
            queryset = queryset.filter(jednostka_id=jednostka_id)

        wydzial_id = request.GET.get("wydzial")
        if wydzial_id:
            queryset = queryset.filter(jednostka__wydzial_id=wydzial_id)

        dyscyplina_id = request.GET.get("dyscyplina")
        if dyscyplina_id:
            queryset = queryset.filter(dyscyplina_naukowa_id=dyscyplina_id)

        rodzaj_autora = request.GET.get("rodzaj_autora")
        if rodzaj_autora and rodzaj_autora != "":
            queryset = queryset.filter(rodzaj_autora=rodzaj_autora)

        return queryset

    def _apply_sorting_to_queryset(self, queryset, request):
        """Apply sorting to queryset based on request parameter."""
        sort = request.GET.get("sort", "-srednia_za_slot_nazbierana")
        sort_mapping = {
            "procent_wykorzystania_slotow": (
                "procent_wykorzystania_slotow",
                "srednia_za_slot_nazbierana",
            ),
            "-procent_wykorzystania_slotow": (
                "-procent_wykorzystania_slotow",
                "-srednia_za_slot_nazbierana",
            ),
            "srednia_za_slot_nazbierana": (
                "srednia_za_slot_nazbierana",
                "procent_wykorzystania_slotow",
            ),
            "-srednia_za_slot_nazbierana": (
                "-srednia_za_slot_nazbierana",
                "-procent_wykorzystania_slotow",
            ),
        }

        allowed_sorts = list(sort_mapping.keys()) + [
            "autor__nazwisko",
            "-autor__nazwisko",
        ]

        if sort not in allowed_sorts:
            return queryset

        if sort in sort_mapping:
            return queryset.order_by(*sort_mapping[sort])
        return queryset.order_by(sort)

    def _determine_visible_columns(self):
        """Determine which columns should be visible in export."""
        from bpp.models import Dyscyplina_Naukowa

        uczelnia = Uczelnia.objects.get_default()
        uzywa_wydzialow = uczelnia.uzywaj_wydzialow if uczelnia else False

        wszystkie_dyscypliny = Dyscyplina_Naukowa.objects.filter(
            metrykaautora__isnull=False
        ).distinct()
        tylko_jedna_dyscyplina = wszystkie_dyscypliny.count() == 1

        return {
            "uzywa_wydzialow": uzywa_wydzialow,
            "tylko_jedna_dyscyplina": tylko_jedna_dyscyplina,
        }

    def _create_headers(self, visible_columns):
        """Create header list based on visible columns."""
        headers = [
            "Lp.",
            "Autor",
            "Rodzaj autora",
            "ID systemu kadrowego",
            "ORCID",
            "PBN UID ID",
        ]

        if not visible_columns["tylko_jedna_dyscyplina"]:
            headers.append("Dyscyplina")

        if visible_columns["uzywa_wydzialow"]:
            headers.append("Wydział")

        headers.extend(
            [
                "Jednostka",
                "Slot maksymalny",
                "Slot nazbierany",
                "Slot niewykorzystany",
                "% wykorzystania",
                "PKDaut nazbierane",
                "PKDaut wszystkie",
                "Średnia PKDaut/slot (nazbierane)",
                "Średnia PKDaut/slot (wszystkie)",
                "Liczba prac (nazbierane)",
                "Liczba prac (wszystkie)",
                "Rok min",
                "Rok max",
            ]
        )

        return headers

    def _write_headers(self, ws, headers, styles):
        """Write headers to worksheet with styling."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["header_alignment"]
            cell.border = styles["thin_border"]

    def _format_rodzaj_autora(self, metryka):
        """Format rodzaj autora display value."""
        if metryka.rodzaj_autora == " ":
            return "Brak danych"

        try:
            rodzaj = Rodzaj_Autora.objects.get(skrot=metryka.rodzaj_autora)
            return rodzaj.nazwa
        except Rodzaj_Autora.DoesNotExist:
            return metryka.rodzaj_autora

    def _write_cell(self, ws, row, col, value, styles, style_type=None, row_fill=None):
        """Write a single cell with consistent styling."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = styles["thin_border"]

        if style_type:
            cell.style = styles[style_type]

        if row_fill:
            cell.fill = row_fill

        return col + 1

    def _write_data_row(self, ws, row_idx, metryka, visible_columns, styles, row_fill):
        """Write a single data row to worksheet."""
        col = 1

        # Lp.
        col = self._write_cell(ws, row_idx, col, row_idx - 1, styles, row_fill=row_fill)

        # Autor
        col = self._write_cell(
            ws, row_idx, col, str(metryka.autor), styles, row_fill=row_fill
        )

        # Rodzaj autora
        col = self._write_cell(
            ws,
            row_idx,
            col,
            self._format_rodzaj_autora(metryka),
            styles,
            row_fill=row_fill,
        )

        # ID systemu kadrowego
        col = self._write_cell(
            ws,
            row_idx,
            col,
            metryka.autor.system_kadrowy_id or "",
            styles,
            row_fill=row_fill,
        )

        # ORCID
        col = self._write_cell(
            ws, row_idx, col, metryka.autor.orcid or "", styles, row_fill=row_fill
        )

        # PBN UID ID
        col = self._write_cell(
            ws, row_idx, col, metryka.autor.pbn_uid_id or "", styles, row_fill=row_fill
        )

        # Dyscyplina (if shown)
        if not visible_columns["tylko_jedna_dyscyplina"]:
            dyscyplina_value = (
                metryka.dyscyplina_naukowa.nazwa if metryka.dyscyplina_naukowa else "-"
            )
            col = self._write_cell(
                ws, row_idx, col, dyscyplina_value, styles, row_fill=row_fill
            )

        # Wydział (if shown)
        if visible_columns["uzywa_wydzialow"]:
            wydzial_nazwa = "-"
            if metryka.jednostka and metryka.jednostka.wydzial:
                wydzial_nazwa = metryka.jednostka.wydzial.nazwa
            col = self._write_cell(
                ws, row_idx, col, wydzial_nazwa, styles, row_fill=row_fill
            )

        # Jednostka
        jednostka_value = metryka.jednostka.nazwa if metryka.jednostka else "-"
        col = self._write_cell(
            ws, row_idx, col, jednostka_value, styles, row_fill=row_fill
        )

        # Numeric values with proper formatting
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.slot_maksymalny),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.slot_nazbierany),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.slot_niewykorzystany),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.procent_wykorzystania_slotow) / 100,
            styles,
            "percent_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.punkty_nazbierane),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.punkty_wszystkie),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.srednia_za_slot_nazbierana),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.srednia_za_slot_wszystkie),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws, row_idx, col, len(metryka.prace_nazbierane), styles, row_fill=row_fill
        )
        col = self._write_cell(
            ws, row_idx, col, metryka.liczba_prac_wszystkie, styles, row_fill=row_fill
        )
        col = self._write_cell(
            ws, row_idx, col, metryka.rok_min, styles, row_fill=row_fill
        )
        col = self._write_cell(
            ws, row_idx, col, metryka.rok_max, styles, row_fill=row_fill
        )

    def _write_all_data_rows(self, ws, queryset, visible_columns, styles):
        """Write all data rows to worksheet."""
        last_data_row = 1
        for row_idx, metryka in enumerate(queryset, 2):
            last_data_row = row_idx
            row_fill = styles["even_row_fill"] if row_idx % 2 == 0 else None
            self._write_data_row(
                ws, row_idx, metryka, visible_columns, styles, row_fill
            )

        return last_data_row

    def _setup_worksheet_formatting(self, ws, headers, last_data_row):
        """Setup auto-filter, freeze panes and column widths."""
        from openpyxl.utils import get_column_letter

        if last_data_row > 1:
            last_col_letter = get_column_letter(len(headers))
            filter_range = f"A1:{last_col_letter}{last_data_row}"
            ws.auto_filter.ref = filter_range

        ws.freeze_panes = ws["A2"]

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except BaseException:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_summary(self, ws, queryset, last_data_row):
        """Add summary section to worksheet."""
        summary_row = last_data_row + 2 if last_data_row > 1 else 3
        ws.cell(row=summary_row, column=1, value="Podsumowanie:")
        ws.cell(row=summary_row + 1, column=1, value="Liczba wierszy:")
        ws.cell(row=summary_row + 1, column=2, value=queryset.count())

        return summary_row

    def _collect_filter_info(self, request, visible_columns):
        """Collect applied filter information for display."""
        from bpp.models import Dyscyplina_Naukowa

        filter_info = []

        nazwisko = request.GET.get("nazwisko")
        if nazwisko:
            filter_info.append(f"Nazwisko/Imię: {nazwisko}")

        jednostka_id = request.GET.get("jednostka")
        if jednostka_id:
            try:
                jednostka = Jednostka.objects.get(pk=jednostka_id)
                filter_info.append(f"Jednostka: {jednostka.nazwa}")
            except Jednostka.DoesNotExist:
                pass

        wydzial_id = request.GET.get("wydzial")
        if wydzial_id and visible_columns["uzywa_wydzialow"]:
            try:
                wydzial = Wydzial.objects.get(pk=wydzial_id)
                filter_info.append(f"Wydział: {wydzial.nazwa}")
            except Wydzial.DoesNotExist:
                pass

        dyscyplina_id = request.GET.get("dyscyplina")
        if dyscyplina_id and not visible_columns["tylko_jedna_dyscyplina"]:
            try:
                dyscyplina = Dyscyplina_Naukowa.objects.get(pk=dyscyplina_id)
                filter_info.append(f"Dyscyplina: {dyscyplina.nazwa}")
            except Dyscyplina_Naukowa.DoesNotExist:
                pass

        rodzaj_autora = request.GET.get("rodzaj_autora")
        if rodzaj_autora:
            try:
                rodzaj = Rodzaj_Autora.objects.get(skrot=rodzaj_autora)
                filter_info.append(f"Rodzaj autora: {rodzaj.nazwa}")
            except Rodzaj_Autora.DoesNotExist:
                filter_info.append(f"Rodzaj autora: {rodzaj_autora}")

        return filter_info

    def _add_filter_info(self, ws, summary_row, filter_info):
        """Add filter information to worksheet."""
        filters_row = summary_row + 3
        ws.cell(row=filters_row, column=1, value="Zastosowane filtry:")

        if filter_info:
            ws.cell(row=filters_row + 1, column=1, value="; ".join(filter_info))
        else:
            ws.cell(row=filters_row + 1, column=1, value="Brak filtrów")

    def _create_response(self, wb):
        """Create HTTP response with workbook."""
        import datetime

        from django.http import HttpResponse

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = (
            f"metryki_ewaluacyjne_"
            f"{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    def get(self, request):
        from django.db.models import Count, OuterRef, Subquery

        # Setup workbook and styles
        styles = self._setup_workbook_styles()
        ws = styles["ws"]
        wb = styles["wb"]

        # Build queryset with discipline count annotation
        discipline_count = (
            MetrykaAutora.objects.filter(autor=OuterRef("autor"))
            .values("autor")
            .annotate(count=Count("dyscyplina_naukowa"))
            .values("count")
        )

        queryset = MetrykaAutora.objects.select_related(
            "autor", "dyscyplina_naukowa", "jednostka", "jednostka__wydzial"
        ).annotate(autor_discipline_count=Subquery(discipline_count))

        # Apply filters and sorting
        queryset = self._apply_filters_to_queryset(queryset, request)
        queryset = self._apply_sorting_to_queryset(queryset, request)

        # Determine visible columns
        visible_columns = self._determine_visible_columns()

        # Create and write headers
        headers = self._create_headers(visible_columns)
        self._write_headers(ws, headers, styles)

        # Write all data rows
        last_data_row = self._write_all_data_rows(ws, queryset, visible_columns, styles)

        # Setup worksheet formatting
        self._setup_worksheet_formatting(ws, headers, last_data_row)

        # Add summary and filter information
        summary_row = self._add_summary(ws, queryset, last_data_row)
        filter_info = self._collect_filter_info(request, visible_columns)
        self._add_filter_info(ws, summary_row, filter_info)

        # Return response
        return self._create_response(wb)
