from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Avg, Count, Q, Sum
from django.http import JsonResponse
from django.shortcuts import redirect
from django.utils.decorators import method_decorator
from django.views import View
from django.views.generic import DetailView, ListView

from bpp.const import GR_WPROWADZANIE_DANYCH
from bpp.models import Jednostka, Wydzial
from bpp.models.uczelnia import Uczelnia
from ewaluacja_common.models import Rodzaj_Autora

from .models import MetrykaAutora, StatusGenerowania
from .tasks import generuj_metryki_task


def ma_uprawnienia_ewaluacji(user):
    """Sprawdza czy użytkownik ma uprawnienia do ewaluacji"""
    return user.is_superuser or user.groups.filter(name=GR_WPROWADZANIE_DANYCH).exists()


class EwaluacjaRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    """Mixin wymagający uprawnień do ewaluacji"""

    def test_func(self):
        return ma_uprawnienia_ewaluacji(self.request.user)


class MetrykiListView(EwaluacjaRequiredMixin, ListView):
    model = MetrykaAutora
    template_name = "ewaluacja_metryki/lista.html"
    context_object_name = "metryki"
    paginate_by = 20

    def get_paginate_by(self, queryset):
        """Wyłącz paginację dla widoków kółek, metryki i wykresów"""
        widok = self.request.GET.get("widok", "tabela")
        if widok in [
            "kola",
            "pkdaut",
            "mapa_ciepla",
            "mapa_ciepla_bloki",
            "mapa_konturowa",
            "wykres_babelkowy",
            "kontury_z_punktami",
            "efektywnosc_babelki",
            "efektywnosc_mapa_ciepla",
            "przestrzen_pasternaka",  # Wykres 3D: X=sloty, Y=punkty, Z=średni IF
            "kosmos_bpp_3d",  # Wykres 3D: X=sloty, Y=punkty, Z=średni rok prac
        ]:
            return None  # Disable pagination for visualization views
        return self.paginate_by

    def _apply_filters(self, queryset):
        """Apply all filter parameters from GET request to queryset."""
        # Filtrowanie po ID autora (dla przycisku "Kadruj")
        autor_id = self.request.GET.get("autor_id")
        if autor_id:
            queryset = queryset.filter(autor_id=autor_id)

        # Filtrowanie po nazwisku
        nazwisko = self.request.GET.get("nazwisko")
        if nazwisko:
            queryset = queryset.filter(
                Q(autor__nazwisko__icontains=nazwisko)
                | Q(autor__imiona__icontains=nazwisko)
            )

        # Filtrowanie po jednostce
        jednostka_id = self.request.GET.get("jednostka")
        if jednostka_id:
            queryset = queryset.filter(jednostka_id=jednostka_id)

        # Filtrowanie po wydziale
        wydzial_id = self.request.GET.get("wydzial")
        if wydzial_id:
            queryset = queryset.filter(jednostka__wydzial_id=wydzial_id)

        # Filtrowanie po dyscyplinie
        dyscyplina_id = self.request.GET.get("dyscyplina")
        if dyscyplina_id:
            queryset = queryset.filter(dyscyplina_naukowa_id=dyscyplina_id)

        # Filtrowanie po rodzaju autora
        rodzaj_autora = self.request.GET.get("rodzaj_autora")
        if rodzaj_autora and rodzaj_autora != "":
            queryset = queryset.filter(rodzaj_autora=rodzaj_autora)

        return queryset

    def _apply_sorting(self, queryset):
        """Apply sorting based on sort parameter."""
        sort = self.request.GET.get("sort", "-srednia_za_slot_nazbierana")
        allowed_sorts = [
            "srednia_za_slot_nazbierana",
            "-srednia_za_slot_nazbierana",
            "procent_wykorzystania_slotow",
            "-procent_wykorzystania_slotow",
            "autor__nazwisko",
            "-autor__nazwisko",
        ]

        if sort not in allowed_sorts:
            return queryset

        # Define sorting with secondary sort column
        sort_mapping = {
            "procent_wykorzystania_slotow": (
                "procent_wykorzystania_slotow",
                "srednia_za_slot_nazbierana",
            ),
            "-procent_wykorzystania_slotow": (
                "-procent_wykorzystania_slotow",
                "-srednia_za_slot_nazbierana",
            ),
            "srednia_za_slot_nazbierana": (
                "srednia_za_slot_nazbierana",
                "procent_wykorzystania_slotow",
            ),
            "-srednia_za_slot_nazbierana": (
                "-srednia_za_slot_nazbierana",
                "-procent_wykorzystania_slotow",
            ),
        }

        if sort in sort_mapping:
            return queryset.order_by(*sort_mapping[sort])
        return queryset.order_by(sort)

    def get_queryset(self):
        from django.db.models import Count, OuterRef, Subquery

        # Subquery to count disciplines for each author
        discipline_count = (
            MetrykaAutora.objects.filter(autor=OuterRef("autor"))
            .values("autor")
            .annotate(count=Count("dyscyplina_naukowa"))
            .values("count")
        )

        queryset = (
            super()
            .get_queryset()
            .select_related(
                "autor", "dyscyplina_naukowa", "jednostka", "jednostka__wydzial"
            )
            .annotate(autor_discipline_count=Subquery(discipline_count))
        )

        queryset = self._apply_filters(queryset)
        queryset = self._apply_sorting(queryset)

        return queryset

    def _get_filtered_autor_context(self):
        """Get filtered autor information if autor_id is in request."""
        autor_id = self.request.GET.get("autor_id")
        if not autor_id:
            return {}

        from bpp.models import Autor

        try:
            autor = Autor.objects.get(pk=autor_id)
            return {"filtered_autor": autor}
        except Autor.DoesNotExist:
            return {}

    def _get_jednostki_wydzialy_context(self):
        """Get jednostki and wydzialy lists for filters."""
        from bpp.models import Autor

        context = {}

        # Sprawdź czy uczelnia używa wydziałów
        uczelnia = Uczelnia.objects.get_default()
        context["uzywa_wydzialow"] = uczelnia.uzywaj_wydzialow if uczelnia else False

        # Jeśli wydzial jest wybrany, filtruj jednostki tylko z tego wydziału
        wydzial_id = self.request.GET.get("wydzial")
        jednostki_queryset = Jednostka.objects.filter(
            pk__in=Autor.objects.filter(metryki__isnull=False)
            .values_list("aktualna_jednostka", flat=True)
            .distinct()
        ).distinct()

        if wydzial_id:
            jednostki_queryset = jednostki_queryset.filter(wydzial_id=wydzial_id)

        context["jednostki"] = jednostki_queryset.order_by("nazwa")

        if context["uzywa_wydzialow"]:
            # Buduj listę wydziałów na podstawie aktualnych jednostek autorów z metrykami
            context["wydzialy"] = (
                Wydzial.objects.filter(
                    jednostka__in=Autor.objects.filter(metryki__isnull=False)
                    .values_list("aktualna_jednostka", flat=True)
                    .distinct()
                )
                .distinct()
                .order_by("nazwa")
            )
            # Check if there's only one faculty
            context["tylko_jeden_wydzial"] = context["wydzialy"].count() == 1
        else:
            context["tylko_jeden_wydzial"] = False

        return context

    def _get_dyscypliny_context(self):
        """Get dyscypliny list for filters."""
        from bpp.models import Dyscyplina_Naukowa

        dyscypliny = (
            Dyscyplina_Naukowa.objects.filter(metrykaautora__isnull=False)
            .distinct()
            .order_by("nazwa")
        )

        return {
            "dyscypliny": dyscypliny,
            "tylko_jedna_dyscyplina": dyscypliny.count() == 1,
        }

    def _get_statistics_context(self):
        """Get statistics for current queryset."""
        stats = self.get_queryset().aggregate(
            srednia_wykorzystania=Avg("procent_wykorzystania_slotow"),
            srednia_pkd_slot=Avg("srednia_za_slot_nazbierana"),
            liczba_wierszy=Count("id"),
            liczba_autorow=Count("autor", distinct=True),
        )
        return {"statystyki": stats}

    def _get_status_context(self):
        """Get generation status and progress information."""
        status = StatusGenerowania.get_or_create()
        context = {
            "status_generowania": status,
            "dostepne_rodzaje_autorow": Rodzaj_Autora.objects.filter(
                licz_sloty=True
            ).order_by("sort"),
        }

        # Oblicz procent postępu
        if status.w_trakcie and status.liczba_do_przetworzenia > 0:
            context["progress_procent"] = round(
                (status.liczba_przetworzonych / status.liczba_do_przetworzenia * 100), 1
            )
        else:
            context["progress_procent"] = 0

        return context

    def _calculate_3d_metrics(self, metryki, widok):
        """Calculate average IF and average year for 3D visualizations."""
        if widok not in ["przestrzen_pasternaka", "kosmos_bpp_3d"]:
            return

        from bpp.models.cache import Cache_Punktacja_Autora_Query

        # Iterate through all metryki and calculate average IF and average year
        for metryka in metryki:
            # Get all works for this author/discipline
            if metryka.prace_nazbierane:
                prace = Cache_Punktacja_Autora_Query.objects.filter(
                    pk__in=metryka.prace_nazbierane
                ).select_related("rekord")

                # Calculate average Impact Factor
                impact_factors = []
                years = []
                for praca in prace:
                    years.append(praca.rekord.rok)
                    try:
                        if hasattr(praca.rekord, "original") and hasattr(
                            praca.rekord.original, "impact_factor"
                        ):
                            if praca.rekord.original.impact_factor:
                                impact_factors.append(
                                    float(praca.rekord.original.impact_factor)
                                )
                    except (AttributeError, ValueError, TypeError):
                        pass

                # Set average IF (or 0 if no IF data)
                metryka.sredni_if = (
                    sum(impact_factors) / len(impact_factors) if impact_factors else 0.0
                )

                # Set average year
                metryka.sredni_rok = (
                    sum(years) / len(years) if years else metryka.rok_min
                )
            else:
                metryka.sredni_if = 0.0
                metryka.sredni_rok = metryka.rok_min

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Set default widok to "tabela" if not provided
        context["widok"] = self.request.GET.get("widok", "tabela")

        # Zachowaj parametry filtru
        context["request"] = self.request

        # Update context with data from helper methods
        context.update(self._get_filtered_autor_context())
        context.update(self._get_jednostki_wydzialy_context())
        context.update(self._get_dyscypliny_context())
        context.update(self._get_statistics_context())
        context.update(self._get_status_context())

        # Calculate 3D metrics if needed
        self._calculate_3d_metrics(context.get("metryki", []), context["widok"])

        return context


class MetrykaDetailView(EwaluacjaRequiredMixin, DetailView):
    model = MetrykaAutora
    template_name = "ewaluacja_metryki/szczegoly.html"
    context_object_name = "metryka"

    def get_object(self, queryset=None):
        """Override get_object to lookup by autor slug and dyscyplina kod"""
        if queryset is None:
            queryset = self.get_queryset()

        autor_slug = self.kwargs.get("autor_slug")
        dyscyplina_kod = self.kwargs.get("dyscyplina_kod")

        if not autor_slug or not dyscyplina_kod:
            from django.http import Http404

            raise Http404("Autor slug and dyscyplina kod are required")

        try:
            obj = queryset.get(
                autor__slug=autor_slug, dyscyplina_naukowa__kod=dyscyplina_kod
            )
        except self.model.DoesNotExist as err:
            from django.http import Http404

            raise Http404(
                f"MetrykaAutora for autor '{autor_slug}' and dyscyplina '{dyscyplina_kod}' not found"
            ) from err

        return obj

    def _get_discipline_years_context(self, metryka):
        """Get discipline years and calculate averages."""
        from decimal import Decimal

        from bpp.models import Autor_Dyscyplina

        context = {}

        dyscyplina_lata = (
            Autor_Dyscyplina.objects.filter(
                autor=metryka.autor,
                dyscyplina_naukowa=metryka.dyscyplina_naukowa,
                rok__in=[2022, 2023, 2024, 2025],
            )
            .select_related("dyscyplina_naukowa", "subdyscyplina_naukowa")
            .order_by("rok")
        )
        context["dyscyplina_lata"] = dyscyplina_lata

        if dyscyplina_lata:
            suma_etatu = Decimal("0")
            liczba_etatu = 0
            suma_procent = Decimal("0")
            liczba_procent = 0

            for rok_data in dyscyplina_lata:
                if rok_data.wymiar_etatu:
                    suma_etatu += rok_data.wymiar_etatu
                    liczba_etatu += 1

                # Sprawdź czy to subdyscyplina czy dyscyplina główna
                if (
                    rok_data.subdyscyplina_naukowa
                    and rok_data.subdyscyplina_naukowa == metryka.dyscyplina_naukowa
                ):
                    if rok_data.procent_subdyscypliny:
                        suma_procent += rok_data.procent_subdyscypliny
                        liczba_procent += 1
                else:
                    if rok_data.procent_dyscypliny:
                        suma_procent += rok_data.procent_dyscypliny
                        liczba_procent += 1

            context["srednia_etatu"] = (
                (suma_etatu / liczba_etatu) if liczba_etatu > 0 else None
            )
            context["srednia_procent"] = (
                (suma_procent / liczba_procent) if liczba_procent > 0 else None
            )

        return context

    def _get_collected_works_context(self, metryka):
        """Get collected works (nazbierane) with pkdaut/slot calculations."""
        from bpp.models.cache import Cache_Punktacja_Autora_Query

        context = {}

        if metryka.prace_nazbierane:
            prace_nazbierane = (
                Cache_Punktacja_Autora_Query.objects.filter(
                    pk__in=metryka.prace_nazbierane
                )
                .select_related("rekord")
                .order_by("-pkdaut")
            )

            # Calculate pkdaut/slot for each work and find autor_assignment_id
            for praca in prace_nazbierane:
                if praca.slot and praca.slot > 0:
                    praca.pkdaut_per_slot = float(praca.pkdaut) / float(praca.slot)
                else:
                    praca.pkdaut_per_slot = None

                # Find the autor_assignment_id for this work using the original publication
                praca.autor_assignment_id = None
                try:
                    assignment = praca.rekord.original.autorzy_set.filter(
                        autor_id=metryka.autor.id,
                        dyscyplina_naukowa_id=metryka.dyscyplina_naukowa.id,
                        przypieta=True,
                    ).first()
                    if assignment:
                        praca.autor_assignment_id = assignment.pk
                except (AttributeError, Exception):
                    pass

            context["prace_nazbierane"] = prace_nazbierane

        context["liczba_prac_nazbierane"] = (
            len(metryka.prace_nazbierane) if metryka.prace_nazbierane else 0
        )

        return context

    def _get_chart_data_context(self, metryka):
        """Prepare data for visualization charts."""
        import json
        from collections import defaultdict

        from bpp.models.cache import Cache_Punktacja_Autora_Query

        context = {}

        # Pobierz wszystkie prace (nazbierane i nie nazbierane) z latami
        if metryka.prace_wszystkie:
            prace_wszystkie = (
                Cache_Punktacja_Autora_Query.objects.filter(
                    pk__in=metryka.prace_wszystkie
                )
                .select_related("rekord")
                .order_by("rekord__rok", "-pkdaut")
            )

            # Przygotuj dane dla wykresów
            chart_data = []
            year_aggregates = defaultdict(
                lambda: {"punkty": 0, "sloty": 0, "liczba": 0}
            )

            for praca in prace_wszystkie:
                rok = praca.rekord.rok
                pkdaut = float(praca.pkdaut or 0)
                slot = float(praca.slot or 0)
                is_nazbierana = praca.pk in (metryka.prace_nazbierane or [])

                # Get Impact Factor from original publication if available
                impact_factor = 0.0
                try:
                    if hasattr(praca.rekord, "original") and hasattr(
                        praca.rekord.original, "impact_factor"
                    ):
                        if praca.rekord.original.impact_factor:
                            impact_factor = float(praca.rekord.original.impact_factor)
                except (AttributeError, ValueError, TypeError):
                    pass

                chart_data.append(
                    {
                        "rok": rok,
                        "pkdaut": pkdaut,
                        "slot": slot,
                        "tytul": str(praca.rekord.tytul_oryginalny or ""),
                        "is_nazbierana": is_nazbierana,
                        "impact_factor": impact_factor,
                    }
                )

                year_aggregates[rok]["punkty"] += pkdaut
                year_aggregates[rok]["sloty"] += slot
                year_aggregates[rok]["liczba"] += 1

            # Serializuj do JSON
            context["chart_data_json"] = json.dumps(chart_data)
            context["year_aggregates_json"] = json.dumps(dict(year_aggregates))
            context["chart_data"] = chart_data  # Keep for template conditional

        return context

    def _get_all_works_context(self, metryka):
        """Get all works (wszystkie) with pkdaut/slot calculations."""
        from bpp.models.cache import Cache_Punktacja_Autora_Query

        context = {}

        if metryka.prace_wszystkie:
            prace_wszystkie = (
                Cache_Punktacja_Autora_Query.objects.filter(
                    pk__in=metryka.prace_wszystkie
                )
                .select_related("rekord")
                .order_by("-pkdaut")
            )

            # Calculate pkdaut/slot for each work and find autor_assignment_id
            for praca in prace_wszystkie:
                if praca.slot and praca.slot > 0:
                    praca.pkdaut_per_slot = float(praca.pkdaut) / float(praca.slot)
                else:
                    praca.pkdaut_per_slot = None

                # Find the autor_assignment_id for this work using the original publication
                praca.autor_assignment_id = None
                try:
                    assignment = praca.rekord.original.autorzy_set.filter(
                        autor_id=metryka.autor.id,
                        dyscyplina_naukowa_id=metryka.dyscyplina_naukowa.id,
                        przypieta=True,
                    ).first()
                    if assignment:
                        praca.autor_assignment_id = assignment.pk
                except (AttributeError, Exception):
                    pass

            context["prace_wszystkie"] = prace_wszystkie

        return context

    def _get_unpinned_works_context(self, metryka):
        """Get unpinned works (odpiete) from all publication models."""
        from decimal import Decimal

        from bpp.models import (
            Patent_Autor,
            Wydawnictwo_Ciagle_Autor,
            Wydawnictwo_Zwarte_Autor,
        )

        context = {}
        prace_odpiete = []

        # Query all three types of unpinned publications
        for model in [Wydawnictwo_Ciagle_Autor, Wydawnictwo_Zwarte_Autor, Patent_Autor]:
            assignments = model.objects.filter(
                autor=metryka.autor,
                dyscyplina_naukowa=metryka.dyscyplina_naukowa,
                przypieta=False,
                rekord__rok__gte=metryka.rok_min,
                rekord__rok__lte=metryka.rok_max,
            ).select_related("rekord")

            for assignment in assignments:
                rekord = assignment.rekord
                author_count = Decimal(str(rekord.autorzy_set.count() or 1))

                praca = type(
                    "PracaOdpieta",
                    (),
                    {
                        "rekord": rekord,
                        "tytul_oryginalny": rekord.tytul_oryginalny,
                        "rok": rekord.rok,
                        "pkdaut_est": Decimal(str(rekord.punkty_kbn or 0))
                        / author_count,
                        "slot_est": Decimal("1") / author_count,
                        "autor_assignment_id": assignment.pk,
                    },
                )()

                # Calculate estimated PKDaut/slot
                if praca.slot_est and praca.slot_est > 0:
                    praca.pkdaut_per_slot_est = float(praca.pkdaut_est) / float(
                        praca.slot_est
                    )
                else:
                    praca.pkdaut_per_slot_est = None

                prace_odpiete.append(praca)

        # Sort by estimated PKDaut descending
        prace_odpiete.sort(key=lambda x: x.pkdaut_est, reverse=True)

        # Add to context if there are unpinned works
        if prace_odpiete:
            context["prace_odpiete"] = prace_odpiete
            context["prace_odpiete_suma_pkdaut"] = sum(
                p.pkdaut_est for p in prace_odpiete
            )
            context["prace_odpiete_suma_slotow"] = sum(
                p.slot_est for p in prace_odpiete
            )
            context["prace_odpiete_liczba"] = len(prace_odpiete)

        return context

    def _get_position_context(self, metryka):
        """Get author's position in unit rankings."""
        context = {}

        if metryka.jednostka:
            context["pozycja_w_jednostce"] = (
                MetrykaAutora.objects.filter(
                    jednostka=metryka.jednostka,
                    dyscyplina_naukowa=metryka.dyscyplina_naukowa,
                    srednia_za_slot_nazbierana__gt=metryka.srednia_za_slot_nazbierana,
                ).count()
                + 1
            )

            context["liczba_w_jednostce"] = MetrykaAutora.objects.filter(
                jednostka=metryka.jednostka,
                dyscyplina_naukowa=metryka.dyscyplina_naukowa,
            ).count()

        return context

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        metryka = self.object

        # Pobierz inne dyscypliny tego samego autora
        inne_dyscypliny = (
            MetrykaAutora.objects.filter(autor=metryka.autor)
            .exclude(pk=metryka.pk)
            .select_related("dyscyplina_naukowa")
            .order_by("dyscyplina_naukowa__nazwa")
        )
        context["inne_dyscypliny"] = inne_dyscypliny

        # Dodaj dostępne rodzaje autorów dla szablonu
        context["rodzaje_autorow"] = Rodzaj_Autora.objects.all().order_by("skrot")

        # Get all context data from helper methods
        context.update(self._get_discipline_years_context(metryka))
        context.update(self._get_collected_works_context(metryka))
        context.update(self._get_all_works_context(metryka))
        context.update(self._get_unpinned_works_context(metryka))
        context.update(self._get_position_context(metryka))
        context.update(self._get_chart_data_context(metryka))

        return context


class PrzypnijDyscyplineView(EwaluacjaRequiredMixin, View):
    """Handle pinning a discipline for an author in a publication"""

    def post(self, request, autor_assignment_id):
        from django.db import transaction
        from django.shortcuts import redirect

        from bpp.models import (
            Patent_Autor,
            Wydawnictwo_Ciagle_Autor,
            Wydawnictwo_Zwarte_Autor,
        )
        from bpp.models.sloty.core import IPunktacjaCacher
        from ewaluacja_metryki.utils import przelicz_metryki_dla_publikacji

        with transaction.atomic():
            # Find the assignment in different publication types
            wa = None
            for model in [
                Wydawnictwo_Ciagle_Autor,
                Wydawnictwo_Zwarte_Autor,
                Patent_Autor,
            ]:
                try:
                    wa = model.objects.get(pk=autor_assignment_id)
                    break
                except model.DoesNotExist:
                    continue

            if not wa:
                raise Exception("Nie znaleziono przypisania autora do publikacji")

            publikacja = wa.rekord

            # Update przypieta status
            wa.przypieta = True
            wa.save()

            # Rebuild cache punktacja for the publication
            cacher = IPunktacjaCacher(publikacja)
            cacher.removeEntries()
            cacher.rebuildEntries()

            # Recalculate evaluation metrics for all affected authors
            przelicz_metryki_dla_publikacji(publikacja)

        # Redirect back to the detail view
        # Get MetrykaAutora for the author and discipline to redirect to the correct detail page
        from django.urls import reverse

        from .models import MetrykaAutora

        metryka = MetrykaAutora.objects.filter(
            autor=wa.autor, dyscyplina_naukowa=wa.dyscyplina_naukowa
        ).first()

        if metryka:
            url = reverse(
                "ewaluacja_metryki:szczegoly",
                kwargs={
                    "autor_slug": metryka.autor.slug,
                    "dyscyplina_kod": metryka.dyscyplina_naukowa.kod,
                },
            )
            return redirect(url + "#prace-nazbierane")
        return redirect("ewaluacja_metryki:lista")


class OdepnijDyscyplineView(EwaluacjaRequiredMixin, View):
    """Handle unpinning a discipline for an author in a publication"""

    def post(self, request, autor_assignment_id):
        from django.db import transaction
        from django.shortcuts import redirect

        from bpp.models import (
            Patent_Autor,
            Wydawnictwo_Ciagle_Autor,
            Wydawnictwo_Zwarte_Autor,
        )
        from bpp.models.sloty.core import IPunktacjaCacher
        from ewaluacja_metryki.utils import przelicz_metryki_dla_publikacji

        with transaction.atomic():
            # Find the assignment in different publication types
            wa = None
            for model in [
                Wydawnictwo_Ciagle_Autor,
                Wydawnictwo_Zwarte_Autor,
                Patent_Autor,
            ]:
                try:
                    wa = model.objects.get(pk=autor_assignment_id)
                    break
                except model.DoesNotExist:
                    continue

            if not wa:
                raise Exception("Nie znaleziono przypisania autora do publikacji")

            publikacja = wa.rekord

            # Update przypieta status
            wa.przypieta = False
            wa.save()

            # Rebuild cache punktacja for the publication
            cacher = IPunktacjaCacher(publikacja)
            cacher.removeEntries()
            cacher.rebuildEntries()

            # Recalculate evaluation metrics for all affected authors
            przelicz_metryki_dla_publikacji(publikacja)

        # Redirect back to the detail view
        # Get MetrykaAutora for the author and discipline to redirect to the correct detail page
        from django.urls import reverse

        from .models import MetrykaAutora

        metryka = MetrykaAutora.objects.filter(
            autor=wa.autor, dyscyplina_naukowa=wa.dyscyplina_naukowa
        ).first()

        if metryka:
            url = reverse(
                "ewaluacja_metryki:szczegoly",
                kwargs={
                    "autor_slug": metryka.autor.slug,
                    "dyscyplina_kod": metryka.dyscyplina_naukowa.kod,
                },
            )
            return redirect(url + "#prace-nazbierane")
        return redirect("ewaluacja_metryki:lista")


class StatystykiView(EwaluacjaRequiredMixin, ListView):
    model = MetrykaAutora
    template_name = "ewaluacja_metryki/statystyki.html"
    context_object_name = "top_autorzy_pkd"  # Renamed for clarity

    def get_queryset(self):
        # Top 20 autorów wg średniej PKDaut/slot
        return MetrykaAutora.objects.select_related(
            "autor", "dyscyplina_naukowa", "jednostka"
        ).order_by("-srednia_za_slot_nazbierana")[:20]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Keep the old name for backward compatibility in template
        context["top_autorzy"] = context["top_autorzy_pkd"]

        # Top 20 autorów wg slotów wypełnionych (the absolute top - highest slot filling AND highest PKDaut/slot)
        context["top_autorzy_sloty"] = (
            MetrykaAutora.objects.select_related(
                "autor", "dyscyplina_naukowa", "jednostka"
            )
            .filter(slot_nazbierany__gt=0)  # Exclude those with zero slots
            .order_by("-slot_nazbierany", "-srednia_za_slot_nazbierana")[:20]
        )

        # Statystyki globalne
        wszystkie = MetrykaAutora.objects.all()
        context["statystyki_globalne"] = wszystkie.aggregate(
            liczba_wierszy=Count("id"),
            liczba_autorow=Count("autor", distinct=True),
            srednia_wykorzystania=Avg("procent_wykorzystania_slotow"),
            srednia_pkd_slot=Avg("srednia_za_slot_nazbierana"),
            suma_punktow=Sum("punkty_nazbierane"),
            suma_slotow=Sum("slot_nazbierany"),
        )

        # Najniższe 20 autorów wg PKDaut/slot (nie-zerowych)
        context["bottom_autorzy_pkd"] = (
            MetrykaAutora.objects.select_related(
                "autor", "dyscyplina_naukowa", "jednostka"
            )
            .filter(srednia_za_slot_nazbierana__gt=0)
            .order_by("srednia_za_slot_nazbierana")[:20]
        )

        # Najniższe 20 autorów wg slotów wypełnionych (nie-zerowych)
        context["bottom_autorzy_sloty"] = (
            MetrykaAutora.objects.select_related(
                "autor", "dyscyplina_naukowa", "jednostka"
            )
            .filter(slot_nazbierany__gt=0)
            .order_by("slot_nazbierany")[:20]
        )

        # Autorzy zerowi z latami
        from bpp.models import Autor_Dyscyplina

        autorzy_zerowi_raw = (
            MetrykaAutora.objects.select_related(
                "autor", "dyscyplina_naukowa", "jednostka"
            )
            .filter(srednia_za_slot_nazbierana=0)
            .order_by("autor__nazwisko", "autor__imiona")
        )

        # Dodaj informację o latach dla każdego autora zerowego
        # TYLKO dla lat gdzie rodzaj_autora.jest_w_n = True
        autorzy_zerowi = []
        for metryka in autorzy_zerowi_raw:
            # Pobierz lata, w których autor był przypisany do dyscypliny
            # TYLKO dla rodzajów autorów z jest_w_n = True
            lata_dyscypliny = (
                Autor_Dyscyplina.objects.filter(
                    autor=metryka.autor,
                    dyscyplina_naukowa=metryka.dyscyplina_naukowa,
                    rok__gte=metryka.rok_min,
                    rok__lte=metryka.rok_max,
                    rodzaj_autora__jest_w_n=True,
                )
                .values_list("rok", flat=True)
                .order_by("rok")
            )

            lata_list = list(lata_dyscypliny)
            # Dodaj do listy tylko jeśli autor ma przynajmniej jeden rok z jest_w_n=True
            if lata_list:
                metryka.lata_zerowe = lata_list
                autorzy_zerowi.append(metryka)

        context["autorzy_zerowi"] = autorzy_zerowi

        # Statystyki wg jednostek
        jednostki_stats = (
            MetrykaAutora.objects.values("jednostka__nazwa", "jednostka__skrot")
            .annotate(
                liczba_autorow=Count("id"),
                srednia_wykorzystania=Avg("procent_wykorzystania_slotow"),
                srednia_pkd_slot=Avg("srednia_za_slot_nazbierana"),
                suma_punktow=Sum("punkty_nazbierane"),
            )
            .order_by("-srednia_pkd_slot")[:10]
        )
        context["jednostki_stats"] = jednostki_stats

        # Statystyki wg dyscyplin
        dyscypliny_stats = (
            MetrykaAutora.objects.values(
                "dyscyplina_naukowa__nazwa", "dyscyplina_naukowa__kod"
            )
            .annotate(
                liczba_autorow=Count("id"),
                srednia_wykorzystania=Avg("procent_wykorzystania_slotow"),
                srednia_pkd_slot=Avg("srednia_za_slot_nazbierana"),
                suma_punktow=Sum("punkty_nazbierane"),
            )
            .order_by("-srednia_pkd_slot")
        )
        context["dyscypliny_stats"] = dyscypliny_stats

        # Rozkład wykorzystania slotów
        context["wykorzystanie_ranges"] = {
            "0-25%": wszystkie.filter(procent_wykorzystania_slotow__lt=25).count(),
            "25-50%": wszystkie.filter(
                procent_wykorzystania_slotow__gte=25,
                procent_wykorzystania_slotow__lt=50,
            ).count(),
            "50-75%": wszystkie.filter(
                procent_wykorzystania_slotow__gte=50,
                procent_wykorzystania_slotow__lt=75,
            ).count(),
            "75-99%": wszystkie.filter(
                procent_wykorzystania_slotow__gte=75,
                procent_wykorzystania_slotow__lt=99,
            ).count(),
            "99-100%": wszystkie.filter(procent_wykorzystania_slotow__gte=99).count(),
        }

        return context


@method_decorator(user_passes_test(ma_uprawnienia_ewaluacji), name="dispatch")
class UruchomGenerowanieView(View):
    """Widok do uruchamiania generowania metryk przez Celery task"""

    def post(self, request, *args, **kwargs):
        from django.http import HttpResponse
        from django.shortcuts import render

        # Dodatkowe sprawdzenie uprawnień do generowania (tylko staff)
        if not request.user.is_staff:
            if request.headers.get("HX-Request"):
                # Dla HTMX zwróć błąd
                response = HttpResponse(
                    "Nie masz uprawnień do uruchomienia generowania."
                )
                response["HX-Trigger"] = '{"showError": "No permissions"}'
                return response
            messages.error(request, "Nie masz uprawnień do uruchomienia generowania.")
            return redirect("ewaluacja_metryki:lista")

        # Sprawdź czy generowanie nie jest już w trakcie
        status = StatusGenerowania.get_or_create()
        if status.w_trakcie:
            if request.headers.get("HX-Request"):
                # Dla HTMX zwróć aktualny status
                return render(
                    request,
                    "ewaluacja_metryki/partials/status_box.html",
                    {"status_generowania": status, "progress_procent": 0},
                )
            messages.warning(
                request,
                f"Generowanie jest już w trakcie (rozpoczęte: {status.data_rozpoczecia.strftime('%Y-%m-%d %H:%M:%S')})",
            )
            return redirect("ewaluacja_metryki:lista")

        # Pobierz parametry z formularza (jeśli są)
        rok_min = int(request.POST.get("rok_min", 2022))
        rok_max = int(request.POST.get("rok_max", 2025))
        minimalny_pk = float(request.POST.get("minimalny_pk", 0.01))
        nadpisz = request.POST.get("nadpisz", "on") == "on"

        # Pobierz zaznaczone rodzaje autorów
        rodzaje_autora = request.POST.getlist("rodzaj_autora")
        if not rodzaje_autora:
            # Domyślnie wszystkie rodzaje z licz_sloty=True jeśli nic nie zaznaczono
            rodzaje_autora = list(
                Rodzaj_Autora.objects.filter(licz_sloty=True).values_list(
                    "skrot", flat=True
                )
            )

        # Oblicz liczbę autorów do przetworzenia (aby ustawić status od razu)
        from ewaluacja_liczba_n.models import IloscUdzialowDlaAutoraZaCalosc

        total_count = IloscUdzialowDlaAutoraZaCalosc.objects.all().count()

        # Uruchom task (z domyślnym przeliczaniem liczby N)
        result = generuj_metryki_task.delay(
            rok_min=rok_min,
            rok_max=rok_max,
            minimalny_pk=minimalny_pk,
            nadpisz=nadpisz,
            przelicz_liczbe_n=True,  # Zawsze przeliczaj liczbę N przy generowaniu metryk
            rodzaje_autora=rodzaje_autora,
        )

        # KLUCZOWE: Ustaw status w_trakcie=True OD RAZU w widoku
        # (task będzie tylko aktualizował postęp, nie rozpoczynał od nowa)
        status.rozpocznij_generowanie(
            task_id=str(result.id), liczba_do_przetworzenia=total_count
        )

        # Jeśli to żądanie HTMX, zwróć fragment HTML ze statusem
        if request.headers.get("HX-Request"):
            # Status już jest ustawiony powyżej
            context = {
                "status_generowania": status,
                "progress_procent": 0,
            }
            response = render(
                request, "ewaluacja_metryki/partials/status_box.html", context
            )
            # Używamy angielskiego triggerera aby uniknąć problemów z kodowaniem MIME
            # Komunikat po polsku będzie generowany po stronie JavaScript
            import json

            trigger_data = {"metricsStarted": True, "taskId": str(result.id)}
            response["HX-Trigger"] = json.dumps(trigger_data)
            return response

        # Dla zwykłego żądania zachowaj kompatybilność wsteczną
        messages.success(
            request,
            f"Generowanie metryk zostało uruchomione (ID zadania: {result.id}). "
            "Odśwież stronę za chwilę, aby zobaczyć postęp.",
        )

        return redirect("ewaluacja_metryki:lista")

    def get(self, request, *args, **kwargs):
        # GET przekierowuje do listy
        return redirect("ewaluacja_metryki:lista")


@method_decorator(user_passes_test(ma_uprawnienia_ewaluacji), name="dispatch")
class StatusGenerowaniaView(View):
    """Widok zwracający status generowania jako JSON (dla AJAX)"""

    def get(self, request, *args, **kwargs):
        status = StatusGenerowania.get_or_create()

        return JsonResponse(
            {
                "w_trakcie": status.w_trakcie,
                "data_rozpoczecia": (
                    status.data_rozpoczecia.isoformat()
                    if status.data_rozpoczecia
                    else None
                ),
                "data_zakonczenia": (
                    status.data_zakonczenia.isoformat()
                    if status.data_zakonczenia
                    else None
                ),
                "liczba_przetworzonych": status.liczba_przetworzonych,
                "liczba_bledow": status.liczba_bledow,
                "ostatni_komunikat": status.ostatni_komunikat,
                "procent": (
                    round(
                        (
                            status.liczba_przetworzonych
                            / status.liczba_do_przetworzenia
                            * 100
                        ),
                        1,
                    )
                    if status.w_trakcie and status.liczba_do_przetworzenia > 0
                    else 0
                ),
                "liczba_do_przetworzenia": status.liczba_do_przetworzenia,
            }
        )


@method_decorator(user_passes_test(ma_uprawnienia_ewaluacji), name="dispatch")
class StatusGenerowaniaPartialView(View):
    """Widok zwracający częściowy HTML ze statusem generowania dla HTMX"""

    def get(self, request, *args, **kwargs):
        from django.shortcuts import render

        # Pobierz status generowania
        status = StatusGenerowania.get_or_create()

        # Sprawdź czy poprzedni status był w trakcie (dla odświeżenia strony po zakończeniu)
        previous_status_was_running = (
            request.session.get("generation_was_running", False)
            if not status.w_trakcie
            else False
        )

        # Zapamiętaj aktualny stan
        request.session["generation_was_running"] = status.w_trakcie

        # Oblicz procent postępu
        progress_procent = 0
        if status.w_trakcie and status.liczba_do_przetworzenia > 0:
            progress_procent = round(
                (status.liczba_przetworzonych / status.liczba_do_przetworzenia * 100), 1
            )

        # Kontekst dla szablonu
        context = {
            "status_generowania": status,
            "progress_procent": progress_procent,
            "previous_status_was_running": previous_status_was_running,
        }

        return render(request, "ewaluacja_metryki/partials/status_box.html", context)


@method_decorator(user_passes_test(ma_uprawnienia_ewaluacji), name="dispatch")
class ExportStatystykiXLSX(View):
    """Export statistics tables to XLSX format"""

    def _setup_workbook_and_styles(self):
        """Create workbook and define styles."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        wb = Workbook()
        ws = wb.active

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        return wb, ws, header_font, header_fill, header_alignment, thin_border

    def _create_response(self, wb, table_type):
        """Create HTTP response with workbook."""
        import datetime

        from django.http import HttpResponse

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"metryki_statystyki_{table_type}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    def get(self, request, table_type):
        from django.http import HttpResponse

        from .export_helpers import (
            auto_adjust_column_widths,
            export_bottom_pkd,
            export_bottom_sloty,
            export_dyscypliny,
            export_globalne_stats,
            export_jednostki,
            export_top_autorzy,
            export_top_sloty,
            export_wykorzystanie,
            export_zerowi,
        )

        # Dispatch table to appropriate export handler
        table_handlers = {
            "globalne": export_globalne_stats,
            "top-autorzy": export_top_autorzy,
            "top-sloty": export_top_sloty,
            "bottom-pkd": export_bottom_pkd,
            "bottom-sloty": export_bottom_sloty,
            "zerowi": export_zerowi,
            "jednostki": export_jednostki,
            "dyscypliny": export_dyscypliny,
            "wykorzystanie": export_wykorzystanie,
        }

        if table_type not in table_handlers:
            return HttpResponse("Nieznany typ tabeli", status=400)

        wb, ws, header_font, header_fill, header_alignment, thin_border = (
            self._setup_workbook_and_styles()
        )

        # Call the appropriate handler
        handler = table_handlers[table_type]
        if table_type == "globalne":
            handler(ws, header_font, header_fill, header_alignment)
        else:
            handler(ws, header_font, header_fill, header_alignment, thin_border)

        auto_adjust_column_widths(ws)

        return self._create_response(wb, table_type)


@method_decorator(user_passes_test(ma_uprawnienia_ewaluacji), name="dispatch")
class ExportListaXLSX(View):
    """Export the main metrics list to XLSX format with filters applied"""

    def _setup_workbook_styles(self):
        """Setup workbook, worksheet and define all styles."""
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            NamedStyle,
            PatternFill,
            Side,
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Metryki ewaluacyjne"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        even_row_fill = PatternFill(
            start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"
        )

        # Create named styles for numbers
        try:
            percent_style = NamedStyle(name="percent_style")
            percent_style.number_format = "0.00%"
            percent_style.alignment = Alignment(horizontal="right")
            wb.add_named_style(percent_style)
        except ValueError:
            percent_style = "percent_style"

        try:
            decimal_style = NamedStyle(name="decimal_style")
            decimal_style.number_format = "0.00"
            decimal_style.alignment = Alignment(horizontal="right")
            wb.add_named_style(decimal_style)
        except ValueError:
            decimal_style = "decimal_style"

        return {
            "wb": wb,
            "ws": ws,
            "header_font": header_font,
            "header_fill": header_fill,
            "header_alignment": header_alignment,
            "thin_border": thin_border,
            "even_row_fill": even_row_fill,
            "percent_style": percent_style,
            "decimal_style": decimal_style,
        }

    def _apply_filters_to_queryset(self, queryset, request):
        """Apply all filters from request to queryset."""
        autor_id = request.GET.get("autor_id")
        if autor_id:
            queryset = queryset.filter(autor_id=autor_id)

        nazwisko = request.GET.get("nazwisko")
        if nazwisko:
            queryset = queryset.filter(
                Q(autor__nazwisko__icontains=nazwisko)
                | Q(autor__imiona__icontains=nazwisko)
            )

        jednostka_id = request.GET.get("jednostka")
        if jednostka_id:
            queryset = queryset.filter(jednostka_id=jednostka_id)

        wydzial_id = request.GET.get("wydzial")
        if wydzial_id:
            queryset = queryset.filter(jednostka__wydzial_id=wydzial_id)

        dyscyplina_id = request.GET.get("dyscyplina")
        if dyscyplina_id:
            queryset = queryset.filter(dyscyplina_naukowa_id=dyscyplina_id)

        rodzaj_autora = request.GET.get("rodzaj_autora")
        if rodzaj_autora and rodzaj_autora != "":
            queryset = queryset.filter(rodzaj_autora=rodzaj_autora)

        return queryset

    def _apply_sorting_to_queryset(self, queryset, request):
        """Apply sorting to queryset based on request parameter."""
        sort = request.GET.get("sort", "-srednia_za_slot_nazbierana")
        sort_mapping = {
            "procent_wykorzystania_slotow": (
                "procent_wykorzystania_slotow",
                "srednia_za_slot_nazbierana",
            ),
            "-procent_wykorzystania_slotow": (
                "-procent_wykorzystania_slotow",
                "-srednia_za_slot_nazbierana",
            ),
            "srednia_za_slot_nazbierana": (
                "srednia_za_slot_nazbierana",
                "procent_wykorzystania_slotow",
            ),
            "-srednia_za_slot_nazbierana": (
                "-srednia_za_slot_nazbierana",
                "-procent_wykorzystania_slotow",
            ),
        }

        allowed_sorts = list(sort_mapping.keys()) + [
            "autor__nazwisko",
            "-autor__nazwisko",
        ]

        if sort not in allowed_sorts:
            return queryset

        if sort in sort_mapping:
            return queryset.order_by(*sort_mapping[sort])
        return queryset.order_by(sort)

    def _determine_visible_columns(self):
        """Determine which columns should be visible in export."""
        from bpp.models import Dyscyplina_Naukowa

        uczelnia = Uczelnia.objects.get_default()
        uzywa_wydzialow = uczelnia.uzywaj_wydzialow if uczelnia else False

        wszystkie_dyscypliny = Dyscyplina_Naukowa.objects.filter(
            metrykaautora__isnull=False
        ).distinct()
        tylko_jedna_dyscyplina = wszystkie_dyscypliny.count() == 1

        return {
            "uzywa_wydzialow": uzywa_wydzialow,
            "tylko_jedna_dyscyplina": tylko_jedna_dyscyplina,
        }

    def _create_headers(self, visible_columns):
        """Create header list based on visible columns."""
        headers = [
            "Lp.",
            "Autor",
            "Rodzaj autora",
            "ID systemu kadrowego",
            "ORCID",
            "PBN UID ID",
        ]

        if not visible_columns["tylko_jedna_dyscyplina"]:
            headers.append("Dyscyplina")

        if visible_columns["uzywa_wydzialow"]:
            headers.append("Wydział")

        headers.extend(
            [
                "Jednostka",
                "Slot maksymalny",
                "Slot nazbierany",
                "Slot niewykorzystany",
                "% wykorzystania",
                "PKDaut nazbierane",
                "PKDaut wszystkie",
                "Średnia PKDaut/slot (nazbierane)",
                "Średnia PKDaut/slot (wszystkie)",
                "Liczba prac (nazbierane)",
                "Liczba prac (wszystkie)",
                "Rok min",
                "Rok max",
            ]
        )

        return headers

    def _write_headers(self, ws, headers, styles):
        """Write headers to worksheet with styling."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.alignment = styles["header_alignment"]
            cell.border = styles["thin_border"]

    def _format_rodzaj_autora(self, metryka):
        """Format rodzaj autora display value."""
        if metryka.rodzaj_autora == " ":
            return "Brak danych"

        try:
            rodzaj = Rodzaj_Autora.objects.get(skrot=metryka.rodzaj_autora)
            return rodzaj.nazwa
        except Rodzaj_Autora.DoesNotExist:
            return metryka.rodzaj_autora

    def _write_cell(self, ws, row, col, value, styles, style_type=None, row_fill=None):
        """Write a single cell with consistent styling."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = styles["thin_border"]

        if style_type:
            cell.style = styles[style_type]

        if row_fill:
            cell.fill = row_fill

        return col + 1

    def _write_data_row(self, ws, row_idx, metryka, visible_columns, styles, row_fill):
        """Write a single data row to worksheet."""
        col = 1

        # Lp.
        col = self._write_cell(ws, row_idx, col, row_idx - 1, styles, row_fill=row_fill)

        # Autor
        col = self._write_cell(
            ws, row_idx, col, str(metryka.autor), styles, row_fill=row_fill
        )

        # Rodzaj autora
        col = self._write_cell(
            ws,
            row_idx,
            col,
            self._format_rodzaj_autora(metryka),
            styles,
            row_fill=row_fill,
        )

        # ID systemu kadrowego
        col = self._write_cell(
            ws,
            row_idx,
            col,
            metryka.autor.system_kadrowy_id or "",
            styles,
            row_fill=row_fill,
        )

        # ORCID
        col = self._write_cell(
            ws, row_idx, col, metryka.autor.orcid or "", styles, row_fill=row_fill
        )

        # PBN UID ID
        col = self._write_cell(
            ws, row_idx, col, metryka.autor.pbn_uid_id or "", styles, row_fill=row_fill
        )

        # Dyscyplina (if shown)
        if not visible_columns["tylko_jedna_dyscyplina"]:
            dyscyplina_value = (
                metryka.dyscyplina_naukowa.nazwa if metryka.dyscyplina_naukowa else "-"
            )
            col = self._write_cell(
                ws, row_idx, col, dyscyplina_value, styles, row_fill=row_fill
            )

        # Wydział (if shown)
        if visible_columns["uzywa_wydzialow"]:
            wydzial_nazwa = "-"
            if metryka.jednostka and metryka.jednostka.wydzial:
                wydzial_nazwa = metryka.jednostka.wydzial.nazwa
            col = self._write_cell(
                ws, row_idx, col, wydzial_nazwa, styles, row_fill=row_fill
            )

        # Jednostka
        jednostka_value = metryka.jednostka.nazwa if metryka.jednostka else "-"
        col = self._write_cell(
            ws, row_idx, col, jednostka_value, styles, row_fill=row_fill
        )

        # Numeric values with proper formatting
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.slot_maksymalny),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.slot_nazbierany),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.slot_niewykorzystany),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.procent_wykorzystania_slotow) / 100,
            styles,
            "percent_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.punkty_nazbierane),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.punkty_wszystkie),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.srednia_za_slot_nazbierana),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws,
            row_idx,
            col,
            float(metryka.srednia_za_slot_wszystkie),
            styles,
            "decimal_style",
            row_fill,
        )
        col = self._write_cell(
            ws, row_idx, col, len(metryka.prace_nazbierane), styles, row_fill=row_fill
        )
        col = self._write_cell(
            ws, row_idx, col, metryka.liczba_prac_wszystkie, styles, row_fill=row_fill
        )
        col = self._write_cell(
            ws, row_idx, col, metryka.rok_min, styles, row_fill=row_fill
        )
        col = self._write_cell(
            ws, row_idx, col, metryka.rok_max, styles, row_fill=row_fill
        )

    def _write_all_data_rows(self, ws, queryset, visible_columns, styles):
        """Write all data rows to worksheet."""
        last_data_row = 1
        for row_idx, metryka in enumerate(queryset, 2):
            last_data_row = row_idx
            row_fill = styles["even_row_fill"] if row_idx % 2 == 0 else None
            self._write_data_row(
                ws, row_idx, metryka, visible_columns, styles, row_fill
            )

        return last_data_row

    def _setup_worksheet_formatting(self, ws, headers, last_data_row):
        """Setup auto-filter, freeze panes and column widths."""
        from openpyxl.utils import get_column_letter

        if last_data_row > 1:
            last_col_letter = get_column_letter(len(headers))
            filter_range = f"A1:{last_col_letter}{last_data_row}"
            ws.auto_filter.ref = filter_range

        ws.freeze_panes = ws["A2"]

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except BaseException:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_summary(self, ws, queryset, last_data_row):
        """Add summary section to worksheet."""
        summary_row = last_data_row + 2 if last_data_row > 1 else 3
        ws.cell(row=summary_row, column=1, value="Podsumowanie:")
        ws.cell(row=summary_row + 1, column=1, value="Liczba wierszy:")
        ws.cell(row=summary_row + 1, column=2, value=queryset.count())

        return summary_row

    def _collect_filter_info(self, request, visible_columns):
        """Collect applied filter information for display."""
        from bpp.models import Dyscyplina_Naukowa

        filter_info = []

        nazwisko = request.GET.get("nazwisko")
        if nazwisko:
            filter_info.append(f"Nazwisko/Imię: {nazwisko}")

        jednostka_id = request.GET.get("jednostka")
        if jednostka_id:
            try:
                jednostka = Jednostka.objects.get(pk=jednostka_id)
                filter_info.append(f"Jednostka: {jednostka.nazwa}")
            except Jednostka.DoesNotExist:
                pass

        wydzial_id = request.GET.get("wydzial")
        if wydzial_id and visible_columns["uzywa_wydzialow"]:
            try:
                wydzial = Wydzial.objects.get(pk=wydzial_id)
                filter_info.append(f"Wydział: {wydzial.nazwa}")
            except Wydzial.DoesNotExist:
                pass

        dyscyplina_id = request.GET.get("dyscyplina")
        if dyscyplina_id and not visible_columns["tylko_jedna_dyscyplina"]:
            try:
                dyscyplina = Dyscyplina_Naukowa.objects.get(pk=dyscyplina_id)
                filter_info.append(f"Dyscyplina: {dyscyplina.nazwa}")
            except Dyscyplina_Naukowa.DoesNotExist:
                pass

        rodzaj_autora = request.GET.get("rodzaj_autora")
        if rodzaj_autora:
            try:
                rodzaj = Rodzaj_Autora.objects.get(skrot=rodzaj_autora)
                filter_info.append(f"Rodzaj autora: {rodzaj.nazwa}")
            except Rodzaj_Autora.DoesNotExist:
                filter_info.append(f"Rodzaj autora: {rodzaj_autora}")

        return filter_info

    def _add_filter_info(self, ws, summary_row, filter_info):
        """Add filter information to worksheet."""
        filters_row = summary_row + 3
        ws.cell(row=filters_row, column=1, value="Zastosowane filtry:")

        if filter_info:
            ws.cell(row=filters_row + 1, column=1, value="; ".join(filter_info))
        else:
            ws.cell(row=filters_row + 1, column=1, value="Brak filtrów")

    def _create_response(self, wb):
        """Create HTTP response with workbook."""
        import datetime

        from django.http import HttpResponse

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        filename = f"metryki_ewaluacyjne_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    def get(self, request):
        from django.db.models import Count, OuterRef, Subquery

        # Setup workbook and styles
        styles = self._setup_workbook_styles()
        ws = styles["ws"]
        wb = styles["wb"]

        # Build queryset with discipline count annotation
        discipline_count = (
            MetrykaAutora.objects.filter(autor=OuterRef("autor"))
            .values("autor")
            .annotate(count=Count("dyscyplina_naukowa"))
            .values("count")
        )

        queryset = MetrykaAutora.objects.select_related(
            "autor", "dyscyplina_naukowa", "jednostka", "jednostka__wydzial"
        ).annotate(autor_discipline_count=Subquery(discipline_count))

        # Apply filters and sorting
        queryset = self._apply_filters_to_queryset(queryset, request)
        queryset = self._apply_sorting_to_queryset(queryset, request)

        # Determine visible columns
        visible_columns = self._determine_visible_columns()

        # Create and write headers
        headers = self._create_headers(visible_columns)
        self._write_headers(ws, headers, styles)

        # Write all data rows
        last_data_row = self._write_all_data_rows(ws, queryset, visible_columns, styles)

        # Setup worksheet formatting
        self._setup_worksheet_formatting(ws, headers, last_data_row)

        # Add summary and filter information
        summary_row = self._add_summary(ws, queryset, last_data_row)
        filter_info = self._collect_filter_info(request, visible_columns)
        self._add_filter_info(ws, summary_row, filter_info)

        # Return response
        return self._create_response(wb)
