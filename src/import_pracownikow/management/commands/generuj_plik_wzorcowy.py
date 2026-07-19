"""Generator pliku wzorcowego importu pracowników.

Źródło prawdy dla treści i formatowania pliku, który użytkownik pobiera
przyciskiem „pobierz plik wzorcowy". Regeneracja:

    uv run python src/manage.py generuj_plik_wzorcowy

Zapisuje do ``static/import_pracownikow/import_pracownikow_przyklad.xlsx``.
Ten plik jest ZACOMMITOWANĄ binarką (nie generuje się w runtime) — po każdej
zmianie generatora trzeba go zregenerować i zacommitować.
"""

import os

from django.core.management.base import BaseCommand
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Border, Font, Side

import import_pracownikow

# 15 kanonicznych nagłówków — brzmienie oczyszczone ze śmieci formatujących
# (bez końcowej spacji, bez „\n"). Mapują się 1:1 na pola docelowe importu
# (patrz import_pracownikow.mapping._SYNONIMY). Kolejność = kolejność kolumn.
NAGLOWKI = [
    "Numer",
    "Nazwisko",
    "Imię",
    "ORCID",
    "Tytuł/Stopień",
    "Stanowisko",
    "Grupa pracownicza",
    "Nazwa jednostki",
    "Wydział",
    "Data zatrudnienia",
    "Data końca zatrudnienia",
    "Podstawowe miejsce pracy",
    "PBN UUID",
    "BPP ID",
    "Wymiar etatu",
]

# Podpowiedzi w komentarzach komórek nagłówka (klucz = nagłówek).
_KOMENTARZE = {
    "Numer": "Numer ID pracownika z systemu kadrowego (opcjonalny).",
    "Podstawowe miejsce pracy": "Wpisz TAK lub NIE.",
    "PBN UUID": "Identyfikator użytkownika w systemie PBN (opcjonalny).",
    "BPP ID": "Numer ID w Bazie Publikacji Pracowników (opcjonalny).",
}

# 4 wiersze przykładowe — różne przypadki (patrz spec). None = pusta komórka.
# Kolejność wartości odpowiada NAGLOWKI.
_WIERSZE = [
    # pełny etat, zatrudnienie trwa
    [
        9530,
        "Kowalski",
        "Jan",
        None,
        "lek. med.",
        "Asystent",
        "Badawczo-dydaktyczna",
        "Katedra i Klinika Dermatologii",
        "Wydział Lekarski",
        "2016-10-01",
        None,
        "TAK",
        None,
        None,
        "Pełny etat",
    ],
    # część etatu
    [
        9531,
        "Nowak",
        "Maria",
        "0000-0002-2752-5144",
        "dr n. med.",
        "Adiunkt",
        "Dydaktyczna",
        "Katedra Testowa",
        "Wydział Lekarski",
        "2018-02-01",
        None,
        "NIE",
        None,
        None,
        "1/2 etatu",
    ],
    # zatrudnienie zakończone
    [
        9532,
        "Wiśniewski",
        "Adam",
        None,
        "prof. dr hab.",
        "Profesor",
        "Badawczo-dydaktyczna",
        "Katedra Testowa",
        "Wydział Lekarski",
        "2005-10-01",
        "2020-09-30",
        "TAK",
        None,
        None,
        "Pełny etat",
    ],
    # minimum danych (bez ORCID/PBN/BPP ID)
    [
        None,
        "Lubelska",
        "Anna",
        None,
        None,
        None,
        None,
        "Katedra Testowa",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    ],
]

# Zakładka „Opis kolumn" — układ PIONOWY (nazwa | znaczenie | wymagana?).
# Nazwy kolumn jedna pod drugą w kolumnie A — NIGDY ≥2 nazwy w jednym wierszu,
# inaczej find_similar_row wziąłby to za nagłówek i policzył jako drugi arkusz
# danych (→ BadNoOfSheetsException). Chroni to test kontraktowy (Task 2).
_OPIS = [
    ("Numer", "ID z systemu kadrowego", "opcjonalna"),
    ("Nazwisko", "Nazwisko pracownika", "wymagana"),
    ("Imię", "Imię pracownika", "wymagana"),
    ("ORCID", "Identyfikator ORCID", "opcjonalna"),
    ("Tytuł/Stopień", "Tytuł lub stopień naukowy (np. dr, prof.)", "opcjonalna"),
    # UWAGA: opis NIE może po normalizacji dać dokładnego synonimu z
    # _SYNONIMY (np. „funkcja w jednostce" → funkcja_w_jednostce), bo razem
    # z nazwą w kolumnie A dałby 2 trafienia → wiersz wzięty za nagłówek.
    ("Stanowisko", "Nazwa stanowiska/funkcji", "opcjonalna"),
    ("Grupa pracownicza", "Np. badawczo-dydaktyczna", "opcjonalna"),
    ("Nazwa jednostki", "Pełna nazwa jednostki organizacyjnej", "wymagana"),
    ("Wydział", "Nazwa wydziału", "opcjonalna"),
    ("Data zatrudnienia", "Format RRRR-MM-DD", "opcjonalna"),
    ("Data końca zatrudnienia", "Format RRRR-MM-DD; puste = trwa", "opcjonalna"),
    ("Podstawowe miejsce pracy", "TAK lub NIE", "opcjonalna"),
    ("PBN UUID", "Identyfikator użytkownika w PBN", "opcjonalna"),
    ("BPP ID", "ID w Bazie Publikacji Pracowników", "opcjonalna"),
    ("Wymiar etatu", "Np. Pełny etat, 1/2 etatu", "opcjonalna"),
]

SCIEZKA_DOMYSLNA = os.path.join(
    os.path.dirname(import_pracownikow.__file__),
    "static",
    "import_pracownikow",
    "import_pracownikow_przyklad.xlsx",
)

_RAMKA = Border(*(Side(style="thin") for _ in range(4)))


def zbuduj_workbook() -> Workbook:
    """Buduje workbook pliku wzorcowego (arkusz danych + zakładka opisu)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Pracownicy"

    for col, naglowek in enumerate(NAGLOWKI, start=1):
        cell = ws.cell(row=1, column=col, value=naglowek)
        cell.font = Font(bold=True)
        cell.border = _RAMKA
        if naglowek in _KOMENTARZE:
            cell.comment = Comment(_KOMENTARZE[naglowek], "BPP")

    for r, wiersz in enumerate(_WIERSZE, start=2):
        for c, wartosc in enumerate(wiersz, start=1):
            ws.cell(row=r, column=c, value=wartosc)

    ws.column_dimensions["H"].width = 41

    opis = wb.create_sheet("Opis kolumn")
    opis.cell(row=1, column=1, value="Kolumna").font = Font(bold=True)
    opis.cell(row=1, column=2, value="Znaczenie").font = Font(bold=True)
    opis.cell(row=1, column=3, value="Wymagana?").font = Font(bold=True)
    for r, (nazwa, znaczenie, wym) in enumerate(_OPIS, start=2):
        opis.cell(row=r, column=1, value=nazwa)
        opis.cell(row=r, column=2, value=znaczenie)
        opis.cell(row=r, column=3, value=wym)
    opis.column_dimensions["A"].width = 26
    opis.column_dimensions["B"].width = 44

    return wb


class Command(BaseCommand):
    help = "Generuje plik wzorcowy importu pracowników (XLSX)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            default=None,
            help="Ścieżka wyjściowa (domyślnie: plik static pliku wzorcowego).",
        )

    def handle(self, *args, **options):
        sciezka = options["output"] or SCIEZKA_DOMYSLNA
        zbuduj_workbook().save(sciezka)
        self.stdout.write(f"Zapisano plik wzorcowy: {sciezka}")
