from io import BytesIO

import pytest
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.urls import reverse
from model_bakery import baker
from openpyxl import load_workbook

from bpp.const import GR_WPROWADZANIE_DANYCH
from bpp.models import Autor, Autor_Jednostka, Jednostka
from import_pracownikow.eksport import zapisz_snapshot_po_imporcie
from import_pracownikow.models import ImportPracownikow, ImportPracownikowRow
from import_pracownikow.tests._helpers import unikalna_nazwa


def _user_w_grupie(django_user_model, username="entry"):
    u = django_user_model.objects.create_user(username=username, password="pass")
    grupa, _ = Group.objects.get_or_create(name=GR_WPROWADZANIE_DANYCH)
    u.groups.add(grupa)
    return u


def _import_z_plikiem(owner, nazwa="testdata.xlsx"):
    imp = baker.make(ImportPracownikow, owner=owner)
    imp.plik_xls.save(nazwa, SimpleUploadedFile(nazwa, b"PK\x03\x04udawany"), save=True)
    return imp


@pytest.mark.django_db
def test_oryginal_pobiera_wlasciciel_z_grupa(client, django_user_model):
    u = _user_w_grupie(django_user_model)
    client.force_login(u)
    imp = _import_z_plikiem(u)
    resp = client.get(
        reverse("import_pracownikow:pobierz-oryginal", kwargs={"pk": imp.pk})
    )
    assert resp.status_code == 200
    assert "attachment" in resp["Content-Disposition"]
    assert "testdata.xlsx" in resp["Content-Disposition"]


@pytest.mark.django_db
def test_oryginal_bez_grupy_odmowa(client, django_user_model):
    u = django_user_model.objects.create_user(username="plain", password="pass")
    client.force_login(u)
    imp = _import_z_plikiem(u)
    resp = client.get(
        reverse("import_pracownikow:pobierz-oryginal", kwargs={"pk": imp.pk})
    )
    assert resp.status_code != 200  # braces GroupRequiredMixin blokuje


@pytest.mark.django_db
def test_oryginal_cudzy_import_404(client, django_user_model):
    wlasciciel = _user_w_grupie(django_user_model, "wlasciciel")
    obcy = _user_w_grupie(django_user_model, "obcy")
    imp = _import_z_plikiem(wlasciciel)
    client.force_login(obcy)
    resp = client.get(
        reverse("import_pracownikow:pobierz-oryginal", kwargs={"pk": imp.pk})
    )
    assert resp.status_code == 404


@pytest.mark.django_db
def test_oryginal_brak_pliku_404(client, django_user_model):
    u = _user_w_grupie(django_user_model)
    client.force_login(u)
    imp = baker.make(ImportPracownikow, owner=u)  # bez plik_xls
    resp = client.get(
        reverse("import_pracownikow:pobierz-oryginal", kwargs={"pk": imp.pk})
    )
    assert resp.status_code == 404


@pytest.mark.django_db
def test_po_imporcie_przed_finalizacja_404(client, django_user_model):
    u = _user_w_grupie(django_user_model)
    client.force_login(u)
    imp = baker.make(
        ImportPracownikow, owner=u, stan=ImportPracownikow.STAN_PRZEANALIZOWANY
    )
    resp = client.get(
        reverse("import_pracownikow:pobierz-po-imporcie", kwargs={"pk": imp.pk})
    )
    assert resp.status_code == 404


@pytest.mark.django_db
def test_po_imporcie_po_finalizacji_zwraca_xlsx(client, django_user_model):
    u = _user_w_grupie(django_user_model)
    client.force_login(u)
    imp = _import_z_plikiem(u, nazwa="pracownicy_2026.xlsx")
    imp.stan = ImportPracownikow.STAN_ZINTEGROWANY
    imp.mapowanie_kolumn = {
        "Nazwisko": "nazwisko",
        "Imię": "imię",
        "Jednostka": "nazwa_jednostki",
    }
    imp.save()
    resp = client.get(
        reverse("import_pracownikow:pobierz-po-imporcie", kwargs={"pk": imp.pk})
    )
    assert resp.status_code == 200
    assert "pracownicy_2026-po-imporcie.xlsx" in resp["Content-Disposition"]
    ws = load_workbook(BytesIO(resp.getvalue())).active
    assert [c.value for c in ws[1]][:4] == [
        "BPP ID",
        "Nazwisko",
        "Imię",
        "Nazwa jednostki",
    ]


@pytest.mark.django_db
def test_po_imporcie_nazwa_z_polskimi_znakami(client, django_user_model):
    u = _user_w_grupie(django_user_model)
    client.force_login(u)
    imp = _import_z_plikiem(u, nazwa="wydział_lekarski.xlsx")
    imp.stan = ImportPracownikow.STAN_ZINTEGROWANY
    imp.mapowanie_kolumn = {
        "Nazwisko": "nazwisko",
        "Imię": "imię",
        "Jednostka": "nazwa_jednostki",
    }
    imp.save()
    resp = client.get(
        reverse("import_pracownikow:pobierz-po-imporcie", kwargs={"pk": imp.pk})
    )
    assert resp.status_code == 200
    disposition = resp["Content-Disposition"]
    assert "filename*=utf-8''" in disposition.lower()
    assert "=?utf-8?" not in disposition
    assert "%C5%82" in disposition


@pytest.mark.django_db
def test_po_imporcie_cudzy_import_404(client, django_user_model):
    wlasciciel = _user_w_grupie(django_user_model, "wlasciciel")
    obcy = _user_w_grupie(django_user_model, "obcy")
    imp = _import_z_plikiem(wlasciciel, nazwa="pracownicy_2026.xlsx")
    imp.stan = ImportPracownikow.STAN_ZINTEGROWANY
    imp.mapowanie_kolumn = {
        "Nazwisko": "nazwisko",
        "Imię": "imię",
        "Jednostka": "nazwa_jednostki",
    }
    imp.save()
    client.force_login(obcy)
    resp = client.get(
        reverse("import_pracownikow:pobierz-po-imporcie", kwargs={"pk": imp.pk})
    )
    assert resp.status_code == 404


@pytest.mark.django_db
def test_rezultaty_pokazuje_przyciski(client, django_user_model):
    u = _user_w_grupie(django_user_model)
    client.force_login(u)
    imp = _import_z_plikiem(u)
    imp.stan = ImportPracownikow.STAN_ZINTEGROWANY
    imp.save()
    resp = client.get(
        reverse("import_pracownikow:importpracownikow-results", kwargs={"pk": imp.pk})
    )
    tresc = resp.content.decode()
    assert (
        reverse("import_pracownikow:pobierz-oryginal", kwargs={"pk": imp.pk}) in tresc
    )
    assert (
        reverse("import_pracownikow:pobierz-po-imporcie", kwargs={"pk": imp.pk})
        in tresc
    )


@pytest.mark.django_db
def test_po_imporcie_dziala_gdy_plik_xls_pusty_po_cleanupie(client, django_user_model):
    # `usun_stare_pliki_importu_pracownikow` (housekeeping 90-dniowy) czyści
    # `plik_xls` po fakcie, ale ZOSTAWIA import + wiersze. Builder
    # (`eksport.zbuduj_plik_po_imporcie`) nie czyta `plik_xls` — czyta z
    # `Autor`/`Autor_Jednostka` — więc link „po imporcie” musi nadal działać,
    # a nazwa pliku spada do fallbacku `import-<pk>-po-imporcie.xlsx`.
    u = _user_w_grupie(django_user_model)
    client.force_login(u)
    imp = baker.make(
        ImportPracownikow,
        owner=u,
        stan=ImportPracownikow.STAN_ZINTEGROWANY,
        mapowanie_kolumn={
            "Nazwisko": "nazwisko",
            "Imię": "imię",
            "Jednostka": "nazwa_jednostki",
        },
    )
    # Symuluj stan PO housekeepingu — dokładnie to, co robi komenda porządkowa.
    imp.plik_xls = ""
    imp.save(update_fields=["plik_xls"])

    resp = client.get(
        reverse("import_pracownikow:pobierz-po-imporcie", kwargs={"pk": imp.pk})
    )

    assert resp.status_code == 200
    assert f"import-{imp.pk}-po-imporcie.xlsx" in resp["Content-Disposition"]


@pytest.mark.django_db
def test_rezultaty_ukrywa_po_imporcie_przed_finalizacja(client, django_user_model):
    u = _user_w_grupie(django_user_model)
    client.force_login(u)
    imp = _import_z_plikiem(u)
    imp.stan = ImportPracownikow.STAN_PRZEANALIZOWANY
    imp.save()
    resp = client.get(
        reverse("import_pracownikow:importpracownikow-results", kwargs={"pk": imp.pk})
    )
    tresc = resp.content.decode()
    assert (
        reverse("import_pracownikow:pobierz-oryginal", kwargs={"pk": imp.pk}) in tresc
    )  # oryginał zawsze
    assert (
        reverse("import_pracownikow:pobierz-po-imporcie", kwargs={"pk": imp.pk})
        not in tresc
    )  # po-imporcie ukryty


# --- Fix 4: niezalogowany użytkownik nie pobiera plików -------------------


@pytest.mark.django_db
def test_oryginal_niezalogowany_nie_pobiera(client, django_user_model):
    u = _user_w_grupie(django_user_model)
    imp = _import_z_plikiem(u)
    resp = client.get(
        reverse("import_pracownikow:pobierz-oryginal", kwargs={"pk": imp.pk})
    )
    assert resp.status_code != 200  # braces GroupRequiredMixin → redirect login


@pytest.mark.django_db
def test_po_imporcie_niezalogowany_nie_pobiera(client, django_user_model):
    u = _user_w_grupie(django_user_model)
    imp = _import_z_plikiem(u, nazwa="pracownicy_2026.xlsx")
    imp.stan = ImportPracownikow.STAN_ZINTEGROWANY
    imp.mapowanie_kolumn = {
        "Nazwisko": "nazwisko",
        "Imię": "imię",
        "Jednostka": "nazwa_jednostki",
    }
    imp.save()
    resp = client.get(
        reverse("import_pracownikow:pobierz-po-imporcie", kwargs={"pk": imp.pk})
    )
    assert resp.status_code != 200  # braces GroupRequiredMixin → redirect login


# --- Immutable snapshot at finalization ------------------------------------


@pytest.mark.django_db
def test_po_imporcie_serwuje_zamrozony_snapshot_mimo_pozniejszej_edycji_autora(
    client, django_user_model
):
    # Kluczowa własność: plik pobrany PO edycji autora nadal ma STARE
    # (zamrożone przy finalizacji) dane — snapshot jest niezmienny.
    u = _user_w_grupie(django_user_model)
    client.force_login(u)
    j = baker.make(Jednostka, nazwa=unikalna_nazwa("Klinika Immutable"))
    a = baker.make(Autor, nazwisko="Zamrozony", imiona="Jan")
    aj = baker.make(Autor_Jednostka, autor=a, jednostka=j)
    imp = _import_z_plikiem(u, nazwa="pracownicy_2026.xlsx")
    imp.stan = ImportPracownikow.STAN_ZINTEGROWANY
    imp.mapowanie_kolumn = {
        "Nazwisko": "nazwisko",
        "Imię": "imię",
        "Jednostka": "nazwa_jednostki",
    }
    imp.save()
    ImportPracownikowRow.objects.create(
        parent=imp,
        autor=a,
        autor_jednostka=aj,
        zmiany_potrzebne=False,
        dane_z_xls={"__xls_loc_sheet__": 0, "__xls_loc_row__": 0},
    )

    # Zamroź snapshot Z NAZWISKIEM SPRZED edycji.
    zapisz_snapshot_po_imporcie(imp)

    # Edycja PO finalizacji — snapshot nie powinien tego zobaczyć.
    a.nazwisko = "Zmieniony"
    a.save(update_fields=["nazwisko"])

    resp = client.get(
        reverse("import_pracownikow:pobierz-po-imporcie", kwargs={"pk": imp.pk})
    )
    assert resp.status_code == 200
    ws = load_workbook(BytesIO(resp.getvalue())).active
    nazwiska = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "Zamrozony" in nazwiska
    assert "Zmieniony" not in nazwiska


@pytest.mark.django_db
def test_po_imporcie_bez_snapshotu_buduje_w_locie_fallback(client, django_user_model):
    # Import zfinalizowany BEZ snapshotu (sprzed wprowadzenia mechanizmu, albo
    # błąd generacji) musi nadal dawać 200 — degradacja do budowy w locie.
    u = _user_w_grupie(django_user_model)
    client.force_login(u)
    # Nazwa unikalna (nie "pracownicy_2026.xlsx" jak sąsiednie testy) — inaczej
    # współdzielony MEDIA_ROOT workera dedupikuje storage suffiksem i psuje
    # asercję na dosłownej nazwie w Content-Disposition.
    imp = _import_z_plikiem(u, nazwa="pracownicy_2026_fallback.xlsx")
    imp.stan = ImportPracownikow.STAN_ZINTEGROWANY
    imp.mapowanie_kolumn = {
        "Nazwisko": "nazwisko",
        "Imię": "imię",
        "Jednostka": "nazwa_jednostki",
    }
    imp.save()
    assert not imp.plik_po_imporcie

    resp = client.get(
        reverse("import_pracownikow:pobierz-po-imporcie", kwargs={"pk": imp.pk})
    )
    assert resp.status_code == 200
    assert "pracownicy_2026_fallback-po-imporcie.xlsx" in resp["Content-Disposition"]
    ws = load_workbook(BytesIO(resp.getvalue())).active
    assert [c.value for c in ws[1]][:4] == [
        "BPP ID",
        "Nazwisko",
        "Imię",
        "Nazwa jednostki",
    ]
