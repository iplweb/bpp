import uuid
from io import BytesIO

import pytest
from model_bakery import baker
from openpyxl import load_workbook

from bpp.models import Autor, Autor_Jednostka, Jednostka, StopienSluzbowy
from import_common.util import normalize_cell_header
from import_pracownikow.eksport import (
    zapisz_snapshot_po_imporcie,
    zbuduj_plik_po_imporcie,
)
from import_pracownikow.mapping import (
    POLE_POMIN,
    waliduj_mapowanie,
    zaproponuj_mapowanie,
)
from import_pracownikow.models import ImportPracownikow, ImportPracownikowRow
from import_pracownikow.tests._helpers import unikalna_nazwa
from pbn_api.models import Scientist


def _wczytaj(content):
    ws = load_workbook(BytesIO(content)).active
    naglowki = [c.value for c in ws[1]]
    wiersze = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    return naglowki, wiersze


def _import_zintegrowany(**kw):
    return baker.make(
        ImportPracownikow,
        stan=ImportPracownikow.STAN_ZINTEGROWANY,
        mapowanie_kolumn=kw.pop("mapowanie_kolumn", {}),
        **kw,
    )


def _wiersz(imp, *, loc, autor=None, autor_jednostka=None, dane=None):
    return baker.make(
        ImportPracownikowRow,
        parent=imp,
        autor=autor,
        autor_jednostka=autor_jednostka,
        zmiany_potrzebne=False,
        dane_z_xls={"__xls_loc_sheet__": 0, "__xls_loc_row__": loc, **(dane or {})},
    )


@pytest.mark.django_db
def test_builder_pomija_wiersze_bez_autora_i_zachowuje_kolejnosc():
    imp = _import_zintegrowany(
        mapowanie_kolumn={
            "Nazwisko": "nazwisko",
            "Imię": "imię",
            "Jednostka": "nazwa_jednostki",
        }
    )
    j = baker.make(Jednostka, nazwa=unikalna_nazwa("Klinika Poprawna"))
    a2 = baker.make(Autor, nazwisko="Druga", imiona="Anna")
    aj2 = baker.make(Autor_Jednostka, autor=a2, jednostka=j)
    a1 = baker.make(Autor, nazwisko="Pierwszy", imiona="Jan")
    aj1 = baker.make(Autor_Jednostka, autor=a1, jednostka=j)
    # loc rosnąco: a1 (loc=0), a2 (loc=1); wiersz pominięty (loc=2, autor=None)
    _wiersz(imp, loc=0, autor=a1, autor_jednostka=aj1)
    _wiersz(imp, loc=1, autor=a2, autor_jednostka=aj2)
    _wiersz(imp, loc=2, autor=None, autor_jednostka=None)

    naglowki, wiersze = _wczytaj(zbuduj_plik_po_imporcie(imp))

    assert len(wiersze) == 2  # pominięty wypadł
    assert naglowki[:4] == ["BPP ID", "Nazwisko", "Imię", "Nazwa jednostki"]
    assert [w[0] for w in wiersze] == [a1.pk, a2.pk]  # kolejność z pliku
    assert [w[1] for w in wiersze] == ["Pierwszy", "Druga"]


@pytest.mark.django_db
def test_ignorowane_kolumny_znikaja_uzyte_zostaja():
    imp = _import_zintegrowany(
        mapowanie_kolumn={
            "Nazwisko": "nazwisko",
            "Imię": "imię",
            "Jedn org": "nazwa_jednostki",
            "Dyscyplina": "__pomin__",  # ignorowana
            "Tytuł nauk.": "tytuł_stopień",  # użyta
        }
    )
    j = baker.make(Jednostka, nazwa=unikalna_nazwa("Katedra X"))
    a = baker.make(Autor, nazwisko="Nowak", imiona="Ewa")
    aj = baker.make(Autor_Jednostka, autor=a, jednostka=j)
    _wiersz(
        imp,
        loc=0,
        autor=a,
        autor_jednostka=aj,
        dane={"Dyscyplina": "nauki medyczne", "Tytuł nauk.": "dr"},
    )

    naglowki, wiersze = _wczytaj(zbuduj_plik_po_imporcie(imp))

    assert "Tytuł" in naglowki
    assert "Stopień służbowy" not in naglowki  # nieużyty target → brak kolumny
    assert "Dyscyplina" not in naglowki
    # Prawdziwa gwarancja: wartość zignorowanej kolumny nie wycieka
    # do ŻADNEJ komórki danych w wyeksportowanym pliku.
    assert not any(
        "nauki medyczne" in str(cell)
        for wiersz in wiersze
        for cell in wiersz
        if cell is not None
    )


@pytest.mark.django_db
def test_wartosc_skorygowana_wygrywa_z_plikiem():
    # Plik miał błędną nazwę jednostki; baza ma poprawną → w pliku wynikowym
    # jest wartość z BAZY.
    imp = _import_zintegrowany(
        mapowanie_kolumn={
            "Nazwisko": "nazwisko",
            "Imię": "imię",
            "Jednostka": "nazwa_jednostki",
        }
    )
    poprawna = unikalna_nazwa("Klinika Chorób Wewnętrznych")
    j = baker.make(Jednostka, nazwa=poprawna)
    a = baker.make(Autor, nazwisko="Kowalski", imiona="Jan")
    aj = baker.make(Autor_Jednostka, autor=a, jednostka=j)
    _wiersz(
        imp,
        loc=0,
        autor=a,
        autor_jednostka=aj,
        dane={"Jednostka": "klin chor wewn", "Nazwisko": "Kowalksi"},
    )

    naglowki, wiersze = _wczytaj(zbuduj_plik_po_imporcie(imp))

    kol = {n: i for i, n in enumerate(naglowki)}
    assert wiersze[0][kol["Nazwa jednostki"]] == poprawna  # nie "klin chor wewn"
    assert wiersze[0][kol["Nazwisko"]] == "Kowalski"  # nie "Kowalksi"


@pytest.mark.django_db
def test_id_enrich_orcid_gdy_niepusty_mimo_braku_mapowania():
    imp = _import_zintegrowany(
        mapowanie_kolumn={
            "Nazwisko": "nazwisko",
            "Imię": "imię",
            "Jednostka": "nazwa_jednostki",
        }
    )
    j = baker.make(Jednostka, nazwa=unikalna_nazwa("Zakład Y"))
    a = baker.make(Autor, nazwisko="Test", imiona="Orc", orcid="0000-0002-1825-0097")
    aj = baker.make(Autor_Jednostka, autor=a, jednostka=j)
    _wiersz(imp, loc=0, autor=a, autor_jednostka=aj)

    naglowki, wiersze = _wczytaj(zbuduj_plik_po_imporcie(imp))

    assert "ORCID" in naglowki  # niepusty ORCID → kolumna mimo braku w mapowaniu
    kol = {n: i for i, n in enumerate(naglowki)}
    assert wiersze[0][kol["ORCID"]] == "0000-0002-1825-0097"


@pytest.mark.django_db
def test_autor_bez_zatrudnienia_wchodzi_z_pustymi_polami_zatrudnienia():
    imp = _import_zintegrowany(
        mapowanie_kolumn={
            "Nazwisko": "nazwisko",
            "Imię": "imię",
            "Jednostka": "nazwa_jednostki",
            "Etat": "wymiar_etatu_tekst",
        }
    )
    a = baker.make(Autor, nazwisko="Sam", imiona="Autor")
    _wiersz(imp, loc=0, autor=a, autor_jednostka=None)

    naglowki, wiersze = _wczytaj(zbuduj_plik_po_imporcie(imp))

    assert len(wiersze) == 1  # autor wszedł, choć bez zatrudnienia
    kol = {n: i for i, n in enumerate(naglowki)}
    assert wiersze[0][kol["BPP ID"]] == a.pk
    assert wiersze[0][kol["Nazwa jednostki"]] in (None, "")  # AJ None → puste
    assert wiersze[0][kol["Wymiar etatu"]] in (None, "")


BLOK_ZATRUDNIENIA = [
    "Funkcja w jednostce",
    "Stanowisko dydaktyczne",
    "Grupa pracownicza",
    "Wymiar etatu",
    "Data zatrudnienia",
    "Data końca zatrudnienia",
    "Podstawowe miejsce pracy",
]


@pytest.mark.django_db
def test_blok_zatrudnienia_zawsze_obecny_mimo_braku_mapowania_i_zatrudnienia():
    # Kontrakt: cały blok Autor_Jednostka jest ZAWSZE w pliku — user ma gdzie
    # ręcznie wpisać czas pracy od-do, funkcję i stanowisko, nawet gdy plik
    # wejściowy tych kolumn nie zawierał i autor nie ma jeszcze zatrudnienia.
    imp = _import_zintegrowany(
        mapowanie_kolumn={
            "Nazwisko": "nazwisko",
            "Imię": "imię",
            "Jednostka": "nazwa_jednostki",
        }
    )
    a = baker.make(Autor, nazwisko="Bez", imiona="Etatu")
    _wiersz(imp, loc=0, autor=a, autor_jednostka=None)

    naglowki, wiersze = _wczytaj(zbuduj_plik_po_imporcie(imp))

    for kol in BLOK_ZATRUDNIENIA:
        assert kol in naglowki, f"Brak zawsze-obecnej kolumny: {kol}"
    # Bez zatrudnienia komórki są puste („do wpisania"), ale kolumna istnieje.
    idx = {n: i for i, n in enumerate(naglowki)}
    for kol in BLOK_ZATRUDNIENIA:
        assert wiersze[0][idx[kol]] in (None, "")


@pytest.mark.django_db
def test_round_trip_naglowki_auto_mapuja_sie():
    # Plik „po imporcie" musi re-importować się bez ręcznego mapowania:
    # każdy nagłówek rozpoznany + walidacja mapowania bez błędów.
    imp = _import_zintegrowany(
        mapowanie_kolumn={
            "Nazwisko": "nazwisko",
            "Imię": "imię",
            "Jednostka": "nazwa_jednostki",
            "Tytuł": "tytuł_stopień",
            "Stopień sł.": "stopień_służbowy",
            "Funkcja": "stanowisko",
            "St. dyd.": "stanowisko_dydaktyczne",
            "Grupa": "grupa_pracownicza",
            "Etat": "wymiar_etatu_tekst",
            "Od": "data_zatrudnienia",
            "Do": "data_końca_zatrudnienia",
            "Gł.": "podstawowe_miejsce_pracy",
            "Mail": "email",
            "Nr": "numer",
        }
    )
    j = baker.make(Jednostka, nazwa=unikalna_nazwa("Klinika RT"))
    stopien = baker.make(StopienSluzbowy)
    a = baker.make(
        Autor,
        nazwisko="Rt",
        imiona="Test",
        orcid="0000-0002-1825-0097",
        stopien_sluzbowy=stopien,
    )
    aj = baker.make(Autor_Jednostka, autor=a, jednostka=j)
    _wiersz(imp, loc=0, autor=a, autor_jednostka=aj)

    naglowki, _ = _wczytaj(zbuduj_plik_po_imporcie(imp))

    # Realny re-import normalizuje nagłówki PRZED zaproponuj_mapowanie
    # (otworz_zrodlo(...).data() -> find_similar_row_in_rows ->
    # normalize_cell_header) — tu robimy to samo, żeby test odzwierciedlał
    # faktyczną ścieżkę re-importu, a nie porównywał surowy (Title Case)
    # nagłówek pliku wprost z kluczami _SYNONIMY (zawsze lowercase/"_").
    znormalizowane = [normalize_cell_header(h) for h in naglowki]

    mapowanie = zaproponuj_mapowanie(znormalizowane)
    nierozpoznane = [h for h, cel in mapowanie.items() if cel == POLE_POMIN]
    assert nierozpoznane == [], f"Nierozpoznane nagłówki: {nierozpoznane}"
    assert waliduj_mapowanie(mapowanie) == []


# --- Fix 1: sanityzacja formuł XLSX (OWASP Formula/CSV Injection) ---------


@pytest.mark.parametrize("lead", ["=", "+", "-", "@", "\t"])
@pytest.mark.django_db
def test_wartosc_zaczynajaca_sie_od_znaku_formuly_jest_sanityzowana(lead):
    # Nazwisko I nazwa jednostki, oba zaczynające się od znaku, który Excel/
    # LibreOffice interpretuje jako formułę (albo separator wstrzykujący do
    # innej komórki) — builder MUSI je sanityzować (prefiks apostrofu), inaczej
    # otwarcie „skorygowanego" pliku w Excelu wykonuje wstrzykniętą formułę.
    imp = _import_zintegrowany(
        mapowanie_kolumn={
            "Nazwisko": "nazwisko",
            "Imię": "imię",
            "Jednostka": "nazwa_jednostki",
        }
    )
    j = baker.make(Jednostka, nazwa=unikalna_nazwa(f"{lead}Klinika"))
    a = baker.make(Autor, nazwisko=f"{lead}Groźne", imiona="Jan")
    aj = baker.make(Autor_Jednostka, autor=a, jednostka=j)
    _wiersz(imp, loc=0, autor=a, autor_jednostka=aj)

    naglowki, wiersze = _wczytaj(zbuduj_plik_po_imporcie(imp))

    kol = {n: i for i, n in enumerate(naglowki)}
    nazwisko_cell = wiersze[0][kol["Nazwisko"]]
    jednostka_cell = wiersze[0][kol["Nazwa jednostki"]]
    assert nazwisko_cell.startswith("'")
    assert jednostka_cell.startswith("'")
    # Wartość poza apostrofem musi zostać zachowana (bez utraty danych).
    assert nazwisko_cell == f"'{lead}Groźne"
    assert jednostka_cell == f"'{j.nazwa}"


@pytest.mark.django_db
def test_naglowki_nie_sa_sanityzowane():
    # Nagłówki to nasze własne stałe kanoniczne (żaden nie zaczyna się od
    # znaku formuły) i MUSZĄ zostać bajt-w-bajt — inaczej auto-mapowanie przy
    # re-imporcie (rozpoznawanie po dosłownym tekście nagłówka) by się zepsuło.
    imp = _import_zintegrowany(
        mapowanie_kolumn={
            "Nazwisko": "nazwisko",
            "Imię": "imię",
            "Jednostka": "nazwa_jednostki",
        }
    )
    j = baker.make(Jednostka, nazwa=unikalna_nazwa("Klinika Zwykła"))
    a = baker.make(Autor, nazwisko="Zwykły", imiona="Jan")
    aj = baker.make(Autor_Jednostka, autor=a, jednostka=j)
    _wiersz(imp, loc=0, autor=a, autor_jednostka=aj)

    naglowki, _ = _wczytaj(zbuduj_plik_po_imporcie(imp))

    assert naglowki[:4] == ["BPP ID", "Nazwisko", "Imię", "Nazwa jednostki"]
    assert not any(str(h).startswith("'") for h in naglowki)


# --- Fix 3: hardening długości PBN UUID (AutorForm.pbn_uuid wymaga 24) ----


@pytest.mark.django_db
def test_pbn_uuid_24_znaki_jest_emitowany():
    imp = _import_zintegrowany(
        mapowanie_kolumn={
            "Nazwisko": "nazwisko",
            "Imię": "imię",
            "Jednostka": "nazwa_jednostki",
        }
    )
    j = baker.make(Jednostka, nazwa=unikalna_nazwa("Klinika PBN"))
    mongo_id = uuid.uuid4().hex[:24]
    scientist = baker.make(Scientist, mongoId=mongo_id)
    a = baker.make(Autor, nazwisko="Pbn", imiona="Poprawny", pbn_uid=scientist)
    aj = baker.make(Autor_Jednostka, autor=a, jednostka=j)
    _wiersz(imp, loc=0, autor=a, autor_jednostka=aj)

    naglowki, wiersze = _wczytaj(zbuduj_plik_po_imporcie(imp))

    assert "PBN UUID" in naglowki
    kol = {n: i for i, n in enumerate(naglowki)}
    assert wiersze[0][kol["PBN UUID"]] == mongo_id


@pytest.mark.django_db
def test_pbn_uuid_nietypowej_dlugosci_jest_pomijany():
    # Scientist.mongoId dopuszcza inne długości niż 24 (max_length=32);
    # AutorForm.pbn_uuid wymaga DOKŁADNIE 24 — nietypowa wartość musi zostać
    # wyciszona (pusty string), inaczej re-import całego pliku pada fail-fast.
    imp = _import_zintegrowany(
        mapowanie_kolumn={
            "Nazwisko": "nazwisko",
            "Imię": "imię",
            "Jednostka": "nazwa_jednostki",
            "PBN": "pbn_uuid",
        }
    )
    j = baker.make(Jednostka, nazwa=unikalna_nazwa("Klinika PBN Zła"))
    scientist = baker.make(Scientist, mongoId=f"krotki-{uuid.uuid4().hex[:8]}")
    a = baker.make(Autor, nazwisko="Pbn", imiona="Zly", pbn_uid=scientist)
    aj = baker.make(Autor_Jednostka, autor=a, jednostka=j)
    _wiersz(imp, loc=0, autor=a, autor_jednostka=aj)

    naglowki, wiersze = _wczytaj(zbuduj_plik_po_imporcie(imp))

    kol = {n: i for i, n in enumerate(naglowki)}
    assert wiersze[0][kol["PBN UUID"]] in (None, "")


# --- Snapshot „po imporcie" (immutable finalization) -----------------------


@pytest.mark.django_db
def test_zapisz_snapshot_populuje_pole_i_zgadza_sie_z_builderem():
    imp = _import_zintegrowany(
        mapowanie_kolumn={
            "Nazwisko": "nazwisko",
            "Imię": "imię",
            "Jednostka": "nazwa_jednostki",
        }
    )
    j = baker.make(Jednostka, nazwa=unikalna_nazwa("Klinika Snapshot"))
    a1 = baker.make(Autor, nazwisko="Snap", imiona="Jeden")
    aj1 = baker.make(Autor_Jednostka, autor=a1, jednostka=j)
    a2 = baker.make(Autor, nazwisko="Shot", imiona="Dwa")
    aj2 = baker.make(Autor_Jednostka, autor=a2, jednostka=j)
    _wiersz(imp, loc=0, autor=a1, autor_jednostka=aj1)
    _wiersz(imp, loc=1, autor=a2, autor_jednostka=aj2)

    zapisz_snapshot_po_imporcie(imp)
    imp.refresh_from_db()

    assert imp.plik_po_imporcie
    with imp.plik_po_imporcie.open("rb") as f:
        zapisane = f.read()
    # Porownanie SEMANTYCZNE (naglowki + wiersze), nie bajt-w-bajt: openpyxl
    # pieczetuje kazdy zapis czasem — ``dcterms:created`` w docProps/core.xml
    # (rozdzielczosc 1 s) oraz ``date_time`` kazdego wpisu ZIP-a. Dwa buildy
    # tej samej tresci roznia sie wiec bajtami, gdy tylko przekrocza granice
    # sekundy. Lokalnie oba buildy mieszcza sie w tej samej sekundzie i test
    # przechodzil; na obciazonym CI potrafil rozjechac sie o sekunde i pekac
    # jako ``assert b'PK\x03\x04...' == b'PK\x03\x04...'``. Intencja testu to
    # „snapshot zawiera to samo, co zwraca builder" — tresc, nie identycznosc
    # kontenera ZIP.
    naglowki, wiersze = _wczytaj(zapisane)
    assert (naglowki, wiersze) == _wczytaj(zbuduj_plik_po_imporcie(imp))
    # Sanity na kilku komorkach — snapshot odbija autorow z bazy.
    plaskie = {str(c) for w in wiersze for c in w}
    assert {"Snap", "Jeden", "Shot", "Dwa"} <= plaskie
