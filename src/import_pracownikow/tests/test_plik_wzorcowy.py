import os

from import_common.sources import otworz_zrodlo
from import_pracownikow.management.commands.generuj_plik_wzorcowy import (
    SCIEZKA_DOMYSLNA,
)
from import_pracownikow.mapping import (
    MIN_POINTS,
    POLE_POMIN,
    TRY_NAMES,
    sprawdz_pojedynczy_arkusz,
    waliduj_mapowanie,
    zaproponuj_mapowanie,
)

_LOC = ("__xls_loc_sheet__", "__xls_loc_row__")


def _tekst_komentarza(cell):
    """Tekst komentarza komórki albo None — do porównania drift-guardem."""
    return cell.comment.text if cell.comment is not None else None


def _naglowki_ze_zrodla(zrodlo):
    pierwszy = next(iter(zrodlo.data()))
    return [k for k in pierwszy.keys() if k not in _LOC]


def test_plik_wzorcowy_nie_jest_symlinkiem():
    # Rozdzielenie od fixture'a testowego (testdata.xlsx) musi się utrzymać.
    assert not os.path.islink(SCIEZKA_DOMYSLNA)


def test_plik_wzorcowy_ma_dokladnie_jeden_arkusz_z_danymi():
    # Zakładka „Opis kolumn" NIE może wpaść w fuzzy-detekcję nagłówka
    # (inaczej sprawdz_pojedynczy_arkusz podniósłby BadNoOfSheetsException).
    zrodlo = otworz_zrodlo(SCIEZKA_DOMYSLNA, try_names=TRY_NAMES, min_points=MIN_POINTS)
    assert zrodlo.liczba_arkuszy_z_danymi() == 1
    sprawdz_pojedynczy_arkusz(zrodlo)  # nie podnosi wyjątku


def test_plik_wzorcowy_mapuje_wszystkie_kolumny():
    zrodlo = otworz_zrodlo(SCIEZKA_DOMYSLNA, try_names=TRY_NAMES, min_points=MIN_POINTS)
    naglowki = _naglowki_ze_zrodla(zrodlo)
    mapowanie = zaproponuj_mapowanie(naglowki)
    assert POLE_POMIN not in mapowanie.values(), (
        f"nierozpoznane kolumny: {[h for h, c in mapowanie.items() if c == POLE_POMIN]}"
    )


def test_plik_wzorcowy_przechodzi_walidacje_mapowania():
    zrodlo = otworz_zrodlo(SCIEZKA_DOMYSLNA, try_names=TRY_NAMES, min_points=MIN_POINTS)
    naglowki = _naglowki_ze_zrodla(zrodlo)
    mapowanie = zaproponuj_mapowanie(naglowki)
    assert waliduj_mapowanie(mapowanie) == []


def test_naglowki_bez_smieci_formatujacych():
    # Otwarcie źródła nie może podnosić wyjątku (spójność z resztą testów).
    otworz_zrodlo(SCIEZKA_DOMYSLNA, try_names=TRY_NAMES, min_points=MIN_POINTS)
    # Nagłówki są znormalizowane w źródle; sprawdzamy surowe komórki wprost.
    import openpyxl

    ws = openpyxl.load_workbook(SCIEZKA_DOMYSLNA)["Pracownicy"]
    surowe = [ws.cell(row=1, column=c).value for c in range(1, 16)]
    for h in surowe:
        assert "\n" not in h
        assert h == h.strip()


def test_plik_na_dysku_odzwierciedla_generator():
    # Binarka na dysku musi odzwierciedlać aktualny generator — inaczej
    # edycja generatora bez regeneracji przechodzi po cichu (stary plik
    # wciąż się importuje). Porównanie po wartościach komórek, bo bajtowe
    # by nie zadziałało (xlsx pakuje timestampy w zip).
    import openpyxl

    from import_pracownikow.management.commands.generuj_plik_wzorcowy import (
        zbuduj_workbook,
    )

    gen = zbuduj_workbook()
    disk = openpyxl.load_workbook(SCIEZKA_DOMYSLNA)
    assert gen.sheetnames == disk.sheetnames
    for sn in gen.sheetnames:
        g = [[c.value for c in row] for row in gen[sn].iter_rows()]
        d = [[c.value for c in row] for row in disk[sn].iter_rows()]
        assert g == d, f"plik na dysku nieaktualny wzgledem generatora: {sn}"
        # Komentarze komórek to treść dla użytkownika (podpowiedź „TAK/NIE"
        # itp.), przeniesiona ze specjalnie z nagłówków — porównujemy .text,
        # bo sama równość wartości by tego rozjazdu nie złapała.
        gk = [[_tekst_komentarza(c) for c in row] for row in gen[sn].iter_rows()]
        dk = [[_tekst_komentarza(c) for c in row] for row in disk[sn].iter_rows()]
        assert gk == dk, f"komentarze na dysku nieaktualne wzgledem generatora: {sn}"
